import asyncio
import logging
import random
import os
import time
import config  # Assuming this file contains BOT_TOKEN
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode, ChatType
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    FSInputFile,
    BufferedInputFile,
    InputMediaPhoto,
    Chat,
    User,
)
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramBadRequest

# Pillow for image generation
try:
    from PIL import Image, ImageDraw, ImageFont, UnidentifiedImageError
except ImportError:
    Image = None
    ImageDraw = None
    ImageFont = None
    UnidentifiedImageError = None
    print(
        "Pillow library is not installed. Image generation for Stroop Test"
        " Parts 2 & 3, Mental Rotation Test, and Raven Matrices Test (dummy images) will not work."
    )
    print("Please install it using: pip install Pillow")

from io import BytesIO

# --- Globals & Constants ---
bot = Bot(
    config.BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher(storage=MemoryStorage())

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

EXCEL_FILENAME = "persistent_user_data.xlsx"

BASE_HEADERS = ["Telegram ID", "Unique ID", "Name", "Age"]
CORSI_HEADERS = [
    "Corsi - Max Correct Sequence Length",
    "Corsi - Avg Time Per Element (s)",
    "Corsi - Sequence Times Detail",
    "Corsi - Interrupted",
]
STROOP_HEADERS = [
    "Stroop Part1 Time (s)",
    "Stroop Part1 Errors",
    "Stroop Part2 Time (s)",
    "Stroop Part2 Errors",
    "Stroop Part3 Time (s)",
    "Stroop Part3 Errors",
    "Stroop - Interrupted",
]
REACTION_TIME_HEADERS = [
    "ReactionTime_Time_ms",
    "ReactionTime_Attempts",
    "ReactionTime_Status",
    "ReactionTime_Interrupted",
]
VERBAL_FLUENCY_HEADERS = [
    "VerbalFluency_Category",
    "VerbalFluency_Letter",
    "VerbalFluency_WordCount",
    "VerbalFluency_WordsList",
    "VerbalFluency_Interrupted",
]
MENTAL_ROTATION_HEADERS = [
    "MentalRotation_CorrectAnswers",
    "MentalRotation_AverageReactionTime_s",
    "MentalRotation_TotalTime_s",
    "MentalRotation_IndividualResponses",
    "MentalRotation_Interrupted",
]
# START OF RAVEN MATRICES TEST ADDITION (Headers)
RAVEN_MATRICES_HEADERS = [
    "RavenMatrices_CorrectAnswers",
    "RavenMatrices_TotalTime_s",
    "RavenMatrices_AvgTimeCorrect_s",
    "RavenMatrices_IndividualTimes_s",
    "RavenMatrices_Interrupted",
]
# END OF RAVEN MATRICES TEST ADDITION (Headers)

ALL_EXPECTED_HEADERS = (
    BASE_HEADERS
    + CORSI_HEADERS
    + STROOP_HEADERS
    + REACTION_TIME_HEADERS
    + VERBAL_FLUENCY_HEADERS
    + MENTAL_ROTATION_HEADERS
    + RAVEN_MATRICES_HEADERS  # Added Raven headers
)

IKB = InlineKeyboardButton

ACTION_SELECTION_KEYBOARD_NEW = InlineKeyboardMarkup(
    inline_keyboard=[
        [IKB(text="Пройти батарею тестов", callback_data="run_test_battery")],
        [
            IKB(
                text="Выбрать отдельный тест",
                callback_data="select_specific_test",
            )
        ],
    ]
)

ACTION_SELECTION_KEYBOARD_RETURNING = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            IKB(
                text="Пройти батарею тестов заново",
                callback_data="run_test_battery",
            )
        ],
        [
            IKB(
                text="Выбрать отдельный тест заново",
                callback_data="select_specific_test",
            )
        ],
        [IKB(text="Выйти (сбросить профиль)", callback_data="logout_profile")],
    ]
)

# Stroop Test Constants
STROOP_COLORS_DEF = {
    "Красный": {
        "rgb": (220, 20, 60),
        "name": "Красный",
        "emoji": "🟥",
    },
    "Синий": {
        "rgb": (0, 0, 205),
        "name": "Синий",
        "emoji": "🟦",
    },
    "Зеленый": {
        "rgb": (34, 139, 34),
        "name": "Зеленый",
        "emoji": "🟩",
    },
    "Желтый": {
        "rgb": (255, 215, 0),
        "name": "Желтый",
        "emoji": "🟨",
    },
    "Черный": {"rgb": (0, 0, 0), "name": "Черный", "emoji": "⬛"},
}
STROOP_COLOR_NAMES = list(STROOP_COLORS_DEF.keys())
STROOP_ITERATIONS_PER_PART = 6
STROOP_FONT_PATH = "arial.ttf"  # Make sure this font is available
STROOP_IMAGE_SIZE = (300, 150)
STROOP_TEXT_COLOR_ON_PATCH = (255, 255, 255)

# Reaction Time Test Constants
REACTION_TIME_IMAGE_POOL = [f"images/rt_img_{i}.png" for i in range(1, 11)]
REACTION_TIME_MEMORIZATION_S = 10
REACTION_TIME_STIMULUS_INTERVAL_S = 6
REACTION_TIME_MAX_ATTEMPTS = 2
REACTION_TIME_NUM_STIMULI_IN_SEQUENCE = 7
REACTION_TIME_TARGET_REACTION_WINDOW_S = REACTION_TIME_STIMULUS_INTERVAL_S - 1

# Verbal Fluency Test Constants
VERBAL_FLUENCY_DURATION_S = 60
VERBAL_FLUENCY_CATEGORY = "Общие слова"
_USABLE_RUSSIAN_LETTERS_VF = "АБВГДЕЁЖЗИКЛМНОПРСТУФХЦЧШЩЭЮЯ"
VERBAL_FLUENCY_TASK_POOL = []
if not _USABLE_RUSSIAN_LETTERS_VF:
    logger.error(
        "_USABLE_RUSSIAN_LETTERS_VF is empty. Verbal Fluency Test cannot"
        " function."
    )
else:
    for letter_vf in _USABLE_RUSSIAN_LETTERS_VF:
        VERBAL_FLUENCY_TASK_POOL.append(
            {
                "base_category": VERBAL_FLUENCY_CATEGORY,
                "letter": letter_vf.upper(),
            }
        )
if not VERBAL_FLUENCY_TASK_POOL:
    logger.error(
        "VERBAL_FLUENCY_TASK_POOL is empty. Verbal Fluency Test will not be"
        " available."
    )

# Mental Rotation Test Constants
MENTAL_ROTATION_NUM_ITERATIONS = 5
MR_BASE_DIR = os.path.join("images", "mental_rotation")
MR_REFERENCES_DIR = os.path.join(MR_BASE_DIR, "references")
MR_CORRECT_PROJECTIONS_DIR = os.path.join(MR_BASE_DIR, "correct_projections")
MR_DISTRACTORS_DIR = os.path.join(MR_BASE_DIR, "distractors")
MR_REFERENCE_FILES = [
    f"{i}.jpg" for i in range(1, 21)
]  # Example, adjust as needed
MR_CORRECT_PROJECTIONS_MAP = {
    f"{i}.jpg": [f"{i}_R.jpg"] for i in range(1, 21)
}  # Example
MR_CORRECT_PROJECTIONS_MAP["11.jpg"] = [
    "11_R.jpg",
    "11_R (2).jpg",
]  # Example for multiple correct
MR_ALL_DISTRACTORS_FILES = []  # To be populated at startup
MR_COLLAGE_CELL_SIZE = (250, 250)
MR_COLLAGE_BG_COLOR = (255, 255, 255)
MR_FEEDBACK_DISPLAY_TIME_S = 0.75

# START OF RAVEN MATRICES TEST ADDITION (Constants)
RAVEN_NUM_TASKS_TO_PRESENT = 20
RAVEN_TOTAL_AVAILABLE_TASKS_IDEAL = (
    80  # For information, actual count from dir
)
RAVEN_BASE_DIR = os.path.join("images", "raven_matrices")
RAVEN_FEEDBACK_DISPLAY_TIME_S = 0.75  # As per requirement
RAVEN_ALL_TASK_FILES = []  # To be populated at startup


# END OF RAVEN MATRICES TEST ADDITION (Constants)


# --- FSM States ---
class UserData(StatesGroup):
    waiting_for_first_time_response = State()
    waiting_for_name = State()
    waiting_for_age = State()
    waiting_for_unique_id = State()
    waiting_for_test_overwrite_confirmation = State()


class CorsiTestStates(StatesGroup):
    showing_sequence = State()
    waiting_for_user_sequence = State()


class StroopTestStates(StatesGroup):
    initial_instructions = State()
    part1_stimulus_response = State()
    part2_instructions = State()
    part2_stimulus_response = State()
    part3_instructions = State()
    part3_stimulus_response = State()


class ReactionTimeTestStates(StatesGroup):
    initial_instructions = State()
    memorization_display = State()
    reaction_stimulus_display = State()
    awaiting_retry_confirmation = State()


class VerbalFluencyStates(StatesGroup):
    showing_instructions_and_task = State()
    collecting_words = State()


class MentalRotationStates(StatesGroup):
    initial_instructions_mr = State()
    displaying_stimulus_mr = State()
    processing_answer_mr = State()
    inter_iteration_countdown_mr = State()


# START OF RAVEN MATRICES TEST ADDITION (FSM States)
class RavenMatricesStates(StatesGroup):
    initial_instructions_raven = State()
    displaying_task_raven = State()
    processing_feedback_raven = State()


# END OF RAVEN MATRICES TEST ADDITION (FSM States)


# --- Helper Functions ---
def initialize_excel_file():
    if not os.path.exists(EXCEL_FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.append(ALL_EXPECTED_HEADERS)
        wb.save(EXCEL_FILENAME)
        logger.info(f"'{EXCEL_FILENAME}' created with all headers.")
    else:
        try:
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active
            current_headers = []
            if ws.max_row > 0:
                current_headers = [cell.value for cell in ws[1]]

            if not current_headers:
                if ws.max_row > 0:
                    ws.delete_rows(1, ws.max_row)
                ws.append(ALL_EXPECTED_HEADERS)
                logger.info(
                    "Appended all headers to empty/cleared sheet in"
                    f" '{EXCEL_FILENAME}'."
                )
            else:
                new_headers_to_add = [
                    h for h in ALL_EXPECTED_HEADERS if h not in current_headers
                ]
                if new_headers_to_add:
                    header_col_start_index = len(current_headers) + 1
                    for i, header in enumerate(new_headers_to_add):
                        ws.cell(
                            row=1, column=header_col_start_index + i
                        ).value = header
                    logger.info(
                        f"Added missing headers to '{EXCEL_FILENAME}':"
                        f" {new_headers_to_add}"
                    )
            wb.save(EXCEL_FILENAME)
            logger.info(f"'{EXCEL_FILENAME}' checked/updated for headers.")
        except (InvalidFileException, Exception) as e:
            logger.error(
                f"Error initializing/updating Excel file '{EXCEL_FILENAME}':"
                f" {e}. Manual check might be needed."
            )


async def get_active_profile_from_fsm(state: FSMContext) -> dict | None:
    data = await state.get_data()
    if data.get("active_unique_id"):
        return {
            "unique_id": data.get("active_unique_id"),
            "name": data.get("active_name"),
            "age": data.get("active_age"),
            "telegram_id": data.get("active_telegram_id"),
        }
    return None


async def send_main_action_menu(
    trigger_event_or_message: [Message, CallbackQuery],
    keyboard_markup: InlineKeyboardMarkup,
    text: str = "Выберите дальнейшее действие:",
    state: FSMContext = None,
):
    chat_id = None
    if isinstance(trigger_event_or_message, Message):
        chat_id = trigger_event_or_message.chat.id
    elif isinstance(trigger_event_or_message, CallbackQuery):
        chat_id = trigger_event_or_message.message.chat.id
        try:
            await trigger_event_or_message.message.edit_reply_markup(
                reply_markup=None
            )
        except TelegramBadRequest:
            pass

    if chat_id:
        try:
            await bot.send_message(chat_id, text, reply_markup=keyboard_markup)
        except Exception as e:
            logger.error(
                f"Error in send_main_action_menu for chat {chat_id}: {e}"
            )


# --- Menu Command Handler ---
@dp.message(Command("menu"))
async def menu_command_handler(message: Message, state: FSMContext):
    current_fsm_state = await state.get_state()
    is_in_test = False

    if current_fsm_state:
        for test_cfg in TEST_REGISTRY.values():
            if test_cfg.get(
                "fsm_group_class"
            ) and current_fsm_state.startswith(
                test_cfg["fsm_group_class"].__name__
            ):
                is_in_test = True
                break

    if is_in_test:
        await message.answer(
            "Чтобы получить доступ к меню, пожалуйста, завершите или"
            " остановите текущий тест командой /stoptest."
        )
    else:
        profile = await get_active_profile_from_fsm(state)
        keyboard_to_show = (
            ACTION_SELECTION_KEYBOARD_RETURNING
            if profile
            else ACTION_SELECTION_KEYBOARD_NEW
        )
        menu_text = "Главное меню. Выберите действие:"
        await send_main_action_menu(
            message, keyboard_to_show, text=menu_text, state=state
        )


# --- Corsi Test Specific Logic ---
async def cleanup_corsi_messages(
    state: FSMContext, bot_instance: Bot, final_text: str = None
):
    data = await state.get_data()
    chat_id = data.get("corsi_chat_id")
    if not chat_id:
        return
    msg_ids_to_delete = [
        data.get(key)
        for key in [
            "corsi_status_message_id",
            "corsi_feedback_message_id",
        ]
        if data.get(key)
    ]
    for msg_id in msg_ids_to_delete:
        try:
            await bot_instance.delete_message(
                chat_id=chat_id, message_id=msg_id
            )
        except TelegramBadRequest:
            pass
    grid_message_id = data.get("corsi_grid_message_id")
    if grid_message_id:
        try:
            text_to_set = (
                final_text
                if final_text
                else "Тест Корси завершен или отменен."
            )
            await bot_instance.edit_message_text(
                text=text_to_set,
                chat_id=chat_id,
                message_id=grid_message_id,
                reply_markup=None,
            )
        except TelegramBadRequest:
            logger.warning(
                "Corsi cleanup: Failed to edit grid_message_id"
                f" {grid_message_id}"
            )

    current_fsm_data = await state.get_data()
    new_data = {
        k: v for k, v in current_fsm_data.items() if not k.startswith("corsi_")
    }
    for pk in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if pk in current_fsm_data:
            new_data[pk] = current_fsm_data[pk]

    await state.set_data(new_data)


async def show_corsi_sequence(trigger_message: Message, state: FSMContext):
    data = await state.get_data()
    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        return
    current_sequence_length = data.get("current_sequence_length", 2)
    corsi_chat_id = data.get("corsi_chat_id")
    if not corsi_chat_id:
        await state.clear()
        await trigger_message.answer(
            "Ошибка с тестом Корси. Пожалуйста, нажмите /start."
        )
        return

    grid_msg_id = data.get("corsi_grid_message_id")
    restart_btn_row = [
        IKB(
            text="🔄 Остановить Тест Корси",
            callback_data="corsi_stop_this_attempt",
        )
    ]
    indices = list(range(9))
    random.shuffle(indices)
    correct_seq = indices[:current_sequence_length]
    await state.update_data(
        correct_sequence=correct_seq, user_input_sequence=[]
    )
    base_btns = [
        IKB(text="🟪", callback_data=f"corsi_button_{i}") for i in range(9)
    ]
    base_kbd_rows = [base_btns[i : i + 3] for i in range(0, 9, 3)]
    base_kbd_rows.append(restart_btn_row)
    base_markup = InlineKeyboardMarkup(inline_keyboard=base_kbd_rows)

    if grid_msg_id:
        try:
            await bot.edit_message_text(
                chat_id=corsi_chat_id,
                message_id=grid_msg_id,
                text="Тест Корси: Запоминание",
                reply_markup=base_markup,
            )
        except TelegramBadRequest:
            grid_msg_id = None
    if not grid_msg_id:
        grid_msg_id = (
            await bot.send_message(
                corsi_chat_id,
                "Тест Корси: Запоминание",
                reply_markup=base_markup,
            )
        ).message_id
    await state.update_data(corsi_grid_message_id=grid_msg_id)

    status_msg_id = data.get("corsi_status_message_id")
    status_q = (
        ["Приготовьтесь..."]
        + [f"{i}..." for i in range(3, 0, -1)]
        + ["Запоминайте..."]
    )
    for i, text in enumerate(status_q):
        if await state.get_state() != CorsiTestStates.showing_sequence.state:
            return
        if not status_msg_id:
            status_msg_id = (
                await bot.send_message(corsi_chat_id, text)
            ).message_id
            await state.update_data(corsi_status_message_id=status_msg_id)
        else:
            try:
                await bot.edit_message_text(
                    text=text,
                    chat_id=corsi_chat_id,
                    message_id=status_msg_id,
                )
            except TelegramBadRequest:
                status_msg_id = (
                    await bot.send_message(corsi_chat_id, text)
                ).message_id
                await state.update_data(corsi_status_message_id=status_msg_id)

        await asyncio.sleep(1 if i < len(status_q) - 1 else 0.5)

    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        return

    for btn_idx in correct_seq:
        if await state.get_state() != CorsiTestStates.showing_sequence.state:
            return
        flashed_rows = [
            [
                IKB(
                    text="🟨" if r * 3 + c == btn_idx else "🟪",
                    callback_data=f"corsi_button_{r * 3 + c}",
                )
                for c in range(3)
            ]
            for r in range(3)
        ]
        flashed_rows.append(restart_btn_row)
        flashed_markup = InlineKeyboardMarkup(inline_keyboard=flashed_rows)
        try:
            await bot.edit_message_reply_markup(
                chat_id=corsi_chat_id,
                message_id=grid_msg_id,
                reply_markup=flashed_markup,
            )
            await asyncio.sleep(0.5)
            await bot.edit_message_reply_markup(
                chat_id=corsi_chat_id,
                message_id=grid_msg_id,
                reply_markup=base_markup,
            )
            await asyncio.sleep(0.2)
        except TelegramBadRequest:
            logger.warning(
                "Corsi: Failed to flash button, test might be disrupted."
            )
            return

    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        return

    status_msg_id = (await state.get_data()).get("corsi_status_message_id")
    try:
        if status_msg_id:
            await bot.edit_message_text(
                text="Повторите последовательность:",
                chat_id=corsi_chat_id,
                message_id=status_msg_id,
            )
        else:
            status_msg_id = (
                await bot.send_message(
                    corsi_chat_id, "Повторите последовательность:"
                )
            ).message_id
            await state.update_data(corsi_status_message_id=status_msg_id)
    except TelegramBadRequest:
        logger.warning(
            "Corsi: Failed to update status message for user input phase."
        )

    await state.update_data(sequence_start_time=time.time())
    await state.set_state(CorsiTestStates.waiting_for_user_sequence)


async def handle_corsi_button_press(
    callback: CallbackQuery, state: FSMContext
):
    if (
        await state.get_state()
        != CorsiTestStates.waiting_for_user_sequence.state
    ):
        await callback.answer(
            "Тест был прерван или уже не активен.", show_alert=True
        )
        return

    await callback.answer()
    btn_idx = int(callback.data.split("_")[-1])
    data = await state.get_data()
    user_seq = data.get("user_input_sequence", []) + [btn_idx]
    grid_msg_id = data.get("corsi_grid_message_id")
    chat_id = data.get("corsi_chat_id")

    if not grid_msg_id or not chat_id:
        await callback.message.answer(
            "Произошла ошибка с тестом. Пожалуйста, /start."
        )
        await state.clear()
        return

    new_rows = []
    for r in range(3):
        row = []
        for c in range(3):
            current_button_index = r * 3 + c
            text = "🟨" if current_button_index in user_seq else "🟪"
            row.append(
                IKB(
                    text=text,
                    callback_data=f"corsi_button_{current_button_index}",
                )
            )
        new_rows.append(row)
    new_rows.append(
        [
            IKB(
                text="🔄 Остановить Тест Корси",
                callback_data="corsi_stop_this_attempt",
            )
        ]
    )

    try:
        await bot.edit_message_reply_markup(
            chat_id=chat_id,
            message_id=grid_msg_id,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=new_rows),
        )
    except TelegramBadRequest:
        logger.warning(
            "Corsi: Failed to update button visual feedback for user press."
        )

    await state.update_data(user_input_sequence=user_seq)

    if len(user_seq) == len(data.get("correct_sequence", [])):
        await evaluate_user_sequence(callback.message, state)


async def on_corsi_restart_current_test(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer(text="Тест Корси будет прерван.", show_alert=False)
    await stop_test_command_handler(
        callback.message, state, called_from_test_button=True
    )


async def evaluate_user_sequence(message_context: Message, state: FSMContext):
    current_fsm_state_on_entry = await state.get_state()
    if (
        current_fsm_state_on_entry
        != CorsiTestStates.waiting_for_user_sequence.state
    ):
        return

    data = await state.get_data()
    chat_id = data.get("corsi_chat_id", message_context.chat.id)
    user_seq = data.get("user_input_sequence", [])
    correct_seq = data.get("correct_sequence", [])
    current_sequence_length_attempted = data.get("current_sequence_length", 2)
    error_count_at_this_length = data.get("error_count", 0)
    sequence_times_history = data.get("sequence_times", [])
    sequence_start_time = data.get("sequence_start_time", 0)
    feedback_message_id = data.get("corsi_feedback_message_id")

    time_taken_for_sequence = (
        time.time() - sequence_start_time if sequence_start_time else 0
    )

    async def _delayed_edit_feedback_message(
        msg_id_to_edit: int, new_text_content: str
    ):
        await asyncio.sleep(0.75)
        try:
            await bot.edit_message_text(
                text=new_text_content,
                chat_id=chat_id,
                message_id=msg_id_to_edit,
                parse_mode=None,
            )
        except TelegramBadRequest:
            logger.warning(
                f"Corsi feedback: Delayed edit for msg {msg_id_to_edit} failed (likely deleted or test ended)."
            )
        except Exception as e:
            logger.error(
                f"Corsi feedback: Unexpected error in _delayed_edit_feedback_message for msg {msg_id_to_edit}: {e}"
            )

    next_sequence_length = current_sequence_length_attempted
    current_error_count = error_count_at_this_length
    test_should_continue = True

    if user_seq == correct_seq:
        sequence_times_history.append(
            {
                "len": current_sequence_length_attempted,
                "time": time_taken_for_sequence,
            }
        )
        next_sequence_length = current_sequence_length_attempted + 1
        current_error_count = 0
        immediate_bold_text = "<b>Верно!</b>"
        delayed_normal_text = (
            "Верно! Достигнута максимальная длина. Тест завершен."
            if next_sequence_length > 9
            else "Верно!"
        )
        test_should_continue = next_sequence_length <= 9
    else:
        current_error_count = error_count_at_this_length + 1
        immediate_bold_text = "<b>Ошибка!</b>"
        delayed_normal_text = (
            "Ошибка! Слишком много ошибок. Тест завершен."
            if current_error_count >= 2
            else "Ошибка! Попробуйте эту же длину еще раз."
        )
        test_should_continue = current_error_count < 2

    if feedback_message_id:
        try:
            await bot.edit_message_text(
                immediate_bold_text,
                chat_id=chat_id,
                message_id=feedback_message_id,
                parse_mode=ParseMode.HTML,
            )
        except TelegramBadRequest:
            feedback_message_id = None
    if not feedback_message_id:
        msg = await bot.send_message(
            chat_id, immediate_bold_text, parse_mode=ParseMode.HTML
        )
        feedback_message_id = msg.message_id
        await state.update_data(corsi_feedback_message_id=feedback_message_id)

    if feedback_message_id:
        await asyncio.create_task(
            _delayed_edit_feedback_message(
                feedback_message_id, delayed_normal_text
            )
        )

    await state.update_data(
        current_sequence_length=next_sequence_length,
        error_count=current_error_count,
        sequence_times=sequence_times_history,
        user_input_sequence=[],
    )

    current_fsm_state_after_logic = await state.get_state()
    if (
        current_fsm_state_after_logic
        != CorsiTestStates.waiting_for_user_sequence.state
    ):
        return

    if test_should_continue:
        await state.set_state(CorsiTestStates.showing_sequence)
        await show_corsi_sequence(message_context, state)
    else:
        await save_corsi_results(message_context, state, is_interrupted=False)
        await cleanup_corsi_messages(
            state, bot, final_text="Тест Корси завершен."
        )
        fsm_data_after_test_end = await state.get_data()
        profile_keys = [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        profile_data_to_keep = {
            k: fsm_data_after_test_end.get(k)
            for k in profile_keys
            if fsm_data_after_test_end.get(k)
        }
        await state.set_state(None)
        if profile_data_to_keep.get("active_unique_id"):
            await state.set_data(profile_data_to_keep)
            await send_main_action_menu(
                message_context,
                ACTION_SELECTION_KEYBOARD_RETURNING,
                state=state,
            )
        else:
            await message_context.answer(
                "Тест завершен, но ваш профиль не найден. Пожалуйста, /start."
            )
            await state.clear()


async def start_corsi_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    await state.set_state(CorsiTestStates.showing_sequence)
    await state.update_data(
        unique_id_for_test=profile.get("unique_id"),
        profile_name_for_test=profile.get("name"),
        profile_age_for_test=profile.get("age"),
        profile_telegram_id_for_test=profile.get("telegram_id"),
        current_sequence_length=2,
        error_count=0,
        sequence_times=[],
        correct_sequence=[],
        user_input_sequence=[],
        sequence_start_time=0,
        corsi_grid_message_id=None,
        corsi_status_message_id=None,
        corsi_chat_id=msg_ctx.chat.id,
        corsi_feedback_message_id=None,
    )
    await show_corsi_sequence(msg_ctx, state)


async def save_corsi_results(
    trigger_msg: Message, state: FSMContext, is_interrupted: bool = False
):
    data = await state.get_data()
    uid = data.get("unique_id_for_test")
    p_tgid = None
    p_name = None
    p_age = None

    if not uid:
        logger.warning("Corsi save: unique_id_for_test not found in FSM data.")
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile:
            uid = active_profile.get("unique_id")
            p_tgid = active_profile.get("telegram_id")
            p_name = active_profile.get("name")
            p_age = active_profile.get("age")
            logger.info(
                f"Corsi save: Using active_profile UID {uid} as fallback."
            )
        else:
            if await state.get_state() is not None and hasattr(
                trigger_msg, "chat"
            ):
                await trigger_msg.answer(
                    "Тест Корси: ошибка сохранения (ID пользователя для теста не"
                    " найден)."
                )
            return
    else:
        p_tgid = data.get("profile_telegram_id_for_test")
        p_name = data.get("profile_name_for_test")
        p_age = data.get("profile_age_for_test")

    seq_times = data.get("sequence_times", [])
    max_len = max(i["len"] for i in seq_times) if seq_times else 0
    avg_t = 0.0
    valid_seqs = [i for i in seq_times if i["len"] > 0 and "time" in i]
    if valid_seqs:
        try:
            avg_t = sum(i["time"] / i["len"] for i in valid_seqs) / len(
                valid_seqs
            )
        except ZeroDivisionError:
            avg_t = 0.0

    detail_str = "; ".join(
        [
            f"Дл. {i['len']} – {i['time']:.2f} сек"
            for i in seq_times
            if "time" in i
        ]
    )
    intr_status = "Да" if is_interrupted else "Нет"

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        for idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = idx
                break

        if row_num == -1:
            logger.warning(
                f"Corsi save: UID {uid} not found in Excel. Appending new"
                " row."
            )
            new_row_data = [""] * len(ALL_EXPECTED_HEADERS)
            new_row_data[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row_data[uid_col_idx] = uid
            new_row_data[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row_data[ALL_EXPECTED_HEADERS.index("Age")] = (
                p_age if p_age else ""
            )
            ws.append(new_row_data)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num,
            column=h.index("Corsi - Max Correct Sequence Length") + 1,
        ).value = max_len
        ws.cell(
            row=row_num,
            column=h.index("Corsi - Avg Time Per Element (s)") + 1,
        ).value = round(avg_t, 2)
        ws.cell(
            row=row_num, column=h.index("Corsi - Sequence Times Detail") + 1
        ).value = detail_str
        ws.cell(
            row=row_num, column=h.index("Corsi - Interrupted") + 1
        ).value = intr_status
        wb.save(EXCEL_FILENAME)

        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            status_text = (
                "ПРЕРВАНЫ И СОХРАНЕНЫ"
                if is_interrupted
                else "УСПЕШНО СОХРАНЕНЫ"
            )
            if is_interrupted and max_len == 0 and not seq_times:
                summary = (
                    "Тест Корси был <b>ПРЕРВАН</b> досрочно. Результаты не"
                    " зафиксированы."
                )
            else:
                summary_lines = [
                    f"Результаты Теста Корси <b>{status_text}</b> для UID"
                    f" {uid}:",
                    f"- Максимальная верная последовательность: {max_len}",
                    f"- Среднее время на элемент: {round(avg_t, 2)} сек",
                    "- Детализация серий:"
                    f" {detail_str if detail_str else 'N/A'}",
                ]
                summary = "\n".join(summary_lines)
            await trigger_msg.answer(summary, parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(
            f"Corsi results save error for UID {uid}: {e}", exc_info=True
        )
        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            await trigger_msg.answer(
                "Произошла ошибка при сохранении результатов Теста Корси."
            )


# --- Stroop Test Logic ---
def _get_stroop_font(size: int = 40):
    if not ImageFont:
        return None
    try:
        return ImageFont.truetype(STROOP_FONT_PATH, size)
    except IOError:
        logger.warning(
            f"Font {STROOP_FONT_PATH} not found. Using Pillow default."
        )
        try:
            return ImageFont.load_default(size=size)
        except AttributeError:  # Older Pillow versions
            return ImageFont.load_default()
        except Exception as e_def:
            logger.error(f"Pillow default font error: {e_def}")
            return None


def _generate_stroop_part2_image(
    patch_color_name: str, text_on_patch_name: str
) -> BufferedInputFile | None:
    if not Image or not ImageDraw or not ImageFont:
        return None
    patch_rgb = STROOP_COLORS_DEF[patch_color_name]["rgb"]
    text_rgb = STROOP_TEXT_COLOR_ON_PATCH
    img = Image.new("RGB", STROOP_IMAGE_SIZE, color=patch_rgb)
    draw = ImageDraw.Draw(img)
    font = _get_stroop_font(40)
    text_to_draw = STROOP_COLORS_DEF[text_on_patch_name]["name"]
    if font:
        try:
            bbox = draw.textbbox((0, 0), text_to_draw, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
        except Exception as e_bbox:
            logger.warning(
                f"Stroop P2: textbbox failed ({e_bbox}), using estimate."
            )
            tw, th = (
                STROOP_IMAGE_SIZE[0] * 0.8,
                STROOP_IMAGE_SIZE[1] * 0.5,
            )

        x = (STROOP_IMAGE_SIZE[0] - tw) / 2
        y = (STROOP_IMAGE_SIZE[1] - th) / 2
        draw.text((x, y), text_to_draw, fill=text_rgb, font=font)
    else:
        draw.text((10, 10), "Font Error", fill=text_rgb)
    bio = BytesIO()
    bio.name = "s_p2.png"
    img.save(bio, "PNG")
    bio.seek(0)
    return BufferedInputFile(bio.read(), filename=bio.name)


def _generate_stroop_part3_image(
    word_name: str, ink_name: str
) -> BufferedInputFile | None:
    if not Image or not ImageDraw or not ImageFont:
        return None
    ink_rgb = STROOP_COLORS_DEF[ink_name]["rgb"]
    bg_rgb = (255, 255, 255)
    img = Image.new("RGB", STROOP_IMAGE_SIZE, color=bg_rgb)
    draw = ImageDraw.Draw(img)
    font = _get_stroop_font(50)
    text_to_draw = STROOP_COLORS_DEF[word_name]["name"]
    if font:
        try:
            bbox = draw.textbbox((0, 0), text_to_draw, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
        except Exception as e_bbox:
            logger.warning(
                f"Stroop P3: textbbox failed ({e_bbox}), using estimate."
            )
            tw, th = STROOP_IMAGE_SIZE[0] * 0.8, STROOP_IMAGE_SIZE[1] * 0.5

        x = (STROOP_IMAGE_SIZE[0] - tw) / 2
        y = (STROOP_IMAGE_SIZE[1] - th) / 2
        stroke_width = (
            1 if ink_name == "Желтый" else 0
        )  # Add stroke for light colors like yellow
        stroke_fill = (
            (128, 128, 128) if stroke_width > 0 else None
        )  # Grey stroke

        draw.text(
            (x, y),
            text_to_draw,
            fill=ink_rgb,
            font=font,
            stroke_width=stroke_width,
            stroke_fill=stroke_fill,
        )
    else:
        draw.text((10, 10), "Font Error", fill=ink_rgb)
    bio = BytesIO()
    bio.name = "s_p3.png"
    img.save(bio, "PNG")
    bio.seek(0)
    return BufferedInputFile(bio.read(), filename=bio.name)


STROOP_INSTRUCTION_TEXT_PART1 = (
    "Добро пожаловать в <b>Тест Струпа!</b>\n\n"
    "Этот тест оценивает вашу способность подавлять когнитивную"
    " интерференцию. Он состоит из трех частей.\n\n"
    "<b>Часть 1: Слова</b>\n"
    "Вам будут показаны названия цветов, написанные черным жирным шрифтом."
    " Ваша задача – как можно быстрее нажать на <b>цветной квадрат</b>"
    " (кнопку-эмодзи), соответствующий <b>написанному названию"
    " цвета</b>.\n\n"
    "Приготовьтесь. Нажмите 'Понятно', чтобы начать Часть 1."
)
STROOP_INSTRUCTION_TEXT_PART2 = (
    "<b>Часть 2: Цветные Плашки</b>\n"
    "Теперь вам будут показаны цветные прямоугольники. На каждом"
    " прямоугольнике белыми буквами будет написано случайное название цвета"
    " (оно не имеет значения для задачи).\n\n"
    "Ваша задача – как можно быстрее нажать на кнопку с <b>названием"
    " цвета</b>, соответствующим <b>цвету самого прямоугольника"
    " (фона)</b>.\n\n"
    "Приготовьтесь. Нажмите 'Понятно', чтобы начать Часть 2."
)
STROOP_INSTRUCTION_TEXT_PART3 = (
    "<b>Часть 3: Интерференция</b>\n"
    "В этой части вам снова будут показаны слова, обозначающие цвета."
    " Однако теперь сами слова будут написаны <b>цветными чернилами</b>,"
    " причем цвет чернил НЕ будет совпадать со значением слова.\n\n"
    "Ваша задача – как можно быстрее нажать на кнопку с <b>названием"
    " цвета</b>, соответствующим <b>цвету чернил</b>, которым написано"
    " слово (игнорируйте значение слова).\n\n"
    "Приготовьтесь. Нажмите 'Понятно', чтобы начать Часть 3."
)


async def _send_stroop_instruction_message(
    chat_id: int, part: int, state: FSMContext, bot_instance: Bot
):
    text, cb_data = "", ""
    if part == 1:
        text, cb_data = STROOP_INSTRUCTION_TEXT_PART1, "stroop_ack_part1"
    elif part == 2:
        text, cb_data = STROOP_INSTRUCTION_TEXT_PART2, "stroop_ack_part2"
    elif part == 3:
        text, cb_data = STROOP_INSTRUCTION_TEXT_PART3, "stroop_ack_part3"

    markup = InlineKeyboardMarkup(
        inline_keyboard=[[IKB(text="Понятно", callback_data=cb_data)]]
    )
    msg = await bot_instance.send_message(
        chat_id, text, reply_markup=markup, parse_mode=ParseMode.HTML
    )
    await state.update_data(stroop_instruction_message_id=msg.message_id)


async def _display_next_stroop_stimulus(
    chat_id: int, state: FSMContext, bot_instance: Bot
):
    data = await state.get_data()
    current_part = data.get("stroop_current_part")
    current_iteration = data.get("stroop_current_iteration")
    stimulus_msg_id = data.get("stroop_stimulus_message_id")
    current_stimulus_type = data.get("stroop_stimulus_type")  # text or photo

    stimulus_text_for_part1 = ""
    image_to_send = None  # BufferedInputFile
    correct_answer_color_name = ""
    all_colors = list(STROOP_COLOR_NAMES)
    new_stimulus_type = ""  # To track if message type needs to change

    if current_part == 1:
        stimulus_word_color = random.choice(all_colors)
        stimulus_text_for_part1 = stimulus_word_color
        correct_answer_color_name = stimulus_word_color
        new_stimulus_type = "text"
    elif current_part == 2:
        patch_color = random.choice(all_colors)
        text_on_patch_options = [c for c in all_colors if c != patch_color]
        text_on_patch_color = (
            random.choice(text_on_patch_options)
            if text_on_patch_options
            else patch_color  # Fallback if only one color (should not happen)
        )
        image_to_send = _generate_stroop_part2_image(
            patch_color, text_on_patch_color
        )
        correct_answer_color_name = patch_color
        new_stimulus_type = "photo"
    elif current_part == 3:
        word_text_color = random.choice(all_colors)
        possible_ink_colors = [c for c in all_colors if c != word_text_color]
        if not possible_ink_colors:  # Should not happen with 5 colors
            possible_ink_colors = [word_text_color]
        ink_color = random.choice(possible_ink_colors)
        image_to_send = _generate_stroop_part3_image(
            word_text_color, ink_color
        )
        correct_answer_color_name = ink_color
        new_stimulus_type = "photo"
    else:
        logger.error(
            f"Stroop: Invalid part {current_part} in _display_next_stroop_stimulus"
        )
        return

    # Handle Pillow not being available for parts 2 & 3
    if (new_stimulus_type == "photo") and (not Image or not image_to_send):
        await bot_instance.send_message(
            chat_id, "Ошибка генерации изображения. Тест Струпа прерван."
        )
        # Mock message for save/cleanup context if needed
        mock_chat = Chat(
            id=chat_id, type=ChatType.PRIVATE
        )  # Assuming private chat
        mock_msg = Message(
            message_id=0,
            date=int(time.time()),
            chat=mock_chat,
            from_user=bot.id,
        )  # Minimal mock
        await save_stroop_results(mock_msg, state, is_interrupted=True)
        await cleanup_stroop_ui(
            state, bot, "Тест Струпа прерван (ошибка изображения)."
        )
        await state.set_state(None)  # Clear test state
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile:  # Preserve profile data
            profile_keys_to_keep = {
                "active_unique_id": active_profile.get("unique_id"),
                "active_name": active_profile.get("name"),
                "active_age": active_profile.get("age"),
                "active_telegram_id": active_profile.get("telegram_id"),
            }
            await state.set_data(profile_keys_to_keep)
        await send_main_action_menu(
            mock_msg, ACTION_SELECTION_KEYBOARD_RETURNING, state=state
        )
        return

    await state.update_data(stroop_correct_answer=correct_answer_color_name)
    # Create buttons: 1 correct, 3 distractors (or fewer if not enough colors)
    button_options = list(all_colors)
    distractors = [c for c in button_options if c != correct_answer_color_name]
    random.shuffle(distractors)
    num_button_distractors = min(len(distractors), 3)
    chosen_buttons_names = [correct_answer_color_name] + distractors[
        :num_button_distractors
    ]
    random.shuffle(chosen_buttons_names)

    buttons_grid = []
    row = []
    for i, name in enumerate(chosen_buttons_names):
        btn_text = (
            STROOP_COLORS_DEF[name]["emoji"]
            if current_part == 1
            else STROOP_COLORS_DEF[name]["name"]
        )
        row.append(IKB(text=btn_text, callback_data=f"stroop_answer_{name}"))
        if (
            len(row) == 2 or i == len(chosen_buttons_names) - 1
        ):  # 2 buttons per row
            buttons_grid.append(row)
            row = []
    reply_markup = InlineKeyboardMarkup(inline_keyboard=buttons_grid)

    base_txt = (
        f"<b>Тест Струпа</b>\nЧасть {current_part}, Итерация"
        f" {current_iteration}/{STROOP_ITERATIONS_PER_PART}\n\n"
    )
    instruction = ""
    stimulus_display_text = ""  # Only for part 1 text
    if current_part == 1:
        instruction = "Нажмите на <b>цветной квадрат</b>, соответствующий <b>написанному названию</b>:"
        stimulus_display_text = f"<b>{stimulus_text_for_part1}</b>"
    elif current_part == 2:
        instruction = "Нажмите кнопку с названием цвета, соответствующим <b>цвету прямоугольника</b>:"
    elif current_part == 3:
        instruction = "Нажмите кнопку с названием цвета, соответствующим <b>цвету чернил слова</b>:"

    final_caption = f"{base_txt}{instruction}\n{stimulus_display_text if current_part == 1 else ''}".strip()

    try:
        if (
            stimulus_msg_id is None
            or current_stimulus_type != new_stimulus_type
        ):
            # Delete old message if type changes (text -> photo or photo -> text)
            if stimulus_msg_id and current_stimulus_type:
                try:
                    await bot_instance.delete_message(chat_id, stimulus_msg_id)
                except TelegramBadRequest:
                    pass
                stimulus_msg_id = None  # Force sending new

            if new_stimulus_type == "photo":
                msg = await bot_instance.send_photo(
                    chat_id,
                    photo=image_to_send,
                    caption=final_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
            else:  # text
                msg = await bot_instance.send_message(
                    chat_id,
                    final_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
            await state.update_data(
                stroop_stimulus_message_id=msg.message_id,
                stroop_stimulus_type=new_stimulus_type,
            )
        else:  # Edit existing message of the same type
            if new_stimulus_type == "photo":
                media = InputMediaPhoto(
                    media=image_to_send,
                    caption=final_caption,
                    parse_mode=ParseMode.HTML,
                )
                await bot_instance.edit_message_media(
                    media=media,
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    reply_markup=reply_markup,
                )
            else:  # text
                await bot_instance.edit_message_text(
                    final_caption,
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
    except TelegramBadRequest as e:
        logger.error(
            f"Stroop stimulus UI edit/send failed: {e}. Attempting to send new."
        )
        # Fallback: try to send as a new message if edit failed badly
        try:
            if new_stimulus_type == "photo":
                msg = await bot_instance.send_photo(
                    chat_id,
                    photo=image_to_send,
                    caption=final_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
            else:
                msg = await bot_instance.send_message(
                    chat_id,
                    final_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
            await state.update_data(
                stroop_stimulus_message_id=msg.message_id,
                stroop_stimulus_type=new_stimulus_type,
            )
        except Exception as e_fallback:
            logger.critical(
                f"Stroop stimulus UI critical failure on fallback send: {e_fallback}"
            )
            await bot_instance.send_message(
                chat_id,
                "Критическая ошибка отображения стимула. Тест Струпа прерван.",
            )
            mock_chat_fallback = Chat(id=chat_id, type=ChatType.PRIVATE)
            mock_msg_fallback = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat_fallback,
                from_user=bot.id,
            )
            await save_stroop_results(
                mock_msg_fallback, state, is_interrupted=True
            )
            await cleanup_stroop_ui(
                state,
                bot_instance,
                "Тест Струпа прерван (критическая ошибка UI).",
            )
            await state.set_state(None)
            active_profile_fallback = await get_active_profile_from_fsm(state)
            if active_profile_fallback:
                profile_data_to_keep_fb = {
                    "active_unique_id": active_profile_fallback.get(
                        "unique_id"
                    ),
                    "active_name": active_profile_fallback.get("name"),
                    "active_age": active_profile_fallback.get("age"),
                    "active_telegram_id": active_profile_fallback.get(
                        "telegram_id"
                    ),
                }
                await state.set_data(profile_data_to_keep_fb)
            await send_main_action_menu(
                mock_msg_fallback,
                ACTION_SELECTION_KEYBOARD_RETURNING,
                state=state,
            )
            return

    if current_part == 1:
        await state.set_state(StroopTestStates.part1_stimulus_response)
    elif current_part == 2:
        await state.set_state(StroopTestStates.part2_stimulus_response)
    elif current_part == 3:
        await state.set_state(StroopTestStates.part3_stimulus_response)


async def start_stroop_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    logger.info(f"Starting Stroop Test for UID: {profile.get('unique_id')}")
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    await state.set_state(StroopTestStates.initial_instructions)
    await state.update_data(
        unique_id_for_test=profile.get("unique_id"),
        profile_name_for_test=profile.get("name"),
        profile_age_for_test=profile.get("age"),
        profile_telegram_id_for_test=profile.get("telegram_id"),
        stroop_chat_id=msg_ctx.chat.id,
        stroop_instruction_message_id=None,
        stroop_stimulus_message_id=None,
        stroop_stimulus_type=None,  # 'text' or 'photo'
        stroop_current_part=0,  # Will be 1, 2, or 3
        stroop_current_iteration=0,  # 1 to STROOP_ITERATIONS_PER_PART
        stroop_part1_errors=0,
        stroop_part2_errors=0,
        stroop_part3_errors=0,
        stroop_part1_start_time=None,
        stroop_part2_start_time=None,
        stroop_part3_start_time=None,
        stroop_part1_total_time_s=None,
        stroop_part2_total_time_s=None,
        stroop_part3_total_time_s=None,
        stroop_correct_answer=None,  # Stores the name of the correct color
    )
    await _send_stroop_instruction_message(msg_ctx.chat.id, 1, state, bot)


@dp.callback_query(
    F.data == "stroop_ack_part1", StroopTestStates.initial_instructions
)
async def handle_stroop_ack_part1(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    instruction_msg_id = data.get("stroop_instruction_message_id")
    if instruction_msg_id:
        try:
            await bot.delete_message(
                callback.message.chat.id, instruction_msg_id
            )
        except TelegramBadRequest:
            pass  # Message might already be gone
    await state.update_data(
        stroop_instruction_message_id=None,  # Clear since it's deleted
        stroop_current_part=1,
        stroop_current_iteration=1,
        stroop_part1_start_time=time.time(),
    )
    await _display_next_stroop_stimulus(callback.message.chat.id, state, bot)


@dp.callback_query(
    F.data == "stroop_ack_part2", StroopTestStates.part2_instructions
)
async def handle_stroop_ack_part2(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    instruction_msg_id = data.get("stroop_instruction_message_id")
    if instruction_msg_id:
        try:
            await bot.delete_message(
                callback.message.chat.id, instruction_msg_id
            )
        except TelegramBadRequest:
            pass
    await state.update_data(
        stroop_instruction_message_id=None,
        stroop_current_part=2,
        stroop_current_iteration=1,
        stroop_part2_start_time=time.time(),
    )
    await _display_next_stroop_stimulus(callback.message.chat.id, state, bot)


@dp.callback_query(
    F.data == "stroop_ack_part3", StroopTestStates.part3_instructions
)
async def handle_stroop_ack_part3(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    instruction_msg_id = data.get("stroop_instruction_message_id")
    if instruction_msg_id:
        try:
            await bot.delete_message(
                callback.message.chat.id, instruction_msg_id
            )
        except TelegramBadRequest:
            pass
    await state.update_data(
        stroop_instruction_message_id=None,
        stroop_current_part=3,
        stroop_current_iteration=1,
        stroop_part3_start_time=time.time(),
    )
    await _display_next_stroop_stimulus(callback.message.chat.id, state, bot)


@dp.callback_query(
    F.data.startswith("stroop_answer_"),
    StateFilter(
        StroopTestStates.part1_stimulus_response,
        StroopTestStates.part2_stimulus_response,
        StroopTestStates.part3_stimulus_response,
    ),
)
async def handle_stroop_stimulus_response(
    callback: CallbackQuery, state: FSMContext
):
    data = await state.get_data()
    chosen_color_name = callback.data.split("stroop_answer_")[-1]
    correct_answer = data.get("stroop_correct_answer")
    current_part = data.get("stroop_current_part")
    current_iteration = data.get("stroop_current_iteration")
    feedback_text = ""

    if chosen_color_name == correct_answer:
        feedback_text = "Верно!"
    else:
        feedback_text = f"Ошибка! Правильный ответ: {correct_answer}"
        error_key = f"stroop_part{current_part}_errors"
        await state.update_data({error_key: data.get(error_key, 0) + 1})

    await callback.answer(
        text=feedback_text, show_alert=False
    )  # Quick feedback
    current_iteration += 1

    if current_iteration > STROOP_ITERATIONS_PER_PART:
        # Part finished, record time
        part_start_time_key = f"stroop_part{current_part}_start_time"
        part_start_time = data.get(part_start_time_key)
        if part_start_time:
            time_taken_s = round(time.time() - part_start_time, 2)
            await state.update_data(
                {f"stroop_part{current_part}_total_time_s": time_taken_s}
            )

        # Move to next part or finish test
        current_iteration = 1  # Reset for next part
        current_part += 1
        await state.update_data(
            stroop_current_part=current_part,
            stroop_current_iteration=current_iteration,
        )

        if current_part == 2:
            await state.set_state(StroopTestStates.part2_instructions)
            await _send_stroop_instruction_message(
                callback.message.chat.id, 2, state, bot
            )
        elif current_part == 3:
            await state.set_state(StroopTestStates.part3_instructions)
            await _send_stroop_instruction_message(
                callback.message.chat.id, 3, state, bot
            )
        else:  # Test finished (current_part will be > 3)
            await save_stroop_results(
                callback.message, state, is_interrupted=False
            )
            await cleanup_stroop_ui(
                state, bot, final_text="Тест Струпа успешно завершен!"
            )
            fsm_data_after_test = (
                await state.get_data()
            )  # Get data after cleanup might have cleared some
            profile_keys = [
                "active_unique_id",
                "active_name",
                "active_age",
                "active_telegram_id",
            ]
            profile_data_to_keep = {
                k: fsm_data_after_test.get(k)
                for k in profile_keys
                if fsm_data_after_test.get(k)
            }

            await state.set_state(None)  # Clear Stroop state
            if profile_data_to_keep.get("active_unique_id"):
                await state.set_data(profile_data_to_keep)  # Restore profile
                await send_main_action_menu(
                    callback.message,
                    ACTION_SELECTION_KEYBOARD_RETURNING,
                    state=state,
                )
            else:  # Should not happen if profile was required
                await callback.message.answer(
                    "Тест Струпа завершен. Профиль не найден, пожалуйста /start."
                )
                await state.clear()
    else:
        # Continue current part
        await state.update_data(stroop_current_iteration=current_iteration)
        await _display_next_stroop_stimulus(
            callback.message.chat.id, state, bot
        )


async def save_stroop_results(
    trigger_msg: Message, state: FSMContext, is_interrupted: bool = False
):
    logger.info(f"Saving Stroop results. Interrupted: {is_interrupted}")
    data = await state.get_data()
    uid = data.get("unique_id_for_test")
    p_tgid = data.get(
        "profile_telegram_id_for_test", data.get("active_telegram_id")
    )
    p_name = data.get("profile_name_for_test", data.get("active_name"))
    p_age = data.get("profile_age_for_test", data.get("active_age"))

    if not uid:
        logger.warning("Stroop save: unique_id_for_test not found.")
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile:
            uid = active_profile.get("unique_id")
            p_tgid = active_profile.get("telegram_id")
            p_name = active_profile.get("name")
            p_age = active_profile.get("age")
            logger.info(
                f"Stroop save: Using active_profile UID {uid} as fallback."
            )
        else:
            if await state.get_state() is not None and hasattr(
                trigger_msg, "chat"
            ):  # Only message if still in a Stroop state
                await trigger_msg.answer(
                    "Тест Струпа: ошибка сохранения (ID пользователя для теста не найден)."
                )
            return

    # If interrupted mid-part, calculate time for that part up to interruption
    current_part_on_interrupt = data.get("stroop_current_part", 0)
    if is_interrupted and 1 <= current_part_on_interrupt <= 3:
        part_start_time_key = (
            f"stroop_part{current_part_on_interrupt}_start_time"
        )
        part_total_time_key = (
            f"stroop_part{current_part_on_interrupt}_total_time_s"
        )
        if data.get(part_start_time_key) and not data.get(
            part_total_time_key
        ):  # Time started but not finished
            time_taken_s = round(
                time.time() - data.get(part_start_time_key), 2
            )
            await state.update_data({part_total_time_key: time_taken_s})
            data = await state.get_data()  # Refresh data after update

    p1t = data.get("stroop_part1_total_time_s")
    p1e = data.get("stroop_part1_errors")
    p2t = data.get("stroop_part2_total_time_s")
    p2e = data.get("stroop_part2_errors")
    p3t = data.get("stroop_part3_total_time_s")
    p3e = data.get("stroop_part3_errors")
    intr_status = "Да" if is_interrupted else "Нет"

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = r_idx
                break
        if row_num == -1:
            logger.warning(
                f"Stroop save: UID {uid} not in Excel. Appending new row."
            )
            new_row = [""] * len(ALL_EXPECTED_HEADERS)
            new_row[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row[uid_col_idx] = uid
            new_row[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row[ALL_EXPECTED_HEADERS.index("Age")] = p_age if p_age else ""
            ws.append(new_row)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num, column=h.index("Stroop Part1 Time (s)") + 1
        ).value = p1t
        ws.cell(
            row=row_num, column=h.index("Stroop Part1 Errors") + 1
        ).value = p1e
        ws.cell(
            row=row_num, column=h.index("Stroop Part2 Time (s)") + 1
        ).value = p2t
        ws.cell(
            row=row_num, column=h.index("Stroop Part2 Errors") + 1
        ).value = p2e
        ws.cell(
            row=row_num, column=h.index("Stroop Part3 Time (s)") + 1
        ).value = p3t
        ws.cell(
            row=row_num, column=h.index("Stroop Part3 Errors") + 1
        ).value = p3e
        ws.cell(
            row=row_num, column=h.index("Stroop - Interrupted") + 1
        ).value = intr_status
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Stroop results for UID {uid} saved. Interrupted: {is_interrupted}"
        )

        # Send summary to user only if test was active (not from silent save on bot stop)
        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            status_text_display = (
                "ПРЕРВАНЫ И СОХРАНЕНЫ"
                if is_interrupted
                else "УСПЕШНО СОХРАНЕНЫ"
            )
            res_summary_lines = [
                f"Часть 1: Время {p1t if p1t is not None else 'N/A'} сек, Ошибок: {p1e if p1e is not None else 'N/A'}",
                f"Часть 2: Время {p2t if p2t is not None else 'N/A'} сек, Ошибок: {p2e if p2e is not None else 'N/A'}",
                f"Часть 3: Время {p3t if p3t is not None else 'N/A'} сек, Ошибок: {p3e if p3e is not None else 'N/A'}",
            ]
            summary_txt_final = (
                f"Результаты Теста Струпа <b>{status_text_display}</b> для UID {uid}:\n"
                + "\n".join(res_summary_lines)
            )
            # If interrupted and no data recorded at all
            if is_interrupted and all(
                val is None for val in [p1t, p1e, p2t, p2e, p3t, p3e]
            ):
                summary_txt_final = f"Тест Струпа был <b>ПРЕРВАН</b> досрочно для UID {uid}. Данные не зафиксированы."
            await trigger_msg.answer(
                summary_txt_final, parse_mode=ParseMode.HTML
            )

    except Exception as e_save:
        logger.error(
            f"Stroop results save error for UID {uid}: {e_save}", exc_info=True
        )
        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            await trigger_msg.answer(
                "Произошла ошибка при сохранении результатов Теста Струпа."
            )


async def cleanup_stroop_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: str = "Тест Струпа завершен или отменен.",
):
    data = await state.get_data()
    chat_id = data.get("stroop_chat_id")
    instruction_msg_id = data.get("stroop_instruction_message_id")
    stimulus_msg_id = data.get("stroop_stimulus_message_id")

    if chat_id:
        if instruction_msg_id:
            try:
                await bot_instance.delete_message(chat_id, instruction_msg_id)
            except TelegramBadRequest:
                pass

        if stimulus_msg_id:
            try:
                stimulus_type = data.get("stroop_stimulus_type")
                if stimulus_type == "photo":
                    await bot_instance.edit_message_caption(
                        chat_id=chat_id,
                        message_id=stimulus_msg_id,
                        caption=final_text,
                        reply_markup=None,
                        parse_mode=ParseMode.HTML,
                    )
                else:  # text
                    await bot_instance.edit_message_text(
                        text=final_text,
                        chat_id=chat_id,
                        message_id=stimulus_msg_id,
                        reply_markup=None,
                        parse_mode=ParseMode.HTML,
                    )
            except TelegramBadRequest:  # If edit fails, try sending new
                try:
                    await bot_instance.send_message(
                        chat_id, final_text, parse_mode=ParseMode.HTML
                    )
                except Exception as e_final_send:
                    logger.error(
                        f"Stroop cleanup final send error: {e_final_send}"
                    )
        elif final_text:  # No stimulus message to edit, but final_text exists
            try:
                await bot_instance.send_message(
                    chat_id, final_text, parse_mode=ParseMode.HTML
                )
            except Exception as e_fresh_final_send:
                logger.error(
                    f"Stroop cleanup fresh final send error: {e_fresh_final_send}"
                )

    # Clean FSM data specific to Stroop test
    current_fsm_data = await state.get_data()
    new_data_after_stroop_clear = {
        k: v
        for k, v in current_fsm_data.items()
        if not k.startswith("stroop_")
    }
    # Ensure profile data is preserved if it was also in FSM
    for pk_generic in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if (
            pk_generic in current_fsm_data
            and pk_generic not in new_data_after_stroop_clear
        ):
            new_data_after_stroop_clear[pk_generic] = current_fsm_data[
                pk_generic
            ]
    await state.set_data(new_data_after_stroop_clear)
    logger.info(
        f"Cleaned Stroop FSM keys for chat {chat_id if chat_id else 'N/A'}"
    )


# --- Reaction Time Test Logic ---
async def start_reaction_time_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    logger.info(
        f"Starting Reaction Time Test for UID: {profile.get('unique_id')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    await state.set_state(ReactionTimeTestStates.initial_instructions)
    await state.update_data(
        rt_unique_id_for_test=profile.get("unique_id"),
        rt_profile_name_for_test=profile.get("name"),
        rt_profile_age_for_test=profile.get("age"),
        rt_profile_telegram_id_for_test=profile.get("telegram_id"),
        rt_chat_id=chat_id,
        rt_current_attempt=1,
        rt_reaction_time_ms=None,
        rt_status="Pending",  # "Passed", "Failed", "Interrupted"
        rt_instruction_message_id=None,
        rt_memorization_image_message_id=None,
        rt_reaction_stimulus_message_id=None,
        rt_memorization_task=None,  # asyncio.Task for memorization phase
        rt_reaction_cycle_task=None,  # asyncio.Task for reaction stimulus cycle
    )

    instruction_text = (
        "<b>Тест на Скорость Реакции</b>\n\n"
        "1. Сначала вам будет показано изображение-цель на 10 секунд. Запомните его.\n"
        "2. Затем изображение-цель исчезнет.\n"
        "3. После этого начнут появляться другие изображения. Среди них один раз появится ваше целевое изображение.\n"
        "4. Ваша задача – как можно быстрее нажать кнопку 'РЕАГИРОВАТЬ!', как только увидите целевое изображение.\n"
        "   Если вы нажмете на кнопку, когда показано НЕ целевое изображение, это будет считаться ошибкой.\n\n"
        f"У вас будет {REACTION_TIME_MAX_ATTEMPTS} попытки. "
        "Тест начнется после нажатия кнопки 'Начать'."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать Тест", callback_data="rt_ack_instructions")]
        ]
    )
    try:
        msg_id = None
        if isinstance(
            trigger_event, CallbackQuery
        ):  # Edit previous message if it was a callback
            await trigger_event.message.edit_text(
                instruction_text, reply_markup=kbd
            )
            msg_id = trigger_event.message.message_id
        else:  # Send new message
            msg = await bot.send_message(
                chat_id, instruction_text, reply_markup=kbd
            )
            msg_id = msg.message_id
        if msg_id:
            await state.update_data(rt_instruction_message_id=msg_id)
    except TelegramBadRequest as e:
        logger.error(f"RT start: Error sending/editing instructions: {e}")
        await bot.send_message(
            chat_id, "Ошибка при запуске теста. Попробуйте снова."
        )
        await state.clear()  # Or go to main menu if profile exists


@dp.callback_query(
    F.data == "rt_ack_instructions",
    ReactionTimeTestStates.initial_instructions,
)
async def rt_on_instructions_acknowledged(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    instruction_msg_id = data.get("rt_instruction_message_id")

    if instruction_msg_id:
        try:
            await bot.edit_message_text(
                "Подготовка к фазе запоминания...",
                chat_id=chat_id,
                message_id=instruction_msg_id,
                reply_markup=None,
            )
            # Keep instruction_msg_id in FSM, it will be deleted after memorization
        except TelegramBadRequest:
            logger.warning("RT: Could not edit instruction message on ack.")
            # If edit fails, it might be deleted by another process or already gone
            # The old instruction_msg_id will be attempted to be deleted later anyway.

    await state.set_state(ReactionTimeTestStates.memorization_display)

    target_image_path = random.choice(REACTION_TIME_IMAGE_POOL)
    await state.update_data(rt_target_image_path=target_image_path)

    logger.info(
        f"RT UID {data.get('rt_unique_id_for_test')}: Attempt {data.get('rt_current_attempt')}. Target: {target_image_path}"
    )

    try:
        img_file = FSInputFile(target_image_path)
        msg = await bot.send_photo(
            chat_id=chat_id,
            photo=img_file,
            caption=f"Запомните это изображение! (Исчезнет через {REACTION_TIME_MEMORIZATION_S} сек)",
        )
        await state.update_data(
            rt_memorization_image_message_id=msg.message_id
        )
    except Exception as e:
        logger.error(
            f"RT: Failed to send memorization image {target_image_path}: {e}"
        )
        await bot.send_message(
            chat_id,
            "Ошибка: не удалось загрузить изображение для запоминания. Тест прерван.",
        )
        await save_reaction_time_results(
            state,
            is_interrupted=True,
            status_override="Failed due to image error",
        )
        await cleanup_reaction_time_ui(
            state, bot, "Тест прерван (ошибка изображения)."
        )
        await _rt_go_to_main_menu_or_clear(state, callback.message)
        return

    # Start the memorization timer task
    memo_task = asyncio.create_task(_rt_memorization_phase_task(state, bot))
    await state.update_data(rt_memorization_task=memo_task)


async def _rt_memorization_phase_task(state: FSMContext, bot_instance: Bot):
    try:
        await asyncio.sleep(REACTION_TIME_MEMORIZATION_S)

        # Check if state is still correct before proceeding
        if (
            await state.get_state()
            != ReactionTimeTestStates.memorization_display.state
        ):
            logger.info("RT Memorization task: State changed, aborting.")
            return

        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        memo_msg_id = data.get("rt_memorization_image_message_id")
        instruction_msg_id = data.get(
            "rt_instruction_message_id"
        )  # Instruction msg might have been edited

        if memo_msg_id:
            try:
                await bot_instance.delete_message(
                    chat_id=chat_id, message_id=memo_msg_id
                )
                await state.update_data(rt_memorization_image_message_id=None)
            except TelegramBadRequest:
                logger.warning("RT: Failed to delete memorization image.")

        if (
            instruction_msg_id
        ):  # Delete the (possibly edited) instruction message
            try:
                await bot_instance.delete_message(
                    chat_id=chat_id, message_id=instruction_msg_id
                )
                await state.update_data(rt_instruction_message_id=None)
            except TelegramBadRequest:
                logger.warning(
                    "RT: Failed to delete instruction message after memorization."
                )

        # Transition to reaction phase
        await _start_rt_reaction_phase(state, bot_instance)

    except asyncio.CancelledError:
        logger.info("RT Memorization task cancelled.")
        # If cancelled, UI cleanup will be handled by the calling function (e.g., stoptest)
    except Exception as e:
        logger.error(f"RT Memorization task error: {e}", exc_info=True)
        # Try to gracefully end the test if an unexpected error occurs
        data = await state.get_data()  # Re-fetch data in case of changes
        chat_id = data.get("rt_chat_id")
        if chat_id:
            await bot_instance.send_message(
                chat_id, "Произошла ошибка в фазе запоминания. Тест прерван."
            )

        await save_reaction_time_results(
            state,
            is_interrupted=True,
            status_override="Failed in memorization",
        )

        # Mock message for menu navigation
        mock_message = None
        if chat_id:
            mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
            mock_message = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat,
                from_user=bot.id,
            )

        await cleanup_reaction_time_ui(
            state, bot_instance, "Тест прерван (ошибка)."
        )
        if mock_message:
            await _rt_go_to_main_menu_or_clear(state, mock_message)
        else:  # Should not happen if chat_id was present
            await state.clear()


async def _start_rt_reaction_phase(state: FSMContext, bot_instance: Bot):
    await state.set_state(ReactionTimeTestStates.reaction_stimulus_display)
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    target_image_path = data.get("rt_target_image_path")

    # Prepare stimulus sequence
    distractors = [
        p for p in REACTION_TIME_IMAGE_POOL if p != target_image_path
    ]
    random.shuffle(distractors)

    stimuli_sequence = []
    num_distractors_needed = REACTION_TIME_NUM_STIMULI_IN_SEQUENCE - 1
    actual_num_distractors = min(len(distractors), num_distractors_needed)

    chosen_distractors = distractors[:actual_num_distractors]
    stimuli_sequence = [
        {"path": p, "is_target": False} for p in chosen_distractors
    ]

    # Insert target at a random position
    if REACTION_TIME_NUM_STIMULI_IN_SEQUENCE == 1:  # Only target
        target_insert_pos = 0
    elif (
        not stimuli_sequence
    ):  # If no distractors (e.g. pool too small for sequence length 1)
        target_insert_pos = 0
    else:
        min_pos = 0  # Can be at the beginning
        max_pos = len(stimuli_sequence)  # Can be at the end
        target_insert_pos = random.randint(min_pos, max_pos)

    stimuli_sequence.insert(
        target_insert_pos, {"path": target_image_path, "is_target": True}
    )

    # Trim if sequence is longer than desired (e.g. if many distractors were added but target inserted early)
    stimuli_sequence = stimuli_sequence[:REACTION_TIME_NUM_STIMULI_IN_SEQUENCE]

    # Ensure target is actually in the final sequence if NUM_STIMULI_IN_SEQUENCE > 0
    if (
        REACTION_TIME_NUM_STIMULI_IN_SEQUENCE > 0
        and not any(s['is_target'] for s in stimuli_sequence)
        and stimuli_sequence
    ):
        idx_to_replace = random.randrange(len(stimuli_sequence))
        stimuli_sequence[idx_to_replace] = {
            "path": target_image_path,
            "is_target": True,
        }
        logger.warning(
            "RT: Target safeguard triggered in sequence generation."
        )

    if (
        not stimuli_sequence
    ):  # Should not happen with current logic unless NUM_STIMULI_IN_SEQUENCE is 0
        logger.error("RT: Stimulus sequence is empty! Aborting attempt.")
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                "Ошибка генерации последовательности стимулов. Попытка прервана.",
            )
        await _handle_rt_attempt_failure(
            state, bot_instance, "Ошибка генерации стимулов"
        )
        return

    await state.update_data(
        rt_stimuli_sequence=stimuli_sequence,
        rt_current_stimulus_index=0,
        rt_target_displayed_time=None,  # Time when target is shown
        rt_reacted_correctly_this_attempt=False,
        rt_reaction_stimulus_message_id=None,  # ID of the message showing stimuli
    )

    # Start the reaction cycle
    reaction_task = asyncio.create_task(
        _rt_reaction_cycle_task(state, bot_instance)
    )
    await state.update_data(rt_reaction_cycle_task=reaction_task)


async def _rt_reaction_cycle_task(state: FSMContext, bot_instance: Bot):
    try:
        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        stimuli_sequence = data.get("rt_stimuli_sequence", [])
        current_idx = data.get("rt_current_stimulus_index", 0)
        stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")

        # Check if all stimuli in this attempt have been shown
        if current_idx >= len(stimuli_sequence):
            # If target was displayed but not reacted to (miss)
            if data.get("rt_target_displayed_time") and not data.get(
                "rt_reacted_correctly_this_attempt"
            ):
                logger.info(
                    f"RT UID {data.get('rt_unique_id_for_test')}: Target missed (end of sequence)."
                )
                if chat_id:  # Notify user
                    await bot_instance.send_message(
                        chat_id, "Вы пропустили целевое изображение."
                    )
                await _handle_rt_attempt_failure(
                    state, bot_instance, "Цель пропущена"
                )
            # If target was never displayed (e.g. sequence empty or error), or if reacted correctly, this branch is not a miss.
            # This case (end of sequence without miss already handled) usually means success or earlier error.
            return

        current_stimulus = stimuli_sequence[current_idx]
        image_path = current_stimulus["path"]
        is_target = current_stimulus["is_target"]

        # Store if current displayed image is target for button press handler
        await state.update_data(rt_current_displayed_image_is_target=is_target)

        caption_text = "РЕАГИРОВАТЬ!"  # Minimal caption
        kbd = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    IKB(
                        text="💥 РЕАГИРОВАТЬ! 💥",
                        callback_data="rt_react_button_pressed",
                    )
                ]
            ]
        )

        try:
            img_file = FSInputFile(image_path)
            if not stimulus_msg_id:  # First stimulus in this cycle
                msg = await bot_instance.send_photo(
                    chat_id,
                    photo=img_file,
                    caption=caption_text,
                    reply_markup=kbd,
                )
                stimulus_msg_id = msg.message_id
                await state.update_data(
                    rt_reaction_stimulus_message_id=stimulus_msg_id
                )
            else:  # Edit existing message
                media = InputMediaPhoto(
                    media=img_file, caption=caption_text
                )  # Caption is minimal
                await bot_instance.edit_message_media(
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    media=media,
                    reply_markup=kbd,
                )

            if is_target:
                await state.update_data(rt_target_displayed_time=time.time())
                logger.info(
                    f"RT UID {data.get('rt_unique_id_for_test')}: Target '{image_path}' displayed."
                )

        except (
            Exception
        ) as e:  # Broad exception for file not found, telegram errors etc.
            logger.error(
                f"RT: Failed to send/edit stimulus image {image_path}: {e}"
            )
            if chat_id:
                await bot_instance.send_message(
                    chat_id,
                    f"Ошибка отображения стимула '{os.path.basename(image_path)}'. Попытка прервана.",
                )
            await _handle_rt_attempt_failure(
                state, bot_instance, "Ошибка изображения стимула"
            )
            return

        await state.update_data(rt_current_stimulus_index=current_idx + 1)

        # Wait for stimulus interval, but allow cancellation if button pressed
        start_sleep = time.time()
        while time.time() - start_sleep < REACTION_TIME_STIMULUS_INTERVAL_S:
            await asyncio.sleep(0.1)  # Check frequently for state changes
            if (
                await state.get_state()
                != ReactionTimeTestStates.reaction_stimulus_display.state
            ):
                logger.info(
                    "RT Cycle task: State changed during stimulus display, aborting this cycle."
                )
                return

        # If loop finishes without state change (no button press for this stimulus), schedule next
        if (
            await state.get_state()
            == ReactionTimeTestStates.reaction_stimulus_display.state
        ):
            new_reaction_task = asyncio.create_task(
                _rt_reaction_cycle_task(state, bot_instance)
            )
            await state.update_data(rt_reaction_cycle_task=new_reaction_task)

    except asyncio.CancelledError:
        logger.info("RT Reaction cycle task cancelled.")
    except Exception as e:
        logger.error(f"RT Reaction cycle task error: {e}", exc_info=True)
        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        if chat_id:
            await bot_instance.send_message(
                chat_id, "Произошла ошибка в фазе реакции. Тест прерван."
            )

        await save_reaction_time_results(
            state,
            is_interrupted=True,
            status_override="Failed in reaction cycle",
        )

        mock_message = None
        if chat_id:
            mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
            mock_message = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat,
                from_user=bot.id,
            )

        await cleanup_reaction_time_ui(
            state, bot_instance, "Тест прерван (ошибка)."
        )
        if mock_message:
            await _rt_go_to_main_menu_or_clear(state, mock_message)
        else:
            await state.clear()


@dp.callback_query(
    F.data == "rt_react_button_pressed",
    ReactionTimeTestStates.reaction_stimulus_display,
)
async def on_rt_react_button_pressed(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer()  # Acknowledge button press immediately

    # Cancel the reaction cycle task as user has reacted
    data = await state.get_data()
    reaction_cycle_task = data.get("rt_reaction_cycle_task")
    if reaction_cycle_task and not reaction_cycle_task.done():
        reaction_cycle_task.cancel()
        try:
            # Wait briefly for cancellation to complete, but don't block indefinitely
            await asyncio.wait_for(reaction_cycle_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass  # Task either cancelled or timed out, proceed

    chat_id = data.get("rt_chat_id")
    is_target_displayed_now = data.get(
        "rt_current_displayed_image_is_target", False
    )
    target_display_time = data.get("rt_target_displayed_time")
    uid_for_test = data.get('rt_unique_id_for_test')

    if is_target_displayed_now and target_display_time:
        # Correct reaction
        reaction_time_seconds = time.time() - target_display_time

        # Heuristic for Telegram latency - can be adjusted or made configurable
        # This is a very rough estimate. Real latency varies.
        telegram_latency_seconds = (
            0.575  # Example: 575ms, adjust based on observation
        )

        corrected_reaction_time_seconds = (
            reaction_time_seconds - telegram_latency_seconds
        )

        # Ensure reaction time is not negative after correction
        if corrected_reaction_time_seconds < 0:
            logger.warning(
                f"RT UID {uid_for_test}: Corrected reaction time < 0 ({corrected_reaction_time_seconds * 1000:.0f}ms). Clamped."
            )
            corrected_reaction_time_seconds = max(
                0.001, corrected_reaction_time_seconds
            )  # Clamp to 1ms if negative

        reaction_time_ms = int(corrected_reaction_time_seconds * 1000)

        await state.update_data(
            rt_reaction_time_ms=reaction_time_ms,
            rt_status="Passed",
            rt_reacted_correctly_this_attempt=True,
        )
        logger.info(
            f"RT UID {uid_for_test}: Raw: {reaction_time_seconds * 1000:.0f}ms. Latency adj: {telegram_latency_seconds * 1000:.0f}ms. Corrected: {reaction_time_ms}ms."
        )

        # Remove buttons from stimulus message
        stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")
        if stimulus_msg_id:
            try:
                await bot.edit_message_reply_markup(
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    reply_markup=None,
                )
            except TelegramBadRequest:
                pass  # Ignore if already gone

        await bot.send_message(
            chat_id,
            f"<b>Верно!</b> Ваше время реакции: {reaction_time_ms} мс.",
        )

        # Test passed, save results and end
        await save_reaction_time_results(state, is_interrupted=False)
        await cleanup_reaction_time_ui(
            state, bot, "Тест на Скорость Реакции завершен."
        )
        await _rt_go_to_main_menu_or_clear(state, callback.message)

    else:
        # Incorrect reaction (pressed on distractor or too early/late for target)
        logger.info(f"RT UID {uid_for_test}: Incorrect reaction.")
        if chat_id:
            await bot.send_message(
                chat_id,
                "<b>Ошибка!</b> Вы нажали на кнопку, когда было показано НЕ целевое изображение, или слишком рано/поздно.",
            )
        await _handle_rt_attempt_failure(state, bot, "Неверная реакция")


async def _handle_rt_attempt_failure(
    state: FSMContext, bot_instance: Bot, reason: str
):
    data = await state.get_data()
    current_attempt = data.get("rt_current_attempt", 1)
    chat_id = data.get("rt_chat_id")

    # Ensure any ongoing reaction cycle (showing stimuli) is stopped
    reaction_cycle_task = data.get("rt_reaction_cycle_task")
    if reaction_cycle_task and not reaction_cycle_task.done():
        reaction_cycle_task.cancel()
        try:
            await asyncio.wait_for(reaction_cycle_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    # Clean up the stimulus message if it exists
    stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")
    if stimulus_msg_id and chat_id:
        try:
            await bot_instance.delete_message(chat_id, stimulus_msg_id)
            await state.update_data(rt_reaction_stimulus_message_id=None)
        except TelegramBadRequest:
            logger.warning(
                f"RT: Failed to delete stimulus msg {stimulus_msg_id} on fail."
            )

    current_attempt += 1
    await state.update_data(rt_current_attempt=current_attempt)

    if current_attempt <= REACTION_TIME_MAX_ATTEMPTS:
        # Offer retry
        await state.set_state(
            ReactionTimeTestStates.awaiting_retry_confirmation
        )
        retry_text = (
            f"Причина: {reason}. Попытка {current_attempt - 1} из {REACTION_TIME_MAX_ATTEMPTS} не удалась.\n"
            f"Хотите попробовать еще раз (осталось {REACTION_TIME_MAX_ATTEMPTS - (current_attempt - 1)} попыток)?"
        )
        kbd = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    IKB(
                        text="Да, попробовать снова",
                        callback_data="rt_retry_yes",
                    )
                ],
                [IKB(text="Нет, завершить тест", callback_data="rt_retry_no")],
            ]
        )
        if chat_id:
            msg = await bot_instance.send_message(
                chat_id, retry_text, reply_markup=kbd
            )
            await state.update_data(
                rt_retry_confirmation_message_id=msg.message_id
            )
    else:
        # Max attempts reached
        await state.update_data(rt_status="Failed")
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                f"Причина: {reason}. Максимальное количество попыток ({REACTION_TIME_MAX_ATTEMPTS}) исчерпано. Тест не пройден.",
            )
        await save_reaction_time_results(
            state, is_interrupted=False
        )  # Not interrupted, but failed
        await cleanup_reaction_time_ui(
            state,
            bot_instance,
            "Тест на Скорость Реакции завершен (не пройден).",
        )

        mock_message = None
        if chat_id:  # Create mock message for navigation
            mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
            mock_message = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat,
                from_user=bot.id,
            )
        if mock_message:
            await _rt_go_to_main_menu_or_clear(state, mock_message)
        else:
            await state.clear()


@dp.callback_query(
    F.data == "rt_retry_yes",
    ReactionTimeTestStates.awaiting_retry_confirmation,
)
async def on_rt_retry_yes(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    retry_msg_id = data.get("rt_retry_confirmation_message_id")
    if retry_msg_id:
        try:
            await bot.edit_message_text(
                "Готовим новую попытку...",
                chat_id=callback.message.chat.id,
                message_id=retry_msg_id,
                reply_markup=None,
            )
            # This message becomes the new "instruction message" effectively for cleanup
            await state.update_data(
                rt_instruction_message_id=retry_msg_id,
                rt_retry_confirmation_message_id=None,
            )
        except TelegramBadRequest:
            # If edit failed, just clear the ID, a new instruction message will be sent if needed by rt_on_instructions_acknowledged
            await state.update_data(
                rt_instruction_message_id=None,
                rt_retry_confirmation_message_id=None,
            )

    # Reset relevant FSM data for a new attempt, but keep attempt counter
    await state.update_data(
        rt_target_image_path=None,
        rt_memorization_image_message_id=None,
        rt_reaction_stimulus_message_id=None,
        rt_target_displayed_time=None,
        rt_reacted_correctly_this_attempt=False,
        rt_memorization_task=None,
        rt_reaction_cycle_task=None,
        # rt_status remains "Pending" or as is until a pass/fail
    )
    # Go back to instruction phase to start a new attempt cycle
    await state.set_state(ReactionTimeTestStates.initial_instructions)
    await rt_on_instructions_acknowledged(
        callback, state
    )  # This will re-trigger the memorization phase


@dp.callback_query(
    F.data == "rt_retry_no", ReactionTimeTestStates.awaiting_retry_confirmation
)
async def on_rt_retry_no(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.update_data(rt_status="Failed")  # User chose not to retry

    data = await state.get_data()
    retry_msg_id = data.get("rt_retry_confirmation_message_id")
    if retry_msg_id:
        try:
            await bot.delete_message(
                chat_id=callback.message.chat.id, message_id=retry_msg_id
            )
        except TelegramBadRequest:
            pass

    await bot.send_message(
        callback.message.chat.id, "Тест завершен по вашему выбору."
    )
    await save_reaction_time_results(
        state, is_interrupted=False
    )  # Not interrupted, but failed by choice
    await cleanup_reaction_time_ui(
        state, bot, "Тест на Скорость Реакции завершен (не пройден)."
    )
    await _rt_go_to_main_menu_or_clear(state, callback.message)


async def save_reaction_time_results(
    state: FSMContext,
    is_interrupted: bool = False,
    status_override: str = None,
):
    data = await state.get_data()
    uid = data.get("rt_unique_id_for_test")
    p_tgid, p_name, p_age = None, None, None

    if not uid:
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile:
            uid = active_profile.get("unique_id")
            p_tgid = active_profile.get("telegram_id")
            p_name = active_profile.get("name")
            p_age = active_profile.get("age")
        else:
            logger.warning("RT save: UID not found and no active profile.")
            # Cannot send message here as no trigger_msg context
            return
    else:  # Use test-specific profile info if available
        p_tgid = data.get(
            "rt_profile_telegram_id_for_test", data.get("active_telegram_id")
        )
        p_name = data.get("rt_profile_name_for_test", data.get("active_name"))
        p_age = data.get("rt_profile_age_for_test", data.get("active_age"))

    time_ms = data.get("rt_reaction_time_ms")
    attempts = data.get("rt_current_attempt", 1)  # Number of attempts made
    current_status_from_fsm = data.get("rt_status", "Unknown")
    final_status = current_status_from_fsm

    if status_override:
        final_status = status_override
    elif is_interrupted and final_status not in [
        "Passed",
        "Failed",
    ]:  # If interrupted before pass/fail
        final_status = "Interrupted"

    interrupted_by_command_col_val = (
        "Да" if is_interrupted and final_status == "Interrupted" else "Нет"
    )

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        for idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = idx
                break
        if row_num == -1:  # New user or UID not found, append
            new_row_data = [""] * len(ALL_EXPECTED_HEADERS)
            new_row_data[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row_data[uid_col_idx] = uid
            new_row_data[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row_data[ALL_EXPECTED_HEADERS.index("Age")] = (
                p_age if p_age else ""
            )
            ws.append(new_row_data)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num, column=h.index("ReactionTime_Time_ms") + 1
        ).value = (time_ms if time_ms is not None else "N/A")
        ws.cell(
            row=row_num, column=h.index("ReactionTime_Attempts") + 1
        ).value = attempts
        ws.cell(
            row=row_num, column=h.index("ReactionTime_Status") + 1
        ).value = final_status
        ws.cell(
            row=row_num, column=h.index("ReactionTime_Interrupted") + 1
        ).value = interrupted_by_command_col_val
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"RT results for UID {uid} saved. Status: {final_status}, Time: {time_ms}ms, Attempts: {attempts}, Interrupted by cmd: {interrupted_by_command_col_val}"
        )

    except Exception as e:
        logger.error(
            f"RT results save error for UID {uid}: {e}", exc_info=True
        )
        # Cannot send message to user here as no trigger_msg context
        # Error should be logged, and test flow should handle user notification if possible.
        chat_id_for_error_msg = data.get("rt_chat_id")
        if (
            chat_id_for_error_msg and await state.get_state() is not None
        ):  # If test is somewhat active
            try:  # Try sending error message to user if chat_id is known
                await bot.send_message(
                    chat_id_for_error_msg,
                    "Ошибка сохранения результатов Теста на Скорость Реакции.",
                )
            except Exception as send_err:
                logger.error(
                    f"RT save_results: Failed to send error message to user {chat_id_for_error_msg}: {send_err}"
                )


async def cleanup_reaction_time_ui(
    state: FSMContext, bot_instance: Bot, final_text: str
):
    logger.info(f"Cleaning up Reaction Time UI. Final text: {final_text}")
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")

    # Cancel any ongoing tasks
    memo_task = data.get("rt_memorization_task")
    if memo_task and not memo_task.done():
        memo_task.cancel()
        try:
            await asyncio.wait_for(memo_task, timeout=0.5)  # Brief wait
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    reaction_cycle_task = data.get("rt_reaction_cycle_task")
    if reaction_cycle_task and not reaction_cycle_task.done():
        reaction_cycle_task.cancel()
        try:
            await asyncio.wait_for(reaction_cycle_task, timeout=0.5)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    # Identify messages to clean up
    instruction_msg_id = data.get("rt_instruction_message_id")
    memorization_msg_id = data.get("rt_memorization_image_message_id")
    stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")
    retry_confirm_msg_id = data.get("rt_retry_confirmation_message_id")

    msg_to_edit_id = None
    is_photo_msg = (
        False  # Track if the message to edit is a photo for caption editing
    )

    # Prioritize editing the last relevant message
    if stimulus_msg_id:
        msg_to_edit_id = stimulus_msg_id
        is_photo_msg = True
    elif retry_confirm_msg_id:
        msg_to_edit_id = retry_confirm_msg_id
    elif instruction_msg_id:  # Could be text or edited to "Preparing..."
        msg_to_edit_id = instruction_msg_id
    elif memorization_msg_id:  # Photo
        msg_to_edit_id = memorization_msg_id
        is_photo_msg = True

    all_rt_msg_ids = {
        instruction_msg_id,
        memorization_msg_id,
        stimulus_msg_id,
        retry_confirm_msg_id,
    }
    if chat_id:
        for m_id in all_rt_msg_ids:
            if m_id and m_id != msg_to_edit_id:  # Delete others
                try:
                    await bot_instance.delete_message(chat_id, m_id)
                except TelegramBadRequest:
                    pass

        if msg_to_edit_id:
            try:
                if is_photo_msg:  # Edit caption if it was a photo
                    await bot_instance.edit_message_caption(
                        chat_id=chat_id,
                        message_id=msg_to_edit_id,
                        caption=final_text,
                        reply_markup=None,
                    )
                else:  # Edit text
                    await bot_instance.edit_message_text(
                        text=final_text,
                        chat_id=chat_id,
                        message_id=msg_to_edit_id,
                        reply_markup=None,
                    )
            except TelegramBadRequest:  # If edit fails, send new
                await bot_instance.send_message(
                    chat_id, final_text, reply_markup=None
                )
        elif final_text:  # No message to edit, but text to send
            await bot_instance.send_message(
                chat_id, final_text, reply_markup=None
            )

    # Clean FSM data related to RT test
    current_fsm_data = await state.get_data()
    new_data = {
        k: v for k, v in current_fsm_data.items() if not k.startswith("rt_")
    }
    # Preserve general profile data
    for pk in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if pk in current_fsm_data and pk not in new_data:
            new_data[pk] = current_fsm_data[pk]
    await state.set_data(new_data)


async def check_if_reaction_time_results_exist(
    profile_unique_id: str | int,
) -> bool:
    if not profile_unique_id:
        return False
    try:
        uid = int(profile_unique_id)
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        status_idx = ALL_EXPECTED_HEADERS.index("ReactionTime_Status")
        time_idx = ALL_EXPECTED_HEADERS.index("ReactionTime_Time_ms")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[uid_col_idx] is not None and int(row[uid_col_idx]) == uid:
                status_val = row[status_idx] if status_idx < len(row) else None
                time_val = row[time_idx] if time_idx < len(row) else None
                if status_val is not None or (
                    time_val is not None and time_val != "N/A"
                ):
                    return True
        return False
    except FileNotFoundError:
        return False
    except ValueError:
        return False  # Header not found etc.
    except Exception as e:
        logger.error(
            f"RT check_results_exist error UID {profile_unique_id}: {e}"
        )
        return False


async def _rt_go_to_main_menu_or_clear(
    state: FSMContext, trigger_message: Message
):
    """Clears RT test state and navigates to main menu if profile active, or clears all if not."""
    fsm_data = (
        await state.get_data()
    )  # Get current data which might include profile
    # Extract only profile keys to preserve them
    profile_data = {
        key: fsm_data.get(key)
        for key in [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        if fsm_data.get(key)
    }

    await state.set_state(None)  # Clear current FSM state (e.g., RT states)

    if profile_data.get("active_unique_id"):
        await state.set_data(profile_data)  # Restore only profile data
        await send_main_action_menu(
            trigger_message, ACTION_SELECTION_KEYBOARD_RETURNING, state=state
        )
    else:  # No active profile found in FSM
        if (
            hasattr(trigger_message, 'chat') and trigger_message.chat
        ):  # Check if trigger_message is valid
            await trigger_message.answer(
                "Профиль не активен. Пожалуйста, /start для начала."
            )
        await state.clear()  # Clear all FSM data


# --- Verbal Fluency Test Logic ---
async def start_verbal_fluency_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    logger.info(
        f"Starting Verbal Fluency Test for UID: {profile.get('unique_id')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    if not VERBAL_FLUENCY_TASK_POOL:
        await bot.send_message(
            chat_id,
            "Ошибка: Пул заданий для теста вербальной беглости пуст. Тест не может быть запущен.",
        )
        logger.error(
            "Verbal Fluency Test: Task pool is empty. Cannot start test."
        )
        # Gracefully return to menu
        await state.set_state(None)
        active_profile = await get_active_profile_from_fsm(
            state
        )  # Re-check profile
        keyboard = (
            ACTION_SELECTION_KEYBOARD_RETURNING
            if active_profile
            else ACTION_SELECTION_KEYBOARD_NEW
        )
        await send_main_action_menu(msg_ctx, keyboard, state=state)
        return

    chosen_task = random.choice(VERBAL_FLUENCY_TASK_POOL)
    task_letter = chosen_task["letter"]

    await state.set_state(VerbalFluencyStates.showing_instructions_and_task)
    await state.update_data(
        vf_unique_id_for_test=profile.get("unique_id"),
        vf_profile_name_for_test=profile.get("name"),
        vf_profile_age_for_test=profile.get("age"),
        vf_profile_telegram_id_for_test=profile.get("telegram_id"),
        vf_chat_id=chat_id,
        vf_task_base_category=VERBAL_FLUENCY_CATEGORY,  # Store for record
        vf_task_letter=task_letter,
        vf_collected_words=set(),  # Use a set for unique words
        vf_timer_task=None,  # To store the asyncio.Task for the timer
        vf_task_message_id=None,  # Message ID of the main task/timer display
        vf_trigger_event_for_stop=msg_ctx,  # Store original event for menu nav on stop
    )

    instruction_text = (
        f"<b>Тест на вербальную беглость</b>\n\n"
        f"Вам будет дана буква. Ваша задача – назвать как можно больше слов, "
        f"начинающихся на эту букву.\n"
        f"На выполнение задания даётся {VERBAL_FLUENCY_DURATION_S} секунд.\n"
        f"Слова можно писать в одном или нескольких сообщениях. Каждое слово должно быть не менее двух букв.\n\n"
        f"Нажмите 'Начать', чтобы увидеть букву и запустить таймер."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать", callback_data="vf_start_test_confirmed")]
        ]
    )

    try:
        sent_message = None
        if isinstance(trigger_event, CallbackQuery):  # Edit previous message
            await trigger_event.message.edit_text(
                instruction_text, reply_markup=kbd
            )
            sent_message = trigger_event.message
        else:  # Send new
            sent_message = await bot.send_message(
                chat_id, instruction_text, reply_markup=kbd
            )
        if sent_message:
            await state.update_data(vf_task_message_id=sent_message.message_id)
    except TelegramBadRequest as e:
        logger.error(
            f"Verbal Fluency start: Error sending/editing instructions: {e}"
        )
        await bot.send_message(
            chat_id, "Ошибка при запуске теста. Попробуйте снова."
        )
        await state.clear()  # Or navigate to menu


async def _verbal_fluency_timer_task(state: FSMContext, bot_instance: Bot):
    data = await state.get_data()
    chat_id = data.get("vf_chat_id")
    task_message_id = data.get("vf_task_message_id")
    task_letter = data.get("vf_task_letter")
    last_displayed_text = ""  # To avoid "message is not modified" errors

    if not all([chat_id, task_message_id, task_letter]):
        logger.error(
            "Verbal Fluency timer: Missing critical data from FSM for timer."
        )
        # Attempt to end the test if possible, though context might be lost
        await _end_verbal_fluency_test(
            state, bot_instance, interrupted=True, trigger_event=None
        )
        return

    base_task_text_for_timer = f"Задание: Назовите как можно больше слов, начинающихся на букву <b>'{task_letter}'</b>.\n"

    try:
        for i in range(
            VERBAL_FLUENCY_DURATION_S, -1, -1
        ):  # Countdown from duration to 0
            # Check if state is still correct before editing message or sleeping
            if (
                await state.get_state()
                != VerbalFluencyStates.collecting_words.state
            ):
                logger.info(
                    "Verbal Fluency timer: State changed, aborting timer."
                )
                return

            current_timer_display = f"Осталось: {i} сек."
            full_message_content = f"{base_task_text_for_timer}{current_timer_display}\n\nВводите слова."

            if (
                full_message_content != last_displayed_text
            ):  # Only edit if text changed
                try:
                    await bot_instance.edit_message_text(
                        text=full_message_content,
                        chat_id=chat_id,
                        message_id=task_message_id,
                        parse_mode=ParseMode.HTML,
                    )
                    last_displayed_text = full_message_content
                except TelegramBadRequest as e:
                    if "message is not modified" not in str(e).lower():
                        logger.warning(
                            f"Verbal Fluency timer: edit_message_text (ID: {task_message_id}) failed: {e}."
                        )
                    # If other error, loop continues, next second will try again or error out.

            if i == 0:  # Time is up
                break
            await asyncio.sleep(1)

        # After loop, if state is still collecting_words, time is up.
        if (
            await state.get_state()
            == VerbalFluencyStates.collecting_words.state
        ):
            logger.info("Verbal Fluency timer: Time is up.")
            trigger_event_for_menu = data.get("vf_trigger_event_for_stop")
            # If original trigger_event is not available, create a mock one for menu navigation
            if not trigger_event_for_menu and chat_id:
                mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
                trigger_event_for_menu = Message(
                    message_id=0,
                    date=int(time.time()),
                    chat=mock_chat,
                    from_user=bot.id,
                )

            await _end_verbal_fluency_test(
                state,
                bot_instance,
                interrupted=False,
                trigger_event=trigger_event_for_menu,
            )

    except asyncio.CancelledError:
        logger.info("Verbal Fluency timer task explicitly cancelled.")
        # Cleanup (like unpinning) will be handled by the function that cancelled it (e.g. stoptest)
    except Exception as e:
        logger.error(
            f"Verbal Fluency timer task unexpected error: {e}", exc_info=True
        )
        # Try to end the test gracefully
        trigger_event_for_menu = data.get("vf_trigger_event_for_stop")
        if not trigger_event_for_menu and chat_id:
            mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
            trigger_event_for_menu = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat,
                from_user=bot.id,
            )
        await _end_verbal_fluency_test(
            state,
            bot_instance,
            interrupted=True,
            trigger_event=trigger_event_for_menu,
        )


async def _end_verbal_fluency_test(
    state: FSMContext,
    bot_instance: Bot,
    interrupted: bool,
    trigger_event: [
        Message,
        CallbackQuery,
    ] = None,  # Original event for menu navigation
):
    current_fsm_state_str = await state.get_state()
    # Ensure this function only runs if we are actually in a VF test state
    if not current_fsm_state_str or not current_fsm_state_str.startswith(
        VerbalFluencyStates.__name__
    ):
        logger.info(
            "VF _end_test: Called but test not active or already ended."
        )
        return

    logger.info(f"VF: Entering _end_test. Interrupted: {interrupted}")
    data = await state.get_data()
    chat_id = data.get("vf_chat_id")
    task_message_id = data.get("vf_task_message_id")
    timer_task = data.get("vf_timer_task")

    # Cancel timer task if it's running
    if timer_task and not timer_task.done():
        timer_task.cancel()
        try:
            await asyncio.wait_for(timer_task, timeout=0.5)  # Give it a moment
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass
    await state.update_data(vf_timer_task=None)  # Clear the task from FSM

    collected_words = data.get("vf_collected_words", set())
    word_count = len(collected_words)

    # Save results regardless of interruption (might save 0 words if interrupted early)
    await save_verbal_fluency_results(state, is_interrupted=interrupted)

    result_message_text = ""
    if chat_id:  # Only interact with chat if chat_id is known
        if task_message_id:  # Unpin and delete the task message
            try:
                await bot_instance.unpin_chat_message(
                    chat_id=chat_id, message_id=task_message_id
                )
            except TelegramBadRequest:
                pass  # Ignore if not pinned or already unpinned
            try:
                await bot_instance.delete_message(
                    chat_id=chat_id, message_id=task_message_id
                )
            except TelegramBadRequest:
                logger.warning(
                    f"VF _end_test: Failed to delete task msg {task_message_id}"
                )
        await state.update_data(
            vf_task_message_id=None
        )  # Clear msg_id from FSM

        if interrupted:
            result_message_text = (
                "Тест на вербальную беглость был <b>ПРЕРВАН</b>.\n"
                f"Сохраненный результат: {word_count} слов(а)."
            )
        else:  # Time is up normally
            result_message_text = (
                "Время вышло! Тест на вербальную беглость завершен.\n"
                f"Я сохранил результат. Количество названных (уникальных) слов: {word_count}.\n"
                f"Общее время выполнения: {VERBAL_FLUENCY_DURATION_S} сек."
            )
        try:
            await bot_instance.send_message(
                chat_id, result_message_text, parse_mode=ParseMode.HTML
            )
        except Exception as e_send_res:
            logger.error(
                f"VF _end_test: Fail to send result msg: {e_send_res}"
            )

    # Call general cleanup to remove vf_ keys from FSM, but final_text is already sent
    await cleanup_verbal_fluency_ui(state, bot_instance, final_text=None)

    # Navigate to main menu
    current_data_after_cleanup = (
        await state.get_data()
    )  # Get data after vf_ keys are cleaned
    profile_data_to_keep = {
        k: current_data_after_cleanup.get(k)
        for k in [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        if current_data_after_cleanup.get(k)
    }
    await state.set_state(None)  # Clear VF state

    if profile_data_to_keep.get("active_unique_id"):
        await state.set_data(profile_data_to_keep)  # Restore profile

        event_for_menu = (
            trigger_event  # Use the passed event for menu navigation
        )
        if (
            not event_for_menu and chat_id
        ):  # Fallback if no event passed but chat_id exists
            mock_chat_obj = Chat(id=chat_id, type=ChatType.PRIVATE)
            event_for_menu = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat_obj,
                from_user=bot.id,
            )

        if event_for_menu:
            await send_main_action_menu(
                event_for_menu,
                ACTION_SELECTION_KEYBOARD_RETURNING,
                state=state,
            )
        elif (
            not chat_id
        ):  # Cannot send menu if no chat_id and no trigger_event
            logger.warning(
                "VF _end_test: No chat_id or trigger_event available to send main menu."
            )

    else:  # No active profile
        if chat_id:  # If chat_id was known, inform user
            await bot_instance.send_message(
                chat_id, "Профиль не активен. Пожалуйста, /start для начала."
            )
        await state.clear()  # Clear all FSM if no profile
    logger.info("Verbal Fluency: Exiting _end_verbal_fluency_test.")


async def save_verbal_fluency_results(state: FSMContext, is_interrupted: bool):
    data = await state.get_data()
    uid = data.get("vf_unique_id_for_test")
    p_tgid, p_name, p_age = None, None, None

    if not uid:  # Fallback to active profile if test-specific UID missing
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile:
            uid = active_profile.get("unique_id")
            p_tgid = active_profile.get("telegram_id")
            p_name = active_profile.get("name")
            p_age = active_profile.get("age")
        else:
            logger.error(
                "VF save: UID not found in FSM or active profile. Cannot save."
            )
            # Cannot send message here as no trigger_msg context
            return
    else:
        p_tgid = data.get(
            "vf_profile_telegram_id_for_test", data.get("active_telegram_id")
        )
        p_name = data.get("vf_profile_name_for_test", data.get("active_name"))
        p_age = data.get("vf_profile_age_for_test", data.get("active_age"))

    letter = data.get("vf_task_letter", "N/A")
    collected_words = data.get("vf_collected_words", set())
    word_count = len(collected_words)
    words_list_str = ", ".join(
        sorted(list(collected_words))
    )  # Alphabetical list
    interrupted_status = "Да" if is_interrupted else "Нет"
    excel_category_display = (
        f"Слова на букву {letter}"  # More descriptive for Excel
    )

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        for idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = idx
                break
        if row_num == -1:  # New user or UID not found, append
            new_row_data = [""] * len(ALL_EXPECTED_HEADERS)
            new_row_data[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row_data[uid_col_idx] = uid
            new_row_data[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row_data[ALL_EXPECTED_HEADERS.index("Age")] = (
                p_age if p_age else ""
            )
            ws.append(new_row_data)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_Category") + 1
        ).value = excel_category_display
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_Letter") + 1
        ).value = letter
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_WordCount") + 1
        ).value = word_count
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_WordsList") + 1
        ).value = words_list_str
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_Interrupted") + 1
        ).value = interrupted_status
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"VF results for UID {uid} saved. Cat: {excel_category_display}, L: {letter}, Cnt: {word_count}, Int: {interrupted_status}"
        )
    except Exception as e:
        logger.error(f"VF results save error UID {uid}: {e}", exc_info=True)
        chat_id_for_err = data.get("vf_chat_id")
        if (
            chat_id_for_err and await state.get_state() is not None
        ):  # If VF test state is somewhat active
            try:
                await bot.send_message(
                    chat_id_for_err,
                    "Ошибка сохранения результатов теста вербальной беглости.",
                )
            except Exception as send_err_vf:
                logger.error(
                    f"VF save_results: Failed to send error message to user {chat_id_for_err}: {send_err_vf}"
                )


async def cleanup_verbal_fluency_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: str = None,  # Text to display in the task message before cleaning FSM
):
    logger.info(f"VF: Entering cleanup_ui. Final text: '{final_text}'")
    data = await state.get_data()
    chat_id = data.get("vf_chat_id")
    task_message_id = data.get("vf_task_message_id")
    timer_task = data.get("vf_timer_task")

    # Ensure timer is cancelled if cleanup is called externally (e.g. /stoptest)
    if timer_task and not timer_task.done():
        timer_task.cancel()
        try:
            await asyncio.wait_for(timer_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    # If final_text is provided, try to edit the task message
    if final_text and chat_id and task_message_id:
        try:
            await bot_instance.unpin_chat_message(
                chat_id=chat_id, message_id=task_message_id
            )
        except TelegramBadRequest:
            pass
        try:
            await bot_instance.edit_message_text(
                text=final_text,
                chat_id=chat_id,
                message_id=task_message_id,
                reply_markup=None,  # Remove any buttons
                parse_mode=ParseMode.HTML,
            )
        except TelegramBadRequest:  # If edit fails, try sending new
            try:
                await bot_instance.send_message(
                    chat_id, final_text, parse_mode=ParseMode.HTML
                )
            except Exception as e_sf:
                logger.error(f"VF cleanup send final_text err: {e_sf}")
    elif final_text and chat_id:  # No task_message_id, but final_text exists
        try:
            await bot_instance.send_message(
                chat_id, final_text, parse_mode=ParseMode.HTML
            )
        except Exception as e_sf_alt:
            logger.error(f"VF cleanup send final_text (alt) err: {e_sf_alt}")

    # Clean FSM data specific to Verbal Fluency
    current_fsm_data = await state.get_data()
    new_data = {
        k: v for k, v in current_fsm_data.items() if not k.startswith("vf_")
    }
    # Preserve general profile data
    for pk in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if pk in current_fsm_data and pk not in new_data:
            new_data[pk] = current_fsm_data[pk]
    await state.set_data(new_data)
    logger.info("Verbal Fluency cleanup: VF-specific FSM data cleaned.")


async def check_if_verbal_fluency_results_exist(
    profile_unique_id: str | int,
) -> bool:
    if not profile_unique_id:
        return False
    try:
        uid = int(profile_unique_id)
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        word_count_col_idx = ALL_EXPECTED_HEADERS.index(
            "VerbalFluency_WordCount"
        )
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[uid_col_idx] is not None and int(row[uid_col_idx]) == uid:
                if (
                    word_count_col_idx < len(row)
                    and row[word_count_col_idx] is not None
                ):
                    return True
        return False
    except FileNotFoundError:
        return False
    except ValueError:
        return False  # Header not found
    except Exception as e:
        logger.error(f"VF check_results_exist UID {profile_unique_id}: {e}")
        return False


# --- Mental Rotation Test Logic ---
async def _get_mr_stimulus_for_iteration(
    state: FSMContext,
) -> tuple[str | None, list[str] | None, int | None, str | None]:
    """
    Selects stimulus images for one iteration of the Mental Rotation Test.
    Returns: (reference_image_path, option_paths_list, correct_option_idx_0_based, error_message)
    """
    data = await state.get_data()
    used_references = data.get("mr_used_references", [])

    available_references = [
        ref for ref in MR_REFERENCE_FILES if ref not in used_references
    ]
    if not available_references:
        return None, None, None, "No more unique reference images available."

    selected_reference_filename = random.choice(available_references)
    selected_reference_path = os.path.join(
        MR_REFERENCES_DIR, selected_reference_filename
    )
    if not os.path.exists(selected_reference_path):
        return (
            None,
            None,
            None,
            f"Reference image not found: {selected_reference_path}",
        )

    used_references.append(selected_reference_filename)
    await state.update_data(mr_used_references=used_references)

    correct_projection_filenames = MR_CORRECT_PROJECTIONS_MAP.get(
        selected_reference_filename, []
    )
    if not correct_projection_filenames:
        return (
            None,
            None,
            None,
            f"No correct projection mapping for {selected_reference_filename}",
        )
    chosen_correct_proj_filename = random.choice(correct_projection_filenames)
    correct_projection_path = os.path.join(
        MR_CORRECT_PROJECTIONS_DIR, chosen_correct_proj_filename
    )
    if not os.path.exists(correct_projection_path):
        return (
            None,
            None,
            None,
            f"Correct projection image not found: {correct_projection_path}",
        )

    if not MR_ALL_DISTRACTORS_FILES:
        return (
            None,
            None,
            None,
            "Distractor image pool is empty. Check setup.",
        )

    num_distractors_to_select = 3  # 1 correct + 3 distractors = 4 options
    if len(MR_ALL_DISTRACTORS_FILES) < num_distractors_to_select:
        return (
            None,
            None,
            None,
            f"Not enough distractors available (need {num_distractors_to_select}, have {len(MR_ALL_DISTRACTORS_FILES)}).",
        )

    selected_distractor_paths = random.sample(
        MR_ALL_DISTRACTORS_FILES, num_distractors_to_select
    )
    for (
        dp
    ) in (
        selected_distractor_paths
    ):  # Should be pre-filtered but double check path validity
        if not os.path.exists(dp):
            return None, None, None, f"Distractor image not found: {dp}"

    options_paths = [correct_projection_path] + selected_distractor_paths
    random.shuffle(options_paths)
    correct_option_index = options_paths.index(
        correct_projection_path
    )  # 0-based

    return (
        selected_reference_path,
        options_paths,
        correct_option_index,
        None,  # No error
    )


async def _generate_mr_collage(
    option_image_paths: list[str],  # List of 4 image paths
) -> BufferedInputFile | None:
    if not Image or not UnidentifiedImageError:  # Pillow not available
        logger.error(
            "Pillow not available for MR collage generation. Cannot proceed."
        )
        return None

    images_to_collage = []
    for path in option_image_paths:
        try:
            img = Image.open(path)
            # Resize to fit collage cell, maintaining aspect ratio might be better if source images vary wildly
            img = img.resize(MR_COLLAGE_CELL_SIZE, Image.Resampling.LANCZOS)
            images_to_collage.append(img)
        except FileNotFoundError:
            logger.error(f"MR Collage: Image file not found: {path}")
            return None
        except UnidentifiedImageError:
            logger.error(f"MR Collage: Cannot identify image file: {path}")
            return None
        except Exception as e:
            logger.error(
                f"MR Collage: Error opening/resizing image {path}: {e}"
            )
            return None

    if len(images_to_collage) != 4:  # Expecting 4 images for a 2x2 collage
        logger.error(
            f"MR Collage: Expected 4 images, got {len(images_to_collage)}"
        )
        return None

    # Create a 2x2 collage
    collage_width = MR_COLLAGE_CELL_SIZE[0] * 2
    collage_height = MR_COLLAGE_CELL_SIZE[1] * 2
    collage = Image.new(
        "RGB", (collage_width, collage_height), MR_COLLAGE_BG_COLOR
    )

    # Paste images into collage grid (top-left, top-right, bottom-left, bottom-right)
    collage.paste(images_to_collage[0], (0, 0))
    collage.paste(images_to_collage[1], (MR_COLLAGE_CELL_SIZE[0], 0))
    collage.paste(images_to_collage[2], (0, MR_COLLAGE_CELL_SIZE[1]))
    collage.paste(
        images_to_collage[3],
        (MR_COLLAGE_CELL_SIZE[0], MR_COLLAGE_CELL_SIZE[1]),
    )

    bio = BytesIO()
    bio.name = "mr_collage.png"  # Filename for Telegram
    collage.save(bio, "PNG")
    bio.seek(0)
    return BufferedInputFile(bio.read(), filename=bio.name)


async def start_mental_rotation_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    logger.info(
        f"Starting Mental Rotation Test for UID: {profile.get('unique_id')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    await state.set_state(MentalRotationStates.initial_instructions_mr)
    await state.update_data(
        mr_unique_id_for_test=profile.get("unique_id"),
        mr_profile_name_for_test=profile.get("name"),
        mr_profile_age_for_test=profile.get("age"),
        mr_profile_telegram_id_for_test=profile.get("telegram_id"),
        mr_chat_id=chat_id,
        mr_current_iteration=0,  # 0-indexed internally, 1-indexed for display
        mr_iteration_results=[],
        # List of dicts: {iteration, is_correct, reaction_time_s, selected_option, correct_option}
        mr_used_references=[],  # List of reference image filenames used in this session
        mr_test_start_time=None,  # Set when actual test starts (after instructions)
        mr_reference_message_id=None,
        mr_options_message_id=None,
        mr_countdown_message_id=None,
        mr_feedback_message_id=None,
        mr_inter_iteration_countdown_task_ref=None,  # Store countdown task
        mr_triggering_event_for_menu=msg_ctx,  # Store original event for menu navigation
    )

    instruction_text = (
        "<b>Тест умственного вращения</b>\n\n"
        "Вам будет показан 3D объект и 4 варианта 2D проекций. "
        "Выберите номер той проекции, через которую мог бы пройти 3D объект.\n"
        "Объект можно вращать мысленно перед 'просовыванием'.\n\n"
        f"Тест состоит из {MENTAL_ROTATION_NUM_ITERATIONS} заданий."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать тест", callback_data="mr_ack_instructions")]
        ]
    )
    try:
        if isinstance(trigger_event, CallbackQuery):  # Edit previous message
            await trigger_event.message.edit_text(
                instruction_text, reply_markup=kbd
            )
        else:  # Send new message
            await bot.send_message(chat_id, instruction_text, reply_markup=kbd)
    except TelegramBadRequest as e:
        logger.error(f"MR start: Error sending/editing instructions: {e}")
        await bot.send_message(
            chat_id, "Ошибка при запуске теста. Попробуйте /start."
        )
        await state.clear()  # Or navigate to menu


@dp.callback_query(
    F.data == "mr_ack_instructions",
    MentalRotationStates.initial_instructions_mr,
)
async def mr_ack_instructions_callback(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer()
    await state.update_data(
        mr_test_start_time=time.time()
    )  # Start timer for whole test

    try:  # Delete the instruction message
        await callback.message.delete()
    except TelegramBadRequest:
        pass  # Already deleted or cannot be

    await _display_mr_stimulus(callback.message.chat.id, state, bot)


async def _display_mr_stimulus(
    chat_id: int,
    state: FSMContext,
    bot_instance: Bot,
    is_editing: bool = False,
):
    data = await state.get_data()
    current_iteration = (
        data.get("mr_current_iteration", 0) + 1
    )  # Increment for current task
    await state.update_data(mr_current_iteration=current_iteration)

    ref_path, opt_paths, correct_idx, err_msg = (
        await _get_mr_stimulus_for_iteration(state)
    )

    if err_msg or not ref_path or not opt_paths or correct_idx is None:
        logger.error(f"MR Stimulus Error: {err_msg}")
        await bot_instance.send_message(
            chat_id,
            f"Ошибка подготовки задания: {err_msg if err_msg else 'Неизвестная ошибка'}. Тест прерван.",
        )
        await _finish_mental_rotation_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    collage_file = await _generate_mr_collage(opt_paths)
    if not collage_file:  # Pillow or image file issue
        logger.error("MR Collage generation failed.")
        await bot_instance.send_message(
            chat_id, "Ошибка генерации коллажа. Тест прерван."
        )
        await _finish_mental_rotation_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    await state.update_data(
        mr_correct_option_index_for_current_iter=correct_idx
    )  # Store 0-based correct index

    # --- Send/Edit Reference Image ---
    ref_msg_id = data.get("mr_reference_message_id")
    try:
        if is_editing and ref_msg_id:
            await bot_instance.edit_message_media(
                chat_id=chat_id,
                message_id=ref_msg_id,
                media=InputMediaPhoto(media=FSInputFile(ref_path)),
            )
        else:  # Send new (or if edit failed and ref_msg_id became None)
            if (
                ref_msg_id
            ):  # Delete old if not editing (e.g. fallback from failed edit)
                try:
                    await bot_instance.delete_message(chat_id, ref_msg_id)
                except TelegramBadRequest:
                    pass
            msg = await bot_instance.send_photo(chat_id, FSInputFile(ref_path))
            await state.update_data(mr_reference_message_id=msg.message_id)
    except (TelegramBadRequest, FileNotFoundError) as e:
        logger.error(f"MR: Error sending/editing reference image: {e}")
        # Attempt to send as new if edit failed
        if is_editing:
            try:
                msg = await bot_instance.send_photo(
                    chat_id, FSInputFile(ref_path)
                )
                await state.update_data(mr_reference_message_id=msg.message_id)
            except Exception as e_new:
                logger.error(f"MR: Fallback send reference failed: {e_new}")
                await bot_instance.send_message(
                    chat_id, "Ошибка отображения эталона. Тест прерван."
                )
                await _finish_mental_rotation_test(
                    state,
                    bot_instance,
                    chat_id,
                    is_interrupted=True,
                    error_occurred=True,
                )
                return
        else:  # Failed on initial send
            await bot_instance.send_message(
                chat_id, "Ошибка отображения эталона. Тест прерван."
            )
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id,
                is_interrupted=True,
                error_occurred=True,
            )
            return

    # --- Send/Edit Options Collage with Buttons ---
    options_msg_id = data.get("mr_options_message_id")
    buttons = [  # 2x2 grid of buttons labeled 1-4
        [
            IKB(text="1", callback_data="mr_answer_1"),
            IKB(text="2", callback_data="mr_answer_2"),
        ],
        [
            IKB(text="3", callback_data="mr_answer_3"),
            IKB(text="4", callback_data="mr_answer_4"),
        ],
    ]
    reply_markup = InlineKeyboardMarkup(inline_keyboard=buttons)

    try:
        if is_editing and options_msg_id:
            await bot_instance.edit_message_media(
                chat_id=chat_id,
                message_id=options_msg_id,
                media=InputMediaPhoto(media=collage_file),
                reply_markup=reply_markup,
            )
        else:
            if options_msg_id:  # Delete old if not editing
                try:
                    await bot_instance.delete_message(chat_id, options_msg_id)
                except TelegramBadRequest:
                    pass
            msg = await bot_instance.send_photo(
                chat_id, collage_file, reply_markup=reply_markup
            )
            await state.update_data(mr_options_message_id=msg.message_id)
    except TelegramBadRequest as e:
        logger.error(f"MR: Error sending/editing options collage: {e}")
        if is_editing:  # Fallback if edit failed
            try:
                msg = await bot_instance.send_photo(
                    chat_id, collage_file, reply_markup=reply_markup
                )
                await state.update_data(mr_options_message_id=msg.message_id)
            except Exception as e_new_collage:
                logger.error(
                    f"MR: Fallback send collage failed: {e_new_collage}"
                )
                await bot_instance.send_message(
                    chat_id, "Ошибка отображения вариантов. Тест прерван."
                )
                await _finish_mental_rotation_test(
                    state,
                    bot_instance,
                    chat_id,
                    is_interrupted=True,
                    error_occurred=True,
                )
                return
        else:  # Failed on initial send
            await bot_instance.send_message(
                chat_id, "Ошибка отображения вариантов. Тест прерван."
            )
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id,
                is_interrupted=True,
                error_occurred=True,
            )
            return

    await state.update_data(
        mr_iteration_start_time=time.time()
    )  # Start reaction timer for this iteration
    await state.set_state(MentalRotationStates.displaying_stimulus_mr)


# Убедитесь, что этот импорт есть в начале вашего файла:
# from aiogram.types import User

# >>> НАЧАЛО ЗАМЕНЯЕМОГО БЛОКА <<<


async def _mr_schedule_feedback_revert(
    chat_id: int,
    message_id: int,
    normal_text: str,
    bot_instance: Bot,
    state_context: FSMContext,
):
    """Асинхронно изменяет текст сообщения на обычный (нежирный) после задержки."""
    try:
        await asyncio.sleep(MR_FEEDBACK_DISPLAY_TIME_S)

        current_fsm_data = await state_context.get_data()
        if (
            await state_context.get_state() is not None
            and current_fsm_data.get("mr_feedback_message_id") == message_id
        ):
            try:
                await bot_instance.edit_message_text(
                    text=normal_text,
                    chat_id=chat_id,
                    message_id=message_id,
                    parse_mode=None,
                )
            except TelegramBadRequest as e_edit:
                if (
                    "message is not modified" not in str(e_edit).lower()
                    and "message to edit not found" not in str(e_edit).lower()
                ):
                    logger.warning(
                        f"MR feedback revert (msg {message_id}): Edit failed: {e_edit}"
                    )
        else:
            logger.info(
                f"MR feedback revert (msg {message_id}): State changed or msg_id mismatch. Skipping revert."
            )

    except asyncio.CancelledError:
        logger.info(
            f"MR feedback revert task for msg {message_id} was cancelled."
        )
    except Exception as e:
        logger.error(
            f"MR feedback revert (msg {message_id}): Unexpected error: {e}",
            exc_info=True,
        )


async def start_mental_rotation_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    logger.info(
        f"Starting Mental Rotation Test for UID: {profile.get('unique_id')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    await state.set_state(MentalRotationStates.initial_instructions_mr)
    await state.update_data(
        mr_unique_id_for_test=profile.get("unique_id"),
        mr_profile_name_for_test=profile.get("name"),
        mr_profile_age_for_test=profile.get("age"),
        mr_profile_telegram_id_for_test=profile.get("telegram_id"),
        mr_chat_id=chat_id,
        mr_current_iteration=0,
        mr_iteration_results=[],
        mr_used_references=[],
        mr_test_start_time=None,
        mr_reference_message_id=None,
        mr_options_message_id=None,
        mr_countdown_message_id=None,
        mr_feedback_message_id=None,
        mr_inter_iteration_countdown_task_ref=None,
        mr_feedback_revert_task_ref=None,  # <--- Инициализация нового ключа
        mr_triggering_event_for_menu=msg_ctx,
    )

    instruction_text = (
        "<b>Тест умственного вращения</b>\n\n"
        "Вам будет показан 3D объект и 4 варианта 2D проекций. "
        "Выберите номер той проекции, через которую мог бы пройти 3D объект.\n"
        "Объект можно вращать мысленно перед 'просовыванием'.\n\n"
        f"Тест состоит из {MENTAL_ROTATION_NUM_ITERATIONS} заданий."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать тест", callback_data="mr_ack_instructions")]
        ]
    )
    try:
        if isinstance(trigger_event, CallbackQuery):
            await trigger_event.message.edit_text(
                instruction_text, reply_markup=kbd
            )
        else:
            await bot.send_message(chat_id, instruction_text, reply_markup=kbd)
    except TelegramBadRequest as e:
        logger.error(f"MR start: Error sending/editing instructions: {e}")
        await bot.send_message(
            chat_id, "Ошибка при запуске теста. Попробуйте /start."
        )
        await state.clear()


@dp.callback_query(
    F.data.startswith("mr_answer_"),
    MentalRotationStates.displaying_stimulus_mr,
)
async def mr_answer_callback(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("mr_chat_id")

    reaction_time_s = round(
        time.time() - data.get("mr_iteration_start_time", time.time()), 2
    )
    selected_option_num = int(callback.data.split("_")[-1])
    selected_option_idx = selected_option_num - 1
    correct_option_idx = data.get("mr_correct_option_index_for_current_iter")
    is_correct = selected_option_idx == correct_option_idx

    iteration_data = {
        "iteration": data.get("mr_current_iteration"),
        "is_correct": is_correct,
        "reaction_time_s": reaction_time_s,
        "selected_option": selected_option_num,
        "correct_option": (
            correct_option_idx + 1 if correct_option_idx is not None else "N/A"
        ),
    }
    current_results = data.get("mr_iteration_results", [])
    current_results.append(iteration_data)
    await state.update_data(mr_iteration_results=current_results)

    # --- Логика "мигающего" фидбека ---
    feedback_text_bold = f"<b>{'Верно!' if is_correct else 'Неверно!'}</b>"
    feedback_text_normal = f"{'Верно!' if is_correct else 'Неверно!'}"
    feedback_msg_id = data.get("mr_feedback_message_id")

    previous_revert_task = data.get("mr_feedback_revert_task_ref")
    if previous_revert_task and not previous_revert_task.done():
        previous_revert_task.cancel()
        try:
            await asyncio.wait_for(previous_revert_task, timeout=0.05)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    try:
        if feedback_msg_id:
            await bot.edit_message_text(
                text=feedback_text_bold,
                chat_id=chat_id,
                message_id=feedback_msg_id,
                parse_mode=ParseMode.HTML,
            )
        else:
            msg = await bot.send_message(
                chat_id, feedback_text_bold, parse_mode=ParseMode.HTML
            )
            feedback_msg_id = msg.message_id
            await state.update_data(mr_feedback_message_id=feedback_msg_id)

        if feedback_msg_id:
            revert_task = asyncio.create_task(
                _mr_schedule_feedback_revert(
                    chat_id, feedback_msg_id, feedback_text_normal, bot, state
                )
            )
            await state.update_data(mr_feedback_revert_task_ref=revert_task)
        else:
            logger.error(
                "MR feedback: feedback_msg_id is None after send/edit. Cannot schedule revert."
            )

    except TelegramBadRequest as e:
        if "message is not modified" not in str(e).lower():
            logger.error(f"MR feedback (bold set) error: {e}")
    except Exception as e_bold:
        logger.error(
            f"MR feedback (bold set) general error: {e_bold}", exc_info=True
        )

    options_msg_id = data.get("mr_options_message_id")
    if options_msg_id and chat_id:
        try:
            await bot.edit_message_reply_markup(
                chat_id=chat_id, message_id=options_msg_id, reply_markup=None
            )
        except TelegramBadRequest:
            logger.warning(
                f"MR: Failed to remove buttons from options msg_id {options_msg_id}"
            )
            pass

    await state.set_state(MentalRotationStates.processing_answer_mr)

    if chat_id:
        await _mr_proceed_to_next_iteration_or_finish(state, bot, chat_id)
    else:
        logger.error("MR: chat_id is missing after answer. Test might stall.")
        await _finish_mental_rotation_test(
            state, bot, None, is_interrupted=True, error_occurred=True
        )


async def _mr_proceed_to_next_iteration_or_finish(
    state: FSMContext, bot_instance: Bot, chat_id: int
):
    data = await state.get_data()
    current_iteration = data.get("mr_current_iteration", 0)

    if current_iteration < MENTAL_ROTATION_NUM_ITERATIONS:
        await state.set_state(
            MentalRotationStates.inter_iteration_countdown_mr
        )
        countdown_task = asyncio.create_task(
            _mr_inter_iteration_countdown_task(state, bot_instance, chat_id)
        )
        await state.update_data(
            mr_inter_iteration_countdown_task_ref=countdown_task
        )
    else:
        await _finish_mental_rotation_test(
            state, bot_instance, chat_id, is_interrupted=False
        )


async def _mr_inter_iteration_countdown_task(
    state: FSMContext, bot_instance: Bot, chat_id: int
):
    await asyncio.sleep(0.2)

    current_fsm_state = await state.get_state()
    if (
        current_fsm_state
        != MentalRotationStates.inter_iteration_countdown_mr.state
    ):
        logger.info(
            f"MR Countdown: State changed to {current_fsm_state} before countdown could start. Aborting."
        )
        return

    countdown_msg_id_local = None
    try:
        if not chat_id:
            logger.error(
                "MR Countdown: chat_id is missing. Cannot start countdown."
            )
            if (
                await state.get_state()
                == MentalRotationStates.inter_iteration_countdown_mr.state
            ):
                await _finish_mental_rotation_test(
                    state,
                    bot_instance,
                    None,
                    is_interrupted=True,
                    error_occurred=True,
                )
            return

        countdown_msg = await bot_instance.send_message(
            chat_id, "Следующее задание через: 3..."
        )
        countdown_msg_id_local = countdown_msg.message_id
        await state.update_data(mr_countdown_message_id=countdown_msg_id_local)

        for i in range(2, 0, -1):
            await asyncio.sleep(1)
            if (
                await state.get_state()
                != MentalRotationStates.inter_iteration_countdown_mr.state
            ):
                return
            await bot_instance.edit_message_text(
                text=f"Следующее задание через: {i}...",
                chat_id=chat_id,
                message_id=countdown_msg_id_local,
            )

        await asyncio.sleep(1)
        if (
            await state.get_state()
            != MentalRotationStates.inter_iteration_countdown_mr.state
        ):
            return

        if chat_id and countdown_msg_id_local:
            try:
                await bot_instance.delete_message(
                    chat_id, countdown_msg_id_local
                )
            except TelegramBadRequest:
                pass
        await state.update_data(mr_countdown_message_id=None)

        if chat_id:
            await _display_mr_stimulus(
                chat_id, state, bot_instance, is_editing=True
            )
        else:
            logger.error(
                "MR Countdown: chat_id missing before calling _display_mr_stimulus."
            )
            if (
                await state.get_state()
                == MentalRotationStates.inter_iteration_countdown_mr.state
            ):
                await _finish_mental_rotation_test(
                    state,
                    bot_instance,
                    None,
                    is_interrupted=True,
                    error_occurred=True,
                )

    except TelegramBadRequest as e:
        logger.error(f"MR Countdown error (TelegramBadRequest): {e}")
        await asyncio.sleep(0.5)
        if (
            await state.get_state()
            == MentalRotationStates.inter_iteration_countdown_mr.state
        ):
            if chat_id:
                await _display_mr_stimulus(
                    chat_id, state, bot_instance, is_editing=True
                )
            else:
                logger.error("MR Countdown recovery: chat_id missing.")
                await _finish_mental_rotation_test(
                    state,
                    bot_instance,
                    None,
                    is_interrupted=True,
                    error_occurred=True,
                )
    except asyncio.CancelledError:
        logger.info("MR Countdown task cancelled.")
        data_on_cancel = await state.get_data(default={})
        chat_id_on_cancel = data_on_cancel.get("mr_chat_id")
        countdown_msg_id_on_cancel = data_on_cancel.get(
            "mr_countdown_message_id"
        )
        if countdown_msg_id_on_cancel and chat_id_on_cancel:
            try:
                await bot_instance.delete_message(
                    chat_id_on_cancel, countdown_msg_id_on_cancel
                )
            except TelegramBadRequest:
                pass
    except Exception as e:
        logger.error(f"MR Countdown unexpected error: {e}", exc_info=True)
        data_on_exc = await state.get_data(default={})
        chat_id_on_exc = data_on_exc.get("mr_chat_id")
        if (
            await state.get_state()
            == MentalRotationStates.inter_iteration_countdown_mr.state
        ):
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id_on_exc,
                is_interrupted=True,
                error_occurred=True,
            )
    finally:
        await state.update_data(mr_inter_iteration_countdown_task_ref=None)


async def _finish_mental_rotation_test(
    state: FSMContext,
    bot_instance: Bot,
    chat_id: int | None,
    is_interrupted: bool,
    error_occurred: bool = False,
):
    current_fsm_state = await state.get_state()
    if not current_fsm_state or not current_fsm_state.startswith(
        MentalRotationStates.__name__
    ):
        logger.info(
            "MR _finish_test: Called but test not active or already ended."
        )
        return

    logger.info(
        f"Finishing Mental Rotation Test. Interrupted: {is_interrupted}, Error: {error_occurred}"
    )
    data = await state.get_data()

    fsm_chat_id = data.get("mr_chat_id")
    effective_chat_id = fsm_chat_id if fsm_chat_id is not None else chat_id

    # Отменяем задачу "потухания" фидбека
    feedback_revert_task = data.get("mr_feedback_revert_task_ref")
    if feedback_revert_task and not feedback_revert_task.done():
        feedback_revert_task.cancel()
        try:
            await asyncio.wait_for(feedback_revert_task, timeout=0.05)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass
    await state.update_data(mr_feedback_revert_task_ref=None)

    # Отменяем задачу обратного отсчета
    countdown_task = data.get("mr_inter_iteration_countdown_task_ref")
    if countdown_task and not countdown_task.done():
        countdown_task.cancel()
        try:
            await asyncio.wait_for(countdown_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass
    await state.update_data(mr_inter_iteration_countdown_task_ref=None)

    results = data.get("mr_iteration_results", [])
    total_iterations_done = len(results)
    correct_answers = sum(1 for r in results if r["is_correct"])

    total_test_time_s = 0
    if data.get("mr_test_start_time"):
        total_test_time_s = round(
            time.time() - data.get("mr_test_start_time"), 2
        )

    avg_reaction_time_s = 0.0
    correct_reaction_times = [
        r["reaction_time_s"] for r in results if r["is_correct"]
    ]
    if correct_reaction_times:
        avg_reaction_time_s = round(
            sum(correct_reaction_times) / len(correct_reaction_times), 2
        )

    individual_responses_parts = []
    for r in results:
        status = "Прав" if r["is_correct"] else "Неправ"
        rt = f"{r['reaction_time_s']:.2f}с"
        individual_responses_parts.append(f"И{r['iteration']}:{status},{rt}")
    individual_responses_str = (
        "; ".join(individual_responses_parts)
        if individual_responses_parts
        else "N/A"
    )

    await state.update_data(
        mr_final_correct_answers=correct_answers,
        mr_final_avg_reaction_time_s=avg_reaction_time_s,
        mr_final_total_test_time_s=total_test_time_s,
        mr_final_individual_responses_str=individual_responses_str,
        mr_final_interrupted_status=is_interrupted,
    )

    mock_msg_for_save = None
    if effective_chat_id is not None:
        # from aiogram.types import User # Уже должно быть импортировано глобально
        mock_user = User(
            id=bot_instance.id, is_bot=True, first_name=str(bot_instance.id)
        )
        mock_msg_for_save = Message(
            message_id=0,
            date=int(time.time()),
            chat=Chat(id=effective_chat_id, type=ChatType.PRIVATE),
            from_user=mock_user,
        )

    await save_mental_rotation_results(
        mock_msg_for_save, state, is_interrupted=is_interrupted
    )

    final_text_to_user = ""
    if is_interrupted:
        if error_occurred:
            final_text_to_user = (
                "Тест умственного вращения был прерван из-за ошибки."
            )
        else:
            final_text_to_user = "Тест умственного вращения был прерван."
        if results:
            final_text_to_user += f"\nЧастичные результаты сохранены: {correct_answers}/{total_iterations_done} правильных."
    else:
        final_text_to_user = (
            "Тест умственного вращения завершен!\n"
            f"Правильных ответов: {correct_answers}/{MENTAL_ROTATION_NUM_ITERATIONS}\n"
            f"Среднее время реакции на правильные ответы: {avg_reaction_time_s:.2f} сек\n"
            f"Общее время теста: {total_test_time_s:.2f} сек\n"
            # f"Детализация: {individual_responses_str}" # Раскомментировано для вывода результатов
        )

    if effective_chat_id and final_text_to_user:
        try:
            await bot_instance.send_message(
                effective_chat_id,
                final_text_to_user,
                parse_mode=ParseMode.HTML,
            )
        except Exception as e_send_final:
            logger.error(
                f"MR _finish_test: Error sending final results to user: {e_send_final}"
            )

    await cleanup_mental_rotation_ui(state, bot_instance, final_text=None)

    profile_keys = [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]
    profile_data_to_keep = {
        k: data.get(k) for k in profile_keys if data.get(k)
    }

    await state.set_state(None)
    if profile_data_to_keep.get("active_unique_id"):
        await state.set_data(profile_data_to_keep)

        trigger_event_for_menu = data.get("mr_triggering_event_for_menu")
        if not trigger_event_for_menu and mock_msg_for_save:
            trigger_event_for_menu = mock_msg_for_save
        elif not trigger_event_for_menu and effective_chat_id:
            trigger_event_for_menu = Message(
                message_id=0,
                date=int(time.time()),
                chat=Chat(id=effective_chat_id, type=ChatType.PRIVATE),
                from_user=User(
                    id=bot_instance.id,
                    is_bot=True,
                    first_name=str(bot_instance.id),
                ),
            )

        if trigger_event_for_menu:
            await send_main_action_menu(
                trigger_event_for_menu,
                ACTION_SELECTION_KEYBOARD_RETURNING,
                state=state,
            )
        elif effective_chat_id:
            logger.warning(
                "MR _finish_test: trigger_event_for_menu was None, sending menu directly."
            )
            await bot_instance.send_message(
                effective_chat_id,
                "Тест завершен. Выберите действие:",
                reply_markup=ACTION_SELECTION_KEYBOARD_RETURNING,
            )

    elif effective_chat_id:
        await bot_instance.send_message(
            effective_chat_id,
            "Тест завершен. Профиль не найден, пожалуйста /start.",
        )
        await state.clear()
    else:
        await state.clear()


# >>> КОНЕЦ ЗАМЕНЯЕМОГО БЛОКА <<<


async def save_mental_rotation_results(
    trigger_msg_context: Message | None,  # Can be None if called internally
    state: FSMContext,
    is_interrupted: bool = False,
):
    logger.info(
        f"Saving Mental Rotation results. Interrupted: {is_interrupted}"
    )
    data = await state.get_data()

    uid = data.get(
        "mr_unique_id_for_test", data.get("active_unique_id")
    )  # Prioritize test-specific UID
    p_tgid = data.get(
        "mr_profile_telegram_id_for_test", data.get("active_telegram_id")
    )
    p_name = data.get("mr_profile_name_for_test", data.get("active_name"))
    p_age = data.get("mr_profile_age_for_test", data.get("active_age"))

    if not uid:
        logger.error(
            "MR save: UID not found in FSM or active profile. Cannot save."
        )
        if (
            trigger_msg_context
            and hasattr(trigger_msg_context, 'chat')
            and trigger_msg_context.chat
        ):
            await bot.send_message(
                trigger_msg_context.chat.id,
                "MR Тест: Ошибка сохранения (ID пользователя не найден).",
            )
        return

    correct_ans = data.get("mr_final_correct_answers", "N/A")
    avg_rt = data.get("mr_final_avg_reaction_time_s", "N/A")
    total_time = data.get("mr_final_total_test_time_s", "N/A")
    ind_resp_str = data.get("mr_final_individual_responses_str", "N/A")
    # Use interruption status from FSM if available (set by _finish_test), else use arg
    interrupted_status_val = (
        "Да"
        if data.get("mr_final_interrupted_status", is_interrupted)
        else "Нет"
    )

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = r_idx
                break
        if row_num == -1:
            new_row = [""] * len(ALL_EXPECTED_HEADERS)
            new_row[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row[uid_col_idx] = uid
            new_row[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row[ALL_EXPECTED_HEADERS.index("Age")] = p_age if p_age else ""
            ws.append(new_row)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num, column=h.index("MentalRotation_CorrectAnswers") + 1
        ).value = correct_ans
        ws.cell(
            row=row_num,
            column=h.index("MentalRotation_AverageReactionTime_s") + 1,
        ).value = avg_rt
        ws.cell(
            row=row_num, column=h.index("MentalRotation_TotalTime_s") + 1
        ).value = total_time
        ws.cell(
            row=row_num,
            column=h.index("MentalRotation_IndividualResponses") + 1,
        ).value = ind_resp_str
        ws.cell(
            row=row_num, column=h.index("MentalRotation_Interrupted") + 1
        ).value = interrupted_status_val

        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Mental Rotation results for UID {uid} saved. Interrupted: {interrupted_status_val}"
        )

    except Exception as e_save:
        logger.error(
            f"MR results save error for UID {uid}: {e_save}", exc_info=True
        )
        if (
            trigger_msg_context
            and hasattr(trigger_msg_context, 'chat')
            and trigger_msg_context.chat
        ):
            await bot.send_message(
                trigger_msg_context.chat.id,
                "Произошла ошибка при сохранении результатов Теста умственного вращения.",
            )


async def cleanup_mental_rotation_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: (
        str | None
    ) = None,  # If None, no final message sent, just cleanup FSM keys
):
    logger.info(f"Cleaning up Mental Rotation UI. Final text: '{final_text}'")
    data = await state.get_data()
    chat_id = data.get("mr_chat_id")

    # Cancel countdown task if it's somehow still running (should be handled by _finish_test)
    countdown_task = data.get("mr_inter_iteration_countdown_task_ref")
    if countdown_task and not countdown_task.done():
        countdown_task.cancel()
        try:
            await asyncio.wait_for(countdown_task, timeout=0.1)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    msg_ids_to_handle = [  # Collect all potential message IDs
        data.get("mr_reference_message_id"),
        data.get("mr_options_message_id"),
        data.get("mr_countdown_message_id"),
        data.get("mr_feedback_message_id"),
    ]

    # Determine which message to edit with final_text (prefer options or reference message)
    msg_to_edit_with_final_text = data.get(
        "mr_options_message_id"
    ) or data.get("mr_reference_message_id")

    if chat_id:
        # Edit the primary message if final_text is provided
        if msg_to_edit_with_final_text and final_text:
            try:
                # Try editing as photo caption first, then as text
                is_photo = True  # Assume options/reference were photos
                if (
                    is_photo
                ):  # This logic might need refinement if msg_to_edit could be non-photo
                    await bot_instance.edit_message_caption(
                        chat_id=chat_id,
                        message_id=msg_to_edit_with_final_text,
                        caption=final_text,
                        reply_markup=None,
                    )
                else:
                    await bot_instance.edit_message_text(
                        text=final_text,
                        chat_id=chat_id,
                        message_id=msg_to_edit_with_final_text,
                        reply_markup=None,
                    )
            except (
                TelegramBadRequest
            ):  # If edit fails, try deleting and sending new
                try:
                    await bot_instance.delete_message(
                        chat_id, msg_to_edit_with_final_text
                    )
                    await bot_instance.send_message(chat_id, final_text)
                except (
                    TelegramBadRequest
                ):  # If delete also fails, just try sending new
                    try:
                        await bot_instance.send_message(chat_id, final_text)
                    except Exception as e_send:
                        logger.error(
                            f"MR cleanup: Final send failed: {e_send}"
                        )
        elif final_text:  # No primary message to edit, but final_text exists
            try:
                await bot_instance.send_message(chat_id, final_text)
            except Exception as e_send_alt:
                logger.error(
                    f"MR cleanup: Alt final send failed: {e_send_alt}"
                )

        # Delete all other messages associated with the test
        for msg_id in msg_ids_to_handle:
            if msg_id and (
                not final_text or msg_id != msg_to_edit_with_final_text
            ):  # Don't delete if it was just edited
                try:
                    await bot_instance.delete_message(chat_id, msg_id)
                except TelegramBadRequest:
                    pass

                # Clean FSM data
    current_fsm_data = await state.get_data()
    new_data = {
        k: v for k, v in current_fsm_data.items() if not k.startswith("mr_")
    }
    for pk in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if pk in current_fsm_data and pk not in new_data:
            new_data[pk] = current_fsm_data[pk]
    await state.set_data(new_data)
    logger.info("Mental Rotation cleanup: MR-specific FSM data cleaned.")


async def check_if_mental_rotation_results_exist(
    profile_unique_id: str | int,
) -> bool:
    if not profile_unique_id:
        return False
    try:
        uid = int(profile_unique_id)
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        mr_correct_idx = ALL_EXPECTED_HEADERS.index(
            "MentalRotation_CorrectAnswers"
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[uid_col_idx] is not None and int(row[uid_col_idx]) == uid:
                if (
                    mr_correct_idx < len(row)
                    and row[mr_correct_idx] is not None
                ):
                    return True
        return False
    except FileNotFoundError:
        return False
    except ValueError:
        return False  # Header not found
    except Exception as e:
        logger.error(
            f"MR check_results_exist error UID {profile_unique_id}: {e}"
        )
        return False


# START OF RAVEN MATRICES TEST ADDITION (Helper and Test Logic Functions)
def _parse_raven_filename(
    filename: str,
) -> tuple[str | None, int | None, int | None]:
    """Parses Raven image filename (e.g., X_Y_Z.jpg) into (image_id_str, correct_option_1_based, num_total_options)."""
    try:
        name_part = os.path.splitext(filename)[0]  # Remove extension
        parts = name_part.split('_')
        if len(parts) == 3:
            image_id_str = parts[0]  # Can be alphanumeric, treat as string
            correct_option = int(
                parts[1]
            )  # This should be the 1-based correct answer number
            num_options = int(
                parts[2]
            )  # Total number of choices (e.g., 6 or 8)
            if not (1 <= correct_option <= num_options and num_options > 1):
                logger.warning(
                    f"Raven filename {filename} has invalid option numbers: correct={correct_option}, total={num_options}"
                )
                return None, None, None
            return image_id_str, correct_option, num_options
        else:
            logger.warning(
                f"Raven filename {filename} has unexpected format. Expected X_Y_Z.jpg/png, got {len(parts)} parts."
            )
            return None, None, None
    except ValueError:  # If int conversion fails
        logger.warning(
            f"Raven filename {filename} parts (Y or Z) could not be converted to int."
        )
        return None, None, None
    except Exception as e:
        logger.error(f"Error parsing Raven filename {filename}: {e}")
        return None, None, None


async def start_raven_matrices_test(
    trigger_event: [Message, CallbackQuery],
    state: FSMContext,
    profile: dict,
):
    logger.info(
        f"Starting Raven Matrices Test for UID: {profile.get('unique_id')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    if not RAVEN_ALL_TASK_FILES:
        await bot.send_message(
            chat_id,
            "Ошибка: Файлы заданий для Теста Матриц Равена не найдены или не загружены. Тест не может быть запущен.",
        )
        logger.error(
            "Raven Matrices Test: RAVEN_ALL_TASK_FILES is empty. Cannot start test."
        )
        await state.set_state(None)
        active_profile = await get_active_profile_from_fsm(state)
        keyboard = (
            ACTION_SELECTION_KEYBOARD_RETURNING
            if active_profile
            else ACTION_SELECTION_KEYBOARD_NEW
        )
        await send_main_action_menu(msg_ctx, keyboard, state=state)
        return

    num_tasks_for_session = min(
        RAVEN_NUM_TASKS_TO_PRESENT, len(RAVEN_ALL_TASK_FILES)
    )
    if len(RAVEN_ALL_TASK_FILES) < RAVEN_NUM_TASKS_TO_PRESENT:
        logger.warning(
            f"Raven Matrices: Not enough unique tasks ({len(RAVEN_ALL_TASK_FILES)}) "
            f"to meet requirement of {RAVEN_NUM_TASKS_TO_PRESENT}. Using all {num_tasks_for_session} available."
        )

    session_task_filenames = random.sample(
        RAVEN_ALL_TASK_FILES, num_tasks_for_session
    )
    random.shuffle(session_task_filenames)

    await state.set_state(RavenMatricesStates.initial_instructions_raven)
    await state.update_data(
        raven_unique_id_for_test=profile.get("unique_id"),
        raven_profile_name_for_test=profile.get("name"),
        raven_profile_age_for_test=profile.get("age"),
        raven_profile_telegram_id_for_test=profile.get("telegram_id"),
        raven_chat_id=chat_id,
        raven_session_task_filenames=session_task_filenames,
        raven_current_iteration_num=0,
        raven_iteration_results=[],
        raven_total_test_start_time=None,
        raven_current_task_start_time=None,
        raven_task_message_id=None,  # Stores ID of the message showing the task image and buttons
        raven_feedback_message_id=None,  # Stores ID of the separate message for "Верно/Неверно"
        raven_feedback_blinker_task=None,
        raven_triggering_event_for_menu=msg_ctx,  # Store original event for menu navigation
    )

    instruction_text = (
        "<b>Тест Прогрессивных Матриц Равена</b>\n\n"
        "Вам будет показана матрица с пропущенным элементом и несколько вариантов для его заполнения. "
        "Ваша задача - выбрать наиболее подходящий вариант, чтобы завершить матрицу, следуя логике ее построения.\n\n"
        f"Тест состоит из {num_tasks_for_session} заданий. "
        "Постарайтесь отвечать не только правильно, но и быстро. Время ответа учитывается."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать тест", callback_data="raven_ack_instructions")]
        ]
    )
    try:
        if isinstance(trigger_event, CallbackQuery):
            await trigger_event.message.edit_text(
                instruction_text, reply_markup=kbd
            )
        else:
            await bot.send_message(chat_id, instruction_text, reply_markup=kbd)
    except TelegramBadRequest as e:
        logger.error(f"Raven start: Error sending/editing instructions: {e}")
        await bot.send_message(
            chat_id, "Ошибка при запуске теста. Попробуйте /start."
        )
        await state.clear()


@dp.callback_query(
    F.data == "raven_ack_instructions",
    RavenMatricesStates.initial_instructions_raven,
)
async def raven_ack_instructions_callback(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer()
    await state.update_data(
        raven_total_test_start_time=time.time()
    )  # Start total test timer

    try:
        await callback.message.delete()  # Delete the instruction message
    except TelegramBadRequest:
        pass

    await _display_raven_task(callback.message.chat.id, state, bot)


async def _display_raven_task(
    chat_id: int, state: FSMContext, bot_instance: Bot
):
    data = await state.get_data()
    current_iter_idx = data.get(
        "raven_current_iteration_num", 0
    )  # 0-indexed for list access
    session_tasks = data.get("raven_session_task_filenames", [])

    if current_iter_idx >= len(session_tasks):
        logger.info(
            "Raven: All tasks displayed or index out of bounds. Finishing up."
        )
        await _finish_raven_matrices_test(
            state, bot_instance, chat_id, is_interrupted=False
        )  # Normal completion
        return

    task_filename_only = session_tasks[current_iter_idx]
    task_image_full_path = os.path.join(RAVEN_BASE_DIR, task_filename_only)

    _, correct_option_num_1_based, num_total_options = _parse_raven_filename(
        task_filename_only
    )

    if (
        not os.path.exists(task_image_full_path)
        or correct_option_num_1_based is None
        or num_total_options is None
    ):
        logger.error(
            f"Raven: Invalid task file or parsing error for '{task_filename_only}' at path '{task_image_full_path}'"
        )
        await bot_instance.send_message(
            chat_id,
            f"Ошибка загрузки задания {current_iter_idx + 1}. Тест будет прерван.",
        )
        await _finish_raven_matrices_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    await state.update_data(
        raven_correct_option_for_current_task=correct_option_num_1_based,  # Store 1-based
        raven_num_options_for_current_task=num_total_options,
        raven_current_task_filename=task_filename_only,  # Store just filename for records
    )

    # --- Generate Buttons ---
    buttons_row = []
    buttons_grid = []
    # Flexible button layout: try to fit 3 per row, or 4 if 8 options, else whatever fits
    buttons_per_row = 3
    if num_total_options == 8:
        buttons_per_row = 4
    elif num_total_options == 4:
        buttons_per_row = 2  # For 2x2
    elif num_total_options == 2:
        buttons_per_row = 2

    for i in range(1, num_total_options + 1):
        buttons_row.append(IKB(text=str(i), callback_data=f"raven_answer_{i}"))
        if len(buttons_row) == buttons_per_row or i == num_total_options:
            buttons_grid.append(list(buttons_row))
            buttons_row.clear()
    if buttons_row:  # Should be empty if logic is correct
        buttons_grid.append(list(buttons_row))
    reply_markup = InlineKeyboardMarkup(inline_keyboard=buttons_grid)

    task_message_id = data.get("raven_task_message_id")
    caption_text = f"Задание {current_iter_idx + 1} из {len(session_tasks)}"

    try:
        if task_message_id:  # Edit existing message for subsequent tasks
            media = InputMediaPhoto(
                media=FSInputFile(task_image_full_path), caption=caption_text
            )
            await bot_instance.edit_message_media(
                chat_id=chat_id,
                message_id=task_message_id,
                media=media,
                reply_markup=reply_markup,
            )
        else:  # Send new message for the first task
            msg = await bot_instance.send_photo(
                chat_id=chat_id,
                photo=FSInputFile(task_image_full_path),
                caption=caption_text,
                reply_markup=reply_markup,
            )
            await state.update_data(raven_task_message_id=msg.message_id)
    except (TelegramBadRequest, FileNotFoundError) as e:
        logger.error(
            f"Raven: Error sending/editing task image '{task_filename_only}': {e}"
        )
        await bot_instance.send_message(
            chat_id, "Ошибка отображения задания. Тест прерван."
        )
        await _finish_raven_matrices_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    await state.update_data(
        raven_current_task_start_time=time.time()
    )  # Start timer for this specific task
    await state.set_state(RavenMatricesStates.displaying_task_raven)


async def _raven_delayed_feedback_revert(
    chat_id: int,
    message_id: int,
    normal_text: str,
    bot_instance: Bot,
    state_at_call: State,  # Передаем состояние на момент вызова
):
    """Асинхронно изменяет текст сообщения на обычный (нежирный) после задержки."""
    try:
        await asyncio.sleep(
            RAVEN_FEEDBACK_DISPLAY_TIME_S
        )  # RAVEN_FEEDBACK_DISPLAY_TIME_S = 0.75

        # Проверяем, актуально ли еще это сообщение для фидбека.
        # Это важно, чтобы не изменить сообщение, если тест был прерван,
        # или пользователь уже ответил на следующий вопрос и feedback_message_id обновился.
        current_fsm_data = (
            await state_at_call.get_data()
        )  # Используем состояние, переданное при вызове
        if current_fsm_data.get("raven_feedback_message_id") == message_id:
            await bot_instance.edit_message_text(
                text=normal_text,
                chat_id=chat_id,
                message_id=message_id,
                parse_mode=None,  # Обычный текст
            )
        else:
            logger.info(
                f"Raven delayed feedback revert (msg {message_id}): Message ID is no longer current. Skipping revert."
            )

    except asyncio.CancelledError:
        logger.info(
            f"Raven delayed feedback revert task for msg {message_id} cancelled."
        )
        raise
    except TelegramBadRequest as e:
        if (
            "message to edit not found" in str(e).lower()
            or "message is not modified" in str(e).lower()
        ):
            logger.info(
                f"Raven delayed feedback revert (msg {message_id}): Edit failed gracefully ({e})"
            )
        else:
            logger.warning(
                f"Raven delayed feedback revert (msg {message_id}): TelegramBadRequest: {e}"
            )
    except Exception as e:
        logger.error(
            f"Raven delayed feedback revert (msg {message_id}): Unexpected error: {e}",
            exc_info=True,
        )


@dp.callback_query(
    F.data.startswith("raven_answer_"),
    RavenMatricesStates.displaying_task_raven,
)
async def handle_raven_answer_callback(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("raven_chat_id")

    task_start_time = data.get("raven_current_task_start_time", time.time())
    reaction_time_s = round(time.time() - task_start_time, 2)

    user_choice_str = callback.data.split("raven_answer_")[-1]
    user_choice_num_1_based = int(user_choice_str)

    correct_option_1_based = data.get("raven_correct_option_for_current_task")
    is_correct = user_choice_num_1_based == correct_option_1_based

    current_task_filename = data.get("raven_current_task_filename", "N/A")

    iteration_result = {
        "task_filename": current_task_filename,
        "user_choice": user_choice_num_1_based,
        "correct_answer_number": correct_option_1_based,
        "is_correct": is_correct,
        "reaction_time_s": reaction_time_s,
    }
    current_results = data.get("raven_iteration_results", [])
    current_results.append(iteration_result)

    await state.update_data(raven_iteration_results=current_results)

    # --- Логика "мигающего" фидбека ---
    feedback_text_bold = f"<b>{'Верно!' if is_correct else 'Неверно!'}</b>"
    feedback_text_normal = f"{'Верно!' if is_correct else 'Неверно!'}"

    feedback_msg_id = data.get("raven_feedback_message_id")

    # Отменяем предыдущую задачу "потухания", если она была и еще не выполнилась.
    previous_revert_task = data.get(
        "raven_current_feedback_revert_task_ref"
    )  # Используем новое имя ключа
    if previous_revert_task and not previous_revert_task.done():
        previous_revert_task.cancel()
        try:
            await asyncio.wait_for(previous_revert_task, timeout=0.05)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    try:
        if feedback_msg_id:
            await bot.edit_message_text(
                text=feedback_text_bold,
                chat_id=chat_id,
                message_id=feedback_msg_id,
                parse_mode=ParseMode.HTML,
            )
        else:
            msg = await bot.send_message(
                chat_id, feedback_text_bold, parse_mode=ParseMode.HTML
            )
            feedback_msg_id = msg.message_id
            await state.update_data(raven_feedback_message_id=feedback_msg_id)

        if feedback_msg_id:
            # Передаем 'state' в _schedule_raven_feedback_revert для проверки актуальности сообщения
            revert_task = asyncio.create_task(
                _raven_delayed_feedback_revert(
                    chat_id, feedback_msg_id, feedback_text_normal, bot, state
                )
            )
            await state.update_data(
                raven_current_feedback_revert_task_ref=revert_task
            )  # Сохраняем ссылку
        else:
            logger.error(
                "Raven feedback: feedback_msg_id is None, cannot schedule revert."
            )

    except TelegramBadRequest as e:
        if "message is not modified" not in str(e).lower():
            logger.error(f"Raven feedback (initial bold) display error: {e}")
        try:
            if feedback_msg_id:
                await bot.edit_message_text(
                    text=feedback_text_normal,
                    chat_id=chat_id,
                    message_id=feedback_msg_id,
                )
            else:
                msg = await bot.send_message(chat_id, feedback_text_normal)
                await state.update_data(
                    raven_feedback_message_id=msg.message_id
                )
        except Exception as fallback_e:
            logger.error(
                f"Raven feedback (initial bold) fallback display error: {fallback_e}"
            )
    except Exception as e:
        logger.error(
            f"Raven feedback - general error setting bold text: {e}",
            exc_info=True,
        )

    # --- Продолжаем основную логику теста НЕ ОЖИДАЯ завершения "мигания" ---
    current_iteration_idx_0_based = data.get("raven_current_iteration_num", 0)
    session_tasks_total = len(data.get("raven_session_task_filenames", []))

    next_iteration_idx_0_based = current_iteration_idx_0_based + 1
    await state.update_data(
        raven_current_iteration_num=next_iteration_idx_0_based
    )

    if next_iteration_idx_0_based < session_tasks_total:
        await _display_raven_task(chat_id, state, bot)
    else:
        total_test_end_time = time.time()
        await state.update_data(
            raven_total_test_end_time_actual=total_test_end_time
        )

        logger.info("Raven Matrices Test: All iterations completed.")
        await _finish_raven_matrices_test(
            state, bot, chat_id, is_interrupted=False
        )


async def _finish_raven_matrices_test(
    state: FSMContext,
    bot_instance: Bot,
    chat_id: int | None,
    is_interrupted: bool,
    error_occurred: bool = False,
):
    current_fsm_state = await state.get_state()
    if not current_fsm_state or not current_fsm_state.startswith(
        RavenMatricesStates.__name__
    ):
        logger.info(
            "Raven _finish_test: Called but test not active or already ended."
        )
        return

    logger.info(
        f"Finishing Raven Matrices Test. Interrupted: {is_interrupted}, Error: {error_occurred}"
    )
    data = await state.get_data()

    fsm_chat_id = data.get("raven_chat_id")
    effective_chat_id = fsm_chat_id if fsm_chat_id is not None else chat_id

    # Отменяем последнюю задачу "потухания", если она была запущена и тест завершается/прерывается
    revert_task = data.get(
        "raven_current_feedback_revert_task_ref"
    )  # Используем новое имя ключа
    if revert_task and not revert_task.done():
        revert_task.cancel()
        try:
            await asyncio.wait_for(revert_task, timeout=0.05)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass
    await state.update_data(
        raven_current_feedback_revert_task_ref=None
    )  # Очищаем ссылку

    # ... (остальная часть _finish_raven_matrices_test без изменений, как в вашем полном коде)
    iteration_results = data.get("raven_iteration_results", [])
    total_tasks_presented_in_results = len(iteration_results)
    correct_answers_count = sum(
        1 for r in iteration_results if r["is_correct"]
    )

    total_test_time_s = 0.0
    test_start_time = data.get("raven_total_test_start_time")
    test_end_time_actual = data.get(
        "raven_total_test_end_time_actual", time.time()
    )

    if test_start_time:
        total_test_time_s = round(test_end_time_actual - test_start_time, 2)

    individual_times_s_list = [r["reaction_time_s"] for r in iteration_results]
    individual_times_s_str = (
        ", ".join(map(lambda x: f"{x:.2f}", individual_times_s_list))
        if individual_times_s_list
        else "N/A"
    )

    correct_reaction_times = [
        r["reaction_time_s"] for r in iteration_results if r["is_correct"]
    ]
    avg_time_correct_s = 0.0
    if correct_reaction_times:
        avg_time_correct_s = round(
            sum(correct_reaction_times) / len(correct_reaction_times), 2
        )

    await state.update_data(
        raven_final_correct_answers=correct_answers_count,
        raven_final_total_test_time_s=total_test_time_s,
        raven_final_avg_time_correct_s=avg_time_correct_s,
        raven_final_individual_times_s_str=individual_times_s_str,
        raven_final_interrupted_status=is_interrupted,
        raven_final_total_tasks_attempted=total_tasks_presented_in_results,
    )

    mock_msg_for_save = None
    if effective_chat_id is not None:
        from aiogram.types import User  # Убедимся, что User импортирован

        mock_user = User(id=bot_instance.id, is_bot=True, first_name="Bot")
        mock_msg_for_save = Message(
            message_id=0,
            date=int(time.time()),
            chat=Chat(id=effective_chat_id, type=ChatType.PRIVATE),
            from_user=mock_user,
        )

    await save_raven_matrices_results(
        mock_msg_for_save, state, is_interrupted=is_interrupted
    )

    final_text_to_user = ""
    num_tasks_in_session_config = len(
        data.get("raven_session_task_filenames", [])
    )

    if effective_chat_id:
        if is_interrupted:
            if error_occurred:
                final_text_to_user = "Тест Прогрессивных Матриц Равена был прерван из-за ошибки."
            else:
                final_text_to_user = (
                    "Тест Прогрессивных Матриц Равена был прерван."
                )
            if iteration_results:
                final_text_to_user += f"\nЧастичные результаты (за {total_tasks_presented_in_results} заданий) сохранены: {correct_answers_count} правильных."
        else:
            final_text_to_user = (
                "Тест Прогрессивных Матриц Равена завершен!\n"
                f"Правильных ответов: {correct_answers_count} из {num_tasks_in_session_config}.\n"
                f"Общее время: {total_test_time_s:.2f} сек.\n"
            )
            if correct_answers_count > 0:
                final_text_to_user += f"Среднее время на правильный ответ: {avg_time_correct_s:.2f} сек."
            else:
                final_text_to_user += (
                    "Правильных ответов не было, среднее время не рассчитано."
                )

        await cleanup_raven_ui(
            state, bot_instance, final_text=final_text_to_user
        )
    else:
        logger.warning(
            "Raven _finish_test: No effective_chat_id available to send final summary."
        )
        await cleanup_raven_ui(state, bot_instance, final_text=None)

    profile_keys = [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]
    profile_data_to_keep = {
        k: data.get(k) for k in profile_keys if data.get(k)
    }

    await state.set_state(None)
    if profile_data_to_keep.get("active_unique_id"):
        await state.set_data(profile_data_to_keep)

        trigger_event_for_menu = data.get("raven_triggering_event_for_menu")
        if not trigger_event_for_menu and effective_chat_id is not None:
            trigger_event_for_menu = mock_msg_for_save

        if trigger_event_for_menu:
            await send_main_action_menu(
                trigger_event_for_menu,
                ACTION_SELECTION_KEYBOARD_RETURNING,
                state=state,
            )
        elif effective_chat_id is not None:
            await bot_instance.send_message(
                effective_chat_id,
                "Тест завершен. Выберите действие:",
                reply_markup=ACTION_SELECTION_KEYBOARD_RETURNING,
            )
    elif effective_chat_id is not None:
        await bot_instance.send_message(
            effective_chat_id,
            "Тест завершен. Профиль не найден, пожалуйста /start.",
        )
        await state.clear()
    else:
        await state.clear()


async def save_raven_matrices_results(
    trigger_msg_context: Message | None,
    state: FSMContext,
    is_interrupted: bool = False,
):
    logger.info(
        f"Saving Raven Matrices results. Interrupted: {is_interrupted}"
    )
    data = await state.get_data()

    uid = data.get("raven_unique_id_for_test", data.get("active_unique_id"))
    p_tgid = data.get(
        "raven_profile_telegram_id_for_test", data.get("active_telegram_id")
    )
    p_name = data.get("raven_profile_name_for_test", data.get("active_name"))
    p_age = data.get("raven_profile_age_for_test", data.get("active_age"))

    if not uid:
        logger.error("Raven save: UID not found. Cannot save.")
        if (
            trigger_msg_context
            and hasattr(trigger_msg_context, 'chat')
            and trigger_msg_context.chat
        ):
            await bot.send_message(
                trigger_msg_context.chat.id,
                "Тест Матриц Равена: Ошибка сохранения (ID не найден).",
            )
        return

    correct_ans = data.get("raven_final_correct_answers", "N/A")
    total_time = data.get("raven_final_total_test_time_s", "N/A")
    avg_rt_correct = data.get("raven_final_avg_time_correct_s", "N/A")
    ind_times_str = data.get("raven_final_individual_times_s_str", "N/A")
    interrupted_status_val = (
        "Да"
        if data.get("raven_final_interrupted_status", is_interrupted)
        else "Нет"
    )

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")

        for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = r_idx
                break

        if row_num == -1:
            new_row = [""] * len(ALL_EXPECTED_HEADERS)
            new_row[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row[uid_col_idx] = uid
            new_row[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row[ALL_EXPECTED_HEADERS.index("Age")] = p_age if p_age else ""
            ws.append(new_row)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num, column=h.index("RavenMatrices_CorrectAnswers") + 1
        ).value = correct_ans
        ws.cell(
            row=row_num, column=h.index("RavenMatrices_TotalTime_s") + 1
        ).value = total_time
        ws.cell(
            row=row_num, column=h.index("RavenMatrices_AvgTimeCorrect_s") + 1
        ).value = avg_rt_correct
        ws.cell(
            row=row_num, column=h.index("RavenMatrices_IndividualTimes_s") + 1
        ).value = ind_times_str
        ws.cell(
            row=row_num, column=h.index("RavenMatrices_Interrupted") + 1
        ).value = interrupted_status_val

        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Raven Matrices results for UID {uid} saved. Interrupted: {interrupted_status_val}"
        )

    except Exception as e_save:
        logger.error(
            f"Raven Matrices results save error for UID {uid}: {e_save}",
            exc_info=True,
        )
        if (
            trigger_msg_context
            and hasattr(trigger_msg_context, 'chat')
            and trigger_msg_context.chat
        ):
            await bot.send_message(
                trigger_msg_context.chat.id,
                "Ошибка сохранения Теста Матриц Равена.",
            )


async def cleanup_raven_ui(
    state: FSMContext, bot_instance: Bot, final_text: str | None = None
):
    logger.info(f"Cleaning up Raven Matrices UI. Final text: '{final_text}'")
    data = await state.get_data()
    chat_id = data.get("raven_chat_id")

    blinker_task = data.get("raven_feedback_blinker_task")
    if blinker_task and not blinker_task.done():
        blinker_task.cancel()

    task_msg_id = data.get("raven_task_message_id")
    feedback_msg_id = data.get("raven_feedback_message_id")

    if chat_id:
        if task_msg_id:
            if final_text:
                try:
                    await bot_instance.delete_message(
                        chat_id=chat_id, message_id=task_msg_id
                    )
                    await asyncio.sleep(0.3)
                    await bot_instance.send_message(
                        chat_id=chat_id,
                        text=final_text,
                        reply_markup=None,
                    )
                except TelegramBadRequest:
                    try:  # Fallback to delete and send new if edit caption fails
                        await bot_instance.delete_message(chat_id, task_msg_id)
                        await bot_instance.send_message(chat_id, final_text)
                    except (
                        TelegramBadRequest
                    ):  # If delete also fails, just try sending new
                        try:
                            await bot_instance.send_message(
                                chat_id, final_text
                            )
                        except Exception as e_send_f:
                            logger.error(
                                f"Raven cleanup: Final send failed: {e_send_f}"
                            )
            else:
                try:
                    await bot_instance.delete_message(chat_id, task_msg_id)
                except TelegramBadRequest:
                    pass
        elif final_text:
            try:
                await bot_instance.send_message(chat_id, final_text)
            except Exception as e:
                logger.error(f"Raven cleanup: Failed to send final text: {e}")

        if feedback_msg_id:
            try:
                await bot_instance.delete_message(chat_id, feedback_msg_id)
            except TelegramBadRequest:
                pass

    current_fsm_data = await state.get_data()
    new_data = {
        k: v for k, v in current_fsm_data.items() if not k.startswith("raven_")
    }
    for pk in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if pk in current_fsm_data and pk not in new_data:
            new_data[pk] = current_fsm_data[pk]
    await state.set_data(new_data)
    logger.info("Raven Matrices cleanup: Raven-specific FSM data cleaned.")


async def check_if_raven_matrices_results_exist(
    profile_unique_id: str | int,
) -> bool:
    if not profile_unique_id:
        return False
    try:
        uid = int(profile_unique_id)
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        raven_correct_idx = ALL_EXPECTED_HEADERS.index(
            "RavenMatrices_CorrectAnswers"
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[uid_col_idx] is not None and str(row[uid_col_idx]) == str(
                uid
            ):
                if (
                    raven_correct_idx < len(row)
                    and row[raven_correct_idx] is not None
                ):
                    return True
        return False
    except FileNotFoundError:
        return False
    except ValueError:
        return False
    except Exception as e:
        logger.error(
            f"Raven check_results_exist for UID {profile_unique_id}: {e}"
        )
        return False


# END OF RAVEN MATRICES TEST ADDITION


# --- Test Registry, Stoptest, Selection, Registration, Utils ---
async def check_if_corsi_results_exist(
    profile_unique_id: str | int,
) -> bool:
    if not profile_unique_id:
        return False
    try:
        uid = int(profile_unique_id)
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        max_len_idx = ALL_EXPECTED_HEADERS.index(
            "Corsi - Max Correct Sequence Length"
        )
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[uid_col_idx] is not None and int(row[uid_col_idx]) == uid:
                if max_len_idx < len(row) and row[max_len_idx] is not None:
                    return True
        return False
    except FileNotFoundError:
        return False
    except ValueError:  # Header not found or UID conversion error
        return False
    except Exception:  # Catch other potential errors during file access
        return False


async def check_if_stroop_results_exist(
    profile_unique_id: str | int,
) -> bool:
    if not profile_unique_id:
        return False
    try:
        uid = int(profile_unique_id)
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        p1_time_idx = ALL_EXPECTED_HEADERS.index("Stroop Part1 Time (s)")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[uid_col_idx] is not None and int(row[uid_col_idx]) == uid:
                if p1_time_idx < len(row) and row[p1_time_idx] is not None:
                    return True
        return False
    except FileNotFoundError:
        return False
    except ValueError:
        return False
    except Exception:
        return False


TEST_REGISTRY = {
    "initiate_corsi_test": {
        "name": "Тест Корси",
        "fsm_group_class": CorsiTestStates,
        "start_function": start_corsi_test,
        "save_function": save_corsi_results,
        "cleanup_function": cleanup_corsi_messages,
        "results_exist_check": check_if_corsi_results_exist,
        "requires_active_profile": True,
    },
    "initiate_stroop_test": {
        "name": "Тест Струпа",
        "fsm_group_class": StroopTestStates,
        "start_function": start_stroop_test,
        "save_function": save_stroop_results,
        "cleanup_function": cleanup_stroop_ui,
        "results_exist_check": check_if_stroop_results_exist,
        "requires_active_profile": True,
    },
    "initiate_reaction_time_test": {
        "name": "Тест на Скорость Реакции",
        "fsm_group_class": ReactionTimeTestStates,
        "start_function": start_reaction_time_test,
        "save_function": save_reaction_time_results,
        "cleanup_function": cleanup_reaction_time_ui,
        "results_exist_check": check_if_reaction_time_results_exist,
        "requires_active_profile": True,
        "end_test_function": _rt_go_to_main_menu_or_clear,
    },
    "initiate_verbal_fluency_test": {
        "name": "Тест на вербальную беглость",
        "fsm_group_class": VerbalFluencyStates,
        "start_function": start_verbal_fluency_test,
        "save_function": save_verbal_fluency_results,
        "cleanup_function": cleanup_verbal_fluency_ui,
        "results_exist_check": check_if_verbal_fluency_results_exist,
        "requires_active_profile": True,
        "end_test_function": _end_verbal_fluency_test,
    },
    "initiate_mental_rotation_test": {
        "name": "Тест умственного вращения",
        "fsm_group_class": MentalRotationStates,
        "start_function": start_mental_rotation_test,
        "save_function": save_mental_rotation_results,
        "cleanup_function": cleanup_mental_rotation_ui,
        "results_exist_check": check_if_mental_rotation_results_exist,
        "requires_active_profile": True,
        "end_test_function": _finish_mental_rotation_test,
    },
    # START OF RAVEN MATRICES TEST ADDITION (Test Registry)
    "initiate_raven_matrices_test": {
        "name": "Прогрессивные матрицы Равена",
        "fsm_group_class": RavenMatricesStates,
        "start_function": start_raven_matrices_test,
        "save_function": save_raven_matrices_results,
        "cleanup_function": cleanup_raven_ui,
        "results_exist_check": check_if_raven_matrices_results_exist,
        "requires_active_profile": True,
        "end_test_function": _finish_raven_matrices_test,
    },
    # END OF RAVEN MATRICES TEST ADDITION (Test Registry)
}


@dp.message(Command("stoptest"))
async def stop_test_command_handler(
    message: Message, state: FSMContext, called_from_test_button: bool = False
):
    fsm_state_str = await state.get_state()
    active_test_cfg = None
    active_test_key = None

    if fsm_state_str:
        for key, cfg in TEST_REGISTRY.items():
            if cfg.get("fsm_group_class") and fsm_state_str.startswith(
                cfg["fsm_group_class"].__name__
            ):
                active_test_cfg = cfg
                active_test_key = key
                break

    if active_test_cfg:
        test_name = active_test_cfg["name"]
        trigger_event_obj = message  # /stoptest is always a Message

        if (
            not called_from_test_button
        ):  # Don't send if stopped from test's own button
            await trigger_event_obj.answer(
                f"Останавливаю тест: {test_name}..."
            )

        if active_test_cfg.get("end_test_function"):
            logger.info(
                f"Stoptest: Calling specific end_test_function for {test_name} (key: {active_test_key})"
            )
            chat_id_for_end = (
                trigger_event_obj.chat.id
                if hasattr(trigger_event_obj, 'chat')
                else None
            )

            # Adjust parameters based on the specific end_test_function's signature
            if active_test_key in [
                "initiate_mental_rotation_test",
                "initiate_raven_matrices_test",
            ]:
                # These take (state, bot, chat_id, is_interrupted, error_occurred)
                await active_test_cfg["end_test_function"](
                    state,
                    bot,
                    chat_id_for_end,
                    is_interrupted=True,
                    error_occurred=False,
                )
            elif active_test_key == "initiate_verbal_fluency_test":
                # Takes (state, bot, interrupted, trigger_event)
                await active_test_cfg["end_test_function"](
                    state,
                    bot,
                    interrupted=True,
                    trigger_event=trigger_event_obj,
                )
            elif active_test_key == "initiate_reaction_time_test":
                # RT's end_test_function is _rt_go_to_main_menu_or_clear, which needs different handling
                # First, save with interruption, then cleanup, then call its specific end.
                await save_reaction_time_results(
                    state,
                    is_interrupted=True,
                    status_override="Interrupted by user",
                )
                await cleanup_reaction_time_ui(
                    state, bot, f"Тест '{test_name}' был прерван."
                )
                await _rt_go_to_main_menu_or_clear(state, trigger_event_obj)
                return  # Return early as RT test handles its own menu navigation
            else:  # Generic call for other tests if they adopt a similar signature
                # For Corsi, Stroop if they were to have an end_test_function
                # Assuming a signature like (state, bot, chat_id, is_interrupted) for hypothetical example
                # This branch might need adjustment if other tests get end_test_function with different signatures.
                logger.warning(
                    f"Stoptest: end_test_function called for {test_name} with generic signature, review if correct."
                )
                await active_test_cfg["end_test_function"](
                    state, bot, chat_id_for_end, is_interrupted=True
                )

        else:  # Generic handling if no specific end_test_function (e.g. Corsi, Stroop)
            logger.info(
                f"Stoptest: Using generic save/cleanup for {test_name}"
            )
            save_func = active_test_cfg.get("save_function")
            cleanup_func = active_test_cfg.get("cleanup_function")

            if (
                save_func
            ):  # Reaction Time save is different, already handled if it had end_test_function
                await save_func(trigger_event_obj, state, is_interrupted=True)

            if cleanup_func:
                await cleanup_func(
                    state, bot, final_text=f"Тест '{test_name}' был прерван."
                )

            # Generic menu return after generic cleanup (if not handled by end_test_function)
            await state.set_state(None)
            active_profile = await get_active_profile_from_fsm(state)
            if active_profile:
                await state.set_data(
                    {  # Ensure profile data is preserved
                        "active_unique_id": active_profile["unique_id"],
                        "active_name": active_profile["name"],
                        "active_age": active_profile["age"],
                        "active_telegram_id": active_profile["telegram_id"],
                    }
                )
                await send_main_action_menu(
                    trigger_event_obj,
                    ACTION_SELECTION_KEYBOARD_RETURNING,
                    state=state,
                )
            else:
                await trigger_event_obj.answer(
                    "Тест остановлен. Профиль не активен. Пожалуйста, /start."
                )
                await state.clear()

    elif not called_from_test_button:  # No active test found by /stoptest
        await message.answer(
            "Нет активного теста для остановки. Пожалуйста, /start для начала."
        )


@dp.callback_query(F.data == "select_specific_test")
async def on_select_specific_test_callback(
    cb: CallbackQuery, state: FSMContext
):
    profile = await get_active_profile_from_fsm(state)
    if not profile:
        await cb.answer(
            "Профиль не активен. Пожалуйста, /start.", show_alert=True
        )
        try:
            await cb.message.delete()  # Clean up the menu message
        except TelegramBadRequest:
            pass
        return

    btns = [
        [IKB(text=cfg["name"], callback_data=f"select_test_{key}")]
        for key, cfg in TEST_REGISTRY.items()
        if cfg.get(
            "requires_active_profile", True
        )  # Most tests will require this
    ]
    if not btns:
        await cb.message.edit_text("Нет доступных тестов.", reply_markup=None)
        await cb.answer()
        return

    await cb.answer()
    try:
        await cb.message.edit_text(
            "Выберите тест:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
        )
    except TelegramBadRequest:  # If message was deleted or something
        await cb.message.answer(  # Send new
            "Выберите тест:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
        )


@dp.callback_query(F.data.startswith("select_test_"))
async def on_test_selected_callback(cb: CallbackQuery, state: FSMContext):
    test_key_selected = cb.data.replace("select_test_", "")
    if test_key_selected not in TEST_REGISTRY:
        await cb.answer("Выбранный тест не найден.", show_alert=True)
        return

    cfg = TEST_REGISTRY[test_key_selected]
    profile = await get_active_profile_from_fsm(state)

    if not profile and cfg.get("requires_active_profile"):
        await cb.answer(
            "Нужен активный профиль для этого теста. Пожалуйста, /start.",
            show_alert=True,
        )
        try:
            await cb.message.delete()
        except TelegramBadRequest:
            pass
        return

    await cb.answer()  # Acknowledge selection
    await state.update_data(pending_test_key_for_overwrite=test_key_selected)

    results_already_exist = False
    if profile and cfg.get("results_exist_check"):
        results_already_exist = await cfg["results_exist_check"](
            profile.get("unique_id")
        )

    if results_already_exist:
        kbd = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    IKB(
                        text="Да, перезаписать",
                        callback_data="confirm_overwrite_test_results",
                    )
                ],
                [
                    IKB(
                        text="Нет, отмена",
                        callback_data="cancel_overwrite_test_results",
                    )
                ],
            ]
        )
        txt = f"У вас уже есть сохраненные результаты для теста '{cfg['name']}'. Хотите перезаписать их?"
        try:
            msg = await cb.message.edit_text(txt, reply_markup=kbd)
        except TelegramBadRequest:  # Message might have been deleted
            msg = await cb.message.answer(txt, reply_markup=kbd)  # Send new
        await state.update_data(
            overwrite_confirmation_message_id=msg.message_id
        )
        await state.set_state(UserData.waiting_for_test_overwrite_confirmation)
    else:  # No existing results or no check function, proceed to start
        if cb.message:  # Edit the "Select test" message before starting test
            try:
                await cb.message.edit_text(
                    f"Подготовка к тесту: {cfg['name']}...", reply_markup=None
                )
            except TelegramBadRequest:
                pass  # If edit fails, just proceed
        await cfg["start_function"](cb, state, profile)


@dp.callback_query(
    F.data == "confirm_overwrite_test_results",
    UserData.waiting_for_test_overwrite_confirmation,
)
async def handle_confirm_overwrite_test_results(
    cb: CallbackQuery, state: FSMContext
):
    data = await state.get_data()
    key = data.get("pending_test_key_for_overwrite")
    msg_id = data.get("overwrite_confirmation_message_id")

    if not key or key not in TEST_REGISTRY:
        await cb.answer("Ошибка: тест не определен.", show_alert=True)
        await state.set_state(None)
        if msg_id:
            try:
                await bot.delete_message(cb.message.chat.id, msg_id)
            except TelegramBadRequest:
                pass
        active_profile = await get_active_profile_from_fsm(state)
        kb = (
            ACTION_SELECTION_KEYBOARD_RETURNING
            if active_profile
            else ACTION_SELECTION_KEYBOARD_NEW
        )
        await send_main_action_menu(cb, kb, state=state)
        return

    await cb.answer()
    cfg = TEST_REGISTRY[key]
    profile = await get_active_profile_from_fsm(
        state
    )  # Re-fetch profile just in case
    if not profile and cfg.get(
        "requires_active_profile"
    ):  # Should not happen if checks are consistent
        await cb.message.answer("Профиль не активен. Пожалуйста, /start.")
        await state.set_state(None)
        return

    if msg_id:  # Edit "Overwrite?" message to "Starting..."
        try:
            await bot.edit_message_text(
                text=f"Запускаем тест: {cfg['name']} (перезапись)...",
                chat_id=cb.message.chat.id,
                message_id=msg_id,
                reply_markup=None,
            )
        except TelegramBadRequest:
            pass
    await state.update_data(  # Clear these specific keys from FSM
        overwrite_confirmation_message_id=None,
        pending_test_key_for_overwrite=None,
    )
    # The UserData state will be cleared by the test's start_function implicitly or explicitly.
    # If start_function doesn't clear/set a new state, we might need to clear UserData here.
    # For now, assuming start_function handles state transition.
    await cfg["start_function"](cb, state, profile)


@dp.callback_query(
    F.data == "cancel_overwrite_test_results",
    UserData.waiting_for_test_overwrite_confirmation,
)
async def handle_cancel_overwrite_test_results(
    cb: CallbackQuery, state: FSMContext
):
    await cb.answer("Запуск теста отменен.", show_alert=False)
    data = await state.get_data()
    key = data.get("pending_test_key_for_overwrite")
    msg_id = data.get("overwrite_confirmation_message_id")
    name = (
        TEST_REGISTRY[key]["name"] if key and key in TEST_REGISTRY else "теста"
    )

    if msg_id:
        try:
            await bot.edit_message_text(
                text=f"Запуск теста {name} отменен.",
                chat_id=cb.message.chat.id,
                message_id=msg_id,
                reply_markup=None,
            )
        except TelegramBadRequest:  # Message might be deleted
            await cb.message.answer(
                f"Запуск теста {name} отменен."
            )  # Send new confirmation

    await state.update_data(
        overwrite_confirmation_message_id=None,
        pending_test_key_for_overwrite=None,
    )
    await state.set_state(None)  # Clear UserData state
    await send_main_action_menu(
        cb, ACTION_SELECTION_KEYBOARD_RETURNING, state=state
    )  # Back to main menu


@dp.message(CommandStart())
async def start_command_handler(message: Message, state: FSMContext):
    await state.clear()  # Clear any previous state completely on /start
    await state.set_state(UserData.waiting_for_first_time_response)
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Да (новая регистрация)", callback_data="user_is_new")],
            [
                IKB(
                    text="Нет (вход по моему UID)",
                    callback_data="user_is_returning",
                )
            ],
        ]
    )
    await message.answer(
        "Здравствуйте! Вы впервые пользуетесь этим ботом для тестирования?",
        reply_markup=kbd,
    )


@dp.callback_query(
    F.data == "user_is_new", UserData.waiting_for_first_time_response
)
async def handle_user_is_new_callback(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    await cb.message.edit_reply_markup(reply_markup=None)  # Remove buttons
    await state.set_state(UserData.waiting_for_name)
    await cb.message.answer("Рад знакомству! Пожалуйста, введите ваше имя:")


@dp.callback_query(
    F.data == "user_is_returning", UserData.waiting_for_first_time_response
)
async def handle_user_is_returning_callback(
    cb: CallbackQuery, state: FSMContext
):
    await cb.answer()
    await cb.message.edit_reply_markup(reply_markup=None)  # Remove buttons
    await state.set_state(UserData.waiting_for_unique_id)
    await cb.message.answer(
        "Пожалуйста, введите ваш Уникальный Идентификатор (UID):"
    )


@dp.message(UserData.waiting_for_name)
async def process_name_input(message: Message, state: FSMContext):
    name = message.text.strip() if message.text else ""
    if not name or len(name) < 2:  # Basic validation
        await message.answer(
            "Имя не может быть пустым и должно содержать хотя бы 2 символа. Попробуйте еще раз."
        )
        return
    await state.update_data(name_for_registration=name)
    await state.set_state(UserData.waiting_for_age)
    await message.answer(f"Приятно, {name}! Введите ваш возраст (полных лет):")


@dp.message(UserData.waiting_for_age)
async def process_age_input(message: Message, state: FSMContext):
    age_str = message.text.strip() if message.text else ""
    if not age_str.isdigit() or not (
        0 < int(age_str) < 120
    ):  # Basic age validation
        await message.answer(
            "Пожалуйста, введите корректный возраст (например, 25)."
        )
        return

    data = await state.get_data()
    name = data.get("name_for_registration")
    age = int(age_str)
    tgid = message.from_user.id
    new_uid = None

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        existing_ids = {
            r[uid_col_idx]
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[uid_col_idx] is not None and isinstance(r[uid_col_idx], int)
        }

        min_uid, max_uid = 1000000, 9999999  # Define UID range
        if len(existing_ids) >= (
            max_uid - min_uid + 1
        ):  # Check if UID space is exhausted
            logger.critical("UIDs exhausted. Cannot register new user.")
            await message.answer(
                "Ошибка системы UID. Обратитесь к администратору."
            )
            await state.clear()
            return

        attempts = 0
        while (
            attempts < 1000
        ):  # Try to find a unique ID, prevent infinite loop
            candidate_uid = random.randint(min_uid, max_uid)
            if candidate_uid not in existing_ids:
                new_uid = candidate_uid
                break
            attempts += 1
        else:  # Failed after many attempts
            logger.critical(
                f"Failed to generate unique UID for TGID {tgid} after {attempts} tries."
            )
            await message.answer(
                "Ошибка генерации уникального идентификатора. Попробуйте позже."
            )
            await state.clear()
            return

        # Create new row with all headers, filling in known data
        new_row_template = [""] * len(ALL_EXPECTED_HEADERS)
        new_row_template[ALL_EXPECTED_HEADERS.index("Telegram ID")] = tgid
        new_row_template[uid_col_idx] = new_uid
        new_row_template[ALL_EXPECTED_HEADERS.index("Name")] = name
        new_row_template[ALL_EXPECTED_HEADERS.index("Age")] = age
        ws.append(new_row_template)
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"New user registered: TGID {tgid}, UID {new_uid}, Name '{name}', Age {age}"
        )

        profile_data = {
            "active_telegram_id": tgid,
            "active_unique_id": new_uid,
            "active_name": name,
            "active_age": age,
        }
        await state.set_data(profile_data)  # Set active profile in FSM
        await state.set_state(None)  # Clear registration states
        await message.answer(
            f"Спасибо, {name}! Регистрация прошла успешно.\n<b>Ваш Уникальный Идентификатор (UID): {new_uid}</b>\n"
            "Пожалуйста, сохраните его для будущего входа.",
            parse_mode=ParseMode.HTML,
        )
        await send_main_action_menu(
            message, ACTION_SELECTION_KEYBOARD_NEW, state=state
        )  # Show menu for new user

    except Exception as e:
        logger.error(
            f"Registration error for TGID {tgid}, Name '{name}': {e}",
            exc_info=True,
        )
        await message.answer(
            "Произошла ошибка во время регистрации. Пожалуйста, попробуйте /start еще раз."
        )
        await state.clear()


@dp.message(UserData.waiting_for_unique_id)
async def process_unique_id_input(message: Message, state: FSMContext):
    uid_str = message.text.strip() if message.text else ""
    if not uid_str.isdigit():
        await message.answer(
            "UID должен быть числом. Пожалуйста, попробуйте снова."
        )
        return

    uid_val = int(uid_str)
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        found_profile = None

        # Get column indices by header name for robustness
        tg_id_col = ALL_EXPECTED_HEADERS.index("Telegram ID")
        uid_col = ALL_EXPECTED_HEADERS.index("Unique ID")
        name_col = ALL_EXPECTED_HEADERS.index("Name")
        age_col = ALL_EXPECTED_HEADERS.index("Age")

        for row_values in ws.iter_rows(
            min_row=2, values_only=True
        ):  # Start from row 2 (skip headers)
            # Ensure UID column has a value and matches the input UID
            if row_values[uid_col] is not None and str(
                row_values[uid_col]
            ) == str(uid_val):
                found_profile = {
                    "active_unique_id": uid_val,
                    "active_telegram_id": row_values[
                        tg_id_col
                    ],  # Can be None if not recorded
                    "active_name": str(row_values[name_col]),
                    "active_age": str(
                        row_values[age_col]
                    ),  # Store as string, consistent with input
                }
                break  # Found the profile

        if found_profile:
            await state.set_data(found_profile)  # Load profile into FSM
            await state.set_state(None)  # Clear registration states
            await message.answer(
                f"С возвращением, {found_profile['active_name']}!"
            )
            await send_main_action_menu(
                message, ACTION_SELECTION_KEYBOARD_RETURNING, state=state
            )
        else:  # UID not found
            kbd = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        IKB(
                            text="Ввести UID снова",
                            callback_data="try_id_again",
                        )
                    ],
                    [
                        IKB(
                            text="Новая регистрация",
                            callback_data="register_new_after_fail",
                        )
                    ],
                ]
            )
            await message.answer(
                "UID не найден. Пожалуйста, проверьте правильность ввода или зарегистрируйтесь.",
                reply_markup=kbd,
            )
            # State remains UserData.waiting_for_unique_id for "try_id_again" or changes on "register_new_after_fail"

    except FileNotFoundError:
        logger.error(
            f"Excel file '{EXCEL_FILENAME}' not found during UID lookup for '{uid_str}'."
        )
        await message.answer(
            "Ошибка: Файл данных не найден. Свяжитесь с администратором."
        )
        await state.clear()
    except Exception as e:
        logger.error(
            f"Error during UID lookup for '{uid_str}': {e}", exc_info=True
        )
        await message.answer(
            "Произошла ошибка при проверке UID. Пожалуйста, попробуйте позже."
        )
        await state.clear()


@dp.callback_query(F.data == "try_id_again", UserData.waiting_for_unique_id)
async def handle_try_id_again_callback(cb: CallbackQuery, state: FSMContext):
    await cb.answer()
    try:
        await cb.message.edit_reply_markup(
            reply_markup=None
        )  # Clean up buttons
    except TelegramBadRequest:
        pass
    await cb.message.answer("Введите ваш UID еще раз:")
    # State remains UserData.waiting_for_unique_id


@dp.callback_query(
    F.data == "register_new_after_fail", UserData.waiting_for_unique_id
)
async def handle_register_new_after_fail_callback(
    cb: CallbackQuery, state: FSMContext
):
    await cb.answer()
    try:
        await cb.message.edit_reply_markup(
            reply_markup=None
        )  # Clean up buttons
    except TelegramBadRequest:
        pass
    await state.set_state(
        UserData.waiting_for_name
    )  # Switch to name input for new registration
    await cb.message.answer("Хорошо, давайте зарегистрируемся. Как вас зовут?")


@dp.message(Command("mydata"))
async def show_my_data_command(message: Message, state: FSMContext):
    profile = await get_active_profile_from_fsm(state)
    if not profile:
        await message.answer(
            "Профиль не активен. Пожалуйста, /start для входа или регистрации."
        )
        return

    uid = profile.get("unique_id")
    lines = [
        f"Данные для UID: <b>{uid}</b> (Имя: {profile.get('name', 'N/A')}, Возраст: {profile.get('age', 'N/A')})"
    ]

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        found_in_excel = False
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")

        for row_idx, row_tuple in enumerate(
            ws.iter_rows(min_row=1, values_only=True)
        ):  # Read header too
            if row_idx == 0:  # Header row
                excel_headers = list(row_tuple)
                # Update uid_col_idx if "Unique ID" position changed (shouldn't if ALL_EXPECTED_HEADERS is source of truth)
                if "Unique ID" in excel_headers:
                    uid_col_idx = excel_headers.index("Unique ID")
                else:  # Critical if Unique ID header is missing
                    logger.error(
                        "mydata: 'Unique ID' header missing in Excel."
                    )
                    lines.append(
                        "Ошибка: 'Unique ID' столбец не найден в файле Excel."
                    )
                    break
                continue  # Skip to data rows

            # Data rows (row_idx > 0)
            if row_tuple[uid_col_idx] is not None and str(
                row_tuple[uid_col_idx]
            ) == str(uid):
                found_in_excel = True
                lines.append("--- Результаты тестов из файла ---")
                for i, header_name in enumerate(excel_headers):
                    # Skip base profile data already shown from FSM, unless it's different or to confirm
                    if (
                        header_name in BASE_HEADERS
                        and header_name != "Unique ID"
                    ):
                        continue

                    val_from_excel = (
                        row_tuple[i] if i < len(row_tuple) else None
                    )
                    display_val = (
                        val_from_excel
                        if val_from_excel is not None
                        else "нет данных"
                    )

                    if (
                        isinstance(val_from_excel, float)
                        and val_from_excel.is_integer()
                    ):
                        display_val = int(
                            val_from_excel
                        )  # Display whole floats as int

                    if (
                        val_from_excel is not None
                    ):  # Only show if there's actual data for this header
                        lines.append(f"<b>{header_name}:</b> {display_val}")
                break  # Found the user, no need to check further rows

        if (
            not found_in_excel and row_idx > 0
        ):  # Searched data rows but not found
            lines.append(
                "Профиль с таким UID не найден в файле Excel (хотя активен в сессии)."
            )
        elif (
            not found_in_excel and row_idx == 0 and "Ошибка" not in lines[-1]
        ):  # File only had header or was empty
            lines.append("Файл Excel пуст или содержит только заголовки.")

    except FileNotFoundError:
        lines.append(f"Файл данных '{EXCEL_FILENAME}' не найден.")
    except ValueError as ve:  # e.g. "Unique ID" not in ALL_EXPECTED_HEADERS
        lines.append(f"Ошибка конфигурации заголовков: {ve}")
        logger.error(f"mydata command configuration error: {ve}")
    except Exception as e:
        lines.append(f"Произошла ошибка при загрузке данных из Excel: {e}")
        logger.error(f"mydata command error for UID {uid}: {e}")

    await message.answer("\n".join(lines), parse_mode=ParseMode.HTML)


@dp.message(Command("export"))
async def export_data_to_excel_command(message: Message, state: FSMContext):
    # No specific role check for now, assuming any user can request it
    if os.path.exists(EXCEL_FILENAME):
        try:
            await message.reply_document(
                FSInputFile(EXCEL_FILENAME),
                caption="База данных пользователей и результатов.",
            )
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            await message.answer("Не удалось отправить файл Excel.")
    else:
        await message.answer(
            f"Файл данных '{EXCEL_FILENAME}' не найден на сервере."
        )


@dp.message(Command("restart"))
async def command_restart_bot_session_handler(
    message: Message, state: FSMContext
):
    fsm_state_str = await state.get_state()
    active_test_cfg = None
    active_test_key_for_restart = None

    if fsm_state_str:  # Check if any test is active
        for key_iter, cfg_iter in TEST_REGISTRY.items():
            if cfg_iter.get("fsm_group_class") and fsm_state_str.startswith(
                cfg_iter["fsm_group_class"].__name__
            ):
                active_test_cfg = cfg_iter
                active_test_key_for_restart = key_iter
                break

    if active_test_cfg:
        test_name_for_restart = active_test_cfg.get("name", "активный тест")
        logger.info(
            f"/restart called during active test: {test_name_for_restart} (key: {active_test_key_for_restart})"
        )

        # Try to use the specific end_test_function for graceful cleanup if available
        if active_test_cfg.get("end_test_function"):
            logger.info(
                f"/restart: Using end_test_function for {test_name_for_restart}"
            )
            chat_id_for_restart_end = message.chat.id
            # Call end_test_function with interruption=True
            # Signatures vary, so this part needs careful handling per test
            if active_test_key_for_restart in [
                "initiate_mental_rotation_test",
                "initiate_raven_matrices_test",
            ]:
                await active_test_cfg["end_test_function"](
                    state,
                    bot,
                    chat_id_for_restart_end,
                    is_interrupted=True,
                    error_occurred=False,
                )
            elif active_test_key_for_restart == "initiate_verbal_fluency_test":
                await active_test_cfg["end_test_function"](
                    state, bot, interrupted=True, trigger_event=message
                )
            elif active_test_key_for_restart == "initiate_reaction_time_test":
                # For RT, save/cleanup then call its specific menu func
                await save_reaction_time_results(
                    state,
                    is_interrupted=True,
                    status_override="Interrupted by /restart",
                )
                await cleanup_reaction_time_ui(
                    state,
                    bot,
                    f"Тест '{test_name_for_restart}' принудительно остановлен.",
                )
                # _rt_go_to_main_menu_or_clear will clear RT state and attempt menu, but /restart clears all below
            # Add other specific end_test_function calls here if their signatures differ

        elif active_test_cfg.get(
            "cleanup_function"
        ):  # Fallback to generic cleanup
            logger.info(
                f"/restart: Using generic cleanup_function for {test_name_for_restart}"
            )
            await active_test_cfg["cleanup_function"](
                state,
                bot,
                final_text=f"Тест '{test_name_for_restart}' принудительно остановлен командой /restart.",
            )
        # After specific or generic cleanup (which should clear test-specific state but might keep profile)

    # Perform full clear for /restart, including profile data from FSM
    await state.clear()
    await message.answer(
        "Все текущие операции были остановлены, ваш профиль и состояние теста в этой сессии сброшены.\n"
        "Пожалуйста, используйте команду /start для нового сеанса или входа."
    )


@dp.callback_query(F.data == "logout_profile")
async def logout_profile_callback(cb: CallbackQuery, state: FSMContext):
    await cb.answer(
        "Ваш профиль был сброшен из текущей сессии.", show_alert=True
    )
    await state.clear()  # Clear all FSM data including active profile
    try:
        await cb.message.edit_text(
            "Профиль сброшен. Используйте /start для нового входа или регистрации.",
            reply_markup=None,
        )
    except TelegramBadRequest:  # If message already gone or cannot be edited
        await cb.message.answer(  # Send new message
            "Профиль сброшен. Используйте /start для нового входа или регистрации."
        )


@dp.callback_query(F.data == "run_test_battery")
async def on_run_test_battery_callback(cb: CallbackQuery, state: FSMContext):
    # This feature is not yet implemented as per original code structure
    await cb.answer(
        "Функция 'Пройти батарею тестов' находится в разработке.",
        show_alert=True,
    )


# --- Verbal Fluency Test Handlers ---
@dp.callback_query(
    F.data == "vf_start_test_confirmed",
    VerbalFluencyStates.showing_instructions_and_task,
)
async def handle_verbal_fluency_start_ack(
    callback: CallbackQuery, state: FSMContext
):
    await callback.answer()
    data = await state.get_data()
    task_msg_id = data.get("vf_task_message_id")
    task_letter = data.get("vf_task_letter")
    chat_id = data.get("vf_chat_id")

    if not all([task_msg_id, task_letter, chat_id]):
        logger.error("VF: Missing critical data in FSM for start_ack.")
        if chat_id:
            await bot.send_message(
                chat_id,
                "Произошла ошибка при запуске теста. Пожалуйста, /start.",
            )
        await state.clear()  # Or navigate to menu
        return

    task_text = (
        f"Задание: Назовите как можно больше слов, начинающихся на букву <b>'{task_letter}'</b>.\n"
        f"Осталось: {VERBAL_FLUENCY_DURATION_S} сек.\n\nВводите слова."
    )
    current_task_msg_id = (
        task_msg_id  # Use the ID stored from instruction phase
    )
    try:
        await bot.edit_message_text(
            text=task_text,
            chat_id=chat_id,
            message_id=current_task_msg_id,
            reply_markup=None,
            parse_mode=ParseMode.HTML,
        )
    except (
        TelegramBadRequest
    ) as e:  # If original message was deleted or cannot be edited
        logger.error(
            f"VF: Failed to edit task message {current_task_msg_id}: {e}. Sending new."
        )
        try:
            new_msg = await bot.send_message(
                chat_id, task_text, parse_mode=ParseMode.HTML
            )
            await state.update_data(
                vf_task_message_id=new_msg.message_id
            )  # Update with new message ID
            current_task_msg_id = new_msg.message_id  # Use new ID for pinning
        except Exception as send_e:  # If sending new also fails
            logger.critical(
                f"VF: Critical error - failed to send new task message: {send_e}"
            )
            await bot.send_message(
                chat_id,
                "Критическая ошибка отображения задания. Тест прерван.",
            )
            await _end_verbal_fluency_test(
                state, bot, interrupted=True, trigger_event=callback.message
            )
            return

    # Pin the task message
    if (
        current_task_msg_id and chat_id
    ):  # Ensure we have a valid message ID and chat ID
        try:
            await bot.pin_chat_message(
                chat_id=chat_id,
                message_id=current_task_msg_id,
                disable_notification=True,
            )
        except TelegramBadRequest as pin_e:
            logger.error(
                f"VF: Failed to pin task message {current_task_msg_id}: {pin_e}"
            )

    await state.set_state(VerbalFluencyStates.collecting_words)
    timer_task = asyncio.create_task(_verbal_fluency_timer_task(state, bot))
    await state.update_data(vf_timer_task=timer_task)


@dp.message(VerbalFluencyStates.collecting_words, F.text)
async def handle_verbal_fluency_word_input(
    message: Message, state: FSMContext
):
    data = await state.get_data()
    task_letter = data.get(
        "vf_task_letter", ""
    ).lower()  # Ensure lowercase for comparison
    collected_words_set = data.get("vf_collected_words", set())

    if not task_letter:  # Should not happen if test started correctly
        await message.reply(
            "Ошибка: буква для задания не определена. Пожалуйста, сообщите администратору."
        )
        return

    user_words_raw = (
        message.text.lower().split()
    )  # Split by space, convert to lower
    newly_added_count = 0
    for word in user_words_raw:
        # Basic cleaning: remove common punctuation, ensure it's alphabetic (or allow hyphenated if needed)
        processed_word = ''.join(
            filter(str.isalpha, word)
        )  # Keeps only letters
        if len(processed_word) >= 2 and processed_word.startswith(task_letter):
            if processed_word not in collected_words_set:
                collected_words_set.add(processed_word)
                newly_added_count += 1

    if newly_added_count > 0:
        await state.update_data(vf_collected_words=collected_words_set)
        # Optionally, give feedback like " принято X слов" but can be noisy.
        # await message.reply(f"Принято новых слов: {newly_added_count}. Всего: {len(collected_words_set)}.")


# ... (весь остальной код выше этой части остается НЕИЗМЕННЫМ) ...


async def main():
    print("DEBUG: Entered main() function.")  # Отладочный print
    logger.info("DEBUG LOGGER: Entered main() function.")  # Отладочный logger

    if Image is None:
        logger.error(
            "Библиотека Pillow (PIL) не установлена. "
            "Генерация изображений для Теста Струпа (части 2 и 3), "
            "Теста умственного вращения и временных изображений для Теста скорости реакции "
            "и Теста Матриц Равена (если используются заглушки) будет невозможна или ограничена."
        )

    logger.info("DEBUG LOGGER: Calling initialize_excel_file().")
    initialize_excel_file()
    logger.info("DEBUG LOGGER: Finished initialize_excel_file().")

    logger.info(
        "Запуск бота (DEBUG LOGGER: About to populate MR distractors)..."
    )

    # Populate MR distractors (Mental Rotation Test)
    if os.path.isdir(MR_DISTRACTORS_DIR):
        for f_name in os.listdir(MR_DISTRACTORS_DIR):
            if f_name.lower().endswith((".png", ".jpg", ".jpeg")):
                full_path = os.path.join(MR_DISTRACTORS_DIR, f_name)
                if os.path.isfile(full_path):
                    MR_ALL_DISTRACTORS_FILES.append(full_path)
        if MR_ALL_DISTRACTORS_FILES:
            logger.info(
                f"Loaded {len(MR_ALL_DISTRACTORS_FILES)} MR distractor images."
            )
        else:
            logger.warning(
                f"No MR distractor images found in {MR_DISTRACTORS_DIR}."
            )
    else:
        logger.warning(
            f"MR distractor directory not found: {MR_DISTRACTORS_DIR}. Test may not work."
        )
    logger.info("DEBUG LOGGER: Finished populating MR distractors.")

    # START OF RAVEN MATRICES TEST ADDITION (Populate task files)
    logger.info("DEBUG LOGGER: About to populate Raven tasks.")
    if os.path.isdir(RAVEN_BASE_DIR):
        for f_name in os.listdir(RAVEN_BASE_DIR):
            if f_name.lower().endswith((".png", ".jpg", ".jpeg")):
                _, correct_opt, num_opts = _parse_raven_filename(f_name)
                if correct_opt is not None and num_opts is not None:
                    RAVEN_ALL_TASK_FILES.append(f_name)
        if RAVEN_ALL_TASK_FILES:
            logger.info(
                f"Loaded {len(RAVEN_ALL_TASK_FILES)} valid Raven Matrices task files."
            )
        else:
            logger.warning(
                f"No valid Raven Matrices task files (e.g., X_Y_Z.jpg) found in {RAVEN_BASE_DIR}."
            )
    else:
        logger.warning(
            f"Raven Matrices base directory not found: {RAVEN_BASE_DIR}. This test will not be available."
        )
    logger.info("DEBUG LOGGER: Finished populating Raven tasks.")

    # Register Corsi handlers that are not using decorators
    dp.callback_query.register(
        handle_corsi_button_press,
        F.data.startswith("corsi_button_"),
        CorsiTestStates.waiting_for_user_sequence,
    )
    dp.callback_query.register(
        on_corsi_restart_current_test,
        F.data == "corsi_stop_this_attempt",
        StateFilter(CorsiTestStates),
    )

    logger.info("DEBUG LOGGER: About to delete webhook.")
    await bot.delete_webhook(drop_pending_updates=True)
    logger.info("DEBUG LOGGER: Webhook deleted. About to start polling.")
    print("DEBUG: About to start polling...")

    try:
        await dp.start_polling(bot)
        logger.info(
            "DEBUG LOGGER: start_polling has finished (this usually means bot stopped)."
        )  # Это сообщение вряд ли появится при нормальной работе
        print("DEBUG: start_polling has finished.")
    except Exception as e_poll:
        logger.critical(
            f"CRITICAL ERROR during start_polling: {e_poll}", exc_info=True
        )
        print(f"CRITICAL ERROR during start_polling: {e_poll}")


if __name__ == "__main__":
    print(
        "DEBUG: Script execution started in __main__."
    )  # Самое первое сообщение
    # Ensure base images directory exists
    if not os.path.exists("images"):
        os.makedirs("images")
        logger.info(
            "Создана папка 'images'."
        )  # Это должно быть видно, если logging работает

    # Ensure RT dummy images exist or are created if Pillow is available
    # This part is mostly for testing RT if real images are not present.
    if not os.path.exists("images/rt_img_1.png"):
        for i in range(1, 11):  # Create 10 dummy images
            dummy_image_path = f"images/rt_img_{i}.png"
            if not os.path.exists(dummy_image_path):
                try:
                    if (
                        Image and ImageDraw and ImageFont
                    ):  # Check if Pillow is available
                        img = Image.new(
                            'RGB',
                            (100, 100),
                            color=(
                                random.randint(50, 200),
                                random.randint(50, 200),
                                random.randint(50, 200),
                            ),
                        )
                        draw = ImageDraw.Draw(img)
                        try:
                            font = ImageFont.truetype("arial.ttf", 30)
                        except IOError:
                            font = ImageFont.load_default()
                        draw.text(
                            (10, 10), f"RT {i}", fill=(0, 0, 0), font=font
                        )
                        img.save(dummy_image_path)
                        logger.info(
                            f"Создан RT файл-заглушка: {dummy_image_path}"
                        )
                    else:  # Pillow not available, create empty file as placeholder
                        with open(dummy_image_path, "w") as f:
                            f.write("")
                        logger.info(
                            f"Создан пустой RT файл-заглушка: {dummy_image_path} (Pillow недоступна)."
                        )
                except Exception as e_img:
                    logger.error(
                        f"Не удалось создать RT файл-заглушку {dummy_image_path}: {e_img}"
                    )

    # Mental Rotation Test directory checks and .gitkeep creation
    for mr_dir_path_item in [
        MR_REFERENCES_DIR,
        MR_CORRECT_PROJECTIONS_DIR,
        MR_DISTRACTORS_DIR,
    ]:
        if not os.path.isdir(mr_dir_path_item):
            logger.warning(
                f"Директория для Теста умственного вращения не найдена: {mr_dir_path_item}. "
                "Тест может не работать корректно. Пожалуйста, создайте и наполните её соответствующими изображениями."
            )
            try:
                os.makedirs(mr_dir_path_item, exist_ok=True)
                logger.info(f"Создана директория: {mr_dir_path_item}")
            except OSError as e_mkdir:
                logger.error(
                    f"Не удалось создать директорию {mr_dir_path_item}: {e_mkdir}"
                )

        if os.path.isdir(mr_dir_path_item) and not os.listdir(
            mr_dir_path_item
        ):
            gitkeep_path = os.path.join(mr_dir_path_item, ".gitkeep")
            if not os.path.exists(gitkeep_path):
                try:
                    with open(gitkeep_path, "w") as gk_f:
                        gk_f.write("")
                    logger.info(
                        f"Created .gitkeep in empty MR directory: {mr_dir_path_item}"
                    )
                except Exception as e_gk:
                    logger.warning(
                        f"Could not create .gitkeep in {mr_dir_path_item}: {e_gk}"
                    )

    # RAVEN MATRICES TEST (Directory check and .gitkeep)
    if not os.path.isdir(RAVEN_BASE_DIR):
        logger.warning(
            f"Директория для Теста Матриц Равена не найдена: {RAVEN_BASE_DIR}. "
            "Тест не будет доступен. Пожалуйста, создайте директорию и наполните её файлами изображений "
            "в формате 'НомерЗадания_НомерПравильногоОтвета_КоличествоВариантов.расширение' (например, 1_2_6.jpg)."
        )
        try:
            os.makedirs(RAVEN_BASE_DIR, exist_ok=True)
            logger.info(
                f"Создана директория для Матриц Равена: {RAVEN_BASE_DIR}"
            )
        except OSError as e_mkdir_raven:
            logger.error(
                f"Не удалось создать директорию для Матриц Равена {RAVEN_BASE_DIR}: {e_mkdir_raven}"
            )

    if os.path.isdir(RAVEN_BASE_DIR) and not os.listdir(RAVEN_BASE_DIR):
        raven_gitkeep_path = os.path.join(RAVEN_BASE_DIR, ".gitkeep")
        if not os.path.exists(raven_gitkeep_path):
            try:
                with open(raven_gitkeep_path, "w") as gk_f_raven:
                    gk_f_raven.write("")
                logger.info(
                    f"Created .gitkeep in empty Raven Matrices directory: {RAVEN_BASE_DIR}"
                )
            except Exception as e_gk_raven:
                logger.warning(
                    f"Could not create .gitkeep in {RAVEN_BASE_DIR}: {e_gk_raven}"
                )

    print("DEBUG: Pre-asyncio.run(main()) checks completed.")
    logger.info(
        "DEBUG LOGGER: Pre-asyncio.run(main()) checks completed. Calling main()."
    )

    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logging.info("Бот остановлен вручную (KeyboardInterrupt).")
        print("INFO: Bot stopped manually (KeyboardInterrupt).")
    except Exception as e_global:
        logging.critical(
            f"Критическая ошибка при выполнении бота: {e_global}",
            exc_info=True,
        )
        print(f"CRITICAL: Bot execution failed: {e_global}")
    finally:
        print("DEBUG: Script finished in __main__.")
        logger.info("DEBUG LOGGER: Script finished in __main__.")
