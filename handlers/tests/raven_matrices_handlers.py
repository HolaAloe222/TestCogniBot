# handlers/tests/raven_matrices_handlers.py
import asyncio
import logging
import random
import time
import os

from aiogram import Bot, F, Router
from aiogram.enums import ParseMode, ChatType
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    FSInputFile,
    InputMediaPhoto,
    Message,
    Chat,
    User,
)

from fsm_states import RavenMatricesStates
from settings import (
    ALL_EXPECTED_HEADERS,
    EXCEL_FILENAME,
    RAVEN_NUM_TASKS_TO_PRESENT,
    RAVEN_ALL_TASK_FILES,
    RAVEN_BASE_DIR,
    RAVEN_FEEDBACK_DISPLAY_TIME_S,
)
from utils.bot_helpers import (
    send_main_action_menu,
    get_active_profile_from_fsm,
)
from keyboards import (
    ACTION_SELECTION_KEYBOARD_RETURNING,
    ACTION_SELECTION_KEYBOARD_NEW,  # Included for completeness
)

logger = logging.getLogger(__name__)
router = Router()
IKB = InlineKeyboardButton


def _parse_raven_filename(
    filename: str,
) -> tuple[str | None, int | None, int | None]:
    try:
        name_part = os.path.splitext(filename)[0]
        parts = name_part.split('_')
        if len(parts) == 3:
            image_id_str, correct_option_str, num_options_str = parts
            correct_option = int(correct_option_str)
            num_options = int(num_options_str)
            if not (1 <= correct_option <= num_options and num_options > 1):
                logger.warning(
                    f"Raven filename {filename} invalid numbers: correct={correct_option}, total={num_options}"
                )
                return None, None, None
            return image_id_str, correct_option, num_options
        else:
            logger.warning(
                f"Raven filename {filename} unexpected format. Got {len(parts)} parts."
            )
            return None, None, None
    except ValueError:
        logger.warning(
            f"Raven filename {filename} parts not convertible to int."
        )
        return None, None, None
    except Exception as e:
        logger.error(f"Error parsing Raven filename {filename}: {e}")
        return None, None, None


async def _raven_delayed_feedback_revert(
    chat_id: int,
    message_id: int,
    normal_text: str,
    bot_instance: Bot,
    state_at_call: FSMContext,
):
    try:
        await asyncio.sleep(RAVEN_FEEDBACK_DISPLAY_TIME_S)
        current_fsm_data = await state_at_call.get_data()
        if (
            current_fsm_data.get("raven_feedback_message_id") == message_id
            and await state_at_call.get_state()
            is not None  # Check if state is still active
            and (await state_at_call.get_state()).startswith(
                RavenMatricesStates.__name__
            )
        ):  # Check if still in Raven test
            try:
                await bot_instance.edit_message_text(
                    text=normal_text,
                    chat_id=chat_id,
                    message_id=message_id,
                    parse_mode=None,
                )
            except TelegramBadRequest as e_edit:
                if (
                    "message is not modified" not in str(e_edit).lower()
                    and "message to edit not found" not in str(e_edit).lower()
                ):
                    logger.warning(
                        f"Raven delayed feedback (msg {message_id}): Edit failed: {e_edit}"
                    )
            except Exception as e_gen_edit:
                logger.error(
                    f"Raven delayed feedback (msg {message_id}): General error on edit: {e_gen_edit}",
                    exc_info=True,
                )
        else:
            logger.info(
                f"Raven delayed feedback (msg {message_id}): State/msg_id changed or test ended. Skipping revert."
            )
    except asyncio.CancelledError:
        logger.info(
            f"Raven delayed feedback task for msg {message_id} cancelled."
        )
    except Exception as e:
        logger.error(
            f"Raven delayed feedback (msg {message_id}): Unexpected error in task: {e}",
            exc_info=True,
        )


async def _display_raven_task(
    chat_id: int, state: FSMContext, bot_instance: Bot
):
    data = await state.get_data()
    current_iter_idx = data.get("raven_current_iteration_num", 0)
    session_tasks = data.get("raven_session_task_filenames", [])

    if current_iter_idx >= len(session_tasks):
        logger.info(
            "Raven: All tasks displayed. Finishing test via _display_raven_task."
        )
        await _finish_raven_matrices_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=False,
            error_occurred=False,
        )
        return

    task_filename_only = session_tasks[current_iter_idx]
    task_image_full_path = os.path.join(RAVEN_BASE_DIR, task_filename_only)
    _, correct_option_1_based, num_total_options = _parse_raven_filename(
        task_filename_only
    )

    if (
        not os.path.exists(task_image_full_path)
        or correct_option_1_based is None
        or num_total_options is None
    ):
        logger.error(
            f"Raven: Invalid task file or parsing error for: '{task_filename_only}'"
        )
        await bot_instance.send_message(
            chat_id,
            f"Ошибка загрузки задания {current_iter_idx + 1}. Тест Матриц Равена прерван.",
        )
        await _finish_raven_matrices_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    await state.update_data(
        raven_correct_option_for_current_task=correct_option_1_based,
        raven_num_options_for_current_task=num_total_options,
        raven_current_task_filename=task_filename_only,
    )

    buttons_row, buttons_grid = [], []
    buttons_per_row = 3
    if num_total_options == 8:
        buttons_per_row = 4
    elif num_total_options == 4:
        buttons_per_row = 2
    elif num_total_options == 2:
        buttons_per_row = 2

    for i in range(1, num_total_options + 1):
        buttons_row.append(IKB(text=str(i), callback_data=f"raven_answer_{i}"))
        if len(buttons_row) == buttons_per_row or i == num_total_options:
            buttons_grid.append(list(buttons_row))
            buttons_row.clear()

    buttons_grid.append(
        [IKB(text="⏹️ Остановить Тест", callback_data="request_test_stop")]
    )
    reply_markup = InlineKeyboardMarkup(inline_keyboard=buttons_grid)

    task_message_id = data.get("raven_task_message_id")
    caption_text = f"Задание {current_iter_idx + 1} из {len(session_tasks)}"

    try:
        if task_message_id:
            media = InputMediaPhoto(
                media=FSInputFile(task_image_full_path), caption=caption_text
            )
            await bot_instance.edit_message_media(
                chat_id=chat_id,
                message_id=task_message_id,
                media=media,
                reply_markup=reply_markup,
            )
        else:
            msg = await bot_instance.send_photo(
                chat_id=chat_id,
                photo=FSInputFile(task_image_full_path),
                caption=caption_text,
                reply_markup=reply_markup,
            )
            await state.update_data(raven_task_message_id=msg.message_id)
    except (TelegramBadRequest, FileNotFoundError) as e:
        logger.error(
            f"Raven: Error sending/editing task image '{task_filename_only}': {e}",
            exc_info=True,
        )
        await bot_instance.send_message(
            chat_id, "Ошибка отображения задания. Тест Матриц Равена прерван."
        )
        await _finish_raven_matrices_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    await state.update_data(raven_current_task_start_time=time.time())
    await state.set_state(RavenMatricesStates.displaying_task_raven)


async def _finish_raven_matrices_test(
    state: FSMContext,
    bot_instance: Bot,
    chat_id: int | None,
    is_interrupted: bool,
    error_occurred: bool = False,
    called_by_stop_command: bool = False,  # Added to know context
):
    current_fsm_state = await state.get_state()
    if not current_fsm_state or not current_fsm_state.startswith(
        RavenMatricesStates.__name__
    ):
        logger.info(
            "Raven _finish_test: Called but test not in an active Raven state or already finished."
        )
        # If called by stop command, it handles common status message.
        # Otherwise, if test is not active, there's nothing specific to do here regarding common status.
        return

    logger.info(
        f"Finishing Raven Matrices Test. Interrupted: {is_interrupted}, Error: {error_occurred}, Called by Stop: {called_by_stop_command}"
    )
    data = await state.get_data()
    effective_chat_id = data.get(
        "raven_chat_id", chat_id
    )  # Prefer FSM chat_id

    revert_task = data.get("raven_current_feedback_revert_task_ref")
    if revert_task and not revert_task.done():
        revert_task.cancel()
        await asyncio.sleep(0.01)  # Give a tick
    await state.update_data(raven_current_feedback_revert_task_ref=None)

    iteration_results = data.get("raven_iteration_results", [])
    total_tasks_presented_calc = len(iteration_results)
    correct_answers_count_calc = sum(
        1 for r in iteration_results if r.get("is_correct")
    )

    test_start_time = data.get("raven_total_test_start_time")
    # Use actual end time if recorded, otherwise current time
    test_end_time = data.get("raven_total_test_end_time_actual", time.time())
    total_test_time_s_calc = (
        round(test_end_time - test_start_time, 2) if test_start_time else 0.0
    )

    ind_times_s_list_calc = [
        r.get("reaction_time_s", 0.0) for r in iteration_results
    ]  # Default to 0.0 if missing
    ind_times_s_str_calc = (
        ", ".join(map(lambda x: f"{x:.2f}", ind_times_s_list_calc))
        if ind_times_s_list_calc
        else "N/A"
    )

    correct_reaction_times_calc = [
        r["reaction_time_s"]
        for r in iteration_results
        if r.get("is_correct") and "reaction_time_s" in r
    ]
    avg_time_correct_s_calc = (
        round(
            sum(correct_reaction_times_calc)
            / len(correct_reaction_times_calc),
            2,
        )
        if correct_reaction_times_calc
        else 0.0
    )

    await state.update_data(
        raven_final_correct_answers=correct_answers_count_calc,
        raven_final_total_test_time_s=total_test_time_s_calc,
        raven_final_avg_time_correct_s=avg_time_correct_s_calc,
        raven_final_individual_times_s_str=ind_times_s_str_calc,
        raven_final_interrupted_status=(
            is_interrupted or error_occurred
        ),  # Mark interrupted if error
        raven_final_total_tasks_attempted=total_tasks_presented_calc,
    )

    # Create a mock message context if needed for save_results or send_main_action_menu
    mock_msg_for_context = None
    if effective_chat_id:
        mock_user_obj = User(
            id=bot_instance.id if hasattr(bot_instance, 'id') else 1,
            is_bot=True,
            first_name="Bot",
        )
        mock_chat_obj_ctx = Chat(id=effective_chat_id, type=ChatType.PRIVATE)
        mock_msg_for_context = Message(
            message_id=0,
            date=int(time.time()),
            chat=mock_chat_obj_ctx,
            from_user=mock_user_obj,
            text="mock",
        )

    await save_raven_matrices_results(
        mock_msg_for_context,
        state,
        is_interrupted=(is_interrupted or error_occurred),
    )

    # Send final summary text ONLY if not called by stop_test_command (which handles its own menu/message)
    if not called_by_stop_command and effective_chat_id:
        num_tasks_in_session = len(
            data.get("raven_session_task_filenames", [])
        )
        final_text_to_user = ""
        if is_interrupted or error_occurred:
            final_text_to_user = "Тест Матриц Равена был прерван"
            if error_occurred:
                final_text_to_user += " из-за ошибки."
            else:
                final_text_to_user += "."
            if iteration_results:
                final_text_to_user += f"\nЧастичные результаты ({total_tasks_presented_calc} из {num_tasks_in_session} заданий): {correct_answers_count_calc} правильных."
        else:  # Normal completion
            final_text_to_user = (
                "Тест Матриц Равена завершен!\n"
                f"Правильных ответов: {correct_answers_count_calc} из {num_tasks_in_session}.\n"
                f"Общее время: {total_test_time_s_calc:.2f} сек.\n"
            )
            final_text_to_user += (
                f"Среднее время на правильный ответ: {avg_time_correct_s_calc:.2f} сек."
                if correct_answers_count_calc > 0
                else "Правильных ответов не было."
            )

        try:
            await bot_instance.send_message(
                effective_chat_id,
                final_text_to_user,
                parse_mode=ParseMode.HTML,
            )
        except Exception as e_send_final_summary:
            logger.error(
                f"Raven _finish_test: Error sending final summary to user: {e_send_final_summary}"
            )

    # Cleanup UI (this will delete raven_task_message_id and raven_feedback_message_id)
    await cleanup_raven_ui(state, bot_instance)

    # Delete the common "Запускаем тест..." message IF NOT called by stop_test_command_handler
    if not called_by_stop_command:
        # Re-fetch data as cleanup_raven_ui modifies it
        current_data_for_common_msg_del = await state.get_data()
        common_status_msg_id_del = current_data_for_common_msg_del.get(
            "status_message_id_to_delete_later"
        )
        if common_status_msg_id_del and effective_chat_id:
            try:
                await bot_instance.delete_message(
                    effective_chat_id, common_status_msg_id_del
                )
                logger.info(
                    f"Raven Finish (normal/error): Deleted common status message ID: {common_status_msg_id_del}"
                )
            except TelegramBadRequest:
                logger.warning(
                    f"Raven Finish (normal/error): Common status message {common_status_msg_id_del} already deleted."
                )
            except Exception as e_del_common_raven:
                logger.error(
                    f"Raven Finish (normal/error): Error deleting common status message {common_status_msg_id_del}: {e_del_common_raven}"
                )
            await state.update_data(
                status_message_id_to_delete_later=None
            )  # Clear from FSM

    # Navigate to main menu if not handled by stop_test_command_handler
    if not called_by_stop_command:
        profile_keys = [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        # Get fresh data again after all FSM updates
        final_fsm_data_for_nav = await state.get_data()
        profile_data_to_keep_nav = {
            k: final_fsm_data_for_nav.get(k)
            for k in profile_keys
            if final_fsm_data_for_nav.get(k)
        }

        await state.set_state(None)  # Clear Raven state
        if profile_data_to_keep_nav.get("active_unique_id"):
            await state.set_data(
                profile_data_to_keep_nav
            )  # Keep only profile data

            trigger_event_for_menu_nav = (
                data.get("raven_triggering_event_for_menu")
                or mock_msg_for_context
            )
            message_context_for_menu_nav = None
            if isinstance(trigger_event_for_menu_nav, Message):
                message_context_for_menu_nav = trigger_event_for_menu_nav
            elif (
                isinstance(trigger_event_for_menu_nav, CallbackQuery)
                and trigger_event_for_menu_nav.message
            ):
                message_context_for_menu_nav = (
                    trigger_event_for_menu_nav.message
                )

            if message_context_for_menu_nav:
                await send_main_action_menu(
                    bot_instance,
                    message_context_for_menu_nav,
                    ACTION_SELECTION_KEYBOARD_RETURNING,
                )
            elif effective_chat_id:  # Fallback if no valid message context
                await bot_instance.send_message(
                    effective_chat_id,
                    "Тест завершен. Выберите действие:",
                    reply_markup=ACTION_SELECTION_KEYBOARD_RETURNING,
                )
        elif effective_chat_id:
            await bot_instance.send_message(
                effective_chat_id,
                "Тест завершен. Профиль не найден, пожалуйста /start.",
            )
            await state.clear()
        else:  # No chat_id, critical scenario, just clear state
            await state.clear()


async def start_raven_matrices_test(
    trigger_event: Message | CallbackQuery,
    state: FSMContext,
    profile: dict,
    bot_instance: Bot,
):
    logger.info(
        f"Starting Raven Matrices Test for UID: {profile.get('unique_id', 'N/A')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    if not RAVEN_ALL_TASK_FILES:
        logger.error(
            "Raven Start: RAVEN_ALL_TASK_FILES is empty at startup. Cannot start test."
        )
        await bot_instance.send_message(
            chat_id,
            "Ошибка конфигурации: Файлы для Теста Матриц Равена не загружены. Тест не может быть запущен.",
        )
        # Clean up common status message if it was set by common_handlers
        common_status_msg_id_start_err_cfg = (await state.get_data()).get(
            "status_message_id_to_delete_later"
        )
        if common_status_msg_id_start_err_cfg and chat_id:
            try:
                await bot_instance.delete_message(
                    chat_id, common_status_msg_id_start_err_cfg
                )
            except:
                pass
        await state.clear()  # Clear any partial state
        # Consider sending main menu if profile exists, or just end. For now, simple clear.
        return

    num_tasks_for_session = min(
        RAVEN_NUM_TASKS_TO_PRESENT, len(RAVEN_ALL_TASK_FILES)
    )
    if len(RAVEN_ALL_TASK_FILES) < RAVEN_NUM_TASKS_TO_PRESENT:
        logger.warning(
            f"Raven Start: Not enough tasks ({len(RAVEN_ALL_TASK_FILES)}) for configured {RAVEN_NUM_TASKS_TO_PRESENT}. Using available {num_tasks_for_session}."
        )

    session_task_filenames = random.sample(
        RAVEN_ALL_TASK_FILES, num_tasks_for_session
    )

    await state.set_state(RavenMatricesStates.initial_instructions_raven)
    await state.update_data(
        raven_unique_id_for_test=profile.get("unique_id"),
        raven_profile_name_for_test=profile.get("name"),
        raven_profile_age_for_test=profile.get("age"),
        raven_profile_telegram_id_for_test=profile.get("telegram_id"),
        raven_chat_id=chat_id,
        raven_session_task_filenames=session_task_filenames,
        raven_current_iteration_num=0,
        raven_iteration_results=[],
        raven_total_test_start_time=None,
        raven_current_task_start_time=None,
        raven_task_message_id=None,
        raven_feedback_message_id=None,
        raven_current_feedback_revert_task_ref=None,
        raven_triggering_event_for_menu=msg_ctx,  # For navigating back to menu correctly
    )
    instruction_text = (
        "<b>Тест Прогрессивных Матриц Равена</b>\n\n"
        "Вам будет показана матрица с пропущенным элементом и несколько вариантов для его заполнения. "
        "Ваша задача - выбрать наиболее подходящий вариант, чтобы завершить матрицу, следуя логике ее построения.\n\n"
        f"Тест состоит из {num_tasks_for_session} заданий. "
        "Постарайтесь отвечать не только правильно, но и быстро. Время ответа учитывается."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать тест", callback_data="raven_ack_instructions")]
        ]
    )
    try:
        # Always send instructions as a new message to avoid editing unrelated messages
        await bot_instance.send_message(
            chat_id,
            instruction_text,
            reply_markup=kbd,
            parse_mode=ParseMode.HTML,
        )
        if isinstance(trigger_event, CallbackQuery) and trigger_event.message:
            try:
                await trigger_event.message.delete()  # Delete the "Select test" menu message
            except TelegramBadRequest:
                pass
    except Exception as e_start_instr_send:
        logger.error(
            f"Raven Start: Error sending initial instructions: {e_start_instr_send}",
            exc_info=True,
        )
        await bot_instance.send_message(
            chat_id,
            "Ошибка при запуске Теста Матриц Равена. Попробуйте /start.",
        )
        common_status_msg_id_start_fail_instr = (await state.get_data()).get(
            "status_message_id_to_delete_later"
        )
        if common_status_msg_id_start_fail_instr and chat_id:
            try:
                await bot_instance.delete_message(
                    chat_id, common_status_msg_id_start_fail_instr
                )
            except:
                pass
        await state.clear()


@router.callback_query(
    F.data == "raven_ack_instructions",
    RavenMatricesStates.initial_instructions_raven,
)
async def raven_ack_instructions_callback(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    await state.update_data(raven_total_test_start_time=time.time())
    if (
        callback.message
    ):  # Delete the instruction message with "Начать тест" button
        try:
            await callback.message.delete()
        except TelegramBadRequest:
            pass

    chat_id_ack = (await state.get_data()).get("raven_chat_id")
    if chat_id_ack:
        await _display_raven_task(chat_id_ack, state, bot)
    else:
        logger.error(
            "Raven Ack Instr: chat_id missing from FSM. Cannot proceed."
        )
        if callback.message:
            await callback.message.answer(
                "Критическая ошибка: ID чата не найден. Тест прерван."
            )
        await _finish_raven_matrices_test(
            state, bot, None, is_interrupted=True, error_occurred=True
        )


@router.callback_query(
    F.data.startswith("raven_answer_"),
    RavenMatricesStates.displaying_task_raven,
)
async def handle_raven_answer_callback(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("raven_chat_id")
    if not chat_id:
        logger.error(
            "Raven Answer Callback: chat_id missing. Aborting processing."
        )
        if callback.message:
            await callback.message.answer(
                "Критическая ошибка: ID чата не найден. Тест прерван."
            )
        await _finish_raven_matrices_test(
            state, bot, None, is_interrupted=True, error_occurred=True
        )
        return

    task_start_time = data.get(
        "raven_current_task_start_time", time.time()
    )  # Fallback if somehow missing
    reaction_time_s = round(time.time() - task_start_time, 2)
    user_choice_num_1_based = int(callback.data.split("raven_answer_")[-1])
    correct_option_1_based = data.get("raven_correct_option_for_current_task")
    is_correct = user_choice_num_1_based == correct_option_1_based
    current_task_filename_ans = data.get("raven_current_task_filename", "N/A")

    iteration_result_data = {
        "task_filename": current_task_filename_ans,
        "user_choice": user_choice_num_1_based,
        "correct_answer_number": correct_option_1_based,
        "is_correct": is_correct,
        "reaction_time_s": reaction_time_s,
    }
    current_results_list = data.get("raven_iteration_results", [])
    current_results_list.append(iteration_result_data)
    await state.update_data(raven_iteration_results=current_results_list)

    feedback_text_bold_ans = (
        f"<b>{'Верно! ✅' if is_correct else 'Неверно!'}</b>"
    )
    feedback_text_normal_ans = f"{'Верно!' if is_correct else 'Неверно!'}"
    feedback_msg_id_ans = data.get("raven_feedback_message_id")

    previous_revert_task_ans = data.get(
        "raven_current_feedback_revert_task_ref"
    )
    if previous_revert_task_ans and not previous_revert_task_ans.done():
        previous_revert_task_ans.cancel()
        await asyncio.sleep(0.01)

    try:
        if feedback_msg_id_ans:
            await bot.edit_message_text(
                text=feedback_text_bold_ans,
                chat_id=chat_id,
                message_id=feedback_msg_id_ans,
                parse_mode=ParseMode.HTML,
            )
        else:
            msg_fb_ans = await bot.send_message(
                chat_id, feedback_text_bold_ans, parse_mode=ParseMode.HTML
            )
            feedback_msg_id_ans = msg_fb_ans.message_id
            await state.update_data(
                raven_feedback_message_id=feedback_msg_id_ans
            )

        if feedback_msg_id_ans:
            revert_task_ans = asyncio.create_task(
                _raven_delayed_feedback_revert(
                    chat_id,
                    feedback_msg_id_ans,
                    feedback_text_normal_ans,
                    bot,
                    state,
                )
            )
            await state.update_data(
                raven_current_feedback_revert_task_ref=revert_task_ans
            )
    except TelegramBadRequest as e_tb_fb_ans_cb:
        if "message is not modified" not in str(e_tb_fb_ans_cb).lower():
            logger.error(
                f"Raven Answer Callback: Feedback error (TB): {e_tb_fb_ans_cb}"
            )
    except Exception as e_gen_fb_ans_cb:
        logger.error(
            f"Raven Answer Callback: Feedback general error: {e_gen_fb_ans_cb}",
            exc_info=True,
        )

    current_iter_idx_ans = data.get("raven_current_iteration_num", 0)
    session_tasks_len_ans = len(data.get("raven_session_task_filenames", []))
    next_iter_idx_ans = current_iter_idx_ans + 1
    await state.update_data(raven_current_iteration_num=next_iter_idx_ans)

    if next_iter_idx_ans < session_tasks_len_ans:
        await _display_raven_task(chat_id, state, bot)
    else:  # All tasks completed
        await state.update_data(raven_total_test_end_time_actual=time.time())
        logger.info("Raven Matrices Test: All iterations completed by user.")
        await _finish_raven_matrices_test(
            state, bot, chat_id, is_interrupted=False, error_occurred=False
        )


async def save_raven_matrices_results(
    trigger_msg_context: Message | None,
    state: FSMContext,
    is_interrupted: bool = False,
):
    logger.info(
        f"Saving Raven Matrices results. Interrupted: {is_interrupted}"
    )
    data = await state.get_data()
    uid = data.get("raven_unique_id_for_test", data.get("active_unique_id"))
    p_tgid = data.get(
        "raven_profile_telegram_id_for_test", data.get("active_telegram_id")
    )
    p_name = data.get("raven_profile_name_for_test", data.get("active_name"))
    p_age = data.get("raven_profile_age_for_test", data.get("active_age"))

    if not uid:
        logger.error("Raven Save Results: UID not found. Cannot save results.")
        return

    correct_ans_save = data.get("raven_final_correct_answers", 0)
    total_time_save = data.get("raven_final_total_test_time_s", 0.0)
    avg_rt_correct_save = data.get("raven_final_avg_time_correct_s", 0.0)
    ind_times_str_save = data.get("raven_final_individual_times_s_str", "N/A")
    interrupted_status_save = (
        "Да"
        if data.get("raven_final_interrupted_status", is_interrupted)
        else "Нет"
    )

    try:
        from openpyxl import load_workbook

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        if "Unique ID" not in ALL_EXPECTED_HEADERS:
            raise ValueError(
                "'Unique ID' header not in settings.ALL_EXPECTED_HEADERS for Excel."
            )
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")

        for r_idx_save, row_vals_save in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals_save[uid_col_idx] is not None and str(
                row_vals_save[uid_col_idx]
            ) == str(uid):
                row_num = r_idx_save
                break

        if row_num == -1:
            new_row_excel_raven = [""] * len(ALL_EXPECTED_HEADERS)
            if p_tgid and "Telegram ID" in ALL_EXPECTED_HEADERS:
                new_row_excel_raven[
                    ALL_EXPECTED_HEADERS.index("Telegram ID")
                ] = p_tgid
            new_row_excel_raven[uid_col_idx] = uid
            if p_name and "Name" in ALL_EXPECTED_HEADERS:
                new_row_excel_raven[ALL_EXPECTED_HEADERS.index("Name")] = (
                    p_name
                )
            if p_age and "Age" in ALL_EXPECTED_HEADERS:
                new_row_excel_raven[ALL_EXPECTED_HEADERS.index("Age")] = p_age
            ws.append(new_row_excel_raven)
            row_num = ws.max_row

        h_excel = ALL_EXPECTED_HEADERS

        def set_cell_val_raven(hdr_name_raven, val_to_set_raven):
            if hdr_name_raven in h_excel:
                ws.cell(
                    row=row_num, column=h_excel.index(hdr_name_raven) + 1
                ).value = val_to_set_raven

        set_cell_val_raven("RavenMatrices_CorrectAnswers", correct_ans_save)
        set_cell_val_raven("RavenMatrices_TotalTime_s", total_time_save)
        set_cell_val_raven(
            "RavenMatrices_AvgTimeCorrect_s", avg_rt_correct_save
        )
        set_cell_val_raven(
            "RavenMatrices_IndividualTimes_s", ind_times_str_save
        )
        set_cell_val_raven(
            "RavenMatrices_Interrupted", interrupted_status_save
        )

        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Raven Matrices results for UID {uid} saved. Interrupted: {interrupted_status_save}"
        )
    except Exception as e_save_excel_raven:
        logger.error(
            f"Raven Matrices Save Results: Excel save error for UID {uid}: {e_save_excel_raven}",
            exc_info=True,
        )


async def cleanup_raven_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: str | None = None,
    # final_text is now effectively ignored for editing task_msg
):
    data = await state.get_data()
    chat_id = data.get("raven_chat_id")
    logger.info(
        f"Raven Cleanup UI: Chat {chat_id if chat_id else 'N/A'}. (final_text parameter is ignored for task msg edit)."
    )

    revert_task_cleanup = data.get("raven_current_feedback_revert_task_ref")
    if revert_task_cleanup and not revert_task_cleanup.done():
        revert_task_cleanup.cancel()
        await asyncio.sleep(0.01)

    task_msg_id_cleanup = data.get("raven_task_message_id")
    feedback_msg_id_cleanup = data.get("raven_feedback_message_id")

    if chat_id:
        # Always delete task message if ID exists
        if task_msg_id_cleanup:
            try:
                await bot_instance.delete_message(chat_id, task_msg_id_cleanup)
                logger.debug(
                    f"Raven Cleanup: Deleted task_message_id: {task_msg_id_cleanup}"
                )
            except TelegramBadRequest:
                logger.debug(
                    f"Raven Cleanup: Task message {task_msg_id_cleanup} already deleted."
                )
            except Exception as e_del_task:
                logger.error(
                    f"Raven Cleanup: Error deleting task message {task_msg_id_cleanup}: {e_del_task}"
                )

        # Always delete feedback message if ID exists and is different from task message
        if (
            feedback_msg_id_cleanup
            and feedback_msg_id_cleanup != task_msg_id_cleanup
        ):
            try:
                await bot_instance.delete_message(
                    chat_id, feedback_msg_id_cleanup
                )
                logger.debug(
                    f"Raven Cleanup: Deleted feedback_message_id: {feedback_msg_id_cleanup}"
                )
            except TelegramBadRequest:
                logger.debug(
                    f"Raven Cleanup: Feedback message {feedback_msg_id_cleanup} already deleted."
                )
            except Exception as e_del_fb:
                logger.error(
                    f"Raven Cleanup: Error deleting feedback message {feedback_msg_id_cleanup}: {e_del_fb}"
                )

        # If stop_test_command_handler passed a final_text, it will send its own message.
        # This cleanup function no longer sends/edits based on final_text.

    # Clean FSM: preserve profile and common status message ID, remove Raven specific keys
    current_fsm_data_raven_clean = await state.get_data()
    data_to_keep_after_raven_cleanup = {}
    keys_to_preserve_after_raven = [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
        "status_message_id_to_delete_later",
    ]

    for key_raven_clean in keys_to_preserve_after_raven:
        if (
            key_raven_clean in current_fsm_data_raven_clean
            and current_fsm_data_raven_clean[key_raven_clean] is not None
        ):
            data_to_keep_after_raven_cleanup[key_raven_clean] = (
                current_fsm_data_raven_clean[key_raven_clean]
            )

    # Preserve any other non-raven keys
    for (
        key_other_raven,
        val_other_raven,
    ) in current_fsm_data_raven_clean.items():
        if (
            not key_other_raven.startswith("raven_")
            and key_other_raven not in data_to_keep_after_raven_cleanup
        ):
            data_to_keep_after_raven_cleanup[key_other_raven] = val_other_raven

    await state.set_data(data_to_keep_after_raven_cleanup)
    logger.info(
        f"Raven Cleanup UI: FSM data cleaned. Kept keys: {list(data_to_keep_after_raven_cleanup.keys())}"
    )
