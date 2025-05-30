# handlers/tests/stroop_handlers.py
import asyncio
import logging
import random
import time
from typing import Union, Optional, Dict, Any
import os

from aiogram import Bot, F, Router
from aiogram.enums import ParseMode, ChatType
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
    Message,
    Chat,
    User,
)
from aiogram.filters import StateFilter

from fsm_states import StroopTestStates
from settings import (
    ALL_EXPECTED_HEADERS,
    EXCEL_FILENAME,
    STROOP_COLOR_NAMES,
    STROOP_ITERATIONS_PER_PART,
    STROOP_INSTRUCTION_TEXT_PART1,
    STROOP_INSTRUCTION_TEXT_PART2,
    STROOP_INSTRUCTION_TEXT_PART3,
    STROOP_COLORS_DEF,
)
from utils.image_processors import (
    _generate_stroop_part2_image,
    _generate_stroop_part3_image,
)
from utils.bot_helpers import (
    send_main_action_menu,
    get_active_profile_from_fsm,
    _clear_fsm_and_set_profile,
    _safe_delete_message,
)

from keyboards import ACTION_SELECTION_KEYBOARD_RETURNING

logger = logging.getLogger(__name__)
router = Router()
IKB = InlineKeyboardButton


# --- Helper Functions (scoped to Stroop) ---
async def _safe_delete_stroop_specific_message(
    bot: Bot, state: FSMContext, fsm_key: str, context_info: str = ""
):
    """Safely deletes a Stroop-specific message ID stored in FSM."""
    data = await state.get_data()
    message_id = data.get(fsm_key)
    chat_id = data.get("stroop_chat_id")

    if message_id and chat_id:
        try:
            await bot.delete_message(chat_id, message_id)
            logger.debug(
                f"Stroop: Сообщение ID {message_id} (ключ: {fsm_key}) удалено. {context_info}"
            )
        except TelegramBadRequest:
            logger.warning(
                f"Stroop: Не удалось удалить ID {message_id} (ключ: {fsm_key}). {context_info}"
            )
        except Exception as e:
            logger.error(
                f"Stroop: Ошибка удаления ID {message_id} (ключ: {fsm_key}): {e}. {context_info}"
            )
        # Do not clear FSM key here; let cleanup_stroop_ui or final FSM set handle it.


def _create_mock_message_stroop(
    chat_id: int, bot_id: Optional[int]
) -> Message:
    """Creates a mock Message object for error handling in Stroop test."""
    actual_bot_id = bot_id if bot_id is not None else 1
    mock_user = User(id=actual_bot_id, is_bot=True, first_name="BotStroopErr")
    mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
    return Message(
        message_id=0,
        date=int(time.time()),
        chat=mock_chat,
        from_user=mock_user,
        text="Mock Stroop error message",
    )


async def _send_stroop_instruction_message(
    chat_id: int, part: int, state: FSMContext, bot_instance: Bot
):
    text, cb_data = "", ""
    if part == 1:
        text, cb_data = STROOP_INSTRUCTION_TEXT_PART1, "stroop_ack_part1"
    elif part == 2:
        text, cb_data = STROOP_INSTRUCTION_TEXT_PART2, "stroop_ack_part2"
    elif part == 3:
        text, cb_data = STROOP_INSTRUCTION_TEXT_PART3, "stroop_ack_part3"
    else:
        logger.error(f"Stroop: Неверная часть ({part}) для инструкции.")
        return

    markup = InlineKeyboardMarkup(
        inline_keyboard=[[IKB(text="Понятно", callback_data=cb_data)]]
    )
    try:
        msg = await bot_instance.send_message(
            chat_id, text, reply_markup=markup, parse_mode=ParseMode.HTML
        )
        await state.update_data(stroop_instruction_message_id=msg.message_id)
    except Exception as e:
        logger.error(
            f"Stroop: Не удалось отправить инструкцию для части {part}: {e}",
            exc_info=True,
        )


async def _handle_stroop_critical_error(
    chat_id: int,
    state: FSMContext,
    bot_instance: Bot,
    error_context_message: str,
):
    """Handles critical errors during Stroop test."""
    logger.critical(
        f"Stroop: Критическая ошибка в тесте - {error_context_message}"
    )
    try:
        await bot_instance.send_message(
            chat_id,
            f"Критическая ошибка: {error_context_message}. Тест Струпа будет прерван.",
        )
    except Exception as e_send_err:
        logger.error(
            f"Stroop: Не удалось отправить сообщение о критической ошибке: {e_send_err}"
        )

    bot_id_for_mock = (
        bot_instance.id
        if hasattr(bot_instance, "id") and bot_instance.id is not None
        else 1
    )
    mock_msg = _create_mock_message_stroop(chat_id, bot_id_for_mock)

    await save_stroop_results(
        mock_msg, state, bot_instance, is_interrupted=True
    )
    await cleanup_stroop_ui(
        state, bot_instance, f"Тест Струпа прерван ({error_context_message})."
    )

    profile_after_error = await get_active_profile_from_fsm(state)

    fsm_data_for_common_msg_del = await state.get_data()
    common_status_msg_id = fsm_data_for_common_msg_del.get(
        "status_message_id_to_delete_later"
    )
    if common_status_msg_id:
        await _safe_delete_message(
            bot_instance,
            chat_id,
            common_status_msg_id,
            "Stroop critical error common status cleanup",
        )

    await _clear_fsm_and_set_profile(state, profile_after_error)

    if profile_after_error:  # Check if profile still exists
        await send_main_action_menu(
            bot_instance,
            mock_msg,
            ACTION_SELECTION_KEYBOARD_RETURNING,
            text="Тест прерван из-за ошибки.",
        )
    # else: FSM cleared, user will need to /start


async def _display_next_stroop_stimulus(
    chat_id: int, state: FSMContext, bot_instance: Bot
):
    data = await state.get_data()
    current_part = data.get("stroop_current_part")
    current_iteration = data.get("stroop_current_iteration")
    stimulus_msg_id = data.get("stroop_stimulus_message_id")
    current_stimulus_ui_type = data.get("stroop_stimulus_type")

    image_to_send = None
    stimulus_text_for_part1 = ""
    correct_answer_color_name = ""
    all_colors = list(STROOP_COLOR_NAMES)
    new_stimulus_ui_type = ""

    if current_part == 1:
        word_color = random.choice(all_colors)
        stimulus_text_for_part1 = STROOP_COLORS_DEF[word_color]["name"]
        correct_answer_color_name = word_color
        new_stimulus_ui_type = "text"
    elif current_part == 2:
        patch_color = random.choice(all_colors)
        text_choices = [c for c in all_colors if c != patch_color]
        text_on_patch = (
            random.choice(text_choices) if text_choices else patch_color
        )
        image_to_send = _generate_stroop_part2_image(
            patch_color, text_on_patch
        )
        correct_answer_color_name = patch_color
        new_stimulus_ui_type = "photo"
    elif current_part == 3:
        word_name = random.choice(all_colors)
        ink_choices = [c for c in all_colors if c != word_name]
        ink_name = random.choice(ink_choices) if ink_choices else word_name
        image_to_send = _generate_stroop_part3_image(word_name, ink_name)
        correct_answer_color_name = ink_name
        new_stimulus_ui_type = "photo"
    else:
        await _handle_stroop_critical_error(
            chat_id,
            state,
            bot_instance,
            f"Неверная часть теста: {current_part}",
        )
        return

    if new_stimulus_ui_type == "photo" and not image_to_send:
        await _handle_stroop_critical_error(
            chat_id, state, bot_instance, "Ошибка генерации изображения"
        )
        return

    await state.update_data(stroop_correct_answer=correct_answer_color_name)

    distractors = [c for c in all_colors if c != correct_answer_color_name]
    random.shuffle(distractors)
    num_distractors_to_use = min(
        len(distractors),
        3 if len(all_colors) >= 4 else max(0, len(all_colors) - 1),
    )
    button_names = [correct_answer_color_name] + distractors[
        :num_distractors_to_use
    ]
    random.shuffle(button_names)

    buttons_rows, current_row_btns = [], []
    for i, name in enumerate(button_names):
        btn_text = (
            STROOP_COLORS_DEF[name]["emoji"]
            if current_part == 1
            else STROOP_COLORS_DEF[name]["name"]
        )
        current_row_btns.append(
            IKB(text=btn_text, callback_data=f"stroop_answer_{name}")
        )
        if len(current_row_btns) == 2 or i == len(button_names) - 1:
            buttons_rows.append(current_row_btns)
            current_row_btns = []
    reply_markup = InlineKeyboardMarkup(inline_keyboard=buttons_rows)

    base_caption = f"<b>Тест Струпа</b>\nЧасть {current_part}, Повтор {current_iteration}/{STROOP_ITERATIONS_PER_PART}\n\n"
    instruction = ""
    if current_part == 1:
        instruction = (f"\t\t\t<b>{stimulus_text_for_part1}</b>\n\n"
                       f"Нажмите на <b>цветной квадрат</b>, соответствующий написанному названию."
                       )
    elif current_part == 2:
        instruction = "Нажмите кнопку с названием цвета, соответствующим <b>цвету прямоугольника</b>:"
    elif current_part == 3:
        instruction = "Нажмите кнопку с названием цвета, соответствующим <b>цвету чернил слова</b>:"
    full_caption = f"{base_caption}{instruction}".strip()

    try:
        if (
            stimulus_msg_id is None
            or current_stimulus_ui_type != new_stimulus_ui_type
        ):
            if stimulus_msg_id:
                await _safe_delete_stroop_specific_message(
                    bot_instance,
                    state,
                    "stroop_stimulus_message_id",
                    "_display_next type change",
                )
                stimulus_msg_id = None
            msg = await (
                bot_instance.send_photo(
                    chat_id,
                    photo=image_to_send,
                    caption=full_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
                if new_stimulus_ui_type == "photo"
                else bot_instance.send_message(
                    chat_id,
                    full_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
            )
            stimulus_msg_id = msg.message_id
        else:
            if new_stimulus_ui_type == "photo":
                media = InputMediaPhoto(
                    media=image_to_send,
                    caption=full_caption,
                    parse_mode=ParseMode.HTML,
                )
                await bot_instance.edit_message_media(
                    media=media,
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    reply_markup=reply_markup,
                )
            else:
                await bot_instance.edit_message_text(
                    full_caption,
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
        await state.update_data(
            stroop_stimulus_message_id=stimulus_msg_id,
            stroop_stimulus_type=new_stimulus_ui_type,
        )
    except TelegramBadRequest as e_ui:
        logger.warning(f"Stroop: Ошибка UI ({e_ui}). Попытка переотправки.")
        await _safe_delete_stroop_specific_message(
            bot_instance,
            state,
            "stroop_stimulus_message_id",
            "_display_next fallback delete",
        )
        try:
            msg_fb = await (
                bot_instance.send_photo(
                    chat_id,
                    photo=image_to_send,
                    caption=full_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
                if new_stimulus_ui_type == "photo"
                else bot_instance.send_message(
                    chat_id,
                    full_caption,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML,
                )
            )
            await state.update_data(
                stroop_stimulus_message_id=msg_fb.message_id,
                stroop_stimulus_type=new_stimulus_ui_type,
            )
        except Exception as e_fb_send:
            await _handle_stroop_critical_error(
                chat_id,
                state,
                bot_instance,
                f"Критическая ошибка UI при переотправке: {e_fb_send}",
            )
            return

    if current_part == 1:
        await state.set_state(StroopTestStates.part1_stimulus_response)
    elif current_part == 2:
        await state.set_state(StroopTestStates.part2_stimulus_response)
    elif current_part == 3:
        await state.set_state(StroopTestStates.part3_stimulus_response)


# --- Test Lifecycle Functions ---
async def start_stroop_test(
    trigger_event: Union[Message, CallbackQuery],
    state: FSMContext,
    profile: Dict[str, Any],
    bot_instance: Bot,
):
    uid = profile.get("unique_id")
    logger.info(f"Запуск Теста Струпа для UID: {uid}")
    source_message = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = source_message.chat.id

    await state.set_state(StroopTestStates.initial_instructions)
    initial_stroop_data = {
        "unique_id_for_test": uid,
        "profile_name_for_test": profile.get("name"),
        "profile_age_for_test": profile.get("age"),
        "profile_telegram_id_for_test": profile.get("telegram_id"),
        "stroop_chat_id": chat_id,
        "stroop_instruction_message_id": None,
        "stroop_stimulus_message_id": None,
        "stroop_stimulus_type": None,
        "stroop_current_part": 0,
        "stroop_current_iteration": 0,
        "stroop_part1_errors": 0,
        "stroop_part2_errors": 0,
        "stroop_part3_errors": 0,
        "stroop_part1_start_time": None,
        "stroop_part2_start_time": None,
        "stroop_part3_start_time": None,
        "stroop_part1_total_time_s": None,
        "stroop_part2_total_time_s": None,
        "stroop_part3_total_time_s": None,
        "stroop_correct_answer": None,
    }
    # Важно: update_data добавляет/обновляет, не затирая существующие (например, профиль с active_* ключами)
    await state.update_data(**initial_stroop_data)
    await _send_stroop_instruction_message(chat_id, 1, state, bot_instance)


@router.callback_query(
    F.data == "stroop_ack_part1", StroopTestStates.initial_instructions
)
async def handle_stroop_ack_part1(
    cb: CallbackQuery, state: FSMContext, bot: Bot
):
    await cb.answer()
    await _safe_delete_stroop_specific_message(
        bot, state, "stroop_instruction_message_id", "ack_part1"
    )
    await state.update_data(
        stroop_current_part=1,
        stroop_current_iteration=1,
        stroop_part1_start_time=time.time(),
    )
    chat_id = (await state.get_data()).get(
        "stroop_chat_id", cb.message.chat.id if cb.message else cb.from_user.id
    )
    await _display_next_stroop_stimulus(chat_id, state, bot)


@router.callback_query(
    F.data == "stroop_ack_part2", StroopTestStates.part2_instructions
)
async def handle_stroop_ack_part2(
    cb: CallbackQuery, state: FSMContext, bot: Bot
):
    await cb.answer()
    await _safe_delete_stroop_specific_message(
        bot, state, "stroop_instruction_message_id", "ack_part2"
    )
    await state.update_data(
        stroop_current_part=2,
        stroop_current_iteration=1,
        stroop_part2_start_time=time.time(),
    )
    chat_id = (await state.get_data()).get(
        "stroop_chat_id", cb.message.chat.id if cb.message else cb.from_user.id
    )
    await _display_next_stroop_stimulus(chat_id, state, bot)


@router.callback_query(
    F.data == "stroop_ack_part3", StroopTestStates.part3_instructions
)
async def handle_stroop_ack_part3(
    cb: CallbackQuery, state: FSMContext, bot: Bot
):
    await cb.answer()
    await _safe_delete_stroop_specific_message(
        bot, state, "stroop_instruction_message_id", "ack_part3"
    )
    await state.update_data(
        stroop_current_part=3,
        stroop_current_iteration=1,
        stroop_part3_start_time=time.time(),
    )
    chat_id = (await state.get_data()).get(
        "stroop_chat_id", cb.message.chat.id if cb.message else cb.from_user.id
    )
    await _display_next_stroop_stimulus(chat_id, state, bot)


@router.callback_query(
    F.data.startswith("stroop_answer_"),
    StateFilter(
        StroopTestStates.part1_stimulus_response,
        StroopTestStates.part2_stimulus_response,
        StroopTestStates.part3_stimulus_response,
    ),
)
async def handle_stroop_stimulus_response(
    cb: CallbackQuery, state: FSMContext, bot: Bot
):
    data = await state.get_data()
    chosen_color = cb.data.split("stroop_answer_")[-1]
    correct_answer = data.get("stroop_correct_answer")
    current_part = data.get("stroop_current_part")
    current_iter = data.get("stroop_current_iteration", 0)
    chat_id = data.get(
        "stroop_chat_id", cb.message.chat.id if cb.message else cb.from_user.id
    )

    if chosen_color == correct_answer:
        await cb.answer(text="Верно!", show_alert=False)
    else:
        error_fb = (
            f"Ошибка! Правильно: {STROOP_COLORS_DEF[correct_answer]['name']}"
            if correct_answer and correct_answer in STROOP_COLORS_DEF
            else "Ошибка!"
        )
        await cb.answer(text=error_fb, show_alert=False)
        error_key = f"stroop_part{current_part}_errors"
        await state.update_data({error_key: data.get(error_key, 0) + 1})

    current_iter += 1

    if current_iter > STROOP_ITERATIONS_PER_PART:
        part_time_key = f"stroop_part{current_part}_start_time"
        part_start_t = data.get(part_time_key)
        if part_start_t:
            total_t = round(time.time() - part_start_t, 2)
            await state.update_data(
                {f"stroop_part{current_part}_total_time_s": total_t}
            )

        current_part += 1
        current_iter = 1
        await state.update_data(
            stroop_current_part=current_part,
            stroop_current_iteration=current_iter,
        )

        if current_part == 2:
            await state.set_state(StroopTestStates.part2_instructions)
            await _send_stroop_instruction_message(chat_id, 2, state, bot)
        elif current_part == 3:
            await state.set_state(StroopTestStates.part3_instructions)
            await _send_stroop_instruction_message(chat_id, 3, state, bot)
        else:
            logger.info(f"Тест Струпа завершен для чата {chat_id}.")
            await save_stroop_results(
                cb.message, state, bot, is_interrupted=False
            )  # cb.message as trigger_msg
            await cleanup_stroop_ui(
                state, bot, "Тест Струпа штатно завершен (очистка UI)."
            )

            profile_to_set = await get_active_profile_from_fsm(state)

            fsm_data_after_test_ops = await state.get_data()
            common_status_msg_id = fsm_data_after_test_ops.get(
                "status_message_id_to_delete_later"
            )
            if common_status_msg_id and chat_id:
                await _safe_delete_message(
                    bot,
                    chat_id,
                    common_status_msg_id,
                    "Stroop normal completion common status",
                )

            await _clear_fsm_and_set_profile(state, profile_to_set)

            if profile_to_set:
                # cb.message может быть None, если сообщение было удалено. Нужен fallback.
                msg_context_for_menu = (
                    cb.message
                    if cb.message
                    else _create_mock_message_stroop(
                        chat_id, bot.id if hasattr(bot, "id") else None
                    )
                )
                await send_main_action_menu(
                    bot,
                    msg_context_for_menu,
                    ACTION_SELECTION_KEYBOARD_RETURNING,
                    text="Выберите действие:",
                )
            else:
                if cb.message:
                    await cb.message.answer(
                        "Тест Струпа завершен. Профиль не найден, пожалуйста /start."
                    )
                else:
                    await bot.send_message(
                        chat_id,
                        "Тест Струпа завершен. Профиль не найден, пожалуйста /start.",
                    )
                # FSM уже очищен _clear_fsm_and_set_profile
    else:
        await state.update_data(stroop_current_iteration=current_iter)
        await _display_next_stroop_stimulus(chat_id, state, bot)


async def save_stroop_results(
    trigger_msg: Message,
    state: FSMContext,
    bot_instance: Bot,
    is_interrupted: bool = False,
):
    logger.info(
        f"Сохранение результатов Теста Струпа. Прерван: {is_interrupted}"
    )
    data = await state.get_data()

    uid = data.get("unique_id_for_test")
    p_name = data.get(
        "profile_name_for_test"
    )  # These are for new row creation if UID not found
    p_age = data.get("profile_age_for_test")
    p_tgid = data.get("profile_telegram_id_for_test")

    if not uid:
        active_profile = await get_active_profile_from_fsm(
            state
        )  # Returns standardized keys
        if active_profile and active_profile.get("unique_id"):
            uid = active_profile.get("unique_id")
            p_name = active_profile.get("name", p_name)
            p_age = active_profile.get("age", p_age)
            p_tgid = active_profile.get("telegram_id", p_tgid)
            logger.info(
                f"Stroop save: Используются данные из активного профиля для UID {uid}."
            )
        else:
            logger.error("Stroop save: Не найден UID для сохранения.")
            if (
                hasattr(trigger_msg, "chat")
                and await state.get_state() is not None
            ):
                await trigger_msg.answer(
                    "Тест Струпа: ошибка сохранения (ID пользователя не найден)."
                )
            return

    current_part_running = data.get("stroop_current_part", 0)
    if is_interrupted and 1 <= current_part_running <= 3:
        start_t_key = f"stroop_part{current_part_running}_start_time"
        total_t_key = f"stroop_part{current_part_running}_total_time_s"
        if data.get(start_t_key) and not data.get(total_t_key):
            time_taken = round(time.time() - data.get(start_t_key), 2)
            await state.update_data({total_t_key: time_taken})
            data = await state.get_data()

    p1t = data.get("stroop_part1_total_time_s")
    p1e = data.get("stroop_part1_errors", 0)
    p2t = data.get("stroop_part2_total_time_s")
    p2e = data.get("stroop_part2_errors", 0)
    p3t = data.get("stroop_part3_total_time_s")
    p3e = data.get("stroop_part3_errors", 0)
    intr_val = "Да" if is_interrupted else "Нет"

    try:
        from openpyxl import load_workbook

        if not os.path.exists(EXCEL_FILENAME):
            logger.error(
                f"Stroop save: Файл Excel '{EXCEL_FILENAME}' не найден."
            )
            if (
                hasattr(trigger_msg, "chat")
                and await state.get_state() is not None
            ):
                await trigger_msg.answer(
                    f"Ошибка: Файл для сохранения ('{EXCEL_FILENAME}') не найден."
                )
            return

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        excel_headers = [cell.value for cell in ws[1]]
        if "Unique ID" not in excel_headers:
            raise ValueError("Столбец 'Unique ID' не найден в Excel.")

        uid_col_idx = excel_headers.index("Unique ID")
        row_num_excel = -1
        for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if (
                len(row_vals) > uid_col_idx
                and row_vals[uid_col_idx] is not None
                and str(row_vals[uid_col_idx]) == str(uid)
            ):
                row_num_excel = r_idx
                break

        if row_num_excel == -1:
            logger.info(
                f"Stroop save: UID {uid} не найден, добавление новой строки."
            )
            new_row = [""] * len(excel_headers)
            new_row[uid_col_idx] = uid
            if p_name and "Name" in excel_headers:
                new_row[excel_headers.index("Name")] = p_name
            if p_age and "Age" in excel_headers:
                new_row[excel_headers.index("Age")] = p_age
            if p_tgid and "Telegram ID" in excel_headers:
                new_row[excel_headers.index("Telegram ID")] = p_tgid
            ws.append(new_row)
            row_num_excel = ws.max_row

        stroop_headers_map = {
            h: excel_headers.index(h) + 1
            for h in [
                "Stroop Part1 Time (s)",
                "Stroop Part1 Errors",
                "Stroop Part2 Time (s)",
                "Stroop Part2 Errors",
                "Stroop Part3 Time (s)",
                "Stroop Part3 Errors",
                "Stroop - Interrupted",
            ]
            if h in excel_headers
        }

        def set_cell(header, value):
            if header in stroop_headers_map:
                ws.cell(
                    row=row_num_excel, column=stroop_headers_map[header]
                ).value = value
            else:
                logger.warning(
                    f"Stroop save: Заголовок '{header}' не найден в Excel для записи."
                )

        set_cell("Stroop Part1 Time (s)", p1t)
        set_cell("Stroop Part1 Errors", p1e)
        set_cell("Stroop Part2 Time (s)", p2t)
        set_cell("Stroop Part2 Errors", p2e)
        set_cell("Stroop Part3 Time (s)", p3t)
        set_cell("Stroop Part3 Errors", p3e)
        set_cell("Stroop - Interrupted", intr_val)

        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Результаты Теста Струпа для UID {uid} (Прерван: {is_interrupted}) сохранены."
        )

        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            status = (
                "ПРЕРВАНЫ И СОХРАНЕНЫ"
                if is_interrupted
                else "УСПЕШНО СОХРАНЕНЫ"
            )
            summary = [
                f"Результаты Теста Струпа <b>{status}</b> для UID {uid}:",
                f"Часть 1: Время {p1t if p1t is not None else 'N/A'} сек, Ошибок: {p1e}",
                f"Часть 2: Время {p2t if p2t is not None else 'N/A'} сек, Ошибок: {p2e}",
                f"Часть 3: Время {p3t if p3t is not None else 'N/A'} сек, Ошибок: {p3e}",
            ]
            if (
                is_interrupted
                and all(t is None for t in [p1t, p2t, p3t])
                and all(e == 0 for e in [p1e, p2e, p3e])
            ):
                summary = [
                    f"Тест Струпа был <b>ПРЕРВАН</b> досрочно для UID {uid}. Данные не зафиксированы."
                ]
            try:
                await trigger_msg.answer(
                    "\n".join(summary), parse_mode=ParseMode.HTML
                )
            except Exception as e_ans:
                logger.error(
                    f"Stroop save: Не удалось отправить итог: {e_ans}"
                )

    except FileNotFoundError:
        logger.error(
            f"Stroop save: Файл Excel '{EXCEL_FILENAME}' не найден (повторно)."
        )
        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            await trigger_msg.answer(
                f"Критическая ошибка: Файл для сохранения ('{EXCEL_FILENAME}') не найден."
            )
    except ValueError as ve:
        logger.error(
            f"Stroop save: Ошибка конфигурации Excel для UID {uid}: {ve}",
            exc_info=True,
        )
        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            await trigger_msg.answer(
                "Ошибка конфигурации при сохранении Теста Струпа."
            )
    except Exception as e_gen:
        logger.error(
            f"Stroop save: Общая ошибка сохранения в Excel для UID {uid}: {e_gen}",
            exc_info=True,
        )
        if await state.get_state() is not None and hasattr(
            trigger_msg, "chat"
        ):
            await trigger_msg.answer(
                "Непредвиденная ошибка при сохранении Теста Струпа."
            )


async def cleanup_stroop_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: Optional[str] = None,
):
    logger.info(f"Stroop UI Cleanup. Контекст: '{final_text}'")
    data = await state.get_data()
    chat_id = data.get("stroop_chat_id")

    if chat_id:
        await _safe_delete_stroop_specific_message(
            bot_instance,
            state,
            "stroop_instruction_message_id",
            "cleanup_stroop_ui",
        )
        await _safe_delete_stroop_specific_message(
            bot_instance,
            state,
            "stroop_stimulus_message_id",
            "cleanup_stroop_ui",
        )
        # Ключи из FSM будут удалены через _clear_fsm_and_set_profile или stop_test_command_handler
    else:
        logger.warning(
            "Stroop cleanup: stroop_chat_id не найден, невозможно удалить сообщения."
        )

    # Финальная очистка FSM и установка профиля будет выполнена вызывающей функцией
    # (handle_stroop_stimulus_response при штатном завершении или stop_test_command_handler при прерывании)
    # Эта функция только удаляет UI.
    logger.info("Stroop cleanup: UI-специфичные сообщения обработаны.")
