# handlers/tests/corsi_handlers.py
import asyncio
import logging
import os
import random
import time
from typing import (
    Union,
    Coroutine,
    Any,
    Optional,
    Dict,
)  # Added Optional, Dict

from aiogram import Bot, F, Router
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from aiogram.filters import StateFilter

from fsm_states import CorsiTestStates
from settings import ALL_EXPECTED_HEADERS, EXCEL_FILENAME
from utils.bot_helpers import (
    send_main_action_menu,
    get_active_profile_from_fsm,
    _clear_fsm_and_set_profile, # <--- ИЗМЕНЕННЫЙ ИМПОРТ
    _safe_delete_message,       # <--- ИЗМЕНЕННЫЙ ИМПОРТ
)

# Импортируем _clear_fsm_and_set_profile для использования при завершении теста

from keyboards import ACTION_SELECTION_KEYBOARD_RETURNING

logger = logging.getLogger(__name__)
router = Router()
IKB = InlineKeyboardButton

# --- Constants for Corsi Test ---
CORSI_MAX_SEQUENCE_LENGTH = 9
CORSI_BUTTON_CALLBACK_PREFIX = "corsi_button_"
CORSI_STOP_CALLBACK_DATA = "request_test_stop"


# --- Helper for Message Management (scoped to Corsi) ---
async def _safe_delete_corsi_specific_message(
    bot: Bot, state: FSMContext, fsm_key: str, context_info: str = ""
):
    """Safely deletes a Corsi-specific message ID stored in FSM."""
    data = await state.get_data()
    message_id = data.get(fsm_key)
    chat_id = data.get("corsi_chat_id")  # Assumes corsi_chat_id is in FSM

    if message_id and chat_id:
        try:
            await bot.delete_message(chat_id, message_id)
            logger.debug(
                f"Corsi: Сообщение ID {message_id} (ключ: {fsm_key}) удалено. {context_info}"
            )
        except TelegramBadRequest:
            logger.warning(
                f"Corsi: Не удалось удалить сообщение ID {message_id} (ключ: {fsm_key}). {context_info}"
            )
        except Exception as e:
            logger.error(
                f"Corsi: Ошибка удаления ID {message_id} (ключ: {fsm_key}): {e}. {context_info}"
            )
        # Do not clear the key from FSM here, cleanup_corsi_messages will handle FSM data.
        # Or, if called outside full cleanup, it can be cleared:
        # await state.update_data({fsm_key: None})


# --- Test Logic Functions ---
async def cleanup_corsi_messages(
    state: FSMContext, bot_instance: Bot, final_text: Optional[str] = None
):
    logger.info(f"Corsi cleanup. Контекст: '{final_text}'")
    data = await state.get_data()  # Get data once
    chat_id = data.get("corsi_chat_id")

    if chat_id:
        corsi_message_ids_keys = [
            "corsi_status_message_id",
            "corsi_feedback_message_id",
            "corsi_grid_message_id",
        ]
        for key in corsi_message_ids_keys:
            msg_id_to_del = data.get(key)
            if msg_id_to_del:  # Check if msg_id exists before calling helper
                await _safe_delete_corsi_specific_message(
                    bot_instance, state, key, "cleanup_corsi_messages"
                )
                # _safe_delete_corsi_specific_message does not clear FSM keys by default
                # so we clear them here after attempting deletion.
                current_data_for_key_clear = await state.get_data()
                if key in current_data_for_key_clear:
                    del current_data_for_key_clear[key]
                    await state.set_data(current_data_for_key_clear)

    # FSM data cleaning will be handled by _clear_fsm_and_set_profile
    # or by stop_test_command_handler if interrupted.
    # This function should focus on UI cleanup.
    # The caller (evaluate_user_sequence or stop_test_command_handler)
    # will be responsible for the final FSM state.
    logger.info(
        f"Corsi cleanup: UI-специфичные сообщения для чата {chat_id or 'N/A'} обработаны."
    )


async def show_corsi_sequence(
    trigger_source_msg: Message, state: FSMContext, bot_instance: Bot
):
    current_fsm_state = await state.get_state()
    if current_fsm_state != CorsiTestStates.showing_sequence.state:
        logger.warning(
            f"show_corsi_sequence вызван, но состояние FSM '{current_fsm_state}'. Прерывание."
        )
        return

    data = await state.get_data()
    current_sequence_length = data.get("current_sequence_length", 2)
    corsi_chat_id = data.get("corsi_chat_id")
    grid_msg_id = data.get("corsi_grid_message_id")
    status_msg_id = data.get("corsi_status_message_id")

    if not corsi_chat_id:
        logger.error(
            "Corsi (show_sequence): corsi_chat_id не найден в FSM. Тест не может продолжаться."
        )
        # UI cleanup should be done by caller if this fails critically
        await trigger_source_msg.answer(
            "Ошибка конфигурации Теста Корси. Пожалуйста, /start и попробуйте снова."
        )
        await _clear_fsm_and_set_profile(
            state, await get_active_profile_from_fsm(state)
        )  # Clear to profile or empty
        return

    indices = list(range(9))
    random.shuffle(indices)
    correct_seq_to_show = indices[:current_sequence_length]
    await state.update_data(
        correct_sequence=correct_seq_to_show, user_input_sequence=[]
    )

    base_grid_buttons = [
        IKB(text="🟪", callback_data=f"{CORSI_BUTTON_CALLBACK_PREFIX}{i}")
        for i in range(9)
    ]
    base_grid_rows = [base_grid_buttons[i : i + 3] for i in range(0, 9, 3)]
    base_grid_rows.append(
        [
            IKB(
                text="⏹️ Остановить Тест Корси",
                callback_data=CORSI_STOP_CALLBACK_DATA,
            )
        ]
    )
    base_markup = InlineKeyboardMarkup(inline_keyboard=base_grid_rows)
    grid_message_text = "Тест Корси: Запоминание Последовательности"

    try:  # Send/Edit Grid Message
        if grid_msg_id:
            await bot_instance.edit_message_text(
                chat_id=corsi_chat_id,
                message_id=grid_msg_id,
                text=grid_message_text,
                reply_markup=base_markup,
            )
        else:
            grid_msg_obj = await bot_instance.send_message(
                corsi_chat_id, grid_message_text, reply_markup=base_markup
            )
            grid_msg_id = grid_msg_obj.message_id
            await state.update_data(corsi_grid_message_id=grid_msg_id)
    except Exception as e_grid:
        logger.error(
            f"Corsi (show_sequence): Критическая ошибка при отправке/редактировании сетки: {e_grid}",
            exc_info=True,
        )
        await trigger_source_msg.answer(
            "Критическая ошибка в Тесте Корси. Пожалуйста, /start."
        )
        # Cleanup and reset FSM
        await cleanup_corsi_messages(
            state, bot_instance, "Критическая ошибка отображения сетки"
        )
        active_profile = await get_active_profile_from_fsm(state)
        await _clear_fsm_and_set_profile(state, active_profile)
        return

    status_texts = [
        "Приготовьтесь...",
        "3...",
        "2...",
        "1...",
        "Запоминайте...",
    ]
    for i, text in enumerate(status_texts):
        if await state.get_state() != CorsiTestStates.showing_sequence.state:
            return
        try:
            if not status_msg_id:
                status_msg_obj = await bot_instance.send_message(
                    corsi_chat_id, text
                )
                status_msg_id = status_msg_obj.message_id
                await state.update_data(corsi_status_message_id=status_msg_id)
            else:
                await bot_instance.edit_message_text(
                    text=text, chat_id=corsi_chat_id, message_id=status_msg_id
                )
        except TelegramBadRequest as tb_err:
            if "message is not modified" not in str(tb_err).lower():
                logger.warning(
                    f"Corsi: Не удалось отредактировать статус (ID: {status_msg_id}) на '{text}'. Попытка переотправки."
                )
                await _safe_delete_corsi_specific_message(
                    bot_instance,
                    state,
                    "corsi_status_message_id",
                    "status edit fail",
                )
                status_msg_id = None
                try:
                    status_msg_obj_retry = await bot_instance.send_message(
                        corsi_chat_id, text
                    )
                    status_msg_id = status_msg_obj_retry.message_id
                    await state.update_data(
                        corsi_status_message_id=status_msg_id
                    )
                except Exception as e_resend:
                    logger.error(
                        f"Corsi: Не удалось переотправить статус '{text}': {e_resend}"
                    )
        except Exception as e_status:
            logger.error(
                f"Corsi: Ошибка обновления статуса '{text}': {e_status}"
            )
        await asyncio.sleep(1.0 if i < len(status_texts) - 1 else 0.5)

    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        return

    for flash_count, flashed_idx in enumerate(correct_seq_to_show):
        if await state.get_state() != CorsiTestStates.showing_sequence.state:
            return
        flashed_buttons = [
            IKB(
                text=("🟨" if i == flashed_idx else "🟪"),
                callback_data=f"{CORSI_BUTTON_CALLBACK_PREFIX}{i}",
            )
            for i in range(9)
        ]
        flashed_rows = [flashed_buttons[i : i + 3] for i in range(0, 9, 3)]
        flashed_rows.append(
            [
                IKB(
                    text="⏹️ Остановить Тест Корси",
                    callback_data=CORSI_STOP_CALLBACK_DATA,
                )
            ]
        )
        flashed_markup = InlineKeyboardMarkup(inline_keyboard=flashed_rows)
        try:
            if grid_msg_id:
                await bot_instance.edit_message_reply_markup(
                    chat_id=corsi_chat_id,
                    message_id=grid_msg_id,
                    reply_markup=flashed_markup,
                )
                await asyncio.sleep(0.5)
                await bot_instance.edit_message_reply_markup(
                    chat_id=corsi_chat_id,
                    message_id=grid_msg_id,
                    reply_markup=base_markup,
                )
                if flash_count < len(correct_seq_to_show) - 1:
                    await asyncio.sleep(0.2)
            else:
                break
        except (
            Exception
        ) as e_flash:  # Catch broader exceptions during flashing
            logger.error(
                f"Corsi (show_sequence): Ошибка при подсветке: {e_flash}",
                exc_info=True,
            )
            break  # Stop flashing if an error occurs

    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        return

    input_prompt = "Повторите последовательность, нажимая на плитки:"
    status_msg_id_for_input = (await state.get_data()).get(
        "corsi_status_message_id"
    )
    try:
        if status_msg_id_for_input:
            await bot_instance.edit_message_text(
                text=input_prompt,
                chat_id=corsi_chat_id,
                message_id=status_msg_id_for_input,
            )
        else:
            new_status_msg = await bot_instance.send_message(
                corsi_chat_id, input_prompt
            )
            await state.update_data(
                corsi_status_message_id=new_status_msg.message_id
            )
    except Exception as e_prompt:
        logger.error(f"Corsi: Ошибка установки промпта для ввода: {e_prompt}")

    await state.update_data(sequence_start_time=time.time())
    await state.set_state(CorsiTestStates.waiting_for_user_sequence)
    logger.info(
        f"Corsi: Последовательность длиной {current_sequence_length} показана. Ожидание ввода."
    )


@router.callback_query(
    F.data.startswith(CORSI_BUTTON_CALLBACK_PREFIX),
    StateFilter(CorsiTestStates.waiting_for_user_sequence),
)
async def handle_corsi_button_press(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()

    button_idx_pressed = int(callback.data.split("_")[-1])
    data = await state.get_data()
    user_sequence = data.get("user_input_sequence", []) + [button_idx_pressed]
    await state.update_data(user_input_sequence=user_sequence)

    grid_msg_id = data.get("corsi_grid_message_id")
    chat_id = data.get("corsi_chat_id")
    correct_sequence = data.get("correct_sequence", [])

    if not (grid_msg_id and chat_id and correct_sequence is not None):
        logger.error(
            "Corsi (button_press): Важные данные FSM (grid_id, chat_id, correct_sequence) отсутствуют."
        )
        if callback.message:
            await callback.message.answer(
                "Произошла ошибка с Тестом Корси. Пожалуйста, /start."
            )
        # Cleanup UI and FSM
        await cleanup_corsi_messages(
            state,
            bot,
            "Критическая ошибка: отсутствуют данные FSM при нажатии кнопки",
        )
        active_profile = await get_active_profile_from_fsm(state)
        await _clear_fsm_and_set_profile(state, active_profile)
        return

    pressed_buttons_markup_rows = []
    for r in range(3):
        row = [
            IKB(
                text=("🟨" if (r * 3 + c) in user_sequence else "🟪"),
                callback_data=f"{CORSI_BUTTON_CALLBACK_PREFIX}{r * 3 + c}",
            )
            for c in range(3)
        ]
        pressed_buttons_markup_rows.append(row)
    pressed_buttons_markup_rows.append(
        [
            IKB(
                text="⏹️ Остановить Тест Корси",
                callback_data=CORSI_STOP_CALLBACK_DATA,
            )
        ]
    )
    try:
        await bot.edit_message_reply_markup(
            chat_id=chat_id,
            message_id=grid_msg_id,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=pressed_buttons_markup_rows
            ),
        )
    except TelegramBadRequest:
        logger.warning(
            "Corsi (button_press): Не удалось обновить визуализацию сетки."
        )

    if len(user_sequence) == len(correct_sequence):
        await evaluate_user_sequence(callback.message, state, bot)


async def evaluate_user_sequence(
    trigger_message: Message, state: FSMContext, bot_instance: Bot
):
    current_fsm_state = await state.get_state()
    if current_fsm_state != CorsiTestStates.waiting_for_user_sequence.state:
        logger.warning(
            f"evaluate_user_sequence вызван, но состояние '{current_fsm_state}'. Прерывание."
        )
        return

    data = await state.get_data()
    chat_id = data.get("corsi_chat_id", trigger_message.chat.id)
    user_seq = data.get("user_input_sequence", [])
    correct_seq = data.get("correct_sequence", [])
    current_len = data.get("current_sequence_length", 0)
    error_count = data.get("error_count", 0)
    sequence_times_history = data.get("sequence_times", [])
    seq_start_time = data.get("sequence_start_time", 0)
    feedback_msg_id_from_fsm = data.get("corsi_feedback_message_id")

    time_taken = (time.time() - seq_start_time) if seq_start_time > 0 else 0.0

    async def _update_feedback(text: str, is_bold: bool = False) -> None:
        nonlocal feedback_msg_id_from_fsm
        if await state.get_state() not in [
            CorsiTestStates.waiting_for_user_sequence.state,
            CorsiTestStates.showing_sequence.state,
        ]:
            return
        effective_pm = ParseMode.HTML if is_bold else None
        try:
            if feedback_msg_id_from_fsm:
                await bot_instance.edit_message_text(
                    text=text,
                    chat_id=chat_id,
                    message_id=feedback_msg_id_from_fsm,
                    parse_mode=effective_pm,
                )
            else:
                msg = await bot_instance.send_message(
                    chat_id, text, parse_mode=effective_pm
                )
                feedback_msg_id_from_fsm = msg.message_id
                await state.update_data(
                    corsi_feedback_message_id=feedback_msg_id_from_fsm
                )
        except TelegramBadRequest as tb_err:
            if "message is not modified" not in str(tb_err).lower():
                logger.warning(
                    f"Corsi eval: Не удалось изменить фидбэк '{text}'. Попытка переотправки. Ошибка: {tb_err}"
                )
                # Use the main _safe_delete_message from common_handlers, or a local one for Corsi
                await _safe_delete_corsi_specific_message(
                    bot_instance,
                    state,
                    "corsi_feedback_message_id",
                    "feedback update fail",
                )
                feedback_msg_id_from_fsm = None
                try:
                    msg_retry = await bot_instance.send_message(
                        chat_id, text, parse_mode=effective_pm
                    )
                    feedback_msg_id_from_fsm = msg_retry.message_id
                    await state.update_data(
                        corsi_feedback_message_id=feedback_msg_id_from_fsm
                    )
                except Exception as e_resend_fb:
                    logger.error(
                        f"Corsi eval: Не удалось переотправить фидбэк '{text}': {e_resend_fb}"
                    )
        except Exception as e_fb:
            logger.error(
                f"Corsi eval: Ошибка обновления фидбэка '{text}': {e_fb}"
            )

    next_len_to_try = current_len
    next_error_count = error_count
    test_should_continue = True

    if user_seq == correct_seq:
        sequence_times_history.append({"len": current_len, "time": time_taken})
        next_len_to_try = current_len + 1
        next_error_count = 0
        await _update_feedback("<b>Верно!</b> ✅", is_bold=True)
        delayed_msg = "Следующая последовательность..."
        if next_len_to_try > CORSI_MAX_SEQUENCE_LENGTH:
            test_should_continue = False
            delayed_msg = f"Верно! Достигнута максимальная длина ({current_len}). Тест завершен."
    else:
        next_error_count += 1
        await _update_feedback("<b>Ошибка!</b> ❌", is_bold=True)
        delayed_msg = f"Ошибка! Попробуйте еще раз последовательность длиной {current_len}."
        if next_error_count >= 2:
            test_should_continue = False
            delayed_msg = f"Ошибка! ({next_error_count}-я на длине {current_len}). Тест завершен."

    await state.update_data(
        current_sequence_length=next_len_to_try,
        error_count=next_error_count,
        sequence_times=sequence_times_history,
        user_input_sequence=[],
        corsi_feedback_message_id=feedback_msg_id_from_fsm,
    )

    await asyncio.sleep(1.2 if test_should_continue else 1.8)
    if await state.get_state() not in [
        CorsiTestStates.waiting_for_user_sequence.state,
        CorsiTestStates.showing_sequence.state,
    ]:
        return
    await _update_feedback(delayed_msg)
    if (
        await state.get_state()
        != CorsiTestStates.waiting_for_user_sequence.state
    ):
        return

    if test_should_continue:
        await state.set_state(CorsiTestStates.showing_sequence)
        await show_corsi_sequence(trigger_message, state, bot_instance)
    else:
        logger.info(f"Тест Корси завершается для чата {chat_id}.")
        await save_corsi_results(
            trigger_message, state, bot_instance, is_interrupted=False
        )
        await cleanup_corsi_messages(
            state, bot_instance, "Тест Корси штатно завершен."
        )

        # Получаем профиль перед его финальной установкой и удалением common_status_msg
        profile_to_set = await get_active_profile_from_fsm(state)

        # Удаляем общее сообщение "Подготовка к тесту..."
        fsm_data_for_common_msg_del = await state.get_data()
        common_status_msg_id = fsm_data_for_common_msg_del.get(
            "status_message_id_to_delete_later"
        )
        if common_status_msg_id and chat_id:
            # Используем общую функцию _safe_delete_message, т.к. это не специфичное для Корси сообщение
            await _safe_delete_message(
                bot_instance,
                chat_id,
                common_status_msg_id,
                "Corsi normal completion common status",
            )
            # Ключ status_message_id_to_delete_later будет удален _clear_fsm_and_set_profile
            # или если profile_to_set его не содержит.

        # Очищаем FSM и устанавливаем только профиль
        await _clear_fsm_and_set_profile(
            state, profile_to_set
        )  # Устанавливает state=None

        if (
            profile_to_set
        ):  # Проверяем, что профиль валиден после всех операций
            await send_main_action_menu(
                bot_instance,
                trigger_message,
                ACTION_SELECTION_KEYBOARD_RETURNING,
                text="Тест Корси завершен. Выберите следующее действие:",
            )
        else:
            logger.warning(
                f"Corsi eval: Тест завершен, но активный профиль не найден в FSM для чата {chat_id}."
            )
            await trigger_message.answer(
                "Тест Корси завершен. Ваш профиль не найден. Пожалуйста, /start."
            )
            # FSM уже очищен _clear_fsm_and_set_profile


async def start_corsi_test(
    trigger_event: Union[Message, CallbackQuery],
    state: FSMContext,
    profile: Dict[
        str, Any
    ],  # Profile data from get_active_profile_from_fsm (standardized keys)
    bot_instance: Bot,
):
    source_message = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    test_chat_id = source_message.chat.id

    await state.set_state(CorsiTestStates.showing_sequence)

    uid = profile.get("unique_id")  # Используем стандартизированные ключи
    name = profile.get("name")
    age = profile.get("age")
    tg_id = profile.get("telegram_id")

    initial_test_data = {
        "unique_id_for_test": uid,
        "profile_name_for_test": name,
        "profile_age_for_test": age,
        "profile_telegram_id_for_test": tg_id,
        "corsi_chat_id": test_chat_id,
        "current_sequence_length": 2,
        "error_count": 0,
        "sequence_times": [],
        "correct_sequence": [],
        "user_input_sequence": [],
        "sequence_start_time": 0,
        "corsi_grid_message_id": None,
        "corsi_status_message_id": None,
        "corsi_feedback_message_id": None,
    }
    # Добавляем к существующим данным FSM (профиль уже должен быть там с active_* ключами,
    # и status_message_id_to_delete_later от common_handlers)
    await state.update_data(**initial_test_data)
    logger.info(f"Тест Корси запущен для UID {uid} в чате {test_chat_id}.")

    await show_corsi_sequence(source_message, state, bot_instance)


async def save_corsi_results(
    trigger_msg_context: Message,
    state: FSMContext,
    bot_instance: Bot,
    is_interrupted: bool = False,
):
    data = await state.get_data()

    uid = data.get("unique_id_for_test")
    p_name = data.get("profile_name_for_test")
    p_age = data.get("profile_age_for_test")
    p_tgid = data.get("profile_telegram_id_for_test")

    if not uid:
        logger.warning(
            "Corsi save: unique_id_for_test не найден в FSM. Попытка извлечь из активного профиля."
        )
        active_profile = await get_active_profile_from_fsm(
            state
        )  # Вернет стандартизированные ключи
        if active_profile and active_profile.get("unique_id"):
            uid = active_profile.get("unique_id")
            p_name = active_profile.get("name", p_name)
            p_age = active_profile.get("age", p_age)
            p_tgid = active_profile.get("telegram_id", p_tgid)
            logger.info(
                f"Corsi save: Используются данные из активного профиля для UID {uid}."
            )
        else:
            logger.error(
                "Corsi save: КРИТИЧЕСКАЯ ОШИБКА - Не найден UID для сохранения результатов."
            )
            if (
                await state.get_state() is not None
            ):  # Only send if state is somewhat active
                await trigger_msg_context.answer(
                    "Тест Корси: Ошибка сохранения (ID пользователя не найден)."
                )
            return

    seq_times = data.get("sequence_times", [])
    max_len = 0
    if seq_times and all(
        isinstance(item, dict) and "len" in item for item in seq_times
    ):
        max_len = max(item["len"] for item in seq_times) if seq_times else 0

    avg_time_per_el = 0.0
    valid_times = [
        item
        for item in seq_times
        if isinstance(item, dict)
        and item.get("len", 0) > 0
        and isinstance(item.get("time"), (int, float))
        and item.get("time", -1) >= 0
    ]
    if valid_times:
        try:
            avg_time_per_el = sum(
                item["time"] / item["len"] for item in valid_times
            ) / len(valid_times)
        except ZeroDivisionError:
            avg_time_per_el = 0.0

    seq_details = (
        "; ".join(
            [f"Дл.{item['len']}-{item['time']:.2f}с" for item in valid_times]
        )
        or "Нет данных"
    )
    interrupted_str = "Да" if is_interrupted else "Нет"

    try:
        from openpyxl import load_workbook

        if not os.path.exists(EXCEL_FILENAME):
            logger.error(
                f"Corsi save: Файл Excel '{EXCEL_FILENAME}' не найден."
            )
            if await state.get_state() is not None:
                await trigger_msg_context.answer(
                    f"Ошибка: Файл для сохранения ('{EXCEL_FILENAME}') не найден."
                )
            return

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        excel_headers = [cell.value for cell in ws[1]]
        if "Unique ID" not in excel_headers:
            raise ValueError("Столбец 'Unique ID' не найден в Excel.")

        uid_col_excel_idx = excel_headers.index("Unique ID")
        target_row_num_excel = -1
        for r_num_excel, row_data_tuple in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if (
                len(row_data_tuple) > uid_col_excel_idx
                and row_data_tuple[uid_col_excel_idx] is not None
                and str(row_data_tuple[uid_col_excel_idx]) == str(uid)
            ):
                target_row_num_excel = r_num_excel
                break

        if target_row_num_excel == -1:
            logger.info(
                f"Corsi save: UID {uid} не найден, добавление новой строки."
            )
            new_row_values = [""] * len(
                excel_headers
            )  # Base on actual excel headers count
            new_row_values[uid_col_excel_idx] = (
                uid  # Set UID in its actual column
            )
            # Populate other known base headers if they exist in excel_headers
            if p_name and "Name" in excel_headers:
                new_row_values[excel_headers.index("Name")] = p_name
            if p_age and "Age" in excel_headers:
                new_row_values[excel_headers.index("Age")] = p_age
            if p_tgid and "Telegram ID" in excel_headers:
                new_row_values[excel_headers.index("Telegram ID")] = p_tgid
            ws.append(new_row_values)
            target_row_num_excel = ws.max_row

        corsi_headers_map = {}
        for h_name in [
            "Corsi - Max Correct Sequence Length",
            "Corsi - Avg Time Per Element (s)",
            "Corsi - Sequence Times Detail",
            "Corsi - Interrupted",
        ]:
            if h_name in excel_headers:
                corsi_headers_map[h_name] = excel_headers.index(h_name) + 1
            else:
                logger.warning(
                    f"Corsi save: Заголовок '{h_name}' не найден в Excel. Пропуск."
                )

        if "Corsi - Max Correct Sequence Length" in corsi_headers_map:
            ws.cell(
                row=target_row_num_excel,
                column=corsi_headers_map[
                    "Corsi - Max Correct Sequence Length"
                ],
            ).value = max_len
        if "Corsi - Avg Time Per Element (s)" in corsi_headers_map:
            ws.cell(
                row=target_row_num_excel,
                column=corsi_headers_map["Corsi - Avg Time Per Element (s)"],
            ).value = round(avg_time_per_el, 2)
        if "Corsi - Sequence Times Detail" in corsi_headers_map:
            ws.cell(
                row=target_row_num_excel,
                column=corsi_headers_map["Corsi - Sequence Times Detail"],
            ).value = seq_details
        if "Corsi - Interrupted" in corsi_headers_map:
            ws.cell(
                row=target_row_num_excel,
                column=corsi_headers_map["Corsi - Interrupted"],
            ).value = interrupted_str

        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Результаты Теста Корси для UID {uid} (Прерван: {is_interrupted}) сохранены."
        )

    except FileNotFoundError:
        logger.error(
            f"Corsi save: Файл Excel '{EXCEL_FILENAME}' не найден (повторно)."
        )
        if await state.get_state() is not None:
            await trigger_msg_context.answer(
                f"Критическая ошибка: Файл для сохранения ('{EXCEL_FILENAME}') не найден."
            )
    except ValueError as ve_excel:
        logger.error(
            f"Corsi save: Ошибка конфигурации Excel для UID {uid}: {ve_excel}",
            exc_info=True,
        )
        if await state.get_state() is not None:
            await trigger_msg_context.answer(
                "Ошибка конфигурации при сохранении Теста Корси."
            )
    except Exception as e_excel_gen:
        logger.error(
            f"Corsi save: Общая ошибка сохранения в Excel для UID {uid}: {e_excel_gen}",
            exc_info=True,
        )
        if await state.get_state() is not None:
            await trigger_msg_context.answer(
                "Непредвиденная ошибка при сохранении Теста Корси."
            )

    # Send summary message to user if test was not abruptly stopped (i.e., state is still somewhat valid)
    if await state.get_state() is not None:
        summary_status = (
            "ПРЕРВАНЫ И СОХРАНЕНЫ" if is_interrupted else "УСПЕШНО СОХРАНЕНЫ"
        )
        msg_parts = [
            f"Результаты Теста Корси для UID {uid} <b>{summary_status}</b>:"
        ]
        if not (is_interrupted and max_len == 0 and not valid_times):
            msg_parts.extend(
                [
                    f"— Максимальная длина верной последовательности: {max_len}",
                    f"— Среднее время на элемент: {round(avg_time_per_el, 2)} сек",
                ]
            )
            if seq_details != "Нет данных":
                msg_parts.append(f"— Детализация по сериям: {seq_details}")
            elif not is_interrupted and not valid_times:
                msg_parts.append(
                    "— Детализация по сериям: Нет выполненных верных последовательностей."
                )
        else:
            msg_parts = [
                f"Тест Корси для UID {uid} был <b>ПРЕРВАН</b> досрочно. Результаты не зафиксированы."
            ]
        try:
            await trigger_msg_context.answer(
                "\n".join(msg_parts), parse_mode=ParseMode.HTML
            )
        except Exception as e_ans:
            logger.error(
                f"Corsi save: Не удалось отправить итог (UID {uid}): {e_ans}"
            )


@router.callback_query(
    F.data == CORSI_STOP_CALLBACK_DATA, StateFilter(CorsiTestStates)
)  # Changed to generic stop
async def on_corsi_stop_button_generic(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    logger.info(
        f"Тест Корси: запрос на остановку через кнопку (UID: {callback.from_user.id})."
    )
    from handlers.common_handlers import stop_test_command_handler

    await callback.answer("Останавливаю Тест Корси...", show_alert=False)
    await stop_test_command_handler(
        trigger_event=callback,
        state=state,
        bot=bot,
        called_from_test_button=True,
    )
