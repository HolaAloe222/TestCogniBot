# handlers/tests/reaction_time_handlers.py
import asyncio
import logging
import random
import time
import os

from aiogram import Bot, F, Router
from aiogram.enums import ParseMode, ChatType
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    FSInputFile,
    InputMediaPhoto,
    Message,
    Chat,
    User,
)
from aiogram.filters import StateFilter

from fsm_states import ReactionTimeTestStates
from settings import (
    ALL_EXPECTED_HEADERS,
    EXCEL_FILENAME,
    REACTION_TIME_IMAGE_POOL,  # Populated in main_bot.py
    REACTION_TIME_MEMORIZATION_S,
    REACTION_TIME_STIMULUS_INTERVAL_S,
    REACTION_TIME_MAX_ATTEMPTS,
    REACTION_TIME_NUM_STIMULI_IN_SEQUENCE,
)
from utils.bot_helpers import (
    send_main_action_menu,
    get_active_profile_from_fsm,
)
from keyboards import ACTION_SELECTION_KEYBOARD_RETURNING

logger = logging.getLogger(__name__)
router = Router()
IKB = InlineKeyboardButton


async def _delete_common_status_message(
    state: FSMContext, bot_instance: Bot, chat_id: int | None = None
):
    """Helper to delete the common 'status_message_id_to_delete_later'."""
    data = await state.get_data()
    common_status_msg_id = data.get("status_message_id_to_delete_later")
    # Ensure chat_id is available, try from FSM first, then passed arg
    current_chat_id = data.get("rt_chat_id") or chat_id

    if common_status_msg_id and current_chat_id:
        try:
            await bot_instance.delete_message(
                current_chat_id, common_status_msg_id
            )
            logger.info(
                f"RT Common: Deleted common status message ID: {common_status_msg_id} in chat {current_chat_id}"
            )
        except TelegramBadRequest:
            logger.warning(
                f"RT Common: Failed to delete common status message ID: {common_status_msg_id}, already deleted or inaccessible."
            )
        except Exception as e_del_common:
            logger.error(
                f"RT Common: Error deleting common status message ID {common_status_msg_id}: {e_del_common}"
            )
        await state.update_data(status_message_id_to_delete_later=None)


async def _rt_go_to_main_menu_or_clear(
    state: FSMContext, trigger_message: Message, bot_instance: Bot
):
    """Navigates to the main menu or clears state if no profile. Called after all cleanups."""
    fsm_data = (
        await state.get_data()
    )  # Should ideally contain only profile data now
    profile_data = {
        key: fsm_data.get(key)
        for key in [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        if fsm_data.get(key)
    }
    await state.set_state(None)

    if profile_data.get("active_unique_id"):
        await state.set_data(
            profile_data
        )  # Ensure only profile data remains in FSM
        await send_main_action_menu(
            bot_instance, trigger_message, ACTION_SELECTION_KEYBOARD_RETURNING
        )
    else:
        if (
            hasattr(trigger_message, 'chat') and trigger_message.chat
        ):  # Ensure valid message context
            await trigger_message.answer(
                "Профиль не активен. Пожалуйста, /start для начала."
            )
        await state.clear()


async def _rt_memorization_phase_task(state: FSMContext, bot_instance: Bot):
    """Handles the memorization phase countdown and transitions to reaction phase."""
    try:
        await asyncio.sleep(REACTION_TIME_MEMORIZATION_S)
        if (
            await state.get_state()
            != ReactionTimeTestStates.memorization_display.state
        ):
            logger.info("RT Memorization task: State changed, aborting.")
            return

        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        memo_msg_id = data.get("rt_memorization_image_message_id")

        if memo_msg_id and chat_id:
            try:
                await bot_instance.delete_message(
                    chat_id=chat_id, message_id=memo_msg_id
                )
                await state.update_data(rt_memorization_image_message_id=None)
                logger.debug(
                    f"RT Memo task: Deleted memorization image message {memo_msg_id}."
                )
            except TelegramBadRequest:
                logger.debug(
                    f"RT Memo task: Memorization image message {memo_msg_id} already deleted."
                )
            except Exception as e_del_memo:
                logger.error(
                    f"RT Memo task: Error deleting memo_msg_id {memo_msg_id}: {e_del_memo}"
                )

        # The message with "Подготовка к фазе запоминания..." (rt_instruction_message_id)
        # will be deleted by cleanup_reaction_time_ui as part of its general cleanup.

        await _start_rt_reaction_phase(state, bot_instance)
    except asyncio.CancelledError:
        logger.info("RT Memorization task cancelled.")
    except Exception as e:
        logger.error(
            f"RT Memorization task critical error: {e}", exc_info=True
        )
        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                "Произошла критическая ошибка в фазе запоминания. Тест прерван.",
            )

        await save_reaction_time_results(
            state,
            is_interrupted=True,
            status_override="Failed in memorization phase",
        )
        await cleanup_reaction_time_ui(
            state, bot_instance, final_text=None
        )  # Delete all UI
        await _delete_common_status_message(
            state, bot_instance, chat_id
        )  # Delete "Запускаем тест..."

        mock_message = None
        if chat_id:  # Create mock message for menu navigation
            mock_user = User(
                id=bot_instance.id if hasattr(bot_instance, 'id') else 1,
                is_bot=True,
                first_name="Bot",
            )
            mock_chat_obj = Chat(id=chat_id, type=ChatType.PRIVATE)
            mock_message = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat_obj,
                from_user=mock_user,
                text="mock",
            )

        if mock_message:
            await _rt_go_to_main_menu_or_clear(
                state, mock_message, bot_instance
            )
        else:
            await state.clear()


async def _start_rt_reaction_phase(state: FSMContext, bot_instance: Bot):
    """Sets up and starts the reaction stimulus display cycle."""
    await state.set_state(ReactionTimeTestStates.reaction_stimulus_display)
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    target_image_path = data.get("rt_target_image_path")

    if not REACTION_TIME_IMAGE_POOL:
        logger.error(
            "RT Start Reaction Phase: REACTION_TIME_IMAGE_POOL is empty! Cannot proceed."
        )
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                "Ошибка конфигурации: пул изображений для теста пуст. Тест прерван.",
            )
        await _handle_rt_attempt_failure(
            state, bot_instance, "Ошибка конфигурации: пул изображений пуст"
        )
        return

    distractors = [
        p for p in REACTION_TIME_IMAGE_POOL if p != target_image_path
    ]
    random.shuffle(distractors)

    stimuli_sequence = []
    num_distractors_needed = REACTION_TIME_NUM_STIMULI_IN_SEQUENCE - 1
    actual_num_distractors = min(len(distractors), num_distractors_needed)
    chosen_distractors = distractors[:actual_num_distractors]
    stimuli_sequence.extend(
        [{"path": p, "is_target": False} for p in chosen_distractors]
    )

    if (
        REACTION_TIME_NUM_STIMULI_IN_SEQUENCE > 0
    ):  # Only add target if sequence length is > 0
        target_insert_pos = (
            random.randint(0, len(stimuli_sequence)) if stimuli_sequence else 0
        )
        stimuli_sequence.insert(
            target_insert_pos, {"path": target_image_path, "is_target": True}
        )

    stimuli_sequence = stimuli_sequence[
        :REACTION_TIME_NUM_STIMULI_IN_SEQUENCE
    ]  # Trim to max length

    if (
        REACTION_TIME_NUM_STIMULI_IN_SEQUENCE > 0
        and stimuli_sequence
        and not any(s['is_target'] for s in stimuli_sequence)
    ):
        idx_to_replace = random.randrange(len(stimuli_sequence))
        stimuli_sequence[idx_to_replace] = {
            "path": target_image_path,
            "is_target": True,
        }
        logger.warning(
            "RT Start Reaction Phase: Target safeguard triggered (ensured target presence)."
        )

    if not stimuli_sequence and REACTION_TIME_NUM_STIMULI_IN_SEQUENCE > 0:
        logger.error(
            "RT Start Reaction Phase: Stimulus sequence is empty despite non-zero num_stimuli. Aborting."
        )
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                "Критическая ошибка генерации стимулов. Попытка прервана.",
            )
        await _handle_rt_attempt_failure(
            state, bot_instance, "Критическая ошибка генерации стимулов"
        )
        return

    await state.update_data(
        rt_stimuli_sequence=stimuli_sequence,
        rt_current_stimulus_index=0,
        rt_target_displayed_time=None,
        rt_reacted_correctly_this_attempt=False,
        rt_reaction_stimulus_message_id=None,
        rt_target_missed_message_id=None,
    )
    reaction_task = asyncio.create_task(
        _rt_reaction_cycle_task(state, bot_instance)
    )
    await state.update_data(rt_reaction_cycle_task=reaction_task)


async def _rt_reaction_cycle_task(state: FSMContext, bot_instance: Bot):
    """Manages the display of individual stimuli and waits for reactions or interval timeout."""
    try:
        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        stimuli_sequence = data.get("rt_stimuli_sequence", [])
        current_idx = data.get("rt_current_stimulus_index", 0)
        stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")

        if current_idx >= len(stimuli_sequence):  # All stimuli shown
            if data.get("rt_target_displayed_time") and not data.get(
                "rt_reacted_correctly_this_attempt"
            ):
                logger.info(
                    f"RT UID {data.get('rt_unique_id_for_test', 'N/A')}: Target missed (end of sequence)."
                )
                if chat_id:
                    target_missed_msg = await bot_instance.send_message(
                        chat_id, "Вы пропустили целевое изображение."
                    )
                    await state.update_data(
                        rt_target_missed_message_id=target_missed_msg.message_id
                    )
                await _handle_rt_attempt_failure(
                    state, bot_instance, "Цель пропущена"
                )
            return

        current_stimulus = stimuli_sequence[current_idx]
        image_path = current_stimulus["path"]
        is_target = current_stimulus["is_target"]
        await state.update_data(rt_current_displayed_image_is_target=is_target)

        caption_text = "РЕАГИРОВАТЬ!"
        kbd = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    IKB(
                        text="💥 РЕАГИРОВАТЬ! 💥",
                        callback_data="rt_react_button_pressed",
                    )
                ]
            ]
        )

        try:
            img_file = FSInputFile(image_path)
            if not stimulus_msg_id:
                msg = await bot_instance.send_photo(
                    chat_id,
                    photo=img_file,
                    caption=caption_text,
                    reply_markup=kbd,
                )
                stimulus_msg_id = msg.message_id
                await state.update_data(
                    rt_reaction_stimulus_message_id=stimulus_msg_id
                )
            else:
                media = InputMediaPhoto(media=img_file, caption=caption_text)
                await bot_instance.edit_message_media(
                    chat_id=chat_id,
                    message_id=stimulus_msg_id,
                    media=media,
                    reply_markup=kbd,
                )

            if is_target:
                await state.update_data(rt_target_displayed_time=time.time())
                logger.info(
                    f"RT UID {data.get('rt_unique_id_for_test', 'N/A')}: Target '{os.path.basename(image_path)}' displayed."
                )
        except Exception as e:
            logger.error(
                f"RT Reaction Cycle: Failed to send/edit stimulus image {image_path}: {e}",
                exc_info=True,
            )
            if chat_id:
                await bot_instance.send_message(
                    chat_id, "Ошибка отображения стимула. Попытка прервана."
                )
            await _handle_rt_attempt_failure(
                state, bot_instance, "Ошибка отображения стимула"
            )
            return

        await state.update_data(rt_current_stimulus_index=current_idx + 1)

        start_sleep = time.time()
        while time.time() - start_sleep < REACTION_TIME_STIMULUS_INTERVAL_S:
            await asyncio.sleep(0.05)
            if (
                await state.get_state()
                != ReactionTimeTestStates.reaction_stimulus_display.state
            ):
                logger.info(
                    "RT Reaction Cycle: State changed during stimulus display interval, aborting cycle."
                )
                return

        if (
            await state.get_state()
            == ReactionTimeTestStates.reaction_stimulus_display.state
        ):  # If still in reaction phase (no reaction yet)
            new_reaction_task = asyncio.create_task(
                _rt_reaction_cycle_task(state, bot_instance)
            )
            await state.update_data(rt_reaction_cycle_task=new_reaction_task)

    except asyncio.CancelledError:
        logger.info("RT Reaction cycle task cancelled.")
    except Exception as e:
        logger.error(
            f"RT Reaction cycle task critical error: {e}", exc_info=True
        )
        data = await state.get_data()
        chat_id = data.get("rt_chat_id")
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                "Произошла критическая ошибка в фазе реакции. Тест прерван.",
            )

        await save_reaction_time_results(
            state,
            is_interrupted=True,
            status_override="Failed in reaction cycle (critical)",
        )
        await cleanup_reaction_time_ui(state, bot_instance, final_text=None)
        await _delete_common_status_message(state, bot_instance, chat_id)

        mock_message = None
        if chat_id:
            mock_user = User(
                id=bot_instance.id if hasattr(bot_instance, 'id') else 1,
                is_bot=True,
                first_name="Bot",
            )
            mock_chat_obj = Chat(id=chat_id, type=ChatType.PRIVATE)
            mock_message = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat_obj,
                from_user=mock_user,
                text="mock",
            )

        if mock_message:
            await _rt_go_to_main_menu_or_clear(
                state, mock_message, bot_instance
            )
        else:
            await state.clear()


async def _handle_rt_attempt_failure(
    state: FSMContext, bot_instance: Bot, reason: str
):
    """Handles a failed attempt, offering retry or ending the test."""
    data = await state.get_data()
    current_attempt = data.get("rt_current_attempt", 1)
    chat_id = data.get("rt_chat_id")

    reaction_cycle_task = data.get("rt_reaction_cycle_task")
    if reaction_cycle_task and not reaction_cycle_task.done():
        reaction_cycle_task.cancel()
        try:
            await asyncio.wait_for(reaction_cycle_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass
        await state.update_data(rt_reaction_cycle_task=None)

    # Delete "Вы пропустили целевое изображение" message if it exists
    target_missed_msg_id = data.get("rt_target_missed_message_id")
    if target_missed_msg_id and chat_id:
        try:
            await bot_instance.delete_message(chat_id, target_missed_msg_id)
            await state.update_data(rt_target_missed_message_id=None)
            logger.debug(
                f"RT Handle Failure: Deleted target_missed_message_id: {target_missed_msg_id}"
            )
        except TelegramBadRequest:
            logger.debug(
                f"RT Handle Failure: target_missed_message_id {target_missed_msg_id} already deleted."
            )
        except Exception as e_del_missed:
            logger.error(
                f"RT Handle Failure: Error deleting target_missed_message_id {target_missed_msg_id}: {e_del_missed}"
            )

    # Delete the stimulus image message (rt_reaction_stimulus_message_id)
    stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")
    if stimulus_msg_id and chat_id:
        try:
            await bot_instance.delete_message(chat_id, stimulus_msg_id)
            await state.update_data(rt_reaction_stimulus_message_id=None)
            logger.debug(
                f"RT Handle Failure: Deleted reaction_stimulus_message_id: {stimulus_msg_id}"
            )
        except TelegramBadRequest:
            logger.debug(
                f"RT Handle Failure: reaction_stimulus_message_id {stimulus_msg_id} already deleted."
            )
        except Exception as e_del_stim_fail:
            logger.error(
                f"RT Handle Failure: Error deleting reaction_stimulus_message_id {stimulus_msg_id}: {e_del_stim_fail}"
            )

    current_attempt += 1
    await state.update_data(rt_current_attempt=current_attempt)

    if current_attempt <= REACTION_TIME_MAX_ATTEMPTS:
        await state.set_state(
            ReactionTimeTestStates.awaiting_retry_confirmation
        )
        retry_text = (
            f"Причина: {reason}. Попытка {current_attempt - 1} из {REACTION_TIME_MAX_ATTEMPTS} не удалась.\n"
            f"Хотите попробовать еще раз (осталось {REACTION_TIME_MAX_ATTEMPTS - (current_attempt - 1)} попыток)?"
        )
        kbd = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    IKB(
                        text="Да, попробовать снова",
                        callback_data="rt_retry_yes",
                    )
                ],
                [IKB(text="Нет, завершить тест", callback_data="rt_retry_no")],
            ]
        )
        if chat_id:
            try:
                retry_prompt_msg = await bot_instance.send_message(
                    chat_id, retry_text, reply_markup=kbd
                )
                await state.update_data(
                    rt_retry_confirmation_message_id=retry_prompt_msg.message_id
                )
            except Exception as e_send_retry_prompt:
                logger.error(
                    f"RT Handle Failure: Failed to send retry prompt: {e_send_retry_prompt}",
                    exc_info=True,
                )
                # Critical failure in UI, abort test
                await save_reaction_time_results(
                    state,
                    is_interrupted=True,
                    status_override="UI error on retry prompt",
                )
                await cleanup_reaction_time_ui(
                    state, bot_instance, final_text=None
                )
                await _delete_common_status_message(
                    state, bot_instance, chat_id
                )
                mock_message_crit_fail = Message(
                    message_id=0,
                    date=int(time.time()),
                    chat=Chat(id=chat_id, type=ChatType.PRIVATE),
                    from_user=User(id=1, is_bot=True, first_name="Bot"),
                    text="mock",
                )
                await _rt_go_to_main_menu_or_clear(
                    state, mock_message_crit_fail, bot_instance
                )
                return
    else:  # Max attempts exhausted
        logger.info(
            f"RT UID {data.get('rt_unique_id_for_test', 'N/A')}: Max attempts reached. Reason: {reason}."
        )
        await state.update_data(rt_status="Failed")
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                f"Причина: {reason}. Максимум попыток ({REACTION_TIME_MAX_ATTEMPTS}) исчерпано. Тест не пройден.",
            )

        await save_reaction_time_results(
            state, is_interrupted=False
        )  # Test ended due to exhaustion
        await cleanup_reaction_time_ui(
            state, bot_instance, final_text=None
        )  # Delete all test UI
        await _delete_common_status_message(
            state, bot_instance, chat_id
        )  # Delete "Запускаем тест..."

        mock_message_max = Message(
            message_id=0,
            date=int(time.time()),
            chat=Chat(id=chat_id, type=ChatType.PRIVATE),
            from_user=User(id=1, is_bot=True, first_name="Bot"),
            text="mock",
        )
        await _rt_go_to_main_menu_or_clear(
            state, mock_message_max, bot_instance
        )


async def start_reaction_time_test(
    trigger_event: Message | CallbackQuery,
    state: FSMContext,
    profile: dict,
    bot_instance: Bot,
):
    logger.info(
        f"Starting Reaction Time Test for UID: {profile.get('unique_id', 'N/A')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    await state.set_state(ReactionTimeTestStates.initial_instructions)
    await state.update_data(
        rt_unique_id_for_test=profile.get("unique_id"),
        rt_profile_name_for_test=profile.get("name"),
        rt_profile_age_for_test=profile.get("age"),
        rt_profile_telegram_id_for_test=profile.get("telegram_id"),
        rt_chat_id=chat_id,
        rt_current_attempt=1,
        rt_reaction_time_ms=None,
        rt_status="Pending",
        rt_target_image_path=None,
        rt_instruction_message_id=None,
        rt_memorization_image_message_id=None,
        rt_reaction_stimulus_message_id=None,
        rt_retry_confirmation_message_id=None,
        rt_target_missed_message_id=None,
        rt_current_displayed_image_is_target=False,
        rt_target_displayed_time=None,
        rt_reacted_correctly_this_attempt=False,
        rt_stimuli_sequence=[],
        rt_current_stimulus_index=0,
        rt_memorization_task=None,
        rt_reaction_cycle_task=None,
    )
    instruction_text = (
        "<b>Тест на Скорость Реакции</b>\n\n"
        "1. Сначала вам будет показано изображение-цель на 10 секунд. Запомните его.\n"
        "2. Затем изображение-цель исчезнет.\n"
        "3. После этого начнут появляться другие изображения. Среди них один раз появится ваше целевое изображение.\n"
        "4. Ваша задача – как можно быстрее нажать кнопку 'РЕАГИРОВАТЬ!', как только увидите целевое изображение.\n"
        "   Если вы нажмете на кнопку, когда показано НЕ целевое изображение, это будет считаться ошибкой.\n\n"
        f"У вас будет {REACTION_TIME_MAX_ATTEMPTS} попытки. Тест начнется после нажатия кнопки 'Начать'."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать Тест", callback_data="rt_ack_instructions")]
        ]
    )
    try:
        instr_msg = await bot_instance.send_message(
            chat_id,
            instruction_text,
            reply_markup=kbd,
            parse_mode=ParseMode.HTML,
        )
        await state.update_data(rt_instruction_message_id=instr_msg.message_id)
    except Exception as e:
        logger.error(
            f"RT start_reaction_time_test: Error sending initial instructions: {e}",
            exc_info=True,
        )
        await bot_instance.send_message(
            chat_id,
            "Ошибка при запуске теста на Скорость Реакции. Попробуйте снова.",
        )
        await state.clear()


@router.callback_query(
    F.data == "rt_ack_instructions",
    ReactionTimeTestStates.initial_instructions,
)
async def rt_on_instructions_acknowledged(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    instruction_msg_id = data.get(
        "rt_instruction_message_id"
    )  # This is the message with "Начать тест"

    if instruction_msg_id and chat_id:
        try:
            await bot.edit_message_text(
                "Подготовка к фазе запоминания...",
                chat_id=chat_id,
                message_id=instruction_msg_id,
                reply_markup=None,
            )
            # This message is now "Подготовка...", its ID is still rt_instruction_message_id
        except TelegramBadRequest:
            logger.debug(
                f"RT Ack Instr: Failed to edit instruction_msg_id {instruction_msg_id}. Sending new prep message."
            )
            # Original instruction msg might still be there. cleanup_reaction_time_ui should catch it based on FSM key.
            try:
                prep_msg = await bot.send_message(
                    chat_id, "Подготовка к фазе запоминания..."
                )
                await state.update_data(
                    rt_instruction_message_id=prep_msg.message_id
                )  # Update FSM to new prep message
            except Exception as e_send_prep_fallback:
                logger.error(
                    f"RT Ack Instr: Failed to send fallback 'Подготовка...' message: {e_send_prep_fallback}",
                    exc_info=True,
                )
                await _rt_go_to_main_menu_or_clear(
                    state, callback.message, bot
                )  # Abort if can't send critical UI
                return
        except Exception as e_edit_instr_fatal:
            logger.error(
                f"RT Ack Instr: Fatal error editing instruction_msg_id {instruction_msg_id}: {e_edit_instr_fatal}",
                exc_info=True,
            )
            await _rt_go_to_main_menu_or_clear(
                state, callback.message, bot
            )  # Abort
            return

    await state.set_state(ReactionTimeTestStates.memorization_display)
    if not REACTION_TIME_IMAGE_POOL:
        logger.error(
            "RT Ack Instr: REACTION_TIME_IMAGE_POOL is empty! Aborting test."
        )
        if chat_id:
            await bot.send_message(
                chat_id,
                "Ошибка конфигурации: Пул изображений пуст. Тест не может начаться.",
            )
        await _rt_go_to_main_menu_or_clear(state, callback.message, bot)
        return

    target_image_path = random.choice(REACTION_TIME_IMAGE_POOL)
    await state.update_data(rt_target_image_path=target_image_path)
    logger.info(
        f"RT UID {data.get('rt_unique_id_for_test', 'N/A')}: Attempt {data.get('rt_current_attempt', 1)}. Target image: {os.path.basename(target_image_path)}"
    )

    try:
        img_file = FSInputFile(target_image_path)
        memo_img_msg = await bot.send_photo(
            chat_id=chat_id,
            photo=img_file,
            caption=f"Запомните это изображение! (Исчезнет через {REACTION_TIME_MEMORIZATION_S} сек)",
        )
        await state.update_data(
            rt_memorization_image_message_id=memo_img_msg.message_id
        )
    except Exception as e_send_memo_img:
        logger.error(
            f"RT Ack Instr: Failed to send memorization image '{target_image_path}': {e_send_memo_img}",
            exc_info=True,
        )
        if chat_id:
            await bot.send_message(
                chat_id,
                "Ошибка: не удалось загрузить изображение для запоминания. Тест прерван.",
            )
        await save_reaction_time_results(
            state,
            is_interrupted=True,
            status_override="Failed to send memorization image",
        )
        await cleanup_reaction_time_ui(state, bot, final_text=None)
        await _delete_common_status_message(state, bot, chat_id)
        await _rt_go_to_main_menu_or_clear(state, callback.message, bot)
        return

    memo_task = asyncio.create_task(_rt_memorization_phase_task(state, bot))
    await state.update_data(rt_memorization_task=memo_task)


@router.callback_query(
    F.data == "rt_react_button_pressed",
    ReactionTimeTestStates.reaction_stimulus_display,
)
async def on_rt_react_button_pressed(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    data = await state.get_data()

    reaction_cycle_task = data.get("rt_reaction_cycle_task")
    if reaction_cycle_task and not reaction_cycle_task.done():
        reaction_cycle_task.cancel()
        try:
            await asyncio.wait_for(reaction_cycle_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            logger.debug(
                "RT React Pressed: Reaction cycle task cancelled/timed out."
            )
        await state.update_data(rt_reaction_cycle_task=None)

    chat_id = data.get("rt_chat_id")
    is_target_displayed_now = data.get(
        "rt_current_displayed_image_is_target", False
    )
    target_display_time = data.get("rt_target_displayed_time")
    uid_for_test = data.get('rt_unique_id_for_test', 'N/A')

    if is_target_displayed_now and target_display_time:
        reaction_time_seconds = time.time() - target_display_time

        telegram_latency_seconds = 0.350
        corrected_reaction_time_seconds = (
            reaction_time_seconds - telegram_latency_seconds
        )
        if corrected_reaction_time_seconds < 0:
            logger.warning(
                f"RT UID {uid_for_test}: Corrected RT < 0 ({corrected_reaction_time_seconds * 1000:.0f}ms). Clamped to 1ms."
            )
            corrected_reaction_time_seconds = 0.001  # Clamp to 1ms
        reaction_time_ms = int(corrected_reaction_time_seconds * 1000)

        await state.update_data(
            rt_reaction_time_ms=reaction_time_ms,
            rt_status="Passed",
            rt_reacted_correctly_this_attempt=True,
        )
        logger.info(
            f"RT UID {uid_for_test}: Correct reaction. Raw RT: {reaction_time_seconds * 1000:.0f}ms. Corrected RT: {reaction_time_ms}ms."
        )

        # Delete stimulus message (which had the button)
        stimulus_msg_id = data.get("rt_reaction_stimulus_message_id")
        if stimulus_msg_id and chat_id:
            try:
                await bot.delete_message(chat_id, stimulus_msg_id)
                await state.update_data(rt_reaction_stimulus_message_id=None)
            except TelegramBadRequest:
                logger.debug(
                    f"RT React Correct: Stimulus msg {stimulus_msg_id} already deleted."
                )
            except Exception as e_del_stim_ok:
                logger.error(
                    f"RT React Correct: Error deleting stimulus msg {stimulus_msg_id}: {e_del_stim_ok}"
                )

        if chat_id:  # Send result to user
            await bot.send_message(
                chat_id,
                f"<b>Верно!</b> Ваше время реакции: {reaction_time_ms} мс.",
                parse_mode=ParseMode.HTML,
            )

        await save_reaction_time_results(state, is_interrupted=False)
        await cleanup_reaction_time_ui(
            state, bot, final_text=None
        )  # Delete all other test UI
        await _delete_common_status_message(
            state, bot, chat_id
        )  # Delete "Запускаем тест..."

        await _rt_go_to_main_menu_or_clear(state, callback.message, bot)
    else:  # Incorrect reaction
        logger.info(
            f"RT UID {uid_for_test}: Incorrect reaction (pressed on non-target or invalid time)."
        )
        if chat_id:
            await bot.send_message(
                chat_id,
                "<b>Ошибка!</b> Вы нажали на кнопку, когда было показано НЕ целевое изображение, или реакция была за пределами допустимого окна.",
                parse_mode=ParseMode.HTML,
            )
        await _handle_rt_attempt_failure(state, bot, "Неверная реакция")


@router.callback_query(
    F.data == "rt_retry_yes",
    ReactionTimeTestStates.awaiting_retry_confirmation,
)
async def on_rt_retry_yes(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    retry_msg_id = data.get("rt_retry_confirmation_message_id")

    if retry_msg_id and chat_id:
        try:
            await bot.edit_message_text(
                "Готовим новую попытку...",
                chat_id=chat_id,
                message_id=retry_msg_id,
                reply_markup=None,
            )
            await state.update_data(
                rt_instruction_message_id=retry_msg_id,
                rt_retry_confirmation_message_id=None,
            )
        except TelegramBadRequest:
            logger.debug(
                f"RT Retry Yes: Failed to edit retry_msg_id {retry_msg_id}. Sending new."
            )
            await state.update_data(
                rt_retry_confirmation_message_id=None
            )  # Clear old ID
            try:
                new_prep_msg = await bot.send_message(
                    chat_id, "Готовим новую попытку..."
                )
                await state.update_data(
                    rt_instruction_message_id=new_prep_msg.message_id
                )
            except Exception as e_send_new_prep:
                logger.error(
                    f"RT Retry Yes: Failed to send new 'Готовим...' message: {e_send_new_prep}",
                    exc_info=True,
                )
                await _rt_go_to_main_menu_or_clear(
                    state, callback.message, bot
                )  # Abort if critical UI fails
                return
        except Exception as e_edit_retry_fatal:
            logger.error(
                f"RT Retry Yes: Fatal error editing retry_msg_id {retry_msg_id}: {e_edit_retry_fatal}",
                exc_info=True,
            )
            await _rt_go_to_main_menu_or_clear(state, callback.message, bot)
            return

    await state.update_data(  # Reset for new attempt
        rt_target_image_path=None,
        rt_memorization_image_message_id=None,
        rt_reaction_stimulus_message_id=None,
        rt_target_displayed_time=None,
        rt_reacted_correctly_this_attempt=False,
        rt_stimuli_sequence=[],
        rt_current_stimulus_index=0,
        rt_target_missed_message_id=None,
        rt_memorization_task=None,
        rt_reaction_cycle_task=None,
    )
    await state.set_state(ReactionTimeTestStates.initial_instructions)
    await rt_on_instructions_acknowledged(callback, state, bot)


@router.callback_query(
    F.data == "rt_retry_no", ReactionTimeTestStates.awaiting_retry_confirmation
)
async def on_rt_retry_no(callback: CallbackQuery, state: FSMContext, bot: Bot):
    await callback.answer()
    await state.update_data(rt_status="Failed (user declined retry)")

    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    retry_msg_id = data.get("rt_retry_confirmation_message_id")

    if retry_msg_id and chat_id:
        try:
            await bot.delete_message(chat_id=chat_id, message_id=retry_msg_id)
            await state.update_data(rt_retry_confirmation_message_id=None)
        except TelegramBadRequest:
            logger.debug(
                f"RT Retry No: Retry msg {retry_msg_id} already deleted."
            )
        except Exception as e_del_retry_no_msg:
            logger.error(
                f"RT Retry No: Error deleting retry msg {retry_msg_id}: {e_del_retry_no_msg}"
            )

    if chat_id:
        await bot.send_message(
            chat_id, "Тест завершен по вашему выбору (не пройден)."
        )

    await save_reaction_time_results(state, is_interrupted=False)
    await cleanup_reaction_time_ui(
        state, bot, final_text=None
    )  # Delete all test UI
    await _delete_common_status_message(
        state, bot, chat_id
    )  # Delete "Запускаем тест..."

    await _rt_go_to_main_menu_or_clear(state, callback.message, bot)


async def save_reaction_time_results(
    state: FSMContext,
    is_interrupted: bool = False,
    status_override: str = None,
):
    data = await state.get_data()
    uid = data.get("rt_unique_id_for_test")
    p_tgid, p_name, p_age = (
        data.get("rt_profile_telegram_id_for_test"),
        data.get("rt_profile_name_for_test"),
        data.get("rt_profile_age_for_test"),
    )

    if (
        not uid
    ):  # Fallback to general active profile if test-specific is missing
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile and active_profile.get("unique_id"):
            uid, p_tgid, p_name, p_age = (
                active_profile.get("unique_id"),
                active_profile.get("telegram_id"),
                active_profile.get("name"),
                active_profile.get("age"),
            )
        else:
            logger.warning("RT Save Results: UID not found. Cannot save.")
            return

    time_ms = data.get("rt_reaction_time_ms")
    attempts = data.get("rt_current_attempt", 1)
    final_status = status_override or data.get("rt_status", "Unknown")
    if is_interrupted and final_status not in [
        "Passed",
        "Failed",
        "Interrupted by user",
    ]:  # Ensure specific interruption status
        final_status = "Interrupted by user"

    interrupted_col_val = (
        "Да" if final_status == "Interrupted by user" else "Нет"
    )

    logger.info(
        f"RT Save: UID={uid}, Status={final_status}, TimeMs={time_ms}, Attempts={attempts}, Interrupted={interrupted_col_val}"
    )

    try:
        from openpyxl import load_workbook

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        if "Unique ID" not in ALL_EXPECTED_HEADERS:
            raise ValueError("'Unique ID' not in ALL_EXPECTED_HEADERS")
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")

        for idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = idx
                break

        if row_num == -1:
            new_row = [""] * len(ALL_EXPECTED_HEADERS)
            if p_tgid and "Telegram ID" in ALL_EXPECTED_HEADERS:
                new_row[ALL_EXPECTED_HEADERS.index("Telegram ID")] = p_tgid
            new_row[uid_col_idx] = uid
            if p_name and "Name" in ALL_EXPECTED_HEADERS:
                new_row[ALL_EXPECTED_HEADERS.index("Name")] = p_name
            if p_age and "Age" in ALL_EXPECTED_HEADERS:
                new_row[ALL_EXPECTED_HEADERS.index("Age")] = p_age
            ws.append(new_row)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS

        def set_cell_value(header_name, value):
            if header_name in h:
                ws.cell(row=row_num, column=h.index(header_name) + 1).value = (
                    value
                )

        set_cell_value(
            "ReactionTime_Time_ms", time_ms if time_ms is not None else "N/A"
        )
        set_cell_value("ReactionTime_Attempts", attempts)
        set_cell_value("ReactionTime_Status", final_status)
        set_cell_value("ReactionTime_Interrupted", interrupted_col_val)

        wb.save(EXCEL_FILENAME)
    except Exception as e:
        logger.error(
            f"RT Save Results: Error saving to Excel for UID {uid}: {e}",
            exc_info=True,
        )


async def cleanup_reaction_time_ui(
    state: FSMContext, bot_instance: Bot, final_text: str | None
):
    data = await state.get_data()
    chat_id = data.get("rt_chat_id")
    logger.info(
        f"RT Cleanup UI: Chat {chat_id if chat_id else 'N/A'}. Final text directive: '{final_text}'"
    )

    # Cancel active tasks
    for task_key in ["rt_memorization_task", "rt_reaction_cycle_task"]:
        task = data.get(task_key)
        if task and not task.done():
            task.cancel()
            try:
                await asyncio.wait_for(task, timeout=0.1)  # Short timeout
            except (asyncio.CancelledError, asyncio.TimeoutError):
                pass
            await state.update_data(**{task_key: None})

    # Identify all specific RT UI message IDs
    rt_message_ids_keys = [
        "rt_instruction_message_id",
        "rt_memorization_image_message_id",
        "rt_reaction_stimulus_message_id",
        "rt_retry_confirmation_message_id",
        "rt_target_missed_message_id",
    ]

    ids_to_delete_explicitly = set()
    for key in rt_message_ids_keys:
        msg_id = data.get(key)
        if msg_id:
            ids_to_delete_explicitly.add(msg_id)

    last_relevant_msg_id_for_edit = None
    if final_text:  # Only determine last relevant if we intend to edit
        last_relevant_msg_id_for_edit = (
            data.get("rt_retry_confirmation_message_id")
            or data.get("rt_reaction_stimulus_message_id")
            or data.get("rt_memorization_image_message_id")
            or data.get("rt_instruction_message_id")
            or data.get("rt_target_missed_message_id")
        )

    if chat_id:
        # If final_text is None, delete all identified RT UI messages
        if final_text is None:
            for msg_id in ids_to_delete_explicitly:
                try:
                    await bot_instance.delete_message(chat_id, msg_id)
                    logger.debug(
                        f"RT Cleanup (final_text=None): Deleted message ID {msg_id}."
                    )
                except TelegramBadRequest:
                    logger.debug(
                        f"RT Cleanup (final_text=None): Message ID {msg_id} already deleted."
                    )
                except Exception as e_del_all:
                    logger.error(
                        f"RT Cleanup (final_text=None): Error deleting msg ID {msg_id}: {e_del_all}"
                    )

        # If final_text is provided (likely from stop_test_command_handler)
        elif final_text and last_relevant_msg_id_for_edit:
            is_photo = last_relevant_msg_id_for_edit in [
                data.get("rt_reaction_stimulus_message_id"),
                data.get("rt_memorization_image_message_id"),
            ]
            try:
                if is_photo:
                    await bot_instance.edit_message_caption(
                        chat_id=chat_id,
                        message_id=last_relevant_msg_id_for_edit,
                        caption=final_text,
                        reply_markup=None,
                        parse_mode=ParseMode.HTML,
                    )
                else:
                    await bot_instance.edit_message_text(
                        text=final_text,
                        chat_id=chat_id,
                        message_id=last_relevant_msg_id_for_edit,
                        reply_markup=None,
                        parse_mode=ParseMode.HTML,
                    )
                logger.debug(
                    f"RT Cleanup (final_text provided): Edited msg {last_relevant_msg_id_for_edit}."
                )
                # Delete other messages except the one we just edited
                for msg_id in ids_to_delete_explicitly:
                    if msg_id != last_relevant_msg_id_for_edit:
                        try:
                            await bot_instance.delete_message(chat_id, msg_id)
                        except:
                            pass  # Ignore errors for these secondary deletions
            except (
                TelegramBadRequest
            ):  # Failed to edit, try sending new and delete all old
                logger.warning(
                    f"RT Cleanup: Failed to edit msg {last_relevant_msg_id_for_edit}. Sending new final_text and deleting all old."
                )
                try:
                    await bot_instance.send_message(
                        chat_id, final_text, parse_mode=ParseMode.HTML
                    )
                except:
                    pass  # Ignore send error if already cleaning up
                for msg_id in ids_to_delete_explicitly:  # Delete all originals
                    try:
                        await bot_instance.delete_message(chat_id, msg_id)
                    except:
                        pass
            except Exception as e_edit_final_rt:
                logger.error(
                    f"RT Cleanup: Error editing msg {last_relevant_msg_id_for_edit}: {e_edit_final_rt}"
                )

        elif (
            final_text
        ):  # No specific message to edit, but final_text needs to be sent
            try:
                await bot_instance.send_message(
                    chat_id, final_text, parse_mode=ParseMode.HTML
                )
            except:
                pass

    # Clean FSM: remove all "rt_" prefixed keys and specific message ID keys handled here
    # Preserve essential profile and common keys
    current_fsm_data = await state.get_data()
    data_to_keep_after_rt_cleanup = {}
    preserved_keys = [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
        "status_message_id_to_delete_later",
    ]

    for key in preserved_keys:
        if key in current_fsm_data and current_fsm_data[key] is not None:
            data_to_keep_after_rt_cleanup[key] = current_fsm_data[key]

    # For any other non "rt_" keys that might exist and should be preserved (if any).
    for key, value in current_fsm_data.items():
        if (
            not key.startswith("rt_")
            and key not in data_to_keep_after_rt_cleanup
        ):
            data_to_keep_after_rt_cleanup[key] = value

    await state.set_data(data_to_keep_after_rt_cleanup)
    logger.info(
        f"RT Cleanup UI: FSM data cleaned. Kept keys: {list(data_to_keep_after_rt_cleanup.keys())}"
    )
