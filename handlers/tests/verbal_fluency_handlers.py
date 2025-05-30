# handlers/tests/verbal_fluency_handlers.py
import asyncio
import logging
import random
import time

from aiogram import Bot, F, Router
from aiogram.enums import ParseMode, ChatType
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
    Chat,
    User,
)

# StateFilter не используется напрямую в этом файле после рефакторинга кнопки стоп
# from aiogram.filters import StateFilter

from fsm_states import VerbalFluencyStates
from settings import (
    ALL_EXPECTED_HEADERS,
    EXCEL_FILENAME,
    VERBAL_FLUENCY_DURATION_S,
    VERBAL_FLUENCY_TASK_POOL,
    VERBAL_FLUENCY_CATEGORY,
)
from utils.bot_helpers import (
    send_main_action_menu,
    get_active_profile_from_fsm,
)
from keyboards import (
    ACTION_SELECTION_KEYBOARD_RETURNING,
    ACTION_SELECTION_KEYBOARD_NEW,
)

logger = logging.getLogger(__name__)
router = Router()
IKB = InlineKeyboardButton


async def _verbal_fluency_timer_task(state: FSMContext, bot_instance: Bot):
    data = await state.get_data()
    chat_id = data.get("vf_chat_id")
    task_message_id = data.get("vf_task_message_id")
    task_letter = data.get("vf_task_letter")
    last_displayed_text = ""

    if not all([chat_id, task_message_id, task_letter]):
        logger.error("Verbal Fluency timer: Missing critical data from FSM.")
        trigger_event = data.get("vf_trigger_event_for_stop")
        await _end_verbal_fluency_test(
            state, bot_instance, interrupted=True, trigger_event=trigger_event
        )
        return

    base_task_text = f"Задание: Назовите как можно больше слов, начинающихся на букву <b>'{task_letter}'</b>.\n"

    # Добавляем кнопку "Остановить тест" в сообщение с заданием
    stop_button_markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="⏹️ Остановить Тест", callback_data="request_test_stop")]
        ]
    )

    try:
        for i in range(VERBAL_FLUENCY_DURATION_S, -1, -1):
            if (
                await state.get_state()
                != VerbalFluencyStates.collecting_words.state
            ):
                logger.info(
                    "Verbal Fluency timer: State changed, aborting timer."
                )
                return

            current_timer_display = f"Осталось: {i} сек."
            full_message_content = (
                f"{base_task_text}{current_timer_display}\n\nВводите слова."
            )

            if full_message_content != last_displayed_text:
                try:
                    await bot_instance.edit_message_text(
                        text=full_message_content,
                        chat_id=chat_id,
                        message_id=task_message_id,
                        parse_mode=ParseMode.HTML,
                        reply_markup=stop_button_markup,  # Обновляем с кнопкой
                    )
                    last_displayed_text = full_message_content
                except TelegramBadRequest as e:
                    if "message is not modified" not in str(e).lower():
                        logger.warning(
                            f"VF timer: edit_message_text (ID: {task_message_id}) failed: {e}."
                        )
            if i == 0:
                break
            await asyncio.sleep(1)

        if (
            await state.get_state()
            == VerbalFluencyStates.collecting_words.state
        ):  # Time is up
            logger.info("Verbal Fluency timer: Time is up.")
            trigger_event = data.get("vf_trigger_event_for_stop")
            if not trigger_event and chat_id:
                mock_user = User(
                    id=bot_instance.id, is_bot=True, first_name="Bot"
                )
                mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
                trigger_event = Message(
                    message_id=0,
                    date=int(time.time()),
                    chat=mock_chat,
                    from_user=mock_user,
                )
            await _end_verbal_fluency_test(
                state,
                bot_instance,
                interrupted=False,
                trigger_event=trigger_event,
            )

    except asyncio.CancelledError:
        logger.info("Verbal Fluency timer task explicitly cancelled.")
    except Exception as e:
        logger.error(
            f"Verbal Fluency timer task unexpected error: {e}", exc_info=True
        )
        trigger_event = data.get("vf_trigger_event_for_stop")
        if not trigger_event and chat_id:
            mock_user = User(id=bot_instance.id, is_bot=True, first_name="Bot")
            mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
            trigger_event = Message(
                message_id=0,
                date=int(time.time()),
                chat=mock_chat,
                from_user=mock_user,
            )
        await _end_verbal_fluency_test(
            state, bot_instance, interrupted=True, trigger_event=trigger_event
        )


async def _end_verbal_fluency_test(
    state: FSMContext,
    bot_instance: Bot,
    interrupted: bool,
    trigger_event: Message | CallbackQuery | None,
):
    current_fsm_state_str = await state.get_state()
    if not current_fsm_state_str or not current_fsm_state_str.startswith(
        VerbalFluencyStates.__name__
    ):
        logger.info(
            "VF _end_test: Called but test not active or already ended."
        )
        return

    logger.info(f"VF: Entering _end_test. Interrupted: {interrupted}")
    data = await state.get_data()
    chat_id = data.get("vf_chat_id")
    task_message_id = data.get("vf_task_message_id")
    timer_task = data.get("vf_timer_task")

    if timer_task and not timer_task.done():
        timer_task.cancel()
        try:
            await asyncio.wait_for(timer_task, timeout=0.5)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass
    await state.update_data(vf_timer_task=None)

    collected_words = data.get("vf_collected_words", set())
    word_count = len(collected_words)
    await save_verbal_fluency_results(state, is_interrupted=interrupted)

    result_message_text = ""
    if chat_id:
        if (
            task_message_id
        ):  # Try to edit the task message to remove buttons and show final text
            final_task_msg_text = (
                "Тест завершен."  # Default text if other logic fails
            )
            if interrupted:
                final_task_msg_text = f"Тест на вербальную беглость был <b>ПРЕРВАН</b>. Результаты сохранены."
            else:  # Time up
                final_task_msg_text = (
                    f"Время вышло! Тест на вербальную беглость завершен."
                )
            try:
                await bot_instance.edit_message_text(
                    text=final_task_msg_text,
                    chat_id=chat_id,
                    message_id=task_message_id,
                    reply_markup=None,
                    parse_mode=ParseMode.HTML,
                )
            except (
                TelegramBadRequest
            ):  # If edit fails, just try to unpin and delete
                logger.warning(
                    f"VF _end_test: Failed to edit task msg {task_message_id}, will try to delete."
                )
                try:
                    await bot_instance.unpin_chat_message(
                        chat_id=chat_id, message_id=task_message_id
                    )
                except TelegramBadRequest:
                    pass
                try:
                    await bot_instance.delete_message(
                        chat_id=chat_id, message_id=task_message_id
                    )
                except TelegramBadRequest:
                    logger.warning(
                        f"VF _end_test: Failed to delete task msg {task_message_id} after failed edit."
                    )

        # Send separate summary message
        summary_text_for_user = ""
        if interrupted:
            summary_text_for_user = f"Тест на вербальную беглость был <b>ПРЕРВАН</b>.\nСохраненный результат: {word_count} слов(а)."
        else:
            summary_text_for_user = (
                f"Время вышло! Тест на вербальную беглость завершен.\n"
                f"Я сохранил результат. Количество названных (уникальных) слов: {word_count}.\n"
                f"Общее время выполнения: {VERBAL_FLUENCY_DURATION_S} сек."
            )
        try:
            await bot_instance.send_message(
                chat_id, summary_text_for_user, parse_mode=ParseMode.HTML
            )
        except Exception as e_send_res:
            logger.error(
                f"VF _end_test: Fail to send result summary msg: {e_send_res}"
            )

        await state.update_data(
            vf_task_message_id=None
        )  # Clear msg_id from FSM

    await cleanup_verbal_fluency_ui(
        state, bot_instance, final_text=None
    )  # Cleans FSM keys

    current_data_after_cleanup = await state.get_data()
    profile_data_to_keep = {
        k: current_data_after_cleanup.get(k)
        for k in [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        if current_data_after_cleanup.get(k)
    }
    await state.set_state(None)

    effective_trigger_event = trigger_event
    if not effective_trigger_event and chat_id:
        mock_user = User(id=bot_instance.id, is_bot=True, first_name="Bot")
        mock_chat = Chat(id=chat_id, type=ChatType.PRIVATE)
        effective_trigger_event = Message(
            message_id=0,
            date=int(time.time()),
            chat=mock_chat,
            from_user=mock_user,
        )

    if profile_data_to_keep.get("active_unique_id"):
        await state.set_data(profile_data_to_keep)
        if effective_trigger_event:
            await send_main_action_menu(
                bot_instance,
                effective_trigger_event,
                ACTION_SELECTION_KEYBOARD_RETURNING,
            )
        elif chat_id:
            await bot_instance.send_message(
                chat_id,
                "Тест завершен. Выберите действие:",
                reply_markup=ACTION_SELECTION_KEYBOARD_RETURNING,
            )
    else:
        if chat_id:
            await bot_instance.send_message(
                chat_id, "Профиль не активен. Пожалуйста, /start для начала."
            )
        await state.clear()
    logger.info("Verbal Fluency: Exiting _end_verbal_fluency_test.")


async def start_verbal_fluency_test(
    trigger_event: Message | CallbackQuery,
    state: FSMContext,
    profile: dict,
    bot_instance: Bot,
):
    logger.info(
        f"Starting Verbal Fluency Test for UID: {profile.get('unique_id')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    if not VERBAL_FLUENCY_TASK_POOL:
        await bot_instance.send_message(
            chat_id,
            "Ошибка: Пул заданий для теста пуст. Тест не может быть запущен.",
        )
        logger.error("Verbal Fluency Test: Task pool is empty.")
        await state.set_state(None) # Сбросить состояние, если оно было установлено для VF
        # Возврат в меню (нужно передать корректный trigger_event для send_main_action_menu)
        # Если trigger_event - CallbackQuery, его сообщение могло быть удалено.
        # Безопаснее использовать msg_ctx, но убедиться, что оно не удалено.
        # Или создать mock_message, если msg_ctx больше не валиден.
        # Для простоты, если есть chat_id, отправим меню без привязки к старому сообщению.
        active_profile = await get_active_profile_from_fsm(state)
        keyboard = (
            ACTION_SELECTION_KEYBOARD_RETURNING
            if active_profile
            else ACTION_SELECTION_KEYBOARD_NEW
        )
        # Передаем bot_instance и chat_id напрямую, если msg_ctx может быть невалидным
        await bot_instance.send_message(chat_id, "Выберите действие:", reply_markup=keyboard)
        return

    chosen_task = random.choice(VERBAL_FLUENCY_TASK_POOL)
    task_letter = chosen_task["letter"]

    await state.set_state(VerbalFluencyStates.showing_instructions_and_task)
    await state.update_data(
        vf_unique_id_for_test=profile.get("unique_id"),
        vf_profile_name_for_test=profile.get("name"),
        vf_profile_age_for_test=profile.get("age"),
        vf_profile_telegram_id_for_test=profile.get("telegram_id"),
        vf_chat_id=chat_id,
        vf_task_base_category=VERBAL_FLUENCY_CATEGORY,
        vf_task_letter=task_letter,
        vf_collected_words=set(),
        vf_timer_task=None,
        vf_task_message_id=None, # Будет установлен ID нового сообщения
        vf_trigger_event_for_stop=msg_ctx, # Сохраняем оригинальный контекст для _end_test
    )
    instruction_text = (
        f"<b>Тест на вербальную беглость</b>\n\n"
        f"Вам будет дана буква. Ваша задача – назвать как можно больше слов, "
        f"начинающихся на эту букву.\n"
        f"На выполнение задания даётся {VERBAL_FLUENCY_DURATION_S} секунд.\n"
        f"Слова можно писать в одном или нескольких сообщениях. Каждое слово должно быть не менее двух букв.\n\n"
        f"Нажмите 'Начать', чтобы увидеть букву и запустить таймер."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать", callback_data="vf_start_test_confirmed")]
        ]
    )

    try:
        # Всегда отправляем новое сообщение для инструкций
        sent_message = await bot_instance.send_message(
            chat_id, instruction_text, reply_markup=kbd, parse_mode=ParseMode.HTML # Добавлен parse_mode
        )
        await state.update_data(vf_task_message_id=sent_message.message_id)
    except TelegramBadRequest as e: # Этот блок теперь маловероятен для send_message, но оставим для общей обработки
        logger.error(
            f"Verbal Fluency start: Error sending instructions: {e}"
        )
        await bot_instance.send_message(
            chat_id, "Ошибка при запуске теста. Попробуйте снова."
        )
        await state.clear()


@router.callback_query(
    F.data == "vf_start_test_confirmed",
    VerbalFluencyStates.showing_instructions_and_task,
)
async def handle_verbal_fluency_start_ack(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    data = await state.get_data()
    task_msg_id = data.get("vf_task_message_id")
    task_letter = data.get("vf_task_letter")
    chat_id = data.get("vf_chat_id")

    if not all([task_msg_id, task_letter, chat_id]):
        logger.error("VF: Missing critical data in FSM for start_ack.")
        if chat_id:
            await bot.send_message(
                chat_id, "Произошла ошибка. Пожалуйста, /start."
            )
        await state.clear()
        return

    task_text = (
        f"Задание: Назовите как можно больше слов, начинающихся на букву <b>'{task_letter}'</b>.\n"
        f"Осталось: {VERBAL_FLUENCY_DURATION_S} сек.\n\nВводите слова."
    )
    current_task_msg_id = task_msg_id

    # Кнопка остановки
    stop_button_markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="⏹️ Остановить Тест", callback_data="request_test_stop")]
        ]
    )

    try:
        await bot.edit_message_text(
            text=task_text,
            chat_id=chat_id,
            message_id=current_task_msg_id,
            reply_markup=stop_button_markup,
            parse_mode=ParseMode.HTML,  # Добавлена кнопка
        )
    except TelegramBadRequest as e:
        logger.error(
            f"VF: Failed to edit task message {current_task_msg_id}: {e}. Sending new."
        )
        try:
            new_msg = await bot.send_message(
                chat_id,
                task_text,
                parse_mode=ParseMode.HTML,
                reply_markup=stop_button_markup,
            )
            await state.update_data(vf_task_message_id=new_msg.message_id)
            current_task_msg_id = new_msg.message_id
        except Exception as send_e:
            logger.critical(
                f"VF: Critical error - failed to send new task message: {send_e}"
            )
            await bot.send_message(
                chat_id, "Критическая ошибка. Тест прерван."
            )
            await _end_verbal_fluency_test(
                state, bot, interrupted=True, trigger_event=callback.message
            )
            return

    if current_task_msg_id and chat_id:
        try:
            await bot.pin_chat_message(
                chat_id=chat_id,
                message_id=current_task_msg_id,
                disable_notification=True,
            )
        except TelegramBadRequest as pin_e:
            logger.error(
                f"VF: Failed to pin task message {current_task_msg_id}: {pin_e}"
            )

    await state.set_state(VerbalFluencyStates.collecting_words)
    timer_task = asyncio.create_task(_verbal_fluency_timer_task(state, bot))
    await state.update_data(vf_timer_task=timer_task)


@router.message(VerbalFluencyStates.collecting_words, F.text)
async def handle_verbal_fluency_word_input(
    message: Message, state: FSMContext
):
    data = await state.get_data()
    task_letter = data.get("vf_task_letter", "").lower()
    collected_words_set = data.get("vf_collected_words", set())

    if not task_letter:
        await message.reply("Ошибка: буква для задания не определена.")
        return

    user_words_raw = message.text.lower().split()
    newly_added_count = 0
    for word in user_words_raw:
        processed_word = ''.join(filter(str.isalpha, word))
        if len(processed_word) >= 2 and processed_word.startswith(task_letter):
            if processed_word not in collected_words_set:
                collected_words_set.add(processed_word)
                newly_added_count += 1

    if newly_added_count > 0:
        await state.update_data(vf_collected_words=collected_words_set)


async def save_verbal_fluency_results(state: FSMContext, is_interrupted: bool):
    data = await state.get_data()
    uid = data.get("vf_unique_id_for_test")
    p_tgid, p_name, p_age = None, None, None

    if not uid:
        active_profile = await get_active_profile_from_fsm(state)
        if active_profile:
            uid = active_profile.get("unique_id")
            p_tgid = active_profile.get("telegram_id")
            p_name = active_profile.get("name")
            p_age = active_profile.get("age")
        else:
            logger.error("VF save: UID not found. Cannot save.")
            return
    else:
        p_tgid = data.get(
            "vf_profile_telegram_id_for_test", data.get("active_telegram_id")
        )
        p_name = data.get("vf_profile_name_for_test", data.get("active_name"))
        p_age = data.get("vf_profile_age_for_test", data.get("active_age"))

    letter = data.get("vf_task_letter", "N/A")
    collected_words = data.get("vf_collected_words", set())
    word_count = len(collected_words)
    words_list_str = ", ".join(sorted(list(collected_words)))
    interrupted_status = "Да" if is_interrupted else "Нет"
    excel_category_display = f"Слова на букву {letter}"

    try:
        from openpyxl import load_workbook

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")
        for idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = idx
                break
        if row_num == -1:
            new_row_data = [""] * len(ALL_EXPECTED_HEADERS)
            new_row_data[ALL_EXPECTED_HEADERS.index("Telegram ID")] = (
                p_tgid if p_tgid else ""
            )
            new_row_data[uid_col_idx] = uid
            new_row_data[ALL_EXPECTED_HEADERS.index("Name")] = (
                p_name if p_name else ""
            )
            new_row_data[ALL_EXPECTED_HEADERS.index("Age")] = (
                p_age if p_age else ""
            )
            ws.append(new_row_data)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_Category") + 1
        ).value = excel_category_display
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_Letter") + 1
        ).value = letter
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_WordCount") + 1
        ).value = word_count
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_WordsList") + 1
        ).value = words_list_str
        ws.cell(
            row=row_num, column=h.index("VerbalFluency_Interrupted") + 1
        ).value = interrupted_status
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"VF results for UID {uid} saved. Cat: {excel_category_display}, L: {letter}, Cnt: {word_count}, Int: {interrupted_status}"
        )
    except Exception as e:
        logger.error(f"VF results save error UID {uid}: {e}", exc_info=True)
        chat_id_for_err = data.get("vf_chat_id")
        if chat_id_for_err and await state.get_state() is not None:
            # Cannot send message here as bot_instance is not passed to save_results
            logger.warning(
                f"VF save_results: Cannot send error to user {chat_id_for_err} (no bot instance)"
            )


async def cleanup_verbal_fluency_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: str | None = None,
):
    logger.info(f"VF: Entering cleanup_ui. Final text: '{final_text}'")
    data = await state.get_data()
    chat_id = data.get("vf_chat_id")
    task_message_id = data.get("vf_task_message_id")
    timer_task = data.get("vf_timer_task")

    if timer_task and not timer_task.done():
        timer_task.cancel()
        try:
            await asyncio.wait_for(timer_task, timeout=0.2)
        except (asyncio.CancelledError, asyncio.TimeoutError):
            pass

    if final_text and chat_id and task_message_id:
        try:
            await bot_instance.unpin_chat_message(
                chat_id=chat_id, message_id=task_message_id
            )
        except TelegramBadRequest:
            pass
        try:
            await bot_instance.edit_message_text(
                text=final_text,
                chat_id=chat_id,
                message_id=task_message_id,
                reply_markup=None,
                parse_mode=ParseMode.HTML,
            )
        except TelegramBadRequest:
            try:
                await bot_instance.send_message(
                    chat_id, final_text, parse_mode=ParseMode.HTML
                )
            except Exception as e_sf:
                logger.error(f"VF cleanup send final_text err: {e_sf}")
    elif final_text and chat_id:  # No task_message_id, but final_text exists
        try:
            await bot_instance.send_message(
                chat_id, final_text, parse_mode=ParseMode.HTML
            )
        except Exception as e_sf_alt:
            logger.error(f"VF cleanup send final_text (alt) err: {e_sf_alt}")
    elif (
        task_message_id and chat_id
    ):  # No final text, but task message exists, delete it
        try:
            await bot_instance.unpin_chat_message(
                chat_id=chat_id, message_id=task_message_id
            )
        except TelegramBadRequest:
            pass
        try:
            await bot_instance.delete_message(
                chat_id=chat_id, message_id=task_message_id
            )
        except TelegramBadRequest:
            pass

    current_fsm_data = await state.get_data()
    new_data = {
        k: v for k, v in current_fsm_data.items() if not k.startswith("vf_")
    }
    for pk in [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
    ]:
        if pk in current_fsm_data and pk not in new_data:
            new_data[pk] = current_fsm_data[pk]
    await state.set_data(new_data)
    logger.info("Verbal Fluency cleanup: VF-specific FSM data cleaned.")
