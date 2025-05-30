# handlers/tests/mental_rotation_handlers.py
import asyncio
import logging
import random
import time
import os

from aiogram import Bot, F, Router
from aiogram.enums import ParseMode, ChatType
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    FSInputFile,
    InputMediaPhoto,
    Message,
    Chat,
    User,
)

import settings  # Импортируем модуль settings для доступа к его атрибутам

from fsm_states import MentalRotationStates
from settings import (  # Импортируем константы, которые не изменяются динамически
    ALL_EXPECTED_HEADERS,
    EXCEL_FILENAME,
    MENTAL_ROTATION_NUM_ITERATIONS,
    MR_REFERENCES_DIR,
    MR_CORRECT_PROJECTIONS_DIR,
    # MR_DISTRACTORS_DIR, # Если используется, то тоже можно, но он не был в вашем последнем settings.py
    MR_FEEDBACK_DISPLAY_TIME_S,
)
from utils.image_processors import generate_mr_collage
from utils.bot_helpers import (
    send_main_action_menu,
    get_active_profile_from_fsm,
)
from keyboards import (
    ACTION_SELECTION_KEYBOARD_RETURNING,
    ACTION_SELECTION_KEYBOARD_NEW,
)

logger = logging.getLogger(__name__)
router = Router()
IKB = InlineKeyboardButton


async def _get_mr_stimulus_for_iteration(
    state: FSMContext,
) -> tuple[str | None, list[str] | None, int | None, str | None]:
    data = await state.get_data()
    used_references = data.get("mr_used_references", [])

    # Используем settings. для доступа к изменяемым спискам
    if not settings.MR_REFERENCE_FILES:
        logger.error(
            "_get_mr_stimulus_for_iteration: settings.MR_REFERENCE_FILES is empty!"
        )
        return (
            None,
            None,
            None,
            "Пул эталонных изображений пуст. Проверьте настройку.",
        )

    available_references = [
        ref
        for ref in settings.MR_REFERENCE_FILES
        if ref not in used_references
    ]
    if not available_references:
        return None, None, None, "Больше нет уникальных эталонных изображений."

    selected_reference_filename = random.choice(available_references)
    selected_reference_path = os.path.join(
        MR_REFERENCES_DIR,
        selected_reference_filename,  # MR_REFERENCES_DIR - константа пути
    )
    if not os.path.exists(selected_reference_path):
        logger.error(
            f"_get_mr_stimulus_for_iteration: Reference file not found: {selected_reference_path}"
        )
        return (
            None,
            None,
            None,
            f"Эталонное изображение не найдено: {selected_reference_filename}",
        )

    used_references.append(selected_reference_filename)

    if not settings.MR_CORRECT_PROJECTIONS_MAP:
        logger.error(
            "_get_mr_stimulus_for_iteration: settings.MR_CORRECT_PROJECTIONS_MAP is empty!"
        )
        return (
            None,
            None,
            None,
            "Карта правильных проекций пуста. Проверьте настройку.",
        )

    correct_projection_filenames = settings.MR_CORRECT_PROJECTIONS_MAP.get(
        selected_reference_filename, []
    )
    if not correct_projection_filenames:
        logger.error(
            f"_get_mr_stimulus_for_iteration: No correct projections for {selected_reference_filename}"
        )
        return (
            None,
            None,
            None,
            f"Нет карты правильных проекций для {selected_reference_filename}",
        )

    chosen_correct_proj_filename = random.choice(correct_projection_filenames)
    correct_projection_path = os.path.join(
        MR_CORRECT_PROJECTIONS_DIR,
        chosen_correct_proj_filename,  # MR_CORRECT_PROJECTIONS_DIR - константа пути
    )
    if not os.path.exists(correct_projection_path):
        logger.error(
            f"_get_mr_stimulus_for_iteration: Correct projection file not found: {correct_projection_path}"
        )
        return (
            None,
            None,
            None,
            f"Файл правильной проекции не найден: {chosen_correct_proj_filename}",
        )

    if not settings.MR_ALL_DISTRACTORS_FILES:
        logger.error(
            "_get_mr_stimulus_for_iteration: settings.MR_ALL_DISTRACTORS_FILES is empty!"
        )
        return None, None, None, "Пул дистракторов пуст. Проверьте настройку."

    num_distractors_to_select = 3
    if len(settings.MR_ALL_DISTRACTORS_FILES) < num_distractors_to_select:
        logger.error(
            f"_get_mr_stimulus_for_iteration: Not enough distractors. Need {num_distractors_to_select}, have {len(settings.MR_ALL_DISTRACTORS_FILES)}"
        )
        return (
            None,
            None,
            None,
            f"Недостаточно дистракторов (нужно {num_distractors_to_select}, есть {len(settings.MR_ALL_DISTRACTORS_FILES)}).",
        )

    valid_distractors = [
        dp for dp in settings.MR_ALL_DISTRACTORS_FILES if os.path.exists(dp)
    ]
    if len(valid_distractors) < num_distractors_to_select:
        logger.error(
            f"_get_mr_stimulus_for_iteration: Not enough VALID distractors. Need {num_distractors_to_select}, have {len(valid_distractors)}"
        )
        return (
            None,
            None,
            None,
            f"Недостаточно ВАЛИДНЫХ дистракторов (нужно {num_distractors_to_select}, есть {len(valid_distractors)}).",
        )

    selected_distractor_paths = random.sample(
        valid_distractors, num_distractors_to_select
    )
    await state.update_data(mr_used_references=used_references)

    options_paths = [correct_projection_path] + selected_distractor_paths
    random.shuffle(options_paths)
    correct_option_index = options_paths.index(correct_projection_path)

    return selected_reference_path, options_paths, correct_option_index, None


async def _mr_schedule_feedback_revert(
    chat_id: int,
    message_id: int,
    normal_text: str,
    bot_instance: Bot,
    state_context_at_call: FSMContext,
):
    try:
        await asyncio.sleep(MR_FEEDBACK_DISPLAY_TIME_S)
        current_fsm_state_val = await state_context_at_call.get_state()
        current_fsm_data = await state_context_at_call.get_data()

        if (
            current_fsm_state_val is not None
            and current_fsm_state_val.startswith(MentalRotationStates.__name__)
            and current_fsm_data.get("mr_feedback_message_id") == message_id
        ):
            try:
                await bot_instance.edit_message_text(
                    text=normal_text,
                    chat_id=chat_id,
                    message_id=message_id,
                    parse_mode=None,
                )
            except TelegramBadRequest as e_edit:
                if (
                    "message is not modified" not in str(e_edit).lower()
                    and "message to edit not found" not in str(e_edit).lower()
                ):
                    logger.warning(
                        f"MR Feedback Revert (msg {message_id}): Edit failed: {e_edit}"
                    )
            except Exception as e_gen_edit:
                logger.error(
                    f"MR Feedback Revert (msg {message_id}): General error on edit: {e_gen_edit}",
                    exc_info=True,
                )
        else:
            logger.info(
                f"MR Feedback Revert (msg {message_id}): State/msg_id changed or test ended. Skipping revert."
            )
    except asyncio.CancelledError:
        logger.info(
            f"MR Feedback Revert task for msg {message_id} was cancelled."
        )
    except Exception as e:
        logger.error(
            f"MR Feedback Revert (msg {message_id}): Unexpected error in task: {e}",
            exc_info=True,
        )


async def _display_mr_stimulus(
    chat_id: int,
    state: FSMContext,
    bot_instance: Bot,
    is_editing: bool = False,
):
    data = await state.get_data()
    current_iteration = data.get("mr_current_iteration", 0) + 1
    await state.update_data(mr_current_iteration=current_iteration)

    ref_path, opt_paths, correct_idx, err_msg = (
        await _get_mr_stimulus_for_iteration(state)
    )

    if err_msg or not ref_path or not opt_paths or correct_idx is None:
        logger.error(
            f"MR Display Stimulus: Error from _get_mr_stimulus_for_iteration - {err_msg}"
        )
        if chat_id:
            await bot_instance.send_message(
                chat_id,
                f"Ошибка подготовки задания: {err_msg}. Тест умственного вращения прерван.",
            )
        await _finish_mental_rotation_test(
            state,
            bot_instance,
            chat_id,
            is_interrupted=True,
            error_occurred=True,
        )
        return

    collage_input_file = (
        None  # To hold FSInputFile or path string for deletion
    )
    try:
        collage_file_path_or_bytes = await generate_mr_collage(
            opt_paths
        )  # Assuming this returns a path or BytesIO
        if isinstance(collage_file_path_or_bytes, str):  # If it's a path
            collage_input_file = FSInputFile(collage_file_path_or_bytes)
            path_to_delete_collage = collage_file_path_or_bytes
        elif hasattr(
            collage_file_path_or_bytes, 'read'
        ):  # If it's BytesIO or similar file-like object
            collage_input_file = (
                collage_file_path_or_bytes  # Can be passed directly
            )
            path_to_delete_collage = None  # No file path to delete for BytesIO
        else:
            raise ValueError("generate_mr_collage returned an unexpected type")

        if (
            not collage_input_file
        ):  # Should be caught by ValueError above if type is wrong
            logger.error(
                "MR Display Stimulus: Collage generation failed (returned None or empty)."
            )
            if chat_id:
                await bot_instance.send_message(
                    chat_id,
                    "Ошибка генерации коллажа для вариантов. Тест умственного вращения прерван.",
                )
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id,
                is_interrupted=True,
                error_occurred=True,
            )
            return

        await state.update_data(
            mr_correct_option_index_for_current_iter=correct_idx
        )

        ref_msg_id = data.get("mr_reference_message_id")
        options_msg_id = data.get("mr_options_message_id")

        if not is_editing:
            if ref_msg_id and chat_id:
                try:
                    await bot_instance.delete_message(chat_id, ref_msg_id)
                except TelegramBadRequest:
                    pass
                ref_msg_id = None
            if options_msg_id and chat_id:
                try:
                    await bot_instance.delete_message(chat_id, options_msg_id)
                except TelegramBadRequest:
                    pass
                options_msg_id = None

        # Send/Edit Reference Image
        try:
            if is_editing and ref_msg_id and chat_id:
                await bot_instance.edit_message_media(
                    chat_id=chat_id,
                    message_id=ref_msg_id,
                    media=InputMediaPhoto(media=FSInputFile(ref_path)),
                )
            elif chat_id:
                msg_ref = await bot_instance.send_photo(
                    chat_id, FSInputFile(ref_path)
                )
                await state.update_data(
                    mr_reference_message_id=msg_ref.message_id
                )
            else:
                raise ValueError("chat_id is None for reference image.")
        except (TelegramBadRequest, FileNotFoundError, ValueError) as e_ref:
            logger.error(
                f"MR Display Stimulus: Error with reference image: {e_ref}",
                exc_info=True,
            )
            if chat_id:
                await bot_instance.send_message(
                    chat_id,
                    "Ошибка отображения эталонного изображения. Тест прерван.",
                )
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id,
                is_interrupted=True,
                error_occurred=True,
            )
            return

        # Send/Edit Options Collage with Buttons
        buttons = [
            [
                IKB(text="1", callback_data="mr_answer_1"),
                IKB(text="2", callback_data="mr_answer_2"),
            ],
            [
                IKB(text="3", callback_data="mr_answer_3"),
                IKB(text="4", callback_data="mr_answer_4"),
            ],
            [IKB(text="⏹️ Остановить Тест", callback_data="request_test_stop")],
        ]
        reply_markup = InlineKeyboardMarkup(inline_keyboard=buttons)

        try:
            if is_editing and options_msg_id and chat_id:
                await bot_instance.edit_message_media(
                    chat_id=chat_id,
                    message_id=options_msg_id,
                    media=InputMediaPhoto(media=collage_input_file),
                    reply_markup=reply_markup,
                )
            elif chat_id:
                msg_opts = await bot_instance.send_photo(
                    chat_id, collage_input_file, reply_markup=reply_markup
                )
                await state.update_data(
                    mr_options_message_id=msg_opts.message_id
                )
            else:
                raise ValueError("chat_id is None for options collage.")
        except (TelegramBadRequest, ValueError) as e_opts:
            logger.error(
                f"MR Display Stimulus: Error with options collage: {e_opts}",
                exc_info=True,
            )
            if chat_id:
                await bot_instance.send_message(
                    chat_id,
                    "Ошибка отображения вариантов ответа. Тест прерван.",
                )
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id,
                is_interrupted=True,
                error_occurred=True,
            )
            return

        await state.update_data(mr_iteration_start_time=time.time())
        await state.set_state(MentalRotationStates.displaying_stimulus_mr)

    finally:  # Ensure temporary collage file (if path was returned) is deleted
        if (
            'path_to_delete_collage' in locals()
            and path_to_delete_collage
            and os.path.exists(path_to_delete_collage)
        ):
            try:
                os.remove(path_to_delete_collage)
            except Exception as e_del_collage:
                logger.error(
                    f"MR Display Stimulus: Failed to delete temp collage file {path_to_delete_collage}: {e_del_collage}"
                )


async def _mr_proceed_to_next_iteration_or_finish(
    state: FSMContext, bot_instance: Bot, chat_id: int
):
    data = await state.get_data()
    current_iteration = data.get("mr_current_iteration", 0)

    if current_iteration < MENTAL_ROTATION_NUM_ITERATIONS:
        await state.set_state(
            MentalRotationStates.inter_iteration_countdown_mr
        )
        countdown_task = asyncio.create_task(
            _mr_inter_iteration_countdown_task(state, bot_instance, chat_id)
        )
        await state.update_data(
            mr_inter_iteration_countdown_task_ref=countdown_task
        )
    else:
        await _finish_mental_rotation_test(
            state, bot_instance, chat_id, is_interrupted=False
        )


async def _mr_inter_iteration_countdown_task(
    state: FSMContext, bot_instance: Bot, chat_id: int
):
    await asyncio.sleep(0.2)
    current_fsm_state = await state.get_state()
    if (
        current_fsm_state
        != MentalRotationStates.inter_iteration_countdown_mr.state
    ):
        logger.info(
            f"MR Countdown: State changed to {current_fsm_state}. Aborting."
        )
        return

    countdown_msg_id_local = None
    try:
        if not chat_id:
            logger.error("MR Countdown: chat_id is missing. Cannot proceed.")
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                None,
                is_interrupted=True,
                error_occurred=True,
            )
            return

        countdown_msg = await bot_instance.send_message(
            chat_id, "Следующее задание через: 3..."
        )
        countdown_msg_id_local = countdown_msg.message_id
        await state.update_data(mr_countdown_message_id=countdown_msg_id_local)

        for i in range(2, 0, -1):
            await asyncio.sleep(1)
            if (
                await state.get_state()
                != MentalRotationStates.inter_iteration_countdown_mr.state
            ):
                return
            await bot_instance.edit_message_text(
                text=f"Следующее задание через: {i}...",
                chat_id=chat_id,
                message_id=countdown_msg_id_local,
            )

        await asyncio.sleep(1)
        if (
            await state.get_state()
            != MentalRotationStates.inter_iteration_countdown_mr.state
        ):
            return

        if countdown_msg_id_local:
            try:
                await bot_instance.delete_message(
                    chat_id, countdown_msg_id_local
                )
            except TelegramBadRequest:
                pass
        await state.update_data(mr_countdown_message_id=None)

        await _display_mr_stimulus(
            chat_id, state, bot_instance, is_editing=True
        )

    except TelegramBadRequest as e_tb:
        logger.warning(
            f"MR Countdown: TelegramBadRequest: {e_tb}. Attempting recovery."
        )
        await asyncio.sleep(0.5)
        if (
            await state.get_state()
            == MentalRotationStates.inter_iteration_countdown_mr.state
        ):
            if countdown_msg_id_local:
                try:
                    await bot_instance.delete_message(
                        chat_id, countdown_msg_id_local
                    )
                except:
                    pass  # Best effort
            if chat_id:
                await _display_mr_stimulus(
                    chat_id, state, bot_instance, is_editing=True
                )
            else:
                await _finish_mental_rotation_test(
                    state,
                    bot_instance,
                    None,
                    is_interrupted=True,
                    error_occurred=True,
                )
    except asyncio.CancelledError:
        logger.info(f"MR Countdown task for chat {chat_id} was cancelled.")
        data_on_cancel = await state.get_data()
        chat_id_on_cancel = data_on_cancel.get("mr_chat_id")
        countdown_msg_id_on_cancel = data_on_cancel.get(
            "mr_countdown_message_id"
        )
        if countdown_msg_id_on_cancel and chat_id_on_cancel:
            try:
                await bot_instance.delete_message(
                    chat_id_on_cancel, countdown_msg_id_on_cancel
                )
            except:
                pass
    except Exception as e_unexp:
        logger.error(
            f"MR Countdown: Unexpected error: {e_unexp}", exc_info=True
        )
        if (
            await state.get_state()
            == MentalRotationStates.inter_iteration_countdown_mr.state
        ):
            await _finish_mental_rotation_test(
                state,
                bot_instance,
                chat_id,
                is_interrupted=True,
                error_occurred=True,
            )
    finally:
        await state.update_data(mr_inter_iteration_countdown_task_ref=None)


async def _finish_mental_rotation_test(
    state: FSMContext,
    bot_instance: Bot,
    chat_id: int | None,
    is_interrupted: bool,
    error_occurred: bool = False,
    called_by_stop_command: bool = False,
):
    current_fsm_state_on_entry = await state.get_state()
    if (
        not current_fsm_state_on_entry
        or not current_fsm_state_on_entry.startswith(
            MentalRotationStates.__name__
        )
    ):
        logger.info(
            "MR _finish_test: Called but test not in an active MR state or already finished."
        )
        if (
            called_by_stop_command and chat_id
        ):  # If stop_test called this, ensure common status is handled if it exists
            # Check if _delete_common_status_message helper should be local or imported
            data_stop_check = await state.get_data()
            common_status_msg_id_stop = data_stop_check.get(
                "status_message_id_to_delete_later"
            )
            if common_status_msg_id_stop:
                try:
                    await bot_instance.delete_message(
                        chat_id, common_status_msg_id_stop
                    )
                    await state.update_data(
                        status_message_id_to_delete_later=None
                    )
                except:
                    pass  # Best effort during stop
        return

    logger.info(
        f"Finishing Mental Rotation Test. Interrupted: {is_interrupted}, Error: {error_occurred}, Called by Stop: {called_by_stop_command}"
    )
    data = await state.get_data()
    effective_chat_id = data.get("mr_chat_id", chat_id)

    for task_key in [
        "mr_current_feedback_revert_task_ref",
        "mr_inter_iteration_countdown_task_ref",
    ]:
        task = data.get(task_key)
        if task and not task.done():
            task.cancel()
            await asyncio.sleep(0.01)  # Give a tick for cancellation
        await state.update_data(**{task_key: None})

    results_calc = data.get("mr_iteration_results", [])
    correct_answers_calc = sum(1 for r in results_calc if r.get("is_correct"))
    total_iterations_done_calc = len(results_calc)
    start_time = data.get("mr_test_start_time")
    total_test_time_s_calc = (
        round(time.time() - start_time, 2) if start_time else 0.0
    )

    correct_times = [
        r["reaction_time_s"]
        for r in results_calc
        if r.get("is_correct") and "reaction_time_s" in r
    ]
    avg_reaction_time_s_calc = (
        round(sum(correct_times) / len(correct_times), 2)
        if correct_times
        else 0.0
    )

    ind_resp_parts = [
        f"И{r.get('iteration', '?')}:{'Прав' if r.get('is_correct') else 'Неправ'},{r.get('reaction_time_s', 0):.2f}с"
        for r in results_calc
    ]
    ind_resp_str_calc = "; ".join(ind_resp_parts) if ind_resp_parts else "N/A"

    await state.update_data(
        mr_final_correct_answers=correct_answers_calc,
        mr_final_avg_reaction_time_s=avg_reaction_time_s_calc,
        mr_final_total_test_time_s=total_test_time_s_calc,
        mr_final_individual_responses_str=ind_resp_str_calc,
        mr_final_interrupted_status=(is_interrupted or error_occurred),
    )

    mock_msg_for_save = None
    if effective_chat_id:
        mock_user = User(
            id=bot_instance.id if hasattr(bot_instance, 'id') else 1,
            is_bot=True,
            first_name="Bot",
        )
        mock_chat = Chat(id=effective_chat_id, type=ChatType.PRIVATE)
        mock_msg_for_save = Message(
            message_id=0,
            date=int(time.time()),
            chat=mock_chat,
            from_user=mock_user,
            text="mock_finish",
        )

    await save_mental_rotation_results(
        mock_msg_for_save,
        state,
        is_interrupted=(is_interrupted or error_occurred),
    )

    if not called_by_stop_command and effective_chat_id:
        summary_text = ""
        if is_interrupted or error_occurred:
            summary_text = "Тест умственного вращения был прерван"
            if error_occurred:
                summary_text += " из-за ошибки."
            else:
                summary_text += "."
            if results_calc:
                summary_text += f"\nЧастичные результаты: {correct_answers_calc}/{total_iterations_done_calc} правильных."
        else:
            summary_text = (
                "Тест умственного вращения успешно завершен!\n"
                f"Правильных ответов: {correct_answers_calc}/{MENTAL_ROTATION_NUM_ITERATIONS}\n"
                f"Среднее время реакции на правильные: {avg_reaction_time_s_calc:.2f} сек\n"
                f"Общее время теста: {total_test_time_s_calc:.2f} сек"
            )
        try:
            await bot_instance.send_message(
                effective_chat_id, summary_text, parse_mode=ParseMode.HTML
            )
        except Exception as e_send_summary:
            logger.error(
                f"MR Finish: Error sending summary to user: {e_send_summary}"
            )

    await cleanup_mental_rotation_ui(state, bot_instance, final_text=None)

    if (
        not called_by_stop_command
    ):  # stop_test_command_handler handles common status message itself
        common_status_msg_id_finish = data.get(
            "status_message_id_to_delete_later"
        )
        if common_status_msg_id_finish and effective_chat_id:
            try:
                await bot_instance.delete_message(
                    effective_chat_id, common_status_msg_id_finish
                )
                logger.info(
                    f"MR Finish (normal): Deleted common status message ID: {common_status_msg_id_finish}"
                )
            except TelegramBadRequest:
                logger.warning(
                    f"MR Finish (normal): Common status msg {common_status_msg_id_finish} already deleted."
                )
            except Exception as e_del_cs_mr:
                logger.error(
                    f"MR Finish (normal): Error deleting common status msg {common_status_msg_id_finish}: {e_del_cs_mr}"
                )
            await state.update_data(status_message_id_to_delete_later=None)

    if not called_by_stop_command:
        profile_keys = [
            "active_unique_id",
            "active_name",
            "active_age",
            "active_telegram_id",
        ]
        current_data_final_nav = await state.get_data()
        profile_data_to_keep_final_nav = {
            k: current_data_final_nav.get(k)
            for k in profile_keys
            if current_data_final_nav.get(k)
        }

        await state.set_state(None)
        if profile_data_to_keep_final_nav.get("active_unique_id"):
            await state.set_data(profile_data_to_keep_final_nav)
            trigger_event_for_menu = (
                data.get("mr_triggering_event_for_menu") or mock_msg_for_save
            )  # Use original trigger or mock

            message_context_for_menu_final = None
            if isinstance(trigger_event_for_menu, Message):
                message_context_for_menu_final = trigger_event_for_menu
            elif (
                isinstance(trigger_event_for_menu, CallbackQuery)
                and trigger_event_for_menu.message
            ):
                message_context_for_menu_final = trigger_event_for_menu.message

            if message_context_for_menu_final:
                await send_main_action_menu(
                    bot_instance,
                    message_context_for_menu_final,
                    ACTION_SELECTION_KEYBOARD_RETURNING,
                )
            elif effective_chat_id:
                await bot_instance.send_message(
                    effective_chat_id,
                    "Тест завершен. Выберите действие:",
                    reply_markup=ACTION_SELECTION_KEYBOARD_RETURNING,
                )
        elif effective_chat_id:
            await bot_instance.send_message(
                effective_chat_id,
                "Тест завершен. Профиль не найден, пожалуйста /start.",
            )
            await state.clear()
        else:
            await state.clear()


async def start_mental_rotation_test(
    trigger_event: Message | CallbackQuery,
    state: FSMContext,
    profile: dict,
    bot_instance: Bot,
):
    logger.info(
        f"Starting Mental Rotation Test for UID: {profile.get('unique_id', 'N/A')}"
    )
    msg_ctx = (
        trigger_event.message
        if isinstance(trigger_event, CallbackQuery)
        else trigger_event
    )
    chat_id = msg_ctx.chat.id

    await state.set_state(MentalRotationStates.initial_instructions_mr)
    await state.update_data(
        mr_unique_id_for_test=profile.get("unique_id"),
        mr_profile_name_for_test=profile.get("name"),
        mr_profile_age_for_test=profile.get("age"),
        mr_profile_telegram_id_for_test=profile.get("telegram_id"),
        mr_chat_id=chat_id,
        mr_current_iteration=0,
        mr_iteration_results=[],
        mr_used_references=[],
        mr_test_start_time=None,
        mr_reference_message_id=None,
        mr_options_message_id=None,
        mr_countdown_message_id=None,
        mr_feedback_message_id=None,
        mr_inter_iteration_countdown_task_ref=None,
        mr_current_feedback_revert_task_ref=None,
        mr_triggering_event_for_menu=msg_ctx,
    )
    instruction_text = (
        "<b>Тест умственного вращения</b>\n\n"
        "Вам будет показан 3D объект и 4 варианта 2D проекций. "
        "Выберите номер той проекции, через которую мог бы пройти 3D объект.\n"
        "Объект можно вращать мысленно перед 'просовыванием'.\n\n"
        f"Тест состоит из {MENTAL_ROTATION_NUM_ITERATIONS} заданий."
    )
    kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Начать тест", callback_data="mr_ack_instructions")]
        ]
    )

    logger.info(
        f"MR_HANDLER (start_mental_rotation_test): settings.MR_REFERENCE_FILES content: {settings.MR_REFERENCE_FILES}"
    )
    logger.info(
        f"MR_HANDLER (start_mental_rotation_test): id(settings): {id(settings)}"
    )
    logger.info(
        f"MR_HANDLER (start_mental_rotation_test): id(settings.MR_REFERENCE_FILES): {id(settings.MR_REFERENCE_FILES)}"
    )

    if not settings.MR_REFERENCE_FILES:  # Critical check using settings module
        logger.error(
            "MR_HANDLER (start_mental_rotation_test): CRITICAL - settings.MR_REFERENCE_FILES is empty. Test cannot start."
        )
        await bot_instance.send_message(
            chat_id,
            "Критическая ошибка: Пул эталонных изображений для теста пуст. Тест не может быть запущен. Пожалуйста, сообщите администратору.",
        )
        common_status_msg_id_start_err = (await state.get_data()).get(
            "status_message_id_to_delete_later"
        )
        if (
            common_status_msg_id_start_err and chat_id
        ):  # Clean up common status if it was set
            try:
                await bot_instance.delete_message(
                    chat_id, common_status_msg_id_start_err
                )
            except:
                pass
        await state.clear()
        return

    try:
        await bot_instance.send_message(
            chat_id,
            instruction_text,
            reply_markup=kbd,
            parse_mode=ParseMode.HTML,
        )
        if isinstance(trigger_event, CallbackQuery) and trigger_event.message:
            try:
                await trigger_event.message.delete()
            except TelegramBadRequest:
                pass
    except Exception as e_start_instr:
        logger.error(
            f"MR start_mental_rotation_test: Error sending initial instructions: {e_start_instr}",
            exc_info=True,
        )
        await bot_instance.send_message(
            chat_id,
            "Ошибка при запуске теста умственного вращения. Попробуйте /start.",
        )
        common_status_msg_id_start_fail = (await state.get_data()).get(
            "status_message_id_to_delete_later"
        )
        if (
            common_status_msg_id_start_fail and chat_id
        ):  # Clean up common status
            try:
                await bot_instance.delete_message(
                    chat_id, common_status_msg_id_start_fail
                )
            except:
                pass
        await state.clear()


@router.callback_query(
    F.data == "mr_ack_instructions",
    MentalRotationStates.initial_instructions_mr,
)
async def mr_ack_instructions_callback(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    await state.update_data(mr_test_start_time=time.time())
    if callback.message:
        try:
            await callback.message.delete()  # Delete the message with "Начать тест" button
        except TelegramBadRequest:
            pass

    chat_id = (await state.get_data()).get("mr_chat_id")
    if chat_id:
        await _display_mr_stimulus(chat_id, state, bot)
    else:
        logger.error("MR Ack Instr: chat_id missing from FSM. Cannot proceed.")
        if callback.message:
            await callback.message.answer(
                "Ошибка: не удалось получить ID чата. Тест прерван."
            )
        # No specific UI to cleanup here as it hasn't been shown yet, but finish test logic
        await _finish_mental_rotation_test(
            state, bot, None, is_interrupted=True, error_occurred=True
        )


@router.callback_query(
    F.data.startswith("mr_answer_"),
    MentalRotationStates.displaying_stimulus_mr,
)
async def mr_answer_callback(
    callback: CallbackQuery, state: FSMContext, bot: Bot
):
    await callback.answer()
    data = await state.get_data()
    chat_id = data.get("mr_chat_id")
    if not chat_id:
        logger.error("MR Answer Callback: chat_id missing. Aborting.")
        if callback.message:
            await callback.message.answer(
                "Критическая ошибка: ID чата не найден. Тест прерван."
            )
        await _finish_mental_rotation_test(
            state, bot, None, is_interrupted=True, error_occurred=True
        )
        return

    iteration_start_time = data.get("mr_iteration_start_time", time.time())
    reaction_time_s = round(time.time() - iteration_start_time, 2)

    selected_option_num = int(callback.data.split("_")[-1])
    selected_option_idx = selected_option_num - 1
    correct_option_idx = data.get("mr_correct_option_index_for_current_iter")
    is_correct = selected_option_idx == correct_option_idx

    iteration_data = {
        "iteration": data.get("mr_current_iteration"),
        "is_correct": is_correct,
        "reaction_time_s": reaction_time_s,
        "selected_option": selected_option_num,
        "correct_option": (
            correct_option_idx + 1 if correct_option_idx is not None else "N/A"
        ),
    }
    current_results = data.get("mr_iteration_results", [])
    current_results.append(iteration_data)
    await state.update_data(mr_iteration_results=current_results)

    feedback_text_bold = f"<b>{'Верно!' if is_correct else 'Неверно!'}</b>"
    feedback_text_normal = f"{'Верно!' if is_correct else 'Неверно!'}"
    feedback_msg_id = data.get("mr_feedback_message_id")

    previous_revert_task = data.get("mr_current_feedback_revert_task_ref")
    if previous_revert_task and not previous_revert_task.done():
        previous_revert_task.cancel()
        await asyncio.sleep(0.01)

    try:
        if feedback_msg_id:
            await bot.edit_message_text(
                text=feedback_text_bold,
                chat_id=chat_id,
                message_id=feedback_msg_id,
                parse_mode=ParseMode.HTML,
            )
        else:
            msg_fb = await bot.send_message(
                chat_id, feedback_text_bold, parse_mode=ParseMode.HTML
            )
            feedback_msg_id = msg_fb.message_id
            await state.update_data(mr_feedback_message_id=feedback_msg_id)

        if feedback_msg_id:
            revert_task = asyncio.create_task(
                _mr_schedule_feedback_revert(
                    chat_id, feedback_msg_id, feedback_text_normal, bot, state
                )
            )
            await state.update_data(
                mr_current_feedback_revert_task_ref=revert_task
            )
    except TelegramBadRequest as e_tb_fb_ans:
        if "message is not modified" not in str(e_tb_fb_ans).lower():
            logger.error(
                f"MR Answer Callback: Feedback msg error (TB): {e_tb_fb_ans}"
            )
    except Exception as e_gen_fb_ans:
        logger.error(
            f"MR Answer Callback: Feedback msg general error: {e_gen_fb_ans}",
            exc_info=True,
        )

    options_msg_id = data.get("mr_options_message_id")
    if options_msg_id and chat_id:
        try:
            await bot.edit_message_reply_markup(
                chat_id=chat_id, message_id=options_msg_id, reply_markup=None
            )
        except TelegramBadRequest:
            pass

    await state.set_state(MentalRotationStates.processing_answer_mr)
    await _mr_proceed_to_next_iteration_or_finish(state, bot, chat_id)


async def save_mental_rotation_results(
    trigger_msg_context: Message | None,
    state: FSMContext,
    is_interrupted: bool = False,
):
    logger.info(
        f"Saving Mental Rotation results. Interrupted: {is_interrupted}"
    )
    data = await state.get_data()
    uid = data.get("mr_unique_id_for_test", data.get("active_unique_id"))
    p_tgid = data.get(
        "mr_profile_telegram_id_for_test", data.get("active_telegram_id")
    )
    p_name = data.get("mr_profile_name_for_test", data.get("active_name"))
    p_age = data.get("mr_profile_age_for_test", data.get("active_age"))

    if not uid:
        logger.error("MR Save Results: UID not found. Cannot save results.")
        return

    correct_ans = data.get(
        "mr_final_correct_answers", 0
    )  # Default to 0 if missing
    avg_rt = data.get("mr_final_avg_reaction_time_s", 0.0)
    total_time = data.get("mr_final_total_test_time_s", 0.0)
    ind_resp_str = data.get("mr_final_individual_responses_str", "N/A")
    interrupted_status = (
        "Да"
        if data.get("mr_final_interrupted_status", is_interrupted)
        else "Нет"
    )

    try:
        from openpyxl import load_workbook

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_num = -1
        if "Unique ID" not in ALL_EXPECTED_HEADERS:
            raise ValueError(
                "'Unique ID' header not found in settings.ALL_EXPECTED_HEADERS"
            )
        uid_col_idx = ALL_EXPECTED_HEADERS.index("Unique ID")

        for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_vals[uid_col_idx] is not None and str(
                row_vals[uid_col_idx]
            ) == str(uid):
                row_num = r_idx
                break

        if row_num == -1:
            new_row_excel_save = [""] * len(ALL_EXPECTED_HEADERS)
            if p_tgid and "Telegram ID" in ALL_EXPECTED_HEADERS:
                new_row_excel_save[
                    ALL_EXPECTED_HEADERS.index("Telegram ID")
                ] = p_tgid
            new_row_excel_save[uid_col_idx] = uid
            if p_name and "Name" in ALL_EXPECTED_HEADERS:
                new_row_excel_save[ALL_EXPECTED_HEADERS.index("Name")] = p_name
            if p_age and "Age" in ALL_EXPECTED_HEADERS:
                new_row_excel_save[ALL_EXPECTED_HEADERS.index("Age")] = p_age
            ws.append(new_row_excel_save)
            row_num = ws.max_row

        h = ALL_EXPECTED_HEADERS

        def set_cell_val(hdr_name, val_to_set):
            if hdr_name in h:
                ws.cell(row=row_num, column=h.index(hdr_name) + 1).value = (
                    val_to_set
                )

        set_cell_val("MentalRotation_CorrectAnswers", correct_ans)
        set_cell_val("MentalRotation_AverageReactionTime_s", avg_rt)
        set_cell_val("MentalRotation_TotalTime_s", total_time)
        set_cell_val("MentalRotation_IndividualResponses", ind_resp_str)
        set_cell_val("MentalRotation_Interrupted", interrupted_status)

        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Mental Rotation results for UID {uid} saved. Interrupted: {interrupted_status}"
        )
    except Exception as e_save_excel_mr:
        logger.error(
            f"MR Save Results: Excel save error for UID {uid}: {e_save_excel_mr}",
            exc_info=True,
        )


async def cleanup_mental_rotation_ui(
    state: FSMContext,
    bot_instance: Bot,
    final_text: str | None = None,
):
    data = await state.get_data()
    chat_id = data.get("mr_chat_id")
    logger.info(
        f"MR Cleanup UI: Chat {chat_id if chat_id else 'N/A'}. Final text directive (for stop_test): '{final_text}'"
    )

    for task_key in [
        "mr_inter_iteration_countdown_task_ref",
        "mr_current_feedback_revert_task_ref",
    ]:
        task = data.get(task_key)
        if task and not task.done():
            task.cancel()
            await asyncio.sleep(0.01)

    mr_ui_msg_ids_to_clean = {
        data.get("mr_reference_message_id"),
        data.get("mr_options_message_id"),
        data.get("mr_countdown_message_id"),
        data.get("mr_feedback_message_id"),
    }
    mr_ui_msg_ids_to_clean.discard(None)

    if chat_id:
        if (
            final_text
        ):  # Typically for interruption via stop_test_command_handler
            options_msg_id_for_edit = data.get(
                "mr_options_message_id"
            )  # Prefer editing options message
            edited_one_msg = False
            if options_msg_id_for_edit:
                try:
                    # Assuming options message was always a photo for MR
                    await bot_instance.edit_message_caption(
                        chat_id=chat_id,
                        message_id=options_msg_id_for_edit,
                        caption=final_text,
                        reply_markup=None,
                        parse_mode=ParseMode.HTML,
                    )
                    edited_one_msg = True
                    logger.debug(
                        f"MR Cleanup: Edited options_msg_id {options_msg_id_for_edit} with interruption text."
                    )
                except TelegramBadRequest:
                    logger.warning(
                        f"MR Cleanup: Failed to edit options_msg_id {options_msg_id_for_edit}."
                    )
                except Exception as e_edit_mr_stop:
                    logger.error(
                        f"MR Cleanup: Error editing options_msg {options_msg_id_for_edit}: {e_edit_mr_stop}"
                    )

            if (
                not edited_one_msg
            ):  # If edit failed or no options_msg_id, send new final_text
                try:
                    await bot_instance.send_message(
                        chat_id, final_text, parse_mode=ParseMode.HTML
                    )
                except Exception as e_send_final_stop:
                    logger.error(
                        f"MR Cleanup: Failed to send new final_text on stop: {e_send_final_stop}"
                    )

            # Delete all original UI messages (even if one was edited, others might exist)
            for msg_id_del in mr_ui_msg_ids_to_clean:
                if edited_one_msg and msg_id_del == options_msg_id_for_edit:
                    continue  # Don't delete the one we just edited
                try:
                    await bot_instance.delete_message(chat_id, msg_id_del)
                except:
                    pass  # Best effort
        else:  # final_text is None (normal completion), delete all MR UI
            for msg_id_del_norm in mr_ui_msg_ids_to_clean:
                try:
                    await bot_instance.delete_message(chat_id, msg_id_del_norm)
                    logger.debug(
                        f"MR Cleanup (normal): Deleted MR UI message ID {msg_id_del_norm}."
                    )
                except:
                    pass  # Best effort

    current_fsm_data_mr_clean = await state.get_data()
    data_to_keep_mr_clean = {}
    keys_to_preserve_after_mr = [
        "active_unique_id",
        "active_name",
        "active_age",
        "active_telegram_id",
        "status_message_id_to_delete_later",
    ]

    for key_mr_clean in keys_to_preserve_after_mr:
        if (
            key_mr_clean in current_fsm_data_mr_clean
            and current_fsm_data_mr_clean[key_mr_clean] is not None
        ):
            data_to_keep_mr_clean[key_mr_clean] = current_fsm_data_mr_clean[
                key_mr_clean
            ]

    for (
        key_other,
        val_other,
    ) in current_fsm_data_mr_clean.items():  # Preserve any other non-MR keys
        if (
            not key_other.startswith("mr_")
            and key_other not in data_to_keep_mr_clean
        ):
            data_to_keep_mr_clean[key_other] = val_other

    await state.set_data(data_to_keep_mr_clean)
    logger.info(
        f"MR Cleanup UI: FSM data cleaned. Kept keys: {list(data_to_keep_mr_clean.keys())}"
    )
