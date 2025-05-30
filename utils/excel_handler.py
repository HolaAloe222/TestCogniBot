# utils/excel_handler.py
import os
import logging
import random  # For UID generation
from typing import Optional, Dict, Any, Set, Union  # Updated type hints

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from settings import (
    EXCEL_FILENAME,
    ALL_EXPECTED_HEADERS,
    # BASE_HEADERS, # Not directly used here after refactoring common_handlers
)

logger = logging.getLogger(__name__)


def initialize_excel_file():
    """
    Initializes the Excel file.
    Creates it with all expected headers if it doesn't exist.
    If it exists, checks and adds any missing headers from ALL_EXPECTED_HEADERS
    to the end of the first row.
    """
    if not os.path.exists(EXCEL_FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.append(ALL_EXPECTED_HEADERS)
        try:
            wb.save(EXCEL_FILENAME)
            logger.info(
                f"Файл '{EXCEL_FILENAME}' создан со всеми заголовками."
            )
        except Exception as e_save:
            logger.error(
                f"Не удалось сохранить новый файл '{EXCEL_FILENAME}': {e_save}"
            )
    else:
        try:
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active

            if ws.max_row == 0:
                ws.append(ALL_EXPECTED_HEADERS)
                wb.save(EXCEL_FILENAME)
                logger.info(
                    f"Лист в '{EXCEL_FILENAME}' был пуст. Добавлены все заголовки."
                )
                return

            current_header_values = [cell.value for cell in ws[1]]
            existing_header_set = set(filter(None, current_header_values))
            headers_actually_added_to_file = []

            start_col_for_new_headers = 0
            for col_idx_plus_1, cell_val in enumerate(
                current_header_values, 1
            ):
                if cell_val is not None:
                    start_col_for_new_headers = col_idx_plus_1
            start_col_for_new_headers += 1

            if not existing_header_set and not any(
                current_header_values
            ):  # if all are None or empty
                start_col_for_new_headers = 1

            for header_to_add in ALL_EXPECTED_HEADERS:
                if header_to_add not in existing_header_set:
                    # Find the next available column dynamically if there are gaps
                    current_max_col_with_data = 0
                    for cell_idx_in_row1, cell_in_row1 in enumerate(ws[1], 1):
                        if cell_in_row1.value is not None:
                            current_max_col_with_data = cell_idx_in_row1

                    col_to_write = current_max_col_with_data + 1
                    if not existing_header_set and not any(
                        current_header_values
                    ):  # First header in an "empty" row
                        col_to_write = 1

                    # Ensure we are not overwriting existing headers if logic got complex
                    # For now, simple append logic at calculated start_col_for_new_headers
                    # The refined logic: iterate through ALL_EXPECTED_HEADERS, if a header is missing,
                    # find the first empty column in row 1 (or one after last known header) and write it.
                    # This part can be tricky if headers are sparse.

                    # Simplified: Add to the determined 'start_col_for_new_headers' and increment it
                    ws.cell(row=1, column=start_col_for_new_headers).value = (
                        header_to_add
                    )
                    headers_actually_added_to_file.append(header_to_add)
                    existing_header_set.add(
                        header_to_add
                    )  # Add to set to track for this run
                    start_col_for_new_headers += 1

            if headers_actually_added_to_file:
                logger.info(
                    f"В '{EXCEL_FILENAME}' добавлены недостающие заголовки: {headers_actually_added_to_file}"
                )
                wb.save(EXCEL_FILENAME)
            else:
                logger.info(f"Заголовки в '{EXCEL_FILENAME}' актуальны.")

        except InvalidFileException:
            logger.error(
                f"Файл '{EXCEL_FILENAME}' поврежден или не является валидным Excel файлом. "
                "Требуется ручная проверка."
            )
        except Exception as e:
            logger.error(
                f"Ошибка при инициализации/обновлении Excel файла '{EXCEL_FILENAME}': {e}. "
                "Требуется ручная проверка.",
                exc_info=True,
            )


# --- New Profile Management Functions ---
def generate_unique_id(existing_ids: Set[str]) -> Optional[int]:
    """Generates a unique ID (integer) not present in existing_ids (set of strings)."""
    min_uid, max_uid = 1000000, 9999999  # Example range for 7-digit UIDs
    attempts = 0
    max_attempts = 1000  # Prevent infinite loop
    while attempts < max_attempts:
        candidate_uid = random.randint(min_uid, max_uid)
        if str(candidate_uid) not in existing_ids:
            return candidate_uid
        attempts += 1
    logger.critical(
        f"Не удалось сгенерировать уникальный UID после {max_attempts} попыток."
    )
    return None


def create_user_profile_in_excel(
    name: str, age: int, tgid: int
) -> Optional[int]:
    """
    Registers a new user in the Excel file and returns their UID.
    Returns None if registration fails.
    This is a synchronous function.
    """
    try:
        if not os.path.exists(EXCEL_FILENAME):
            logger.error(
                f"Excel файл '{EXCEL_FILENAME}' не найден. Невозможно создать профиль."
            )
            # Initialize if not exists (optional, depends on desired behavior)
            initialize_excel_file()
            if not os.path.exists(
                EXCEL_FILENAME
            ):  # Check again after init attempt
                return None

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active

        # Ensure headers are present (especially 'Unique ID')
        excel_headers = [cell.value for cell in ws[1]]
        if not excel_headers or "Unique ID" not in excel_headers:
            logger.error(
                "Excel: 'Unique ID' заголовок отсутствует или файл пуст. Попытка инициализации..."
            )
            initialize_excel_file()  # Try to re-initialize headers
            wb = load_workbook(EXCEL_FILENAME)  # Re-load after potential init
            ws = wb.active
            excel_headers = [cell.value for cell in ws[1]]
            if "Unique ID" not in excel_headers:
                logger.error(
                    "Excel: 'Unique ID' все еще отсутствует после повторной инициализации."
                )
                return None

        uid_col_idx_excel = excel_headers.index("Unique ID")

        existing_ids_from_excel: Set[str] = set()
        for row_tuple in ws.iter_rows(min_row=2, values_only=True):
            if (
                uid_col_idx_excel < len(row_tuple)
                and row_tuple[uid_col_idx_excel] is not None
            ):
                existing_ids_from_excel.add(str(row_tuple[uid_col_idx_excel]))

        new_uid = generate_unique_id(existing_ids_from_excel)
        if new_uid is None:
            return None  # Failed to generate UID

        # Use ALL_EXPECTED_HEADERS to map data to the correct columns
        new_row_data = [""] * len(ALL_EXPECTED_HEADERS)
        header_to_index_map = {
            header: i for i, header in enumerate(ALL_EXPECTED_HEADERS)
        }

        if "Telegram ID" in header_to_index_map:
            new_row_data[header_to_index_map["Telegram ID"]] = tgid
        if "Unique ID" in header_to_index_map:
            new_row_data[header_to_index_map["Unique ID"]] = new_uid
        if "Name" in header_to_index_map:
            new_row_data[header_to_index_map["Name"]] = name
        if "Age" in header_to_index_map:
            new_row_data[header_to_index_map["Age"]] = age

        ws.append(new_row_data)
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"Новый пользователь зарегистрирован в Excel: UID {new_uid}, TGID {tgid}, Имя '{name}', Возраст {age}"
        )
        return new_uid
    except Exception as e:
        logger.error(
            f"Ошибка регистрации пользователя в Excel (Имя '{name}', TGID {tgid}): {e}",
            exc_info=True,
        )
        return None


def find_user_profile_in_excel(
    uid_to_find: str, current_tgid: Optional[int] = None
) -> Optional[Dict[str, Any]]:
    """
    Finds a user profile by UID in Excel. Optionally updates Telegram ID.
    Returns a dictionary with profile data if found, else None.
    Keys in returned dict: 'unique_id', 'telegram_id', 'name', 'age'.
    This is a synchronous function.
    """
    try:
        if not os.path.exists(EXCEL_FILENAME):
            logger.warning(
                f"Excel файл '{EXCEL_FILENAME}' не найден при поиске профиля UID {uid_to_find}."
            )
            return None

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active

        if ws.max_row < 1:  # No headers
            logger.warning(
                f"Excel файл '{EXCEL_FILENAME}' пуст (нет заголовков) при поиске UID {uid_to_find}."
            )
            return None

        header_row_excel = [cell.value for cell in ws[1]]

        if "Unique ID" not in header_row_excel:
            logger.error(
                "Excel: 'Unique ID' заголовок отсутствует в файле при поиске профиля."
            )
            return None

        # Get column indices based on actual headers in the file
        try:
            uid_col_idx = header_row_excel.index("Unique ID")
            tg_id_col_idx = (
                header_row_excel.index("Telegram ID")
                if "Telegram ID" in header_row_excel
                else -1
            )
            name_col_idx = (
                header_row_excel.index("Name")
                if "Name" in header_row_excel
                else -1
            )
            age_col_idx = (
                header_row_excel.index("Age")
                if "Age" in header_row_excel
                else -1
            )
        except (
            ValueError
        ) as ve:  # Should not happen if "Unique ID" check passed, but for safety
            logger.error(f"Excel: Ошибка поиска индекса заголовка: {ve}")
            return None

        found_profile_data: Optional[Dict[str, Any]] = None
        row_to_update_tgid_excel_num = -1  # 1-based Excel row number

        for row_idx_excel, row_values in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            excel_uid_value = (
                row_values[uid_col_idx]
                if uid_col_idx < len(row_values)
                else None
            )
            if (
                excel_uid_value is not None
                and str(excel_uid_value) == uid_to_find
            ):
                found_profile_data = {
                    "unique_id": str(excel_uid_value),
                    "telegram_id": (
                        str(row_values[tg_id_col_idx])
                        if tg_id_col_idx != -1
                        and tg_id_col_idx < len(row_values)
                        and row_values[tg_id_col_idx] is not None
                        else str(current_tgid or "N/A")
                    ),
                    "name": (
                        str(row_values[name_col_idx])
                        if name_col_idx != -1
                        and name_col_idx < len(row_values)
                        and row_values[name_col_idx] is not None
                        else "N/A"
                    ),
                    "age": (
                        str(row_values[age_col_idx])
                        if age_col_idx != -1
                        and age_col_idx < len(row_values)
                        and row_values[age_col_idx] is not None
                        else "N/A"
                    ),
                }
                if current_tgid and tg_id_col_idx != -1:
                    current_tg_id_in_excel_str = (
                        str(row_values[tg_id_col_idx])
                        if tg_id_col_idx < len(row_values)
                        and row_values[tg_id_col_idx] is not None
                        else None
                    )
                    if current_tg_id_in_excel_str != str(current_tgid):
                        row_to_update_tgid_excel_num = row_idx_excel
                break  # Found the UID

        if (
            row_to_update_tgid_excel_num != -1 and tg_id_col_idx != -1
        ):  # TG ID needs update
            ws.cell(
                row=row_to_update_tgid_excel_num,
                column=tg_id_col_idx + 1,
                value=str(current_tgid),
            )  # +1 for 1-based column
            wb.save(EXCEL_FILENAME)
            logger.info(
                f"Обновлен Telegram ID для UID {uid_to_find} на {current_tgid} в Excel."
            )
            if (
                found_profile_data
            ):  # Update in-memory profile if it was already fetched
                found_profile_data["telegram_id"] = str(current_tgid)

        return found_profile_data
    except (
        FileNotFoundError
    ):  # Should be caught by os.path.exists earlier, but for safety
        logger.error(
            f"Excel файл '{EXCEL_FILENAME}' не найден при поиске профиля (повторно)."
        )
        return None
    except InvalidFileException:
        logger.error(
            f"Excel файл '{EXCEL_FILENAME}' поврежден. Невозможно найти профиль."
        )
        return None
    except Exception as e:
        logger.error(
            f"Ошибка поиска профиля UID {uid_to_find} в Excel: {e}",
            exc_info=True,
        )
        return None


def get_all_user_data_from_excel(uid_to_find: str) -> Dict[str, Any]:
    """
    Fetches all data for a given UID from Excel for display.
    Returns a dictionary of data or a dictionary with an 'error' or 'info' key.
    This is a synchronous function.
    """
    data_for_display: Dict[str, Any] = {}
    try:
        if not os.path.exists(EXCEL_FILENAME):
            return {"error": f"Файл данных '{EXCEL_FILENAME}' не найден."}

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active

        if ws.max_row < 1:
            return {
                "info": f"Файл Excel '{EXCEL_FILENAME}' пуст (нет заголовков)."
            }

        header_row_values = [cell.value for cell in ws[1]]

        if "Unique ID" not in header_row_values:
            return {
                "error": "Столбец 'Unique ID' не найден в заголовках файла Excel."
            }

        uid_col_idx_excel = header_row_values.index("Unique ID")
        found_in_excel = False

        for row_tuple in ws.iter_rows(min_row=2, values_only=True):
            excel_uid_val = (
                row_tuple[uid_col_idx_excel]
                if uid_col_idx_excel < len(row_tuple)
                else None
            )
            if excel_uid_val is not None and str(excel_uid_val) == uid_to_find:
                found_in_excel = True
                for i, header_name in enumerate(header_row_values):
                    if header_name is None:
                        continue  # Skip empty header cells
                    val_from_excel = (
                        row_tuple[i] if i < len(row_tuple) else None
                    )
                    display_val = (
                        val_from_excel
                        if val_from_excel is not None
                        else "нет данных"
                    )
                    if (
                        isinstance(val_from_excel, float)
                        and val_from_excel.is_integer()
                    ):
                        display_val = int(val_from_excel)
                    data_for_display[str(header_name)] = str(display_val)
                break  # Found and processed the user's row

        if not found_in_excel:
            data_for_display["info"] = (
                "Профиль с указанным UID не найден в файле Excel."
                if ws.max_row > 1
                else "Файл Excel пуст (содержит только заголовки или пуст)."
            )

    except InvalidFileException:
        data_for_display["error"] = (
            f"Файл '{EXCEL_FILENAME}' поврежден. Невозможно загрузить данные."
        )
    except ValueError as ve:  # e.g., header "Unique ID" not found by .index()
        data_for_display["error"] = (
            f"Ошибка конфигурации заголовков в файле Excel: {ve}"
        )
    except Exception as e:
        data_for_display["error"] = (
            f"Ошибка при загрузке данных из Excel для UID {uid_to_find}: {e}"
        )

    return data_for_display


# --- Existing Generic Result Check Functions (async) ---
async def check_if_results_exist_generic(
    profile_unique_id: Union[str, int], result_header_to_check: str
) -> bool:
    if not profile_unique_id:
        logger.warning("Excel check: profile_unique_id не предоставлен.")
        return False
    try:
        uid_to_check_str = str(int(profile_unique_id))
    except ValueError:
        logger.error(
            f"Excel check: Неверный формат profile_unique_id: {profile_unique_id}."
        )
        return False

    try:
        # This function is async, but openpyxl is sync.
        # For frequent calls, consider making it sync and calling via to_thread
        # or finding an async Excel library if performance becomes an issue.
        # For now, keeping it as is, assuming calls are not extremely frequent.
        if not os.path.exists(EXCEL_FILENAME):  # Check before loading
            logger.info(
                f"Excel файл {EXCEL_FILENAME} не найден для проверки результатов (generic)."
            )
            return False

        wb = load_workbook(EXCEL_FILENAME)  # This is a sync I/O operation
        ws = wb.active

        if ws.max_row < 2:
            return False

        header_row_values = [cell.value for cell in ws[1]]

        try:
            uid_col_idx = header_row_values.index("Unique ID")
        except ValueError:
            logger.error(
                "Excel check (generic): Заголовок 'Unique ID' не найден."
            )
            return False

        try:
            result_col_idx = header_row_values.index(result_header_to_check)
        except ValueError:
            logger.warning(
                f"Excel check (generic): Заголовок результата '{result_header_to_check}' не найден."
            )
            return False

        for row_idx in range(2, ws.max_row + 1):
            uid_cell = ws.cell(row=row_idx, column=uid_col_idx + 1)
            if (
                uid_cell.value is not None
                and str(uid_cell.value) == uid_to_check_str
            ):
                result_cell = ws.cell(row=row_idx, column=result_col_idx + 1)
                if result_cell.value is not None:
                    return True
        return False
    except (
        FileNotFoundError
    ):  # Should be caught by os.path.exists, but defensive
        logger.info(
            f"Excel файл {EXCEL_FILENAME} не найден для проверки результатов (generic, fallback)."
        )
        return False
    except InvalidFileException:
        logger.error(
            f"Excel файл {EXCEL_FILENAME} поврежден. Невозможно проверить результаты (generic)."
        )
        return False
    except Exception as e:
        logger.error(
            f"Ошибка при проверке результатов (generic) для UID {profile_unique_id}, заголовок '{result_header_to_check}': {e}",
            exc_info=True,
        )
        return False


async def check_if_corsi_results_exist(
    profile_unique_id: Union[str, int],
) -> bool:
    return await check_if_results_exist_generic(
        profile_unique_id, "Corsi - Max Correct Sequence Length"
    )


async def check_if_stroop_results_exist(
    profile_unique_id: Union[str, int],
) -> bool:
    return await check_if_results_exist_generic(
        profile_unique_id, "Stroop Part1 Time (s)"
    )


async def check_if_reaction_time_results_exist(
    profile_unique_id: Union[str, int],
) -> bool:
    return await check_if_results_exist_generic(
        profile_unique_id, "ReactionTime_Status"
    )


async def check_if_verbal_fluency_results_exist(
    profile_unique_id: Union[str, int],
) -> bool:
    return await check_if_results_exist_generic(
        profile_unique_id, "VerbalFluency_WordCount"
    )


async def check_if_mental_rotation_results_exist(
    profile_unique_id: Union[str, int],
) -> bool:
    return await check_if_results_exist_generic(
        profile_unique_id, "MentalRotation_CorrectAnswers"
    )


async def check_if_raven_matrices_results_exist(
    profile_unique_id: Union[str, int],
) -> bool:
    return await check_if_results_exist_generic(
        profile_unique_id, "RavenMatrices_CorrectAnswers"
    )
